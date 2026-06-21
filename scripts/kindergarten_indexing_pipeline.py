"""
Automated Indexing Pipeline for Kindergarten Dataset
====================================================
Scans all Excel files in the dataset folder, identifies/validates index columns,
generates missing indices, and produces clean indexed output files.

Outputs:
    - final_kindergartens_indexed.xlsx   — indexed Excel
    - final_kindergartens_indexed.csv    — indexed CSV
    - final_kindergartens_removed.csv    — removed entries
    - indexing_report.txt                — full report

Usage:
    python scripts/kindergarten_indexing_pipeline.py
"""

from __future__ import annotations

import csv
import hashlib
import os
import re
import sys
import time
import unicodedata
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


# ── Constants ──────────────────────────────────────────────────────────────────

GOV_CODES: Dict[str, str] = {
    "عمان": "AM",
    "اربد": "IR",
    "الزرقاء": "ZA",
    "المفرق": "MA",
    "جرش": "JA",
    "عجلون": "AJ",
    "البلقاء": "BA",
    "مادبا": "MD",
    "الكرك": "KA",
    "الطفيلة": "TA",
    "معان": "MN",
    "العقبة": "AQ",
}

# Partial-match aliases for governorate names
GOV_ALIASES: Dict[str, str] = {
    "محافظة عمان": "عمان",
    "محافظة عمّان": "عمان",
    "عمّان": "عمان",
    "محافظة اربد": "اربد",
    "محافظة إربد": "اربد",
    "إربد": "اربد",
    "محافظة الزرقاء": "الزرقاء",
    "محافظة المفرق": "المفرق",
    "محافظة جرش": "جرش",
    "محافظة عجلون": "عجلون",
    "محافظة البلقاء": "البلقاء",
    "محافظة مادبا": "مادبا",
    "محافظة الكرك": "الكرك",
    "محافظة الطفيلة": "الطفيلة",
    "محافظة معان": "معان",
    "محافظة العقبة": "العقبة",
    "محافظة العقبه": "العقبة",
    "العقبه": "العقبة",
}

INDEX_KEYWORDS: Set[str] = {
    "commercial_no", "registration_id", "license_no", "license_number",
    "commercial_number", "commercial registration no", "license", "register_id",
    "reg_id", "id", "index", "رقم تجاري", "رقم تسجيل", "رقم ترخيص", "رقم السجل",
    "رقم", "رقم الهوية", "رقم الرخصة",
}

DATS_CANDIDATE_PATHS: List[str] = [
    r"C:\Users\waelj\OneDrive - zuj.edu.jo\Desktop\Project-Kinjo-seed\DATS",
    os.getcwd(),
    r"D:\Final Version\DATS",
    str(Path(__file__).resolve().parent.parent),
]

OUTPUT_DIR = Path(__file__).resolve().parent

MISSING_VALUE_PATTERNS: Set[str] = {"", "nan", "n/a", "none", "—", "-", "null", "na"}

ARABIC_DIGIT_MAP: Dict[str, str] = {
    "٠": "0", "١": "1", "٢": "2", "٣": "3", "٤": "4",
    "٥": "5", "٦": "6", "٧": "7", "٨": "8", "٩": "9",
}

MIN_INDEX_LENGTH = 3
MAX_INDEX_LENGTH = 50


# ── Logging Helpers ────────────────────────────────────────────────────────────

def _ts() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def log(msg: str) -> None:
    print(f"[{_ts()}] {msg}")


def log_error(msg: str) -> None:
    print(f"[{_ts()}] ERROR: {msg}")


def log_warning(msg: str) -> None:
    print(f"[{_ts()}] WARNING: {msg}")


# ── Text Normalization ─────────────────────────────────────────────────────────

def normalize_arabic_digits(text: str) -> str:
    """Convert Arabic-Indic digits (٠-٩) to ASCII digits (0-9)."""
    return "".join(ARABIC_DIGIT_MAP.get(c, c) for c in text)


def strip_excel_artifacts(text: str) -> str:
    """Remove trailing .0 from Excel numeric artifacts."""
    text = text.strip()
    if text.endswith(".0") and text.count(".") == 1:
        integer_part = text[:-2]
        if integer_part.isdigit():
            return integer_part
    return text


def normalize_text(text: Optional[str]) -> str:
    if text is None:
        return ""
    text = str(text).strip()
    text = unicodedata.normalize("NFKC", text)
    return " ".join(text.split())


def normalize_index(text: Optional[str]) -> str:
    if text is None:
        return ""
    result = strip_excel_artifacts(normalize_text(str(text)))
    result = normalize_arabic_digits(result)
    return result


def is_missing(value: Optional[str]) -> bool:
    if value is None:
        return True
    cleaned = normalize_text(value)
    return cleaned.lower() in MISSING_VALUE_PATTERNS


# ── Governorate Helpers ────────────────────────────────────────────────────────

def resolve_governorate(raw: Optional[str]) -> Optional[str]:
    if raw is None:
        return None
    cleaned = normalize_text(str(raw))
    if not cleaned:
        return None
    # Direct match
    if cleaned in GOV_CODES:
        return cleaned
    # Alias match
    for alias, canonical in GOV_ALIASES.items():
        if cleaned == alias or cleaned.endswith(alias):
            return canonical
    # Partial substring match
    for gov_name in GOV_CODES:
        if gov_name in cleaned or cleaned in gov_name:
            return gov_name
    return cleaned


def gov_code(gov_name: Optional[str]) -> str:
    if gov_name is None:
        return "XX"
    return GOV_CODES.get(gov_name, "XX")


def name_hash(name: Optional[str], gov: Optional[str]) -> str:
    raw = f"{name or ''}|{gov or ''}"
    return hashlib.md5(raw.encode("utf-8")).hexdigest()[:8].upper()


# ── Index Validation ───────────────────────────────────────────────────────────

def validate_index_format(index: str) -> Tuple[bool, List[str]]:
    issues: List[str] = []
    if not index:
        return False, ["empty"]
    length = len(index)
    if length < MIN_INDEX_LENGTH:
        issues.append(f"too_short ({length} chars)")
    if length > MAX_INDEX_LENGTH:
        issues.append(f"too_long ({length} chars)")
    forbidden = re.compile(r"[^\w\-\.\s/:]")
    if forbidden.search(index):
        issues.append("non_standard_characters")
    return len(issues) == 0, issues


# ── Column Detection ───────────────────────────────────────────────────────────

def detect_index_column(headers: List[str]) -> Optional[str]:
    found_col = None
    for h in headers:
        h_lower = str(h).strip().lower() if h else ""
        for kw in INDEX_KEYWORDS:
            if kw.lower() in h_lower:
                found_col = h
                break
        if found_col:
            break
    # Fallback: any column containing "رقم"
    if found_col is None:
        for h in headers:
            h_str = str(h).strip() if h else ""
            if "رقم" in h_str and ("تجاري" in h_str or "تسجيل" in h_str or
                                   "ترخيص" in h_str or "سجل" in h_str or
                                   "هوية" in h_str or "رخصة" in h_str):
                found_col = h
                break
    # Broad fallback: any column containing "رقم"
    if found_col is None:
        for h in headers:
            h_str = str(h).strip() if h else ""
            if "رقم" in h_str:
                found_col = h
                break
    return found_col


def detect_name_column(headers: List[str]) -> Optional[str]:
    """Find name columns (Arabic name preferred, English name fallback)."""
    for h in headers:
        h_str = str(h).strip() if h else ""
        hl = h_str.lower()
        if ("name_ar" == hl or "اسم" in h_str or "الروضة" in h_str or
                "الحضانة" in h_str or "kindergarten" in hl or "name" in hl):
            return h
    return None


def detect_governorate_column(headers: List[str]) -> Optional[str]:
    for h in headers:
        h_str = str(h).strip() if h else ""
        hl = h_str.lower()
        if ("governorate" in hl or "gover" in hl or "province" in hl or
                "محافظة" in h_str or "المحافظة" in h_str):
            return h
    return None


def detect_columns(headers: List[str]) -> Dict[str, Optional[str]]:
    return {
        "index": detect_index_column(headers),
        "name": detect_name_column(headers),
        "governorate": detect_governorate_column(headers),
    }


# ── DATS Folder Discovery ──────────────────────────────────────────────────────

def find_dats_folder() -> Optional[Path]:
    for path_str in DATS_CANDIDATE_PATHS:
        p = Path(path_str)
        if p.is_dir():
            xlsx_files = list(p.glob("*.xlsx"))
            if xlsx_files:
                log(f"DATS folder found: {p} ({len(xlsx_files)} Excel files)")
                return p
    log_warning("No DATS folder found with Excel files anywhere")
    return None


# ── Excel Reading ──────────────────────────────────────────────────────────────

def read_all_sheets(filepath: Path) -> List[Dict[str, Any]]:
    results: List[Dict[str, Any]] = []
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        log_error(f"Cannot open {filepath.name}: {e}")
        return results

    for sheet_name in wb.sheetnames:
        ws: Worksheet = wb[sheet_name]
        try:
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            log_warning(f"  Sheet '{sheet_name}' in {filepath.name}: error reading rows — {e}")
            continue

        if not rows:
            log(f"  Sheet '{sheet_name}' in {filepath.name}: empty — skipped")
            continue

        results.append({
            "file": filepath,
            "sheet": sheet_name,
            "rows": rows,
            "row_count": len(rows),
        })
    wb.close()
    return results


# ── Main Pipeline ──────────────────────────────────────────────────────────────

def process_dats_entry(
    entry: Dict[str, Any],
    seq_counters: Dict[str, int],
    seen_indices: Set[str],
    removed_entries: List[Dict[str, Any]],
    warn_log: List[str],
) -> List[List[Any]]:
    """
    Process a single sheet of data. Returns the cleaned rows ready for output.

    Args:
        entry: Sheet metadata + rows from read_all_sheets
        seq_counters: per-governorate sequence number tracker
        seen_indices: globally seen unique index values (for collision detection)
        removed_entries: accumulator for entries removed
        warn_log: accumulator for warnings

    Returns:
        List of rows (each row is a list of cell values) that passed indexing.
    """
    rows = entry["rows"]
    filepath: Path = entry["file"]
    sheet_name: str = entry["sheet"]

    # Detect header row
    header = [str(v).strip() if v is not None else "" for v in rows[0]]

    # Empty header check
    if not any(h for h in header):
        warn_log.append(f"[{filepath.name}/{sheet_name}] Empty header row — skipping")
        return []

    col_map = detect_columns(header)
    index_col_name = col_map["index"]
    name_col_name = col_map["name"]
    gov_col_name = col_map["governorate"]

    # Determine column positions
    def col_pos(name: Optional[str]) -> Optional[int]:
        if name is None or name not in header:
            return None
        return header.index(name)

    idx_pos = col_pos(index_col_name)
    name_pos = col_pos(name_col_name)
    gov_pos = col_pos(gov_col_name)

    log(f"  [{filepath.name}/{sheet_name}] idx={index_col_name!r} (pos={idx_pos}), "
        f"name={name_col_name!r} (pos={name_pos}), "
        f"gov={gov_col_name!r} (pos={gov_pos}) — {len(rows) - 1} data rows")

    output_rows: List[List[Any]] = [header]
    data_rows = rows[1:]

    for row_idx, row in enumerate(data_rows, start=2):
        # Skip completely null rows
        if all(v is None for v in row):
            continue

        # Extract values
        raw_index = str(row[idx_pos]).strip() if idx_pos is not None and idx_pos < len(row) and row[idx_pos] is not None else ""
        raw_name = str(row[name_pos]).strip() if name_pos is not None and name_pos < len(row) and row[name_pos] is not None else ""
        raw_gov = str(row[gov_pos]).strip() if gov_pos is not None and gov_pos < len(row) and row[gov_pos] is not None else ""

        # Normalize
        index_val = normalize_index(raw_index) if raw_index else ""
        name_val = normalize_text(raw_name) if raw_name else ""
        gov_val = resolve_governorate(normalize_text(raw_gov) if raw_gov else None) or ""

        # Case: no name at all → skip (can't use)
        if not name_val:
            removed_entries.append({
                "file": filepath.name,
                "sheet": sheet_name,
                "row": row_idx,
                "name": raw_name,
                "governorate": raw_gov,
                "reason": "No name found — cannot generate index",
            })
            continue

        # Case: has valid existing index
        if index_val and not is_missing(index_val):
            valid, issues = validate_index_format(index_val)
            if valid:
                # Check for duplicate index
                if index_val in seen_indices:
                    # Collision: generate hash-based fallback
                    h = name_hash(name_val, gov_val)
                    prefix = f"KG-HASH-{h}"
                    n = 1
                    final_idx = prefix
                    while final_idx in seen_indices:
                        final_idx = f"{prefix}-{n}"
                        n += 1
                    warn_log.append(
                        f"[{filepath.name}/{sheet_name} row {row_idx}] "
                        f"DUPLICATE index '{index_val}' → hash fallback '{final_idx}'"
                    )
                    index_val = final_idx
                else:
                    warn_log.append(
                        f"[{filepath.name}/{sheet_name} row {row_idx}] "
                        f"VALID index '{index_val}'"
                    )
                seen_indices.add(index_val)
            else:
                warn_log.append(
                    f"[{filepath.name}/{sheet_name} row {row_idx}] "
                    f"SUSPICIOUS index '{index_val}' — issues: {', '.join(issues)}. "
                    f"Still using as-is."
                )
                seen_indices.add(index_val)

        else:
            # No valid index — generate one
            generated = None
            if name_val and gov_val and gov_val in GOV_CODES:
                gov_short = gov_code(gov_val)
                seq_counters[gov_short] = seq_counters.get(gov_short, 0) + 1
                generated = f"KG-{gov_short}-{seq_counters[gov_short]:04d}"
                source = f"KG-{gov_short}-sequence"
            elif name_val:
                h = name_hash(name_val, gov_val)
                generated = f"KG-HASH-{h}"
                source = "hash"
            else:
                # Nothing usable — mark for removal
                removed_entries.append({
                    "file": filepath.name,
                    "sheet": sheet_name,
                    "row": row_idx,
                    "name": raw_name,
                    "governorate": raw_gov,
                    "reason": "No name or governorate to generate index",
                })
                continue

            # Ensure uniqueness of generated index
            n = 0
            base_gen = generated
            while generated in seen_indices:
                n += 1
                generated = f"{base_gen}-{n}"
            seen_indices.add(generated)

            warn_log.append(
                f"[{filepath.name}/{sheet_name} row {row_idx}] "
                f"GENERATED index '{generated}' (source={source})"
            )

            index_val = generated

            # Update row with generated index (only if idx_pos exists)
            if idx_pos is not None:
                row_list = list(row)
                if idx_pos < len(row_list):
                    row_list[idx_pos] = index_val
                else:
                    row_list = list(row) + [index_val]
                row = tuple(row_list)

        # Store final value
        row_list = list(row)
        if idx_pos is not None and idx_pos < len(row_list):
            row_list[idx_pos] = index_val
        output_rows.append(row_list)

    return output_rows


def run_pipeline() -> None:
    log("=" * 70)
    log("KINDERGARTEN INDEXING PIPELINE — START")
    log("=" * 70)

    dats_folder = find_dats_folder()
    if dats_folder is None:
        log_error("No DATS folder found. Exiting.")
        sys.exit(1)

    xlsx_files = sorted(dats_folder.glob("*.xlsx"))
    log(f"Found {len(xlsx_files)} Excel files to process")

    # Read all data
    all_sheet_data: List[Dict[str, Any]] = []
    for fp in xlsx_files:
        sheet_data = read_all_sheets(fp)
        if sheet_data:
            all_sheet_data.extend(sheet_data)

    log(f"Total sheets to process: {len(all_sheet_data)}")
    if not all_sheet_data:
        log_error("No sheets with data found. Exiting.")
        sys.exit(1)

    # Process
    all_output_rows: List[List[Any]] = []
    removed_entries: List[Dict[str, Any]] = []
    warn_log: List[str] = []
    seq_counters: Dict[str, int] = {}
    seen_indices: Set[str] = set()
    total_data_rows = 0
    total_removed = 0

    for entry in all_sheet_data:
        filepath: Path = entry["file"]
        sheet_name: str = entry["sheet"]
        rows = entry["rows"]

        log(f"\n  Processing {filepath.name} / '{sheet_name}' ({len(rows) - 1} data rows)")

        # Detect merged cells: cannot process those
        try:
            wb_temp = openpyxl.load_workbook(filepath, read_only=True)
            ws_temp = wb_temp[sheet_name]
            merged = list(ws_temp.merged_cells.ranges)
            wb_temp.close()
            if merged:
                warn_log.append(
                    f"[{filepath.name}/{sheet_name}] "
                    f"WARNING: Contains {len(merged)} merged cell ranges — skipping sheet"
                )
                log_warning(f"{filepath.name}/{sheet_name}: {len(merged)} merged ranges — skipping")
                continue
        except Exception:
            warn_log.append(f"[{filepath.name}/{sheet_name}] Could not check merged cells — processing anyway")
            log_warning(f"Could not check merged cells for {filepath.name}/{sheet_name}")

        output = process_dats_entry(
            entry, seq_counters, seen_indices,
            removed_entries, warn_log
        )

        if output:
            sheet_header = output[0]
            if not all_output_rows:
                all_output_rows.append(sheet_header)
            data_part = output[1:]
            all_output_rows.extend(data_part)
            total_data_rows += len(data_part)
            total_removed_before = total_removed
            total_removed += sum(
                1 for r in removed_entries
                if r["file"] == filepath.name and r["sheet"] == sheet_name
            ) - total_removed_before
            log(f"  → {len(data_part)} rows indexed, cumulative: {total_data_rows}")

    # Post-processing: verify no STILL missing indices remain
    still_missing: List[List[Any]] = []
    final_rows: List[List[Any]] = []
    if all_output_rows:
        final_rows.append(all_output_rows[0])  # header
        for row in all_output_rows[1:]:
            # Find the index in the row
            idx_col_name = detect_index_column(final_rows[0])
            if idx_col_name and idx_col_name in final_rows[0]:
                pos = final_rows[0].index(idx_col_name)
                if pos < len(row):
                    val = normalize_index(str(row[pos])) if row[pos] is not None else ""
                    if is_missing(val):
                        still_missing.append(row)
                        removed_entries.append({
                            "file": "post_processing",
                            "sheet": "—",
                            "row": "—",
                            "name": str(row[0]) if row else "",
                            "governorate": "",
                            "reason": "Still missing index after generation — removed",
                        })
                        continue
            final_rows.append(row)

    if still_missing:
        log_warning(f"{len(still_missing)} entries still missing index post-generation — removed")

    # ── Write Outputs ──────────────────────────────────────────────────────
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    out_xlsx = OUTPUT_DIR / f"final_kindergartens_indexed.xlsx"
    out_csv = OUTPUT_DIR / f"final_kindergartens_indexed.csv"
    out_removed_csv = OUTPUT_DIR / f"final_kindergartens_removed.csv"
    out_report = OUTPUT_DIR / f"indexing_report.txt"

    # Write Excel
    if final_rows:
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = "Indexed Kindergartens"
        for i, row in enumerate(final_rows, start=1):
            for j, val in enumerate(row, start=1):
                cell = ws_out.cell(row=i, column=j)
                cell.value = val
        wb_out.save(str(out_xlsx))
        log(f"Indexed Excel: {out_xlsx} ({len(final_rows) - 1} entries)")
    else:
        log_warning("No rows to write to Excel")

    # Write CSV
    if final_rows:
        with open(out_csv, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerows(final_rows)
        log(f"Indexed CSV: {out_csv} ({len(final_rows) - 1} entries)")
    else:
        log_warning("No rows to write to CSV")

    # Write removed CSV
    removed_fieldnames = ["file", "sheet", "row", "name", "governorate", "reason"]
    with open(out_removed_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=removed_fieldnames)
        writer.writeheader()
        if removed_entries:
            writer.writerows(removed_entries)
    log(f"Removed entries CSV: {out_removed_csv} ({len(removed_entries)} entries)")

    # Write report
    header_count = 1 if final_rows else 0
    indexed_count = len(final_rows) - header_count if len(final_rows) > 1 else 0

    report_lines = [
        "=" * 70,
        "KINDERGARTEN INDEXING PIPELINE — FINAL REPORT",
        "=" * 70,
        f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}",
        f"DATS folder: {dats_folder}",
        f"Excel files scanned: {len(xlsx_files)}",
        f"Sheets processed: {len(all_sheet_data)}",
        "",
        f"Total data rows (before): {sum(len(e['rows']) - 1 for e in all_sheet_data)}",
        f"Indexed entries: {indexed_count}",
        f"Removed entries: {len(removed_entries)}",
        f"Still missing post-generation: {len(still_missing)}",
        "",
        "=== SEQUENCE COUNTERS PER GOVERNORATE ===",
    ]
    for code, count in sorted(seq_counters.items()):
        report_lines.append(f"  {code}: {count}")
    report_lines.append("")

    if warn_log:
        report_lines.append("=== WARNINGS / LOG ===")
        for line in warn_log:
            report_lines.append(f"  {line}")
        report_lines.append("")

    if removed_entries:
        report_lines.append("=== REMOVED ENTRIES DETAIL ===")
        for i, entry in enumerate(removed_entries, start=1):
            report_lines.append(
                f"  {i}. [{entry['file']}/{entry['sheet']} row {entry['row']}] "
                f"'{entry['name']}' ({entry['governorate']}) — {entry['reason']}"
            )
        report_lines.append("")

    report_lines.extend([
        "=" * 70,
        "OUTPUT FILES",
        "=" * 70,
        f"  {out_xlsx.name}     — Indexed Excel ({indexed_count} entries)",
        f"  {out_csv.name}      — Indexed CSV ({indexed_count} entries)",
        f"  {out_removed_csv.name}  — Removed entries ({len(removed_entries)})",
        f"  {out_report.name} — This report",
        "",
        "=== END ===",
    ])

    out_report.write_text("\n".join(report_lines), encoding="utf-8")
    log(f"Report: {out_report}")

    # ── Print Summary ──────────────────────────────────────────────────────
    print()
    print("=" * 70)
    print("  SUMMARY")
    print("=" * 70)
    print(f"  DATS folder:          {dats_folder}")
    print(f"  Excel files scanned:  {len(xlsx_files)}")
    print(f"  Sheets processed:     {len(all_sheet_data)}")
    print(f"  Total rows scanned:   {sum(len(e['rows']) - 1 for e in all_sheet_data)}")
    print(f"  Rows indexed:         {indexed_count}")
    print(f"  Rows removed:         {len(removed_entries)}")
    if seq_counters:
        print(f"\n  Generated indices by governorate:")
        for code in sorted(seq_counters):
            print(f"    {code}: {seq_counters[code]}")
    if still_missing:
        print(f"\n  WARNING: {len(still_missing)} entries still missing index after generation")
    print(f"\n  Output files:")
    print(f"    ✓ {out_xlsx.name}")
    print(f"    ✓ {out_csv.name}")
    print(f"    ✓ {out_removed_csv.name}")
    print(f"    ✓ {out_report.name}")
    print("=" * 70)


def main() -> None:
    try:
        run_pipeline()
    except KeyboardInterrupt:
        log("\nPipeline interrupted by user.")
        sys.exit(1)
    except Exception as e:
        log_error(f"Pipeline failed: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()