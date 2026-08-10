"""
Import kindergartens from the merged_all_uploads.xlsx Excel file into the database.

Usage:
    python scripts/import_kindergartens_excel.py [--dry-run] [--file PATH] [--sheet NAME]

Options:
    --dry-run   Preview what would be imported without writing to the database
    --file      Path to the Excel file (default: merged_all_uploads.xlsx on Desktop)
    --sheet     Optional sheet name; defaults to "Merged" when present, else first sheet
"""
import sys
import os
import argparse
import logging

# Add project root to path so we can import project modules
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from sqlalchemy.orm import Session
from database import SessionLocal, engine
from models import Base, Kindergarten, KindergartenStatus

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

DEFAULT_FILE = r"C:\Users\waelj\OneDrive - zuj.edu.jo\Desktop\Dataset\merged_all_uploads.xlsx"

# Column index mapping (0-based) from the "Merged" sheet
COL_NAME_AR = 0       # اسم الروضة (عربي)
COL_NAME_EN = 1       # اسم الروضة (إنجليزي)
COL_GOVERNORATE = 2   # المحافظة
COL_CITY = 3          # المدينة
COL_AREA = 4          # المنطقة
COL_ADDRESS = 5       # العنوان التفصيلي
COL_PHONE = 6         # رقم الهاتف
COL_LATITUDE = 8
COL_LONGITUDE = 9


def clean(value) -> str:
    """Strip and return a string, or empty string if None."""
    if value is None:
        return ""
    return str(value).strip()


def coordinate(value):
    """Return a valid coordinate, or None for an empty/invalid source cell."""
    try:
        parsed = float(str(value).strip())
    except (TypeError, ValueError):
        return None
    return parsed if parsed != 0 else None


def _resolve_sheet_name(wb: openpyxl.Workbook, preferred_sheet: str = None) -> str:
    """Resolve target sheet name with sane fallback rules."""
    if preferred_sheet:
        if preferred_sheet not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{preferred_sheet}' not found. Available sheets: {wb.sheetnames}"
            )
        return preferred_sheet

    if "Merged" in wb.sheetnames:
        return "Merged"

    # Fallback: use the first sheet to support one-off files like kinderr.xlsx
    return wb.sheetnames[0]


def import_kindergartens(file_path: str, dry_run: bool = False, sheet_name: str = None):
    """Read the Excel file and insert kindergarten rows into the database."""
    logger.info("Opening %s ...", file_path)
    wb = openpyxl.load_workbook(file_path, read_only=True)
    try:
        resolved_sheet = _resolve_sheet_name(wb, preferred_sheet=sheet_name)
    except ValueError as exc:
        logger.error("%s", exc)
        wb.close()
        return

    ws = wb[resolved_sheet]
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    wb.close()

    logger.info("Using sheet '%s'. Found %d data rows.", resolved_sheet, len(rows))

    # Ensure tables exist
    if not dry_run:
        Base.metadata.create_all(bind=engine)

    db: Session = SessionLocal()
    try:
        # Get existing kindergartens to avoid duplicates (match on name_ar + governorate + city)
        existing = set()
        for kg in db.query(Kindergarten.name_ar, Kindergarten.governorate, Kindergarten.district).all():
            existing.add((kg.name_ar, kg.governorate, kg.district))
        logger.info("Existing kindergartens in DB: %d", len(existing))

        inserted = 0
        skipped_dup = 0
        skipped_empty = 0
        errors = 0

        for row_num, row in enumerate(rows, start=2):
            name_ar = clean(row[COL_NAME_AR])
            name_en = clean(row[COL_NAME_EN])
            governorate = clean(row[COL_GOVERNORATE])
            city = clean(row[COL_CITY])
            area = clean(row[COL_AREA])
            address_line = clean(row[COL_ADDRESS])
            phone = clean(row[COL_PHONE])

            # Skip rows with no name
            if not name_ar:
                skipped_empty += 1
                continue

            # Fill required fields with fallback if empty
            if not governorate:
                governorate = "غير محدد"
            if not city:
                city = "غير محدد"
            if not area:
                area = "غير محدد"
            if not address_line:
                address_line = "غير محدد"
            if not phone:
                phone = "غير متوفر"

            # Check duplicate
            key = (name_ar, governorate, city)
            if key in existing:
                skipped_dup += 1
                continue

            if dry_run:
                logger.info(
                    "[DRY-RUN] Row %d: %s | %s | %s / %s / %s",
                    row_num, name_ar, name_en, governorate, city, area,
                )
                inserted += 1
                existing.add(key)
                continue

            try:
                latitude = coordinate(row[COL_LATITUDE]) if len(row) > COL_LATITUDE else None
                longitude = coordinate(row[COL_LONGITUDE]) if len(row) > COL_LONGITUDE else None
                kg = Kindergarten(
                    name_ar=name_ar,
                    name_en=name_en or None,
                    governorate=governorate,
                    district=city,
                    area=area,
                    address_line=address_line,
                    contact_phone=phone,
                    latitude=latitude,
                    longitude=longitude,
                    status=KindergartenStatus.DRAFT,
                )
                db.add(kg)
                existing.add(key)
                inserted += 1

                # Flush in batches
                if inserted % 100 == 0:
                    db.flush()
                    logger.info("  ... flushed %d rows", inserted)

            except (ValueError, TypeError, AttributeError, RuntimeError) as exc:
                logger.error("Row %d error: %s", row_num, exc)
                errors += 1

        if not dry_run:
            db.commit()
            logger.info("Committed to database.")

        logger.info("=" * 50)
        logger.info("Import summary:")
        logger.info("  Inserted:        %d", inserted)
        logger.info("  Skipped (dup):    %d", skipped_dup)
        logger.info("  Skipped (empty):  %d", skipped_empty)
        logger.info("  Errors:           %d", errors)
        logger.info("=" * 50)

    except (RuntimeError, ValueError, TypeError, AttributeError, OSError) as exc:
        db.rollback()
        logger.exception("Import failed: %s", exc)
    finally:
        db.close()


def main():
    parser = argparse.ArgumentParser(description="Import kindergartens from Excel")
    parser.add_argument("--dry-run", action="store_true", help="Preview without writing to DB")
    parser.add_argument("--file", default=DEFAULT_FILE, help="Path to the Excel file")
    parser.add_argument("--sheet", default=None, help="Optional sheet name")
    args = parser.parse_args()

    if not os.path.exists(args.file):
        logger.error("File not found: %s", args.file)
        sys.exit(1)

    import_kindergartens(args.file, dry_run=args.dry_run, sheet_name=args.sheet)


if __name__ == "__main__":
    main()
