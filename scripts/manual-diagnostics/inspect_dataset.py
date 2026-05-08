import openpyxl, os, sys
sys.stdout.reconfigure(encoding="utf-8")

DATASET = r"C:\Users\waelj\OneDrive - zuj.edu.jo\Desktop\Dataset"
for fname in sorted(os.listdir(DATASET)):
    if not fname.endswith(".xlsx"):
        continue
    try:
        wb = openpyxl.load_workbook(os.path.join(DATASET, fname), read_only=True, data_only=True)
        for sh in wb.sheetnames:
            ws = wb[sh]
            rows = list(ws.iter_rows(values_only=True, max_row=4))
            total = ws.max_row
            print(f"\n=== {fname} :: {sh} (rows~{total}) ===")
            for r in rows:
                print([str(v)[:30] if v else None for v in r])
        wb.close()
    except Exception as e:
        print(f"SKIP {fname}: {e}")
