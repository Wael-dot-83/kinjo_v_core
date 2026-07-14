"""Import kindergartens from a cleaned Excel workbook into the database.

Unlike the legacy ``seed_kindergartens_from_excel.py`` (which hard-codes a
private dataset path and a set of filenames), this importer takes the workbook
path as an argument and maps the already-clean ``kindergartens_schema`` columns
directly onto the ``Kindergarten`` model.

Expected columns (header row), any subset is fine as long as name_ar +
governorate are present::

    name_ar, name_en, governorate, city, area, address_line, contact_phone,
    contact_email, operating_hours_start, operating_hours_end, license_number

Usage::

    python scripts/import_kindergartens_from_workbook.py --file path/to/book.xlsx
    python scripts/import_kindergartens_from_workbook.py --file book.xlsx --sheet kindergartens_schema --dry-run

The DB is taken from ``DATABASE_URL`` (same as the app). Rows missing a name or
governorate are skipped; existing kindergartens (matched by ``name_ar``) are
left untouched, so the import is idempotent.
"""
from __future__ import annotations

import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import models  # noqa: E402

# Excel column -> Kindergarten model attribute. Note: the workbook's "city"
# column maps onto the model's "district" field.
COLUMN_MAP = {
    "name_ar": "name_ar",
    "name_en": "name_en",
    "governorate": "governorate",
    "city": "district",
    "district": "district",
    "area": "area",
    "address_line": "address_line",
    "contact_phone": "contact_phone",
    "contact_email": "contact_email",
    "operating_hours_start": "operating_hours_start",
    "operating_hours_end": "operating_hours_end",
    "license_number": "license_number",
}

# Fallbacks for NOT NULL / required columns when the source row omits them.
REQUIRED_DEFAULTS = {
    "area": "غير محدد",
    "address_line": "غير محدد",
    "contact_phone": "0000000000",
}


def _clean(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() in ("none", "nan", "null"):
        return None
    return s


def _normalize_governorate(gov: str) -> str:
    try:
        import validators
        return validators.validate_jordan_governorate(gov)
    except Exception:
        return gov


def map_row(row: dict) -> dict | None:
    """Map one source row (header->value dict) to Kindergarten kwargs.

    Returns ``None`` when the row lacks the minimum required fields
    (``name_ar`` and ``governorate``).
    """
    payload: dict = {}
    for col, attr in COLUMN_MAP.items():
        if col in row:
            val = _clean(row[col])
            if val is not None:
                payload[attr] = val

    if not payload.get("name_ar") or not payload.get("governorate"):
        return None

    payload["governorate"] = _normalize_governorate(payload["governorate"])
    payload.setdefault("district", payload["governorate"])
    for key, default in REQUIRED_DEFAULTS.items():
        payload.setdefault(key, default)
    payload["status"] = models.KindergartenStatus.DRAFT
    return payload


def import_rows(db, rows: list[dict]) -> dict:
    """Insert mapped rows, skipping invalid ones and existing names.

    Returns a stats dict: created / skipped_invalid / duplicate.
    """
    existing = {name for (name,) in db.query(models.Kindergarten.name_ar).all()}
    created = skipped_invalid = duplicate = 0
    for row in rows:
        payload = map_row(row)
        if payload is None:
            skipped_invalid += 1
            continue
        if payload["name_ar"] in existing:
            duplicate += 1
            continue
        db.add(models.Kindergarten(**payload))
        existing.add(payload["name_ar"])
        created += 1
    db.commit()
    return {"created": created, "skipped_invalid": skipped_invalid, "duplicate": duplicate}


def read_workbook(path: str, sheet: str | None = None) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else _pick_sheet(wb)
    it = ws.iter_rows(values_only=True)
    header = [(_clean(h) or "") for h in next(it, [])]
    rows = [dict(zip(header, r)) for r in it]
    wb.close()
    return rows


def _pick_sheet(wb):
    # Prefer a sheet whose header contains name_ar; else the first sheet.
    for name in wb.sheetnames:
        ws = wb[name]
        first = next(ws.iter_rows(values_only=True), None) or ()
        if any((_clean(h) or "") == "name_ar" for h in first):
            return ws
    return wb[wb.sheetnames[0]]


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Import kindergartens from an Excel workbook.")
    ap.add_argument("--file", required=True, help="Path to the .xlsx workbook")
    ap.add_argument("--sheet", default=None, help="Sheet name (default: auto-detect)")
    ap.add_argument("--dry-run", action="store_true", help="Parse + report only, do not write")
    args = ap.parse_args(argv)

    if not os.path.exists(args.file):
        print(f"File not found: {args.file}", file=sys.stderr)
        return 2

    rows = read_workbook(args.file, args.sheet)
    valid = sum(1 for r in rows if map_row(r) is not None)
    print(f"Read {len(rows)} rows; {valid} valid (have name_ar + governorate).")
    if args.dry_run:
        print("Dry run — nothing written.")
        return 0

    from database import SessionLocal

    db = SessionLocal()
    try:
        stats = import_rows(db, rows)
        total = db.query(models.Kindergarten).count()
    finally:
        db.close()
    print(f"created={stats['created']} skipped_invalid={stats['skipped_invalid']} "
          f"duplicate={stats['duplicate']}")
    print(f"Total kindergartens now: {total}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
