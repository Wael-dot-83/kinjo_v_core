"""Import the authoritative Jordan kindergarten dataset from the seed workbook.

Source
------
``Dataset_ kinjo _jordan.xlsx``, sheet ``Merged``. Its own ``Summary`` sheet records
that the sheet is nine overlapping exports concatenated (198 + 36 + 201 + 39 + 36 +
196 + 195 + 193 + 198 = 1292 rows), so the file is a merge, not a register. Three
defects have to be repaired before any of it can be trusted:

1. **Duplication.** 1292 rows describe 196 distinct kindergartens — one name recurs
   twelve times. Rows are collapsed on ``name + city + area``, keeping whichever
   duplicate carries the most populated fields.

2. **Phone numbers destroyed by numeric coercion.** 247 values were read by Excel as
   numbers, so ``0790225163`` became ``790225163.0`` — trailing ``.0`` appended and
   the leading zero lost. Both are reversed. This also inflates naive de-duplication,
   because the same site appears under two spellings of one phone number.

3. **A city in the governorate column.** 14 rows carry ``الرمثا``, which is a city in
   إربد, not a governorate. Governorates resolve through
   ``services.jordan_locations``; anything it rejects is looked up as a city in
   ``settings.JORDAN_CITIES`` and mapped to its parent governorate. Nothing is
   guessed — a value that satisfies neither is reported and skipped.

Existing rows
-------------
``--replace-imported`` removes kindergartens previously loaded from the database
snapshot (any id above ``--keep-ids-through``) before inserting. It refuses to run if
any of those rows owns dependent data, so seeded sites and their classes, incidents,
reports and enrolments can never be caught by it.

Usage
-----
    python scripts/import_kindergartens_from_dataset.py --source <workbook.xlsx> \
        [--replace-imported] [--apply]

Without ``--apply`` it reports what it would do and changes nothing.
"""

from __future__ import annotations

import argparse
import os
import shutil
import sqlite3
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import settings  # noqa: E402
from services.jordan_locations import is_valid_governorate, normalize_governorate  # noqa: E402

LIVE_DB = os.path.join("data", "kinjo.db")
SHEET = "Merged"

# Column order in the Merged sheet.
COL_NAME_AR, COL_NAME_EN, COL_GOV, COL_CITY, COL_AREA, COL_ADDRESS, COL_PHONE = range(7)

# Built once from the canonical city lists so "الرمثا" resolves to "إربد" without a
# second hardcoded table living in this file.
_CITY_TO_GOVERNORATE = {
    city.strip(): gov for gov, cities in settings.JORDAN_CITIES.items() for city in cities
}


def _text(row, index: int) -> str:
    value = row[index] if index < len(row) else None
    return "" if value is None else str(value).strip()


def repair_phone(raw: str) -> str:
    """Undo Excel's numeric coercion of a phone number.

    ``790225163.0`` -> ``0790225163``. A value Excel stored as a number lost its
    leading zero and gained a ``.0``; restoring both recovers the original string.
    """
    value = raw.strip()
    if not value:
        return ""
    if value.endswith(".0"):
        value = value[:-2]
    if value.isdigit() and not value.startswith("0"):
        value = "0" + value
    return value


def resolve_governorate(raw: str) -> str | None:
    """Canonical governorate name, or None when the value is neither gov nor city."""
    value = raw.strip()
    if not value:
        return None
    if is_valid_governorate(value):
        return normalize_governorate(value)
    parent = _CITY_TO_GOVERNORATE.get(value)
    if parent and is_valid_governorate(parent):
        return normalize_governorate(parent)
    return None


def load_rows(path: str) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    if SHEET not in workbook.sheetnames:
        raise SystemExit(f"sheet {SHEET!r} not found; available: {workbook.sheetnames}")
    raw_rows = [r for r in workbook[SHEET].iter_rows(min_row=2, values_only=True) if any(r)]
    workbook.close()

    # Collapse duplicates, preferring the row with the most populated optional fields.
    best: dict[tuple[str, str, str], tuple[int, tuple]] = {}
    for row in raw_rows:
        key = (_text(row, COL_NAME_AR), _text(row, COL_CITY), _text(row, COL_AREA))
        completeness = sum(
            1 for i in (COL_NAME_EN, COL_ADDRESS, COL_PHONE) if _text(row, i)
        )
        if key not in best or completeness > best[key][0]:
            best[key] = (completeness, row)

    return [
        {
            "name_ar": _text(r, COL_NAME_AR),
            "name_en": _text(r, COL_NAME_EN) or None,
            "governorate_raw": _text(r, COL_GOV),
            "district": _text(r, COL_CITY),
            "area": _text(r, COL_AREA),
            "address_line": _text(r, COL_ADDRESS),
            "contact_phone": repair_phone(_text(r, COL_PHONE)),
        }
        for _, r in best.values()
    ]


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", required=True, help="Path to the seed workbook")
    parser.add_argument("--apply", action="store_true", help="Write changes (default: dry run)")
    parser.add_argument(
        "--replace-imported",
        action="store_true",
        help="Delete previously imported kindergartens before inserting",
    )
    parser.add_argument(
        "--keep-ids-through",
        type=int,
        default=5,
        help="Ids at or below this are seeded data and are never deleted (default: 5)",
    )
    args = parser.parse_args()

    if not os.path.exists(args.source):
        print(f"source not found: {args.source}")
        return 1

    records = load_rows(args.source)

    prepared, unresolved = [], []
    for rec in records:
        gov = resolve_governorate(rec["governorate_raw"])
        if gov is None:
            unresolved.append(rec)
            continue
        rec["governorate"] = gov
        prepared.append(rec)

    print(f"distinct kindergartens in workbook : {len(records)}")
    print(f"governorate resolved               : {len(prepared)}")
    print(f"governorate unresolvable (skipped) : {len(unresolved)}")
    for rec in unresolved[:5]:
        print(f"    {rec['name_ar']!r} governorate={rec['governorate_raw']!r}")

    changed = {}
    for rec in prepared:
        if rec["governorate"] != rec["governorate_raw"]:
            pair = (rec["governorate_raw"], rec["governorate"])
            changed[pair] = changed.get(pair, 0) + 1
    print("governorate normalisation:")
    for (was, now), n in sorted(changed.items(), key=lambda kv: -kv[1]):
        print(f"    {was!r} -> {now!r}  x{n}")

    missing_phone = sum(1 for r in prepared if not r["contact_phone"])
    missing_addr = sum(1 for r in prepared if not r["address_line"])
    print(f"rows with no phone   : {missing_phone} (stored empty; column is NOT NULL)")
    print(f"rows with no address : {missing_addr} (stored empty; column is NOT NULL)")

    live = sqlite3.connect(LIVE_DB)
    before = live.execute("SELECT COUNT(*) FROM kindergartens").fetchone()[0]

    doomed = live.execute(
        "SELECT COUNT(*) FROM kindergartens WHERE id > ?", (args.keep_ids_through,)
    ).fetchone()[0]

    dependents = 0
    for table in ("classes", "incidents", "daily_reports", "enrollment_applications", "reports"):
        try:
            dependents += live.execute(
                f"SELECT COUNT(*) FROM {table} WHERE kindergarten_id > ?",
                (args.keep_ids_through,),
            ).fetchone()[0]
        except sqlite3.OperationalError:
            pass

    print()
    print(f"live kindergartens now             : {before}")
    if args.replace_imported:
        print(f"would delete (id > {args.keep_ids_through})              : {doomed}")
        print(f"dependent rows on those            : {dependents}")
        if dependents:
            print("REFUSING: those kindergartens own dependent data. Nothing changed.")
            return 1
    print(f"would insert                       : {len(prepared)}")
    print(f"resulting total                    : "
          f"{(before - doomed if args.replace_imported else before) + len(prepared)}")

    if not args.apply:
        print("\nDRY RUN — database unchanged. Re-run with --apply.")
        return 0

    backup = f"{LIVE_DB}.backup-before-dataset-import-{datetime.now():%Y%m%d-%H%M%S}"
    shutil.copy2(LIVE_DB, backup)
    print(f"\nbackup written: {backup}")

    try:
        live.execute("BEGIN")
        if args.replace_imported:
            live.execute(
                "DELETE FROM kindergartens WHERE id > ?", (args.keep_ids_through,)
            )
        live.executemany(
            "INSERT INTO kindergartens "
            "(name_ar, name_en, governorate, district, area, address_line, contact_phone, status) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, 'ACTIVE')",
            [
                (
                    r["name_ar"],
                    r["name_en"],
                    r["governorate"],
                    r["district"],
                    r["area"],
                    r["address_line"],
                    r["contact_phone"],
                )
                for r in prepared
            ],
        )
        live.commit()
    except Exception:
        live.rollback()
        print("failed — database rolled back, backup retained")
        raise

    after = live.execute("SELECT COUNT(*) FROM kindergartens").fetchone()[0]
    print(f"kindergartens: {before} -> {after}")
    live.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
