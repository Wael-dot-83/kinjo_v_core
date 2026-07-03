"""
Enhanced seed kindergartens from Excel dataset files with web scraping capability.
Reads all .xlsx files in the Dataset folder, deduplicates by Arabic name and coordinates,
enriches missing data via web search, then inserts only records not already present in the database.
"""
try:
    import requests as _requests
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False

import openpyxl
import os
import sys
import time
import re

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database import SessionLocal
import models

DATASET_PATH = r"C:\Users\waelj\OneDrive - zuj.edu.jo\Desktop\Project-Kinjo-seed\DATS"

HEADERS = [
    "اسم الحضانة (عربي)",
    "اسم الحضانة (إنجليزي)",
    "المحافظة",
    "المدينة",
    "المنطقة",
    "العنوان التفصيلي",
    "رقم الهاتف",
    "خط العرض",
    "خط الطول",
    "البريد الإلكتروني",
]

JORDAN_GOVERNORATE_CENTERS = {
    "Amman": {"latitude": 31.95, "longitude": 35.95},
    "Irbid": {"latitude": 32.55, "longitude": 35.85},
    "Zarqa": {"latitude": 32.07, "longitude": 36.10},
    "Mafraq": {"latitude": 32.34, "longitude": 36.20},
    "Jerash": {"latitude": 32.28, "longitude": 35.90},
    "Ajloun": {"latitude": 32.33, "longitude": 35.75},
    "Balqa": {"latitude": 32.04, "longitude": 35.78},
    "Madaba": {"latitude": 31.72, "longitude": 35.79},
    "Karak": {"latitude": 31.18, "longitude": 35.70},
    "Tafileh": {"latitude": 30.83, "longitude": 35.60},
    "Ma'an": {"latitude": 30.20, "longitude": 35.73},
    "Aqaba": {"latitude": 29.53, "longitude": 35.00},
}

PRIORITY_FILES = [
    "حضانات_وحضانات_محدث.xlsx",
    "merged_all_uploads.xlsx",
    "_محدث حضانات_وحضانات_محدث.xlsx",
]

stats = {
    "files_processed": 0,
    "excel_records_found": 0,
    "records_enriched_coords": 0,
    "records_enriched_contact": 0,
    "records_skipped_duplicates": 0,
    "records_inserted": 0,
    "errors": 0,
    "web_errors": 0,
}


def _parse_coord(value) -> float | None:
    if value is None:
        return None
    try:
        coord = float(str(value).strip())
        if coord == 0.0:
            return None
        return coord
    except (ValueError, TypeError):
        return None


def _normalize_governorate(gov: str) -> str:
    if not gov:
        return ""
    gov_clean = gov.strip()
    mapping = {
        "عمان": "Amman",
        "إربد": "Irbid",
        "الزرقاء": "Zarqa",
        "المفرق": "Mafraq",
        "جرش": "Jerash",
        "عجلون": "Ajloun",
        "البلقاء": "Balqa",
        "المداببة": "Madaba",
        "الكرك": "Karak",
        "الطفيلة": "Tafileh",
        "معان": "Ma'an",
        "العقبة": "Aqaba",
    }
    if gov_clean in mapping:
        return mapping[gov_clean]
    if gov_clean in JORDAN_GOVERNORATE_CENTERS:
        return gov_clean
    return gov_clean


def _http_get(url: str, *, params: dict | None = None, headers: dict | None = None, timeout: int = 15):
    if not REQUESTS_AVAILABLE:
        raise RuntimeError("requests library is not installed; cannot perform web searches")
    last_exc = None
    for attempt in range(3):
        try:
            resp = _requests.get(url, params=params, headers=headers, timeout=timeout)
            if resp.status_code == 429:
                wait = 2 ** (attempt + 1)
                time.sleep(wait)
                continue
            resp.raise_for_status()
            return resp
        except Exception as e:
            last_exc = e
            if attempt < 2:
                time.sleep(1)
    raise RuntimeError(f"Failed after 3 attempts: {last_exc}")


def _web_search_coords(name_ar: str, governorate: str, city: str) -> tuple[float | None, float | None]:
    print(f"    [WEB SEARCH] Looking up coordinates for: {name_ar}")
    time.sleep(1.0)
    query = name_ar
    if city:
        query = f"{name_ar}, {city}"
    query = f"{query}, Jordan"
    params = {"q": query, "format": "json", "limit": 1, "addressdetails": 0}
    headers = {"User-Agent": "KinjoKindergartenSeeder/1.0 (admin@kinjo.local)"}
    try:
        resp = _http_get("https://nominatim.openstreetmap.org/search", params=params, headers=headers)
        data = resp.json()
        if data:
            lat = float(data[0]["lat"])
            lon = float(data[0]["lon"])
            print(f"    [WEB SEARCH] Found coordinates via Nominatim: ({lat}, {lon})")
            return lat, lon
    except Exception as e:
        print(f"    [WEB SEARCH] Nominatim geocoding failed: {e}")
    gov_key = _normalize_governorate(governorate)
    if gov_key in JORDAN_GOVERNORATE_CENTERS:
        center = JORDAN_GOVERNORATE_CENTERS[gov_key]
        lat = center["latitude"]
        lon = center["longitude"]
        print(f"    [WEB SEARCH] Falling back to governorate center: ({lat}, {lon})")
        return lat, lon
    print(f"    [WEB SEARCH] No coordinates found")
    return None, None


def _extract_phone_from_text(text: str) -> str | None:
    patterns = [
        r'(\+962\s?\d{2}\s?\d{4}\s?\d{4})',
        r'(0\d{2}\s?\d{3}\s?\d{4})',
        r'(\d{3}[-\s]?\d{3}[-\s]?\d{4})',
        r'(\d{2}\s?\d{3}\s?\d{4})',
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1).replace(" ", "").replace("-", "")
    return None


def _extract_email_from_text(text: str) -> str | None:
    email_pattern = r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})'
    match = re.search(email_pattern, text)
    if match:
        return match.group(1).lower()
    return None


def _web_search_contact(name_ar: str, governorate: str, city: str) -> tuple[str | None, str | None]:
    print(f"    [WEB SEARCH] Looking up contact info for: {name_ar}")
    time.sleep(1.0)
    phone = None
    email = None
    sources_tried = []
    if not REQUESTS_AVAILABLE:
        print(f"    [WEB SEARCH] Skipping contact search: requests library not available")
        return None, None
    queries = [name_ar]
    if city and city not in ("غير محدد", ""):
        queries.append(f"{name_ar} {city}")
    if governorate and governorate not in ("غير محدد", ""):
        queries.append(f"{name_ar} {governorate}")
    queries.append(f"{name_ar} kindergarten Jordan")
    try:
        ddg_params = {
            "q": f"{name_ar} kindergarten contact phone email Jordan",
            "format": "json",
            "no_html": 1,
            "skip_disambig": 1,
        }
        sources_tried.append("DuckDuckGo")
        resp = _http_get("https://api.duckduckgo.com/", params=ddg_params, headers={"Accept": "application/json"}, timeout=20)
        data = resp.json()
        abstract = data.get("Abstract") or data.get("Answer") or ""
        related = data.get("RelatedTopics") or []
        for topic in related:
            if not isinstance(topic, dict):
                continue
            if "Text" in topic or "Abstract" in topic:
                text = topic.get("Text") or topic.get("Abstract") or ""
            else:
                nested_topics = topic.get("Topics") or []
                text = " ".join(
                    subtopic.get("Text") or subtopic.get("Abstract") or ""
                    for subtopic in nested_topics
                    if isinstance(subtopic, dict)
                )
            if text:
                abstract += " " + text
        if abstract:
            found_phone = _extract_phone_from_text(abstract)
            found_email = _extract_email_from_text(abstract)
            if found_phone:
                phone = found_phone
            if found_email:
                email = found_email
    except Exception as e:
        print(f"    [WEB SEARCH] DuckDuckGo contact search failed: {e}")
    if not (phone or email):
        try:
            sources_tried.append("Nominatim")
            query = f"{name_ar}, {city or governorate or 'Jordan'}"
            geo_params = {"q": query, "format": "json", "limit": 1}
            geo_headers = {"User-Agent": "KinjoKindergartenSeeder/1.0 (admin@kinjo.local)"}
            resp = _http_get("https://nominatim.openstreetmap.org/search", params=geo_params, headers=geo_headers)
            data = resp.json()
            if data:
                display = data[0].get("display_name", "")
                found_phone = _extract_phone_from_text(display)
                found_email = _extract_email_from_text(display)
                if found_phone and not phone:
                    phone = found_phone
                if found_email and not email:
                    email = found_email
        except Exception as e:
            print(f"    [WEB SEARCH] Nominatim contact search failed: {e}")
    if phone or email:
        print(f"    [WEB SEARCH] Contact found via {'/'.join(sources_tried)} -> phone={phone}, email={email}")
    else:
        print(f"    [WEB SEARCH] No contact info found (tried: {', '.join(sources_tried) if sources_tried else 'none'})")
    return phone, email


def _get_fallback_coords(governorate: str) -> tuple[float | None, float | None]:
    gov_key = _normalize_governorate(governorate)
    if gov_key in JORDAN_GOVERNORATE_CENTERS:
        center = JORDAN_GOVERNORATE_CENTERS[gov_key]
        return center["latitude"], center["longitude"]
    return None, None


def _is_contact_missing(phone: str | None, email: str | None) -> bool:
    if phone is None or phone.strip() == "" or phone.strip() in ("غير محدد", "None"):
        return True
    if email is None or email.strip() == "" or email.strip() in ("غير محدد", "None"):
        return True
    return False


def collect_from_excel() -> list[dict]:
    all_rows: list[dict] = []
    seen: set[tuple[str, float | None, float | None]] = set()
    all_files = [f for f in os.listdir(DATASET_PATH) if f.endswith(".xlsx")]
    ordered = PRIORITY_FILES + [f for f in all_files if f not in PRIORITY_FILES]

    print(f"\n{'='*60}")
    print(f"PROCESSING EXCEL FILES")
    print(f"{'='*60}")

    for fname in ordered:
        fpath = os.path.join(DATASET_PATH, fname)
        print(f"\n[PROCESSING] File: {fname}")
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception as e:
            print(f"  [SKIP] Cannot open {fname}: {e}")
            stats["errors"] += 1
            continue

        stats["files_processed"] += 1

        for sheet_name in wb.sheetnames:
            if "Summary" in sheet_name or "ورقة" in sheet_name:
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header_row_idx = None
            for i, row in enumerate(rows[:3]):
                if row and any(v and "اسم الحضانة" in str(v) for v in row):
                    header_row_idx = i
                    break
            if header_row_idx is None:
                continue

            header = [str(v).strip() if v else "" for v in rows[header_row_idx]]
            col_map: dict[str, int] = {}
            for j, h in enumerate(header):
                for wanted in HEADERS:
                    if wanted in h:
                        col_map[wanted] = j
                        break

            if "اسم الحضانة (عربي)" not in col_map:
                print(f"  [SKIP] {fname}::{sheet_name} - cannot map Arabic name column")
                wb.close()
                continue

            def get_col(row_data, col_name: str) -> str:
                idx = col_map.get(col_name)
                if idx is None or idx >= len(row_data):
                    return ""
                v = row_data[idx]
                return str(v).strip() if v and str(v).strip() not in ("None", "") else ""

            added_from_sheet = 0
            enriched_coords = 0
            enriched_contact = 0

            for row_data in rows[header_row_idx + 1:]:
                if not row_data:
                    continue

                idx_ar = col_map["اسم الحضانة (عربي)"]
                raw_ar = row_data[idx_ar] if idx_ar < len(row_data) else None
                name_ar = str(raw_ar).strip() if raw_ar and str(raw_ar).strip() != "None" else ""

                if not name_ar or name_ar.startswith("اسم"):
                    continue

                governorate_val = get_col(row_data, "المحافظة")
                city_val = get_col(row_data, "المدينة")
                phone_val = get_col(row_data, "رقم الهاتف")
                email_val = get_col(row_data, "البريد الإلكتروني")

                lat_raw = None
                lon_raw = None
                if "خط العرض" in col_map:
                    idx_lat = col_map["خط العرض"]
                    lat_raw = row_data[idx_lat] if idx_lat < len(row_data) else None
                if "خط الطول" in col_map:
                    idx_lon = col_map["خط الطول"]
                    lon_raw = row_data[idx_lon] if idx_lon < len(row_data) else None
                lat = _parse_coord(lat_raw)
                lon = _parse_coord(lon_raw)

                coords_needed_enrichment = False
                if lat is None or lon is None:
                    coords_needed_enrichment = True
                    lat, lon = _get_fallback_coords(governorate_val)
                    if lat is None or lon is None:
                        lat, lon = _web_search_coords(name_ar, governorate_val, city_val)
                        if lat is not None and lon is not None:
                            stats["records_enriched_coords"] += 1
                            enriched_coords += 1

                contact_enriched = False
                if _is_contact_missing(phone_val, email_val):
                    found_phone, found_email = _web_search_contact(name_ar, governorate_val, city_val)
                    if found_phone and not phone_val:
                        phone_val = found_phone
                        contact_enriched = True
                    if found_email and not email_val:
                        email_val = found_email
                        contact_enriched = True
                    if contact_enriched:
                        stats["records_enriched_contact"] += 1
                        enriched_contact += 1

                key = (name_ar, lat, lon)
                if key in seen:
                    stats["records_skipped_duplicates"] += 1
                    continue
                seen.add(key)
                stats["excel_records_found"] += 1

                all_rows.append({
                    "name_ar": name_ar,
                    "name_en": get_col(row_data, "اسم الحضانة (إنجليزي)") or None,
                    "governorate": governorate_val or "غير محدد",
                    "city": city_val or "غير محدد",
                    "area": get_col(row_data, "المنطقة") or "غير محدد",
                    "address_line": get_col(row_data, "العنوان التفصيلي") or "غير محدد",
                    "contact_phone": phone_val if phone_val else "غير محدد",
                    "contact_email": email_val if email_val else None,
                    "latitude": lat,
                    "longitude": lon,
                    "coords_enriched": enriched_coords > 0 or coords_needed_enrichment,
                })
                added_from_sheet += 1

            suffix = ""
            if enriched_coords:
                suffix += f" | Coords enriched: {enriched_coords}"
            if enriched_contact:
                suffix += f" | Contact enriched: {enriched_contact}"
            print(f"  {fname}::{sheet_name} -> {added_from_sheet} unique records{suffix}")

        wb.close()

    return all_rows


def seed():
    db = SessionLocal()
    try:
        print(f"\n{'='*60}")
        print(f"ENHANCED KINDERGARTEN SEED SCRIPT")
        print(f"{'='*60}")

        existing_keys: set[tuple[str, float | None, float | None]] = set()
        for kg in db.query(
            models.Kindergarten.name_ar,
            models.Kindergarten.latitude,
            models.Kindergarten.longitude,
        ).all():
            key = (kg[0].strip() if kg[0] else "", kg[1], kg[2])
            existing_keys.add(key)
        print(f"\nExisting in DB: {len(existing_keys)} kindergartens")

        excel_records = collect_from_excel()
        print(f"\n{'='*60}")
        print(f"SUMMARY BEFORE INSERTION")
        print(f"{'='*60}")
        print(f"Unique from Excel: {len(excel_records)} kindergartens")
        print(f"Files processed:   {stats['files_processed']}")

        to_insert = [
            r for r in excel_records
            if (r["name_ar"], r.get("latitude"), r.get("longitude")) not in existing_keys
        ]
        print(f"New to insert:     {len(to_insert)} kindergartens")

        if not to_insert:
            print("\nNothing to insert - all records already in database.")
            return

        inserted = 0
        errors = 0

        print(f"\n{'='*60}")
        print(f"INSERTING RECORDS")
        print(f"{'='*60}")

        for rec in to_insert:
            try:
                kg = models.Kindergarten(
                    name_ar=rec["name_ar"],
                    name_en=rec["name_en"],
                    governorate=rec["governorate"],
                    district=rec["city"],
                    area=rec["area"],
                    address_line=rec["address_line"],
                    contact_phone=rec["contact_phone"],
                    contact_email=rec.get("contact_email"),
                    status=models.KindergartenStatus.ACTIVE,
                    latitude=rec.get("latitude"),
                    longitude=rec.get("longitude"),
                )
                db.add(kg)
                inserted += 1
                stats["records_inserted"] += 1
                enriched_note = " (coords enriched)" if rec.get("coords_enriched") else ""
                name = rec["name_ar"]
                print(f"  [INSERTED] {name}{enriched_note}")
            except Exception as e:
                name = rec["name_ar"]
                print(f'  [ERROR] Cannot insert "{name}": {e}')
                errors += 1
                stats["errors"] += 1

        db.commit()

        print(f"\n{'='*60}")
        print(f"FINAL SUMMARY")
        print(f"{'='*60}")
        print(f"Files processed:           {stats['files_processed']}")
        print(f"Records found in Excel:    {stats['excel_records_found']}")
        print(f"Coords enriched via web:   {stats['records_enriched_coords']}")
        print(f"Contact enriched via web:  {stats['records_enriched_contact']}")
        print(f"Duplicates skipped:        {stats['records_skipped_duplicates']}")
        print(f"Records inserted:          {stats['records_inserted']}")
        print(f"Errors encountered:        {stats['errors']}")
        print(f"  [OK] Inserted {inserted} new kindergartens ({errors} errors)")
        total = db.query(models.Kindergarten).count()
        print(f"  [OK] Total kindergartens in DB now: {total}")

    except Exception as e:
        db.rollback()
        print(f"\n[FATAL] {e}")
        stats["errors"] += 1
        raise
    finally:
        db.close()


if __name__ == "__main__":
    seed()
