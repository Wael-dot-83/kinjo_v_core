"""National immunization schedule: Excel template, parsing, and persistence.

Backs the MOH agency report ``vaccination_due_children``. The admin downloads
an Excel template (المطعوم | العمر | الوحدة), fills it with the national
schedule, and uploads it. Rows are validated and stored in
``NationalImmunizationSchedule``; the report then counts children whose age has
reached each vaccine's scheduled age (age-eligibility, aggregated only).
"""
from __future__ import annotations

import io
from typing import Any

from sqlalchemy.orm import Session

import models

# Header labels used in the downloadable template and expected on upload.
COL_VACCINE = "المطعوم"
COL_AGE = "العمر"
COL_UNIT = "الوحدة (يوم/شهر/سنة)"
COL_NOTES = "ملاحظات (اختياري)"

# Approximate day-count per unit for age-eligibility comparisons.
_UNIT_DAYS = {
    models.ImmunizationAgeUnit.DAY: 1,
    models.ImmunizationAgeUnit.MONTH: 30,
    models.ImmunizationAgeUnit.YEAR: 365,
}

# Accepted unit spellings (Arabic + English, singular/plural) → canonical enum.
_UNIT_ALIASES = {
    "يوم": models.ImmunizationAgeUnit.DAY,
    "أيام": models.ImmunizationAgeUnit.DAY,
    "ايام": models.ImmunizationAgeUnit.DAY,
    "day": models.ImmunizationAgeUnit.DAY,
    "days": models.ImmunizationAgeUnit.DAY,
    "شهر": models.ImmunizationAgeUnit.MONTH,
    "أشهر": models.ImmunizationAgeUnit.MONTH,
    "اشهر": models.ImmunizationAgeUnit.MONTH,
    "month": models.ImmunizationAgeUnit.MONTH,
    "months": models.ImmunizationAgeUnit.MONTH,
    "سنة": models.ImmunizationAgeUnit.YEAR,
    "سنوات": models.ImmunizationAgeUnit.YEAR,
    "عام": models.ImmunizationAgeUnit.YEAR,
    "year": models.ImmunizationAgeUnit.YEAR,
    "years": models.ImmunizationAgeUnit.YEAR,
}

_UNIT_LABEL_AR = {
    models.ImmunizationAgeUnit.DAY: "يوم",
    models.ImmunizationAgeUnit.MONTH: "شهر",
    models.ImmunizationAgeUnit.YEAR: "سنة",
}


class ImmunizationScheduleError(ValueError):
    """Raised when an uploaded schedule file cannot be parsed/validated."""


def unit_label_ar(unit: models.ImmunizationAgeUnit) -> str:
    return _UNIT_LABEL_AR.get(unit, "")


def _due_age_days(value: int, unit: models.ImmunizationAgeUnit) -> int:
    return int(value) * _UNIT_DAYS[unit]


def build_template_xlsx() -> bytes:
    """Return the empty immunization-schedule template as .xlsx bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "جدول المطاعيم"
    ws.sheet_view.rightToLeft = True

    headers = [COL_VACCINE, COL_AGE, COL_UNIT, COL_NOTES]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E5AAC")
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Illustrative example rows (Jordan national schedule sample).
    examples = [
        ("BCG - السل", 0, "يوم", "عند الولادة"),
        ("الخماسي - الجرعة الأولى", 60, "يوم", ""),
        ("شلل الأطفال الفموي - الجرعة الأولى", 2, "شهر", ""),
        ("MMR - الحصبة والنكاف والحصبة الألمانية", 12, "شهر", ""),
        ("الجرعة المنشطة DTP", 18, "شهر", ""),
    ]
    for r, row in enumerate(examples, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    widths = [40, 10, 18, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_schedule_xlsx(data: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """Parse uploaded .xlsx bytes into validated schedule rows.

    Returns ``(rows, errors)``. ``rows`` are dicts ready for persistence.
    Rows with problems are skipped and described in ``errors`` (row-numbered).
    A completely empty/blank sheet yields ``([], [])`` — the caller decides
    whether that is acceptable.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - surface a friendly message
        raise ImmunizationScheduleError(
            "تعذّر قراءة الملف. تأكد أنه ملف Excel بصيغة ‎.xlsx"
        ) from exc

    ws = wb.active
    rows: list[dict[str, Any]] = []
    errors: list[str] = []

    first = True
    order = 0
    for idx, raw in enumerate(ws.iter_rows(values_only=True), start=1):
        if raw is None:
            continue
        cells = list(raw) + [None, None, None, None]
        name, age_raw, unit_raw = cells[0], cells[1], cells[2]
        notes = cells[3]

        # Skip a header row (either the template header or any non-numeric first row).
        if first:
            first = False
            header_like = str(name or "").strip() == COL_VACCINE or not _looks_numeric(age_raw)
            if header_like:
                continue

        if _is_blank(name) and _is_blank(age_raw) and _is_blank(unit_raw):
            continue  # spacer row

        if _is_blank(name):
            errors.append(f"الصف {idx}: اسم المطعوم مفقود")
            continue
        unit = _resolve_unit(unit_raw)
        if unit is None:
            errors.append(f"الصف {idx}: وحدة العمر غير صالحة (استخدم يوم/شهر/سنة)")
            continue
        try:
            age_value = int(float(age_raw))
        except (TypeError, ValueError):
            errors.append(f"الصف {idx}: قيمة العمر غير رقمية")
            continue
        if age_value < 0:
            errors.append(f"الصف {idx}: قيمة العمر يجب أن تكون صفرًا أو أكثر")
            continue

        order += 1
        rows.append({
            "vaccine_name": str(name).strip()[:200],
            "age_value": age_value,
            "age_unit": unit,
            "due_age_days": _due_age_days(age_value, unit),
            "notes": (str(notes).strip()[:500] if not _is_blank(notes) else None),
            "sort_order": order,
        })

    wb.close()
    return rows, errors


def replace_schedule(db: Session, rows: list[dict[str, Any]], actor_id: int | None) -> int:
    """Replace the entire stored schedule with ``rows``. Returns rows written."""
    db.query(models.NationalImmunizationSchedule).delete(synchronize_session=False)
    for row in rows:
        db.add(models.NationalImmunizationSchedule(uploaded_by=actor_id, **row))
    db.flush()
    return len(rows)


def get_schedule(db: Session) -> list[models.NationalImmunizationSchedule]:
    return (
        db.query(models.NationalImmunizationSchedule)
        .order_by(models.NationalImmunizationSchedule.sort_order, models.NationalImmunizationSchedule.id)
        .all()
    )


def schedule_count(db: Session) -> int:
    return db.query(models.NationalImmunizationSchedule).count()


# --------------------------------------------------------------------------
# helpers
# --------------------------------------------------------------------------
def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _looks_numeric(value: Any) -> bool:
    if isinstance(value, (int, float)):
        return True
    try:
        float(str(value).strip())
        return True
    except (TypeError, ValueError):
        return False


def _resolve_unit(value: Any) -> models.ImmunizationAgeUnit | None:
    if value is None:
        return None
    key = str(value).strip().lower()
    return _UNIT_ALIASES.get(key)
