"""
Hardened Admin Endpoints Module

This module provides secure admin endpoints with:
- Explicit admin authorization on all endpoints
- Rate limiting on sensitive operations
- Object-level authorization (IDOR protection)
- Standardized error responses with correlation IDs
- Comprehensive audit logging with before/after diffs
- Pagination enforcement
- Bulk operation guardrails with confirmation tokens
- CSV import with per-row validation and error reporting
- Dry-run/preview mode support
- Manager assignment validation (one active manager per kindergarten)
"""
import csv
import io
import os
import enum
import logging
import math
from zipfile import BadZipFile
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone, date
from typing import Iterable, List, Optional, Dict, Any, Set, Tuple, Union

_JORDAN_TZ = timezone(timedelta(hours=3))

from fastapi import APIRouter, Depends, HTTPException, status, Query, Request, UploadFile, File, Form
from services.jordan_locations import governorate_filter
from fastapi.responses import Response, JSONResponse
from pydantic import BaseModel, ConfigDict, Field, ValidationError, model_validator
from admin_reports_api import AdminAlertResponse, AdminAlertsListResponse

from sqlalchemy.orm import Session, selectinload
from sqlalchemy import or_, func, and_, select, case, union_all
from sqlalchemy.exc import IntegrityError, SQLAlchemyError

import models
import validators
from database import get_db
from utils.time_utils import jordan_day_bounds, jordan_date_range_filter, now_amman, to_jordan_date, to_jordan_iso
from export_service import export_service
from dependencies import get_current_user
from rate_limiter import limiter
from config import settings
from auth import change_user_password, get_password_hash, verify_password
from notification_service import (
    create_message_notifications,
    dispatch_message_notification_tasks,
)
from api.users import (
    PASSWORD_RESET_GENERIC_MESSAGE,
    apply_password_reset,
    initiate_password_reset,
)
from messaging_permissions import ACTIVE_ENROLLMENT_STATUSES, ensure_kindergartens_exist
from validators import build_arabic_search_terms
from kindergarten_import_service import KindergartenImportService
from kpi_service import KPIService
from cache_service import cache_service
from csv_utils import escape_csv_formula
from upload_security import validate_xlsx_archive
from audit_actions import AuditAction
from admin_security import (
    # Error handling
    APIError, forbidden_error, unauthenticated_error, validation_error,
    not_found_error, conflict_error,
    ErrorCode,
    # Audit logging
    log_audit_event, model_to_dict, get_correlation_id,
    # Authorization
    can_admin_access_user,
     validate_bulk_targets,
    # Schemas
    UserCreateSchema, UserUpdateSchema, BulkStatusUpdateSchema,
    BulkDeleteSchema, BulkCreateSchema, AdminPasswordResetSchema,
    PasswordResetRequestSchema, PasswordResetConfirmSchema,
    # Bulk operations
    generate_confirmation_token,
    verify_confirmation_token,
    # CSV
    CSVRowError, CSVImportResult, sanitize_csv_cell,
    # Pagination
    enforce_pagination,
)

logger = logging.getLogger(__name__)
_ADMIN_DASHBOARD_CACHE_TTL_SECONDS = 30


def _admin_dashboard_cache_get(key: str):
    if settings.TESTING:
        return None
    try:
        return cache_service.get(key)
    except (RuntimeError, AttributeError, TypeError, ValueError) as e:
        logger.warning(f"Cache get failed for key '{key}': {e}", exc_info=False)
        return None


def _admin_dashboard_cache_set(
    key: str,
    value: Dict[str, Any],
    ttl_seconds: int = _ADMIN_DASHBOARD_CACHE_TTL_SECONDS,
) -> None:
    if settings.TESTING:
        return
    try:
        cache_service.set(key, value, ttl_seconds=ttl_seconds)
    except (RuntimeError, AttributeError, TypeError, ValueError) as e:
        # Best-effort caching should never break dashboard reads.
        logger.warning(f"Cache set failed for key '{key}': {e}", exc_info=False)


# =============================================================================
# Manager Assignment Service Functions
# =============================================================================

def validate_manager_assignment(
    db: Session,
    role: models.UserRole,
    kindergarten_id: Optional[int],
    status_value: models.UserStatus,
    exclude_user_id: Optional[int] = None
) -> None:
    """
    Validate manager assignment rules.

    Business Rules:
    - Manager must be assigned to a kindergarten
    - Each kindergarten can have at most one active manager
    """
    try:
        validators.validate_manager_rules(
            db,
            role=role,
            kindergarten_id=kindergarten_id,
            status_value=status_value,
            exclude_user_id=exclude_user_id
        )
    except validators.ManagerRuleError as exc:
        if exc.status_code == status.HTTP_409_CONFLICT:
            raise conflict_error(exc.message, {"kindergarten_id": exc.message})
        raise validation_error(
            exc.message,
            {"kindergarten_id": "Kindergarten is required for manager role"}
        )


def validate_bulk_manager_assignments(
    db: Session,
    users_data: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:
    """
    Validate manager assignments for bulk operations.

    Returns list of validation errors for each user that fails validation.
    """
    errors = []

    # Group managers by kindergarten to check for duplicates within the batch
    kg_managers = {}
    for i, user_data in enumerate(users_data):
        role = user_data.get('role')
        if isinstance(role, str):
            try:
                role = models.UserRole(role)
            except ValueError as e:
                logger.warning(f"Invalid role '{role}' at record {i}: {e}")
                role = None
        status_value = user_data.get('status', models.UserStatus.ACTIVE)
        if isinstance(status_value, str):
            try:
                status_value = models.UserStatus(status_value)
            except ValueError as e:
                logger.warning(f"Invalid status '{status_value}' at record {i}: {e}")
                status_value = models.UserStatus.INACTIVE
        if role == models.UserRole.MANAGER:
            kg_id = user_data.get('kindergarten_id')
            if kg_id is None:
                errors.append({
                    "row": i + 1,
                    "field": "kindergarten_id",
                    "error": "Manager must be assigned to a kindergarten"
                })
                continue

            if status_value == models.UserStatus.ACTIVE:
                if kg_id in kg_managers:
                    errors.append({
                        "row": i + 1,
                        "field": "kindergarten_id",
                        "error": f"Multiple managers assigned to kindergarten {kg_id} in this batch"
                    })
                    continue
                kg_managers[kg_id] = i + 1

            try:
                validators.validate_manager_rules(
                    db,
                    role=role,
                    kindergarten_id=kg_id,
                    status_value=status_value
                )
            except validators.ManagerRuleError as exc:
                errors.append({
                    "row": i + 1,
                    "field": "kindergarten_id",
                    "error": exc.message
                })

    return errors


# =============================================================================
# Router Definition
# =============================================================================

router = APIRouter(tags=["Admin"])


# =============================================================================
# Authorization Helpers
# =============================================================================

# require_admin and require_admin_or_manager are wrappers that chain to
# admin_security's functions via Depends(get_current_user).
def require_admin(current_user: models.User = Depends(get_current_user)) -> models.User:
    from admin_security import require_admin_role
    return require_admin_role(current_user)


def require_admin_or_manager(current_user: models.User = Depends(get_current_user)) -> models.User:
    from admin_security import require_admin_or_manager_role
    return require_admin_or_manager_role(current_user)


def get_client_ip(request: Request) -> str:
    """Extract client IP from request"""
    if request.client:
        return request.client.host
    return "unknown"


# =============================================================================
# User Management Endpoints (Hardened)
# =============================================================================

@router.get("/users")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def list_users(
    request: Request,
    page: int = Query(1, ge=1, description="Page number (1-indexed)"),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE),
    role: Optional[models.UserRole] = None,
    status_filter: Optional[models.UserStatus] = Query(None, alias="status"),
    kindergarten_id: Optional[int] = None,
    search: Optional[str] = None,
    current_user: models.User = Depends(require_admin_or_manager),
    db: Session = Depends(get_db)
):
    """
    List users with pagination, filtering, and role-based scoping.

    - Admins: See all non-admin users, can filter by any kindergarten
    - Managers: See only users in their kindergarten

    Returns paginated results with enforced page size limits.
    """
    # Enforce pagination limits
    page, page_size, offset = enforce_pagination(page, page_size)

    query = db.query(models.User).filter(models.User.deleted_at.is_(None))

    # Apply role-based scoping
    if current_user.role == models.UserRole.ADMIN:
        # Admins see all except other admins
        query = query.filter(models.User.role != models.UserRole.ADMIN)
        if kindergarten_id:
            query = query.filter(models.User.kindergarten_id == kindergarten_id)
    else:
        # Managers are scoped to their kindergarten
        query = query.filter(models.User.kindergarten_id == current_user.kindergarten_id)
        # Ignore any cross-kindergarten filter attempts
        if kindergarten_id and kindergarten_id != current_user.kindergarten_id:
            log_audit_event(
                db, AuditAction.ACCESS_DENIED, current_user, "User",
                metadata={"attempted_kindergarten_id": kindergarten_id},
                sensitivity_level=2
            )
            db.commit()

    # Apply filters
    if role:
        # Prevent filtering by ADMIN role
        if role == models.UserRole.ADMIN:
            raise forbidden_error("Cannot filter by ADMIN role")
        query = query.filter(models.User.role == role)

    if status_filter:
        query = query.filter(models.User.status == status_filter)

    if search:
        search = search[:100]
        search_term = f"%{search}%"
        query = query.filter(or_(
            models.User.username.ilike(search_term),
            models.User.email.ilike(search_term)
        ))

    # Get total count before pagination
    total = query.count()

    # Apply pagination with stable ordering; eager-load kindergarten to avoid N+1
    users = (
        query
        .options(selectinload(models.User.kindergarten))
        .order_by(models.User.id)
        .offset(offset)
        .limit(page_size)
        .all()
    )

    # Calculate pagination metadata
    total_pages = (total + page_size - 1) // page_size

    return {
        "data": [
            {
                "id": u.id,
                "username": u.username,
                "email": u.email,
                "role": u.role.value,
                "status": u.status.value,
                "kindergarten_id": u.kindergarten_id,
                "created_at": u.created_at.isoformat() if u.created_at else None
            }
            for u in users
        ],
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": total_pages,
            "has_next": page < total_pages,
            "has_prev": page > 1
        },
        "correlation_id": get_correlation_id()
    }


@router.post("/users", status_code=status.HTTP_201_CREATED)
@limiter.limit(settings.RATE_LIMIT_ADMIN_WRITE)
def create_user(
    request: Request,
    user_data: UserCreateSchema,
    current_user: models.User = Depends(require_admin_or_manager),
    db: Session = Depends(get_db)
):
    """
    Create a new user with strict validation.

    Authorization rules:
    - Admins: Can create any non-admin user
    - Managers: Can only create SUPERVISOR/PARENT roles in their kindergarten
    """

    # Authorization check
    if current_user.role == models.UserRole.MANAGER:
        # Manager restrictions
        if user_data.role in [models.UserRole.ADMIN, models.UserRole.MANAGER]:
            log_audit_event(
                db, AuditAction.ACCESS_DENIED, current_user, "User",
                metadata={"attempted_role": user_data.role.value},
                sensitivity_level=3
            )
            db.commit()
            raise forbidden_error("Managers cannot create Admin or Manager accounts")

        # Force kindergarten to manager's kindergarten
        if user_data.kindergarten_id and user_data.kindergarten_id != current_user.kindergarten_id:
            log_audit_event(
                db, AuditAction.ACCESS_DENIED, current_user, "User",
                metadata={"attempted_kindergarten_id": user_data.kindergarten_id},
                sensitivity_level=3
            )
            db.commit()
            raise forbidden_error("Cannot create users for other kindergartens")

        user_data.kindergarten_id = current_user.kindergarten_id
    else:
        # Admin restrictions - cannot create other admins
        if user_data.role == models.UserRole.ADMIN:
            raise forbidden_error("Cannot create admin accounts through this endpoint")

    if user_data.role == models.UserRole.SUPERVISOR and not user_data.kindergarten_id:
        raise validation_error(
            "Supervisor must belong to a kindergarten",
            {"kindergarten_id": "Supervisor accounts require a kindergarten assignment"}
        )

    # Business rule: Manager validation
    validate_manager_assignment(
        db,
        user_data.role,
        user_data.kindergarten_id,
        models.UserStatus.ACTIVE
    )

    # Check for existing username or email (email only if provided)
    # Always check username uniqueness
    existing_username = db.query(models.User).filter(models.User.username == user_data.username).first()
    if existing_username:
        raise conflict_error("Username already exists", {"username": "Username is already taken"})

    # Check email uniqueness only if email is provided
    if user_data.email is not None:
        existing_email = db.query(models.User).filter(models.User.email == user_data.email).first()
        if existing_email:
            raise conflict_error("Email already exists", {"email": "Email is already registered"})

    # Create user
    hashed_password = get_password_hash(user_data.password)

    new_user = models.User(
        username=user_data.username,
        email=user_data.email,
        hashed_password=hashed_password,
        role=user_data.role,
        kindergarten_id=user_data.kindergarten_id,
        status=models.UserStatus.ACTIVE,
        must_change_password=(user_data.role in [
            models.UserRole.MANAGER, models.UserRole.SUPERVISOR
        ]),
    )

    # Set profile fields if provided
    for field in ("full_name", "phone_number", "address", "nationality", "national_id", "passport_number"):
        val = getattr(user_data, field, None)
        if val is not None:
            setattr(new_user, field, val)

    # Validate identity by nationality for managers/supervisors
    if (
        user_data.nationality
        and user_data.role in [models.UserRole.MANAGER, models.UserRole.SUPERVISOR]
    ):
        try:
            validators.validate_identity_by_nationality(
                user_data.nationality,
                user_data.national_id,
                user_data.passport_number,
            )
        except validators.ValidationError as exc:
            raise validation_error(exc.message, {"identity": exc.message})

    try:
        db.add(new_user)
        db.flush()

        if new_user.role == models.UserRole.SUPERVISOR:
            validators.ensure_supervisor_profile(db, new_user, new_user.kindergarten_id)

        # Handle child creation for PARENT role in the same transaction as the
        # user, profile, and audit rows. Flushes allocate IDs without exposing a
        # partially-created account to other transactions.
        if user_data.role == models.UserRole.PARENT and user_data.children:
            parent_profile = models.ParentProfile(
                user_id=new_user.id,
                first_name="",
                last_name="",
                phone_number="",
                gender=models.Gender.MALE,
                nationality="",
                home_governorate="",
                home_district="",
                home_area="",
                home_address_line="",
                correspondence_preference=True,
                profile_complete=False
            )
            db.add(parent_profile)
            db.flush()

            children = []
            # Child age was already validated by ChildCreateSchema.
            for child_data in user_data.children:
                new_child = models.Child(
                    parent_id=parent_profile.id,
                    first_name=child_data.first_name,
                    last_name=child_data.last_name,
                    gender=child_data.gender,
                    date_of_birth=child_data.date_of_birth,
                    father_name=child_data.father_name,
                    mother_first_name=child_data.mother_first_name,
                    mother_second_name=child_data.mother_second_name or "",
                    mother_last_name=child_data.mother_last_name or "",
                    mother_nationality=child_data.mother_nationality or "",
                    media_consent=False,
                    correspondence_flag=True,
                    profile_complete=False
                )
                db.add(new_child)
                children.append(new_child)

            db.flush()
            log_audit_event(
                db, AuditAction.CHILDREN_CREATED, current_user, "Child",
                target_ids=[child.id for child in children],
                metadata={"parent_user_id": new_user.id, "children_count": len(children)},
                sensitivity_level=2
            )

        log_audit_event(
            db, AuditAction.USER_CREATED, current_user, "User",
            target_ids=new_user.id,
            after_state=model_to_dict(new_user),
            sensitivity_level=3
        )
        response_payload = {
            "id": new_user.id,
            "username": new_user.username,
            "email": new_user.email,
            "role": new_user.role.value,
            "status": new_user.status.value,
            "kindergarten_id": new_user.kindergarten_id,
            "full_name": new_user.full_name,
            "phone_number": new_user.phone_number,
            "address": new_user.address,
            "nationality": new_user.nationality,
            "national_id": new_user.national_id,
            "passport_number": new_user.passport_number,
            "correlation_id": get_correlation_id(),
        }
        db.commit()
    except Exception:
        db.rollback()
        raise

    return response_payload


# =============================================================================
# User Export
# =============================================================================

@router.get("/users/export")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def export_users(
    request: Request,
    format: str = Query("csv", pattern="^(csv|json)$"),
    role: Optional[models.UserRole] = None,
    status_filter: Optional[models.UserStatus] = Query(None, alias="status"),
    kindergarten_id: Optional[int] = None,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Export users list with filtering.
    """
    query = db.query(models.User).filter(models.User.role != models.UserRole.ADMIN)

    if kindergarten_id:
        query = query.filter(models.User.kindergarten_id == kindergarten_id)
    if role:
        query = query.filter(models.User.role == role)
    if status_filter:
        query = query.filter(models.User.status == status_filter)

    # P1-B: Hard row limit to prevent OOM on large deployments.
    # Fetch one extra row to detect overflow without loading everything.
    MAX_EXPORT_ROWS = 10_000
    users = query.order_by(models.User.id).limit(MAX_EXPORT_ROWS + 1).all()

    if len(users) > MAX_EXPORT_ROWS:
        raise HTTPException(
            status_code=422,  # 422 literal: Starlette deprecated the ENTITY constant name
            detail=(
                f"Export would return more than {MAX_EXPORT_ROWS:,} rows. "
                "Apply role, status, or kindergarten filters to narrow the result set."
            ),
        )

    # Audit log
    log_audit_event(
        db, AuditAction.USER_EXPORT, current_user, "User",
        metadata={"format": format, "count": len(users)},
        sensitivity_level=2
    )
    db.commit()

    if format == "json":
        data = [
            {
                "id": u.id,
                "username": u.username,
                "email": u.email,
                "role": u.role.value,
                "status": u.status.value,
                "kindergarten_id": u.kindergarten_id,
                "created_at": u.created_at.isoformat() if u.created_at else None
            }
            for u in users
        ]
        return JSONResponse(
            content=data,
            headers={
                "Content-Disposition": f"attachment; filename=users_export_{datetime.now(_JORDAN_TZ).date()}.json"
            }
        )

    # CSV export
    headers = ["ID", "Username", "Email", "Role", "Status", "Kindergarten ID", "Created At"]
    rows = []
    for u in users:
        rows.append([
            u.id,
            u.username,
            u.email,
            u.role.value,
            u.status.value,
            u.kindergarten_id or "N/A",
            u.created_at.isoformat() if u.created_at else ""
        ])

    return export_service.generate_csv_response(
        headers=headers,
        data=rows,
        filename=f"users_export_{datetime.now(_JORDAN_TZ).date()}.csv"
    )


@router.get("/users/{user_id:int}")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def get_user(
    request: Request,
    user_id: int,
    current_user: models.User = Depends(require_admin_or_manager),
    db: Session = Depends(get_db)
):
    """
    Get user details with IDOR protection.
    """
    user = db.query(models.User).filter(
        models.User.id == user_id, models.User.deleted_at.is_(None)
    ).first()

    if not user:
        raise not_found_error("User not found")

    # IDOR check
    if not can_admin_access_user(current_user, user):
        log_audit_event(
            db, AuditAction.ACCESS_DENIED, current_user, "User",
            target_ids=user_id,
            metadata={"reason": "IDOR protection"},
            sensitivity_level=2
        )
        db.commit()
        raise forbidden_error("Not authorized to access this user")

    return {
        "id": user.id,
        "username": user.username,
        "email": user.email,
        "role": user.role.value,
        "status": user.status.value,
        "kindergarten_id": user.kindergarten_id,
        "full_name": user.full_name,
        "phone_number": user.phone_number,
        "address": user.address,
        "nationality": user.nationality,
        "national_id": user.national_id,
        "passport_number": user.passport_number,
        "created_at": user.created_at.isoformat() if user.created_at else None,
        "correlation_id": get_correlation_id()
    }


@router.put("/users/{user_id:int}")
@limiter.limit(settings.RATE_LIMIT_ADMIN_WRITE)
def update_user(
    request: Request,
    user_id: int,
    user_data: UserUpdateSchema,
    current_user: models.User = Depends(require_admin_or_manager),
    db: Session = Depends(get_db)
):
    """
    Update user with IDOR protection and audit logging.
    """

    user = db.query(models.User).filter(
        models.User.id == user_id, models.User.deleted_at.is_(None)
    ).first()

    if not user:
        raise not_found_error("User not found")

    # IDOR check
    if not can_admin_access_user(current_user, user):
        log_audit_event(
            db, AuditAction.ACCESS_DENIED, current_user, "User",
            target_ids=user_id,
            metadata={"reason": "IDOR protection", "action": "update"},
            sensitivity_level=2
        )
        db.commit()
        raise forbidden_error("Not authorized to update this user")

    # Business rule: Manager validation for updates
    target_role = user_data.role if user_data.role is not None else user.role
    target_kindergarten_id = user_data.kindergarten_id if user_data.kindergarten_id is not None else user.kindergarten_id

    target_status = user_data.status if user_data.status is not None else user.status

    # An ACTIVE kindergarten must retain exactly one ACTIVE manager. Replacement
    # inside the same kindergarten is handled atomically by the dedicated
    # assignment flow; demotion, suspension, deletion, or reassignment of its
    # sole manager is rejected until the kindergarten is frozen/inactive.
    if (
        current_user.role == models.UserRole.ADMIN
        and user.role == models.UserRole.MANAGER
        and user.status == models.UserStatus.ACTIVE
        and user.kindergarten_id is not None
    ):
        current_kg = db.query(models.Kindergarten).filter(
            models.Kindergarten.id == user.kindergarten_id
        ).first()
        remains_active_manager = (
            target_role == models.UserRole.MANAGER
            and target_status == models.UserStatus.ACTIVE
            and target_kindergarten_id == user.kindergarten_id
        )
        if current_kg and current_kg.status == models.KindergartenStatus.ACTIVE and not remains_active_manager:
            raise conflict_error(
                "An active kindergarten must retain an active manager. Freeze the kindergarten or replace its manager atomically.",
                {"manager": "active_kindergarten_requires_manager"},
            )

    # Detect a manager assignment/reassignment transition (FRD §4, C3).
    # This covers: promoting a supervisor/other role to manager, or moving an
    # existing manager to a different kindergarten. Only admins can change
    # role/kindergarten, so the cascade only ever runs for admins.
    is_manager_assignment = (
        current_user.role == models.UserRole.ADMIN
        and target_role == models.UserRole.MANAGER
        and (
            user.role != models.UserRole.MANAGER
            or (user_data.kindergarten_id is not None
                and user_data.kindergarten_id != user.kindergarten_id)
        )
    )

    if not is_manager_assignment:
        validate_manager_assignment(
            db,
            target_role,
            target_kindergarten_id,
            target_status,
            user_id
        )

    # Capture before state for audit
    before_state = model_to_dict(user)

    # Apply updates with authorization checks
    if user_data.email is not None:
        # Check uniqueness
        existing = db.query(models.User).filter(
            models.User.email == user_data.email,
            models.User.id != user_id
        ).first()
        if existing:
            raise conflict_error("Email already in use", {"email": "Email is already registered"})
        user.email = user_data.email

    if user_data.password is not None:
        change_user_password(db, user, user_data.password, commit=False)

    # Update profile fields if provided
    for field in ("full_name", "phone_number", "address", "nationality", "national_id", "passport_number"):
        val = getattr(user_data, field, None)
        if val is not None:
            setattr(user, field, val)

    # Validate identity by nationality if nationality changed
    if user_data.nationality and user.role in [models.UserRole.MANAGER, models.UserRole.SUPERVISOR]:
        try:
            validators.validate_identity_by_nationality(
                user_data.nationality,
                user_data.national_id or user.national_id,
                user_data.passport_number or user.passport_number,
            )
        except validators.ValidationError as exc:
            raise validation_error(exc.message, {"identity": exc.message})

    # Only admins can change role and status
    if current_user.role == models.UserRole.ADMIN:
        if is_manager_assignment:
            # FRD §4 (C1–C5): atomic manager assignment cascade — detaches the
            # user from any previous kindergarten, strips every supervisor
            # artifact, optionally vacates the target KG's outgoing manager,
            # and binds the user as the target KG's active manager.
            if target_kindergarten_id is None:
                raise validation_error(
                    "Manager must be assigned to a kindergarten",
                    {"kindergarten_id": "Kindergarten is required for manager role"},
                )
            from manager_assignment_service import assign_user_as_manager
            assign_user_as_manager(
                db,
                user,
                target_kindergarten_id,
                actor_id=current_user.id,
                allow_replace=bool(user_data.replace_existing_manager),
            )
        else:
            if user_data.role is not None:
                # Prevent promotion to admin
                if user_data.role == models.UserRole.ADMIN:
                    raise forbidden_error("Cannot promote users to admin role")
                # Prevent demotion of admins
                if user.role == models.UserRole.ADMIN:
                    raise forbidden_error("Cannot change role of admin users")
                user.role = user_data.role

            if user_data.status is not None:
                # Guard: ensure KG retains at least one supervisor on deactivation
                if (
                    user_data.status in [models.UserStatus.INACTIVE, models.UserStatus.SUSPENDED]
                    and user.role == models.UserRole.SUPERVISOR
                    and user.kindergarten_id
                ):
                    try:
                        validators.validate_kg_has_supervisor(db, user.kindergarten_id, exclude_user_id=user.id)
                    except validators.ValidationError as exc:
                        raise validation_error(exc.message, {"supervisor": exc.message})
                user.status = user_data.status

            if user_data.kindergarten_id is not None:
                user.kindergarten_id = user_data.kindergarten_id

    try:
        db.flush()
        after_state = model_to_dict(user)
        log_audit_event(
            db, AuditAction.USER_UPDATED, current_user, "User",
            target_ids=user.id,
            before_state=before_state,
            after_state=after_state,
            sensitivity_level=3
        )
        response_payload = {
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "role": user.role.value,
            "status": user.status.value,
            "kindergarten_id": user.kindergarten_id,
            "full_name": user.full_name,
            "phone_number": user.phone_number,
            "address": user.address,
            "nationality": user.nationality,
            "national_id": user.national_id,
            "passport_number": user.passport_number,
            "correlation_id": get_correlation_id()
        }
        db.commit()
    except Exception:
        db.rollback()
        raise

    return response_payload


@router.delete("/users/{user_id:int}", status_code=status.HTTP_204_NO_CONTENT)
@limiter.limit(settings.RATE_LIMIT_ADMIN_WRITE)
def delete_user(
    request: Request,
    user_id: int,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Delete user (Admin only) with IDOR protection.
    """

    user = db.query(models.User).filter(
        models.User.id == user_id, models.User.deleted_at.is_(None)
    ).first()

    if not user:
        raise not_found_error("User not found")

    # Prevent self-deletion
    if current_user.id == user_id:
        raise validation_error("Cannot delete your own account")

    # Prevent deleting admins
    if user.role == models.UserRole.ADMIN:
        raise forbidden_error("Cannot delete admin accounts")

    if (
        user.role == models.UserRole.MANAGER
        and user.status == models.UserStatus.ACTIVE
        and user.kindergarten_id is not None
    ):
        kindergarten = db.query(models.Kindergarten).filter(
            models.Kindergarten.id == user.kindergarten_id
        ).first()
        if kindergarten and kindergarten.status == models.KindergartenStatus.ACTIVE:
            raise conflict_error(
                "An active kindergarten must retain an active manager. Freeze the kindergarten or replace its manager atomically.",
                {"manager": "active_kindergarten_requires_manager"},
            )

    # Capture before state for audit
    before_state = model_to_dict(user)

    user.deleted_at = datetime.now(_JORDAN_TZ)
    user.deleted_by = current_user.id
    user.status = models.UserStatus.INACTIVE
    try:
        log_audit_event(
            db, AuditAction.USER_DELETED, current_user, "User",
            target_ids=user_id,
            before_state=before_state,
            sensitivity_level=3
        )
        db.commit()
    except Exception:
        db.rollback()
        raise

    return Response(status_code=status.HTTP_204_NO_CONTENT)


# =============================================================================
# Password Reset Endpoints (Hardened with Rate Limiting)
# =============================================================================

@router.post("/users/{user_id:int}/admin-reset-password")
@limiter.limit(settings.RATE_LIMIT_PASSWORD_RESET)
def admin_reset_password(
    request: Request,
    user_id: int,
    reset_data: AdminPasswordResetSchema,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Admin-initiated password reset with verification.
    Rate limited to 3 requests per minute.
    """

    user = db.query(models.User).filter(
        models.User.id == user_id, models.User.deleted_at.is_(None)
    ).first()

    if not user:
        raise not_found_error("User not found")

    # Prevent resetting admin passwords
    if user.role == models.UserRole.ADMIN:
        raise forbidden_error("Cannot reset admin passwords through this endpoint")

    # Verify admin's own password
    if not verify_password(reset_data.admin_password, current_user.hashed_password):
        log_audit_event(
            db, AuditAction.ADMIN_PASSWORD_RESET_FAILED, current_user, "User",
            target_ids=user_id,
            metadata={"reason": "Admin password verification failed"},
            sensitivity_level=3
        )
        db.commit()
        raise unauthenticated_error("Admin password verification failed")

    # Apply the same lifecycle as self-service reset: timestamps, temporary
    # credential cleanup, and revocation of every previously issued session.
    change_user_password(db, user, reset_data.new_password, commit=False)
    try:
        log_audit_event(
            db, AuditAction.ADMIN_PASSWORD_RESET, current_user, "User",
            target_ids=user_id,
            metadata={"initiated_by": current_user.username},
            sensitivity_level=3
        )
        db.commit()
    except Exception:
        db.rollback()
        raise

    return {
        "message": "Password reset successfully",
        "user_id": user_id,
        "correlation_id": get_correlation_id()
    }


@router.post("/password-reset-request")
@limiter.limit("5/hour")
def request_password_reset(
    request: Request,
    reset_request: PasswordResetRequestSchema,
    db: Session = Depends(get_db)
):
    """
    Request password reset token (self-service).
    Compatibility alias for the canonical self-service reset workflow.
    Always returns success to prevent email enumeration.
    """

    user, token = initiate_password_reset(
        request=request,
        email=str(reset_request.email),
        captcha_token=reset_request.captcha_token,
        db=db,
    )

    if user:
        # In development, token is returned to support local testing.
        if settings.ENVIRONMENT == "development":
            return {
                "message": PASSWORD_RESET_GENERIC_MESSAGE,
                "token": token,  # Only in development!
                "correlation_id": get_correlation_id()
            }

    # Always return same response to prevent enumeration
    return {
        "message": PASSWORD_RESET_GENERIC_MESSAGE,
        "correlation_id": get_correlation_id()
    }


@router.post("/password-reset-confirm")
@limiter.limit("10/hour")
def confirm_password_reset(
    request: Request,
    reset_data: PasswordResetConfirmSchema,
    db: Session = Depends(get_db)
):
    """
    Confirm password reset using token.
    Compatibility alias for the canonical self-service reset workflow.
    """

    try:
        token_record = apply_password_reset(db, reset_data.token, reset_data.new_password)
    except ValueError as exc:
        db.rollback()
        raise validation_error(str(exc)) from exc

    if not token_record:
        raise validation_error("Invalid or expired token")

    try:
        log_audit_event(
            db, AuditAction.PASSWORD_RESET_COMPLETED, token_record.user, "User",
            target_ids=token_record.user_id,
            sensitivity_level=2
        )
        db.commit()
    except Exception:
        db.rollback()
        raise

    return {
        "message": "Password reset successfully",
        "correlation_id": get_correlation_id()
    }


# =============================================================================
# MFA Management Endpoints (Admin Only)
# =============================================================================

class MFABypassSchema(BaseModel):
    """Request schema for MFA bypass (emergency admin action)"""
    user_id: int
    reason: str = "Emergency unlock"
    admin_password: str  # Verify admin identity

    model_config = ConfigDict(json_schema_extra={
        "example": {
            "user_id": 5,
            "reason": "User locked out - lost MFA device",
            "admin_password": "AdminPassword123!"
        }
    })


@router.post("/users/{user_id:int}/mfa-bypass")
@limiter.limit(settings.RATE_LIMIT_ADMIN_WRITE)
def admin_mfa_bypass(
    request: Request,
    user_id: int,
    mfa_request: MFABypassSchema,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Emergency MFA bypass for locked-out users (Admin only).
    Requires admin password verification and logs audit trail.
    
    WARNING: This is an emergency-only endpoint. Use sparingly and audit all usage.
    """

    if mfa_request.user_id != user_id:
        raise validation_error(
            "User ID in request body must match the route",
            {"user_id": "Body and route user IDs must match"},
        )

    # Self-bypass is not permitted — admin must use normal MFA reset flow
    if user_id == current_user.id:
        raise forbidden_error("Cannot bypass your own MFA. Use the standard MFA reset flow.")

    user = db.query(models.User).filter(
        models.User.id == user_id, models.User.deleted_at.is_(None)
    ).first()
    if not user:
        raise not_found_error("User not found")

    if not can_admin_access_user(current_user, user):
        raise forbidden_error("Cannot manage another admin account")

    # Verify admin's own password after confirming target exists
    if not verify_password(mfa_request.admin_password, current_user.hashed_password):
        log_audit_event(
            db, AuditAction.MFA_BYPASS_FAILED_AUTH, current_user, "User",
            target_ids=user_id,
            metadata={"reason": "Admin password verification failed"},
            sensitivity_level=3
        )
        db.commit()
        raise unauthenticated_error("Admin password verification failed")

    # Reset MFA but don't disable it - require re-setup
    user.mfa_secret = None
    user.mfa_enabled = False
    user.mfa_enrolled_at = None
    user.mfa_last_verified_at = None
    try:
        log_audit_event(
            db, AuditAction.MFA_BYPASS_INITIATED, current_user, "User",
            target_ids=user_id,
            metadata={
                "reason": mfa_request.reason,
                "initiated_by": current_user.username,
                "timestamp": datetime.now(_JORDAN_TZ).isoformat()
            },
            sensitivity_level=3
        )
        db.commit()
    except Exception:
        db.rollback()
        raise

    return {
        "message": "MFA bypass completed. User must re-enroll MFA on next login.",
        "user_id": user_id,
        "correlation_id": get_correlation_id()
    }


@router.get("/users/{user_id:int}/mfa-status")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def get_user_mfa_status(
    request: Request,
    user_id: int,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Get MFA status for a user (Admin only)."""
    user = db.query(models.User).filter(
        models.User.id == user_id, models.User.deleted_at.is_(None)
    ).first()
    if not user:
        raise not_found_error("User not found")

    if not can_admin_access_user(current_user, user):
        raise forbidden_error("Cannot access another admin account")

    return {
        "user_id": user.id,
        "username": user.username,
        "mfa_enabled": user.mfa_enabled,
        "mfa_enrolled_at": user.mfa_enrolled_at.isoformat() if user.mfa_enrolled_at else None,
        "mfa_last_verified_at": user.mfa_last_verified_at.isoformat() if user.mfa_last_verified_at else None,
        "mfa_secret_set": bool(user.mfa_secret),
        "correlation_id": get_correlation_id()
    }


# =============================================================================
# Bulk Operations (Hardened with Guardrails)
# =============================================================================

def _managers_orphaning_active_kindergartens(
    db: Session, users: Iterable[models.User]
) -> List[int]:
    """Ids of active managers whose removal would leave an active kindergarten unmanaged.

    Both bulk paths (status change and delete) need this check. They each used to
    walk the batch issuing one Kindergarten lookup per manager; this resolves the
    whole batch in a single query instead.
    """
    candidates = [
        u for u in users
        if u.role == models.UserRole.MANAGER
        and u.status == models.UserStatus.ACTIVE
        and u.kindergarten_id is not None
    ]
    if not candidates:
        return []

    active_kg_ids = {
        row[0] for row in
        db.query(models.Kindergarten.id).filter(
            models.Kindergarten.id.in_({u.kindergarten_id for u in candidates}),
            models.Kindergarten.status == models.KindergartenStatus.ACTIVE,
        ).all()
    }
    return [u.id for u in candidates if u.kindergarten_id in active_kg_ids]


@router.post("/users/bulk-status-update")
@limiter.limit(settings.RATE_LIMIT_BULK_UPDATE)
def bulk_update_status(
    request: Request,
    bulk_data: BulkStatusUpdateSchema,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Bulk update user status with confirmation for large operations.

    Features:
    - Requires confirmation token for operations affecting > 10 users
    - Supports dry-run mode for preview
    - Returns per-user success/failure results
    - IDOR protection for each target
    """

    # Check if confirmation is required
    needs_confirmation = len(bulk_data.user_ids) > settings.BULK_CONFIRMATION_THRESHOLD

    if needs_confirmation and not bulk_data.dry_run:
        if not bulk_data.confirmation_token:
            # Generate confirmation token
            token = generate_confirmation_token("bulk_status_update", bulk_data.user_ids, current_user.id)
            return JSONResponse(
                status_code=status.HTTP_200_OK,
                content={
                    "requires_confirmation": True,
                    "confirmation_token": token,
                    "affected_count": len(bulk_data.user_ids),
                    "message": f"This operation will affect {len(bulk_data.user_ids)} users. Please confirm.",
                    "correlation_id": get_correlation_id()
                }
            )

        # Verify confirmation token
        if not verify_confirmation_token(
            bulk_data.confirmation_token,
            "bulk_status_update",
            bulk_data.user_ids,
            current_user.id
        ):
            raise validation_error("Invalid confirmation token")

    # Validate access to each user
    access_result = validate_bulk_targets(
        db, current_user, bulk_data.user_ids, models.User,
        can_admin_access_user
    )

    if access_result["forbidden"]:
        log_audit_event(
            db, AuditAction.BULK_ACCESS_DENIED, current_user, "User",
            target_ids=access_result["forbidden"],
            metadata={"action": "bulk_status_update"},
            sensitivity_level=2
        )
        db.commit()

    if bulk_data.dry_run:
        return {
            "dry_run": True,
            "would_update": len(access_result["allowed"]),
            "allowed_ids": access_result["allowed"],
            "forbidden_ids": access_result["forbidden"],
            "not_found_ids": access_result["not_found"],
            "correlation_id": get_correlation_id()
        }

    # Validate manager status changes
    if bulk_data.new_status == models.UserStatus.ACTIVE:
        manager_validation_errors = []
        seen_kgs = set()
        # Batch-load all target users to avoid N+1 queries
        target_users = {
            u.id: u for u in
            db.query(models.User).filter(models.User.id.in_(access_result["allowed"])).all()
        }
        for user_id in access_result["allowed"]:
            user = target_users.get(user_id)
            if user and user.role == models.UserRole.MANAGER:
                if user.kindergarten_id is None:  # pragma: no cover — DB CHECK constraint manager_must_have_kindergarten prevents this state in both SQLite and production
                    manager_validation_errors.append({
                        "user_id": user_id,
                        "error": "Manager must be assigned to a kindergarten",
                        "field": "kindergarten_id"
                    })
                    continue
                if user.kindergarten_id in seen_kgs:
                    manager_validation_errors.append({
                        "user_id": user_id,
                        "error": f"Multiple managers in this batch for kindergarten {user.kindergarten_id}",
                        "field": "status"
                    })
                    continue
                seen_kgs.add(user.kindergarten_id)
                try:
                    validate_manager_assignment(
                        db,
                        user.role,
                        user.kindergarten_id,
                        models.UserStatus.ACTIVE,
                        user_id
                    )
                except APIError as e:
                    manager_validation_errors.append({
                        "user_id": user_id,
                        "error": e.message,
                        "field": "status"
                    })

        if manager_validation_errors:
            return JSONResponse(
                status_code=status.HTTP_409_CONFLICT,
                content={
                    "message": "Manager activation would violate kindergarten assignment rules",
                    "errors": manager_validation_errors,
                    "correlation_id": get_correlation_id()
                }
            )

    # Deactivation cannot orphan an active kindergarten.
    if bulk_data.new_status != models.UserStatus.ACTIVE:
        if 'target_users' not in locals():
            target_users = {
                u.id: u for u in
                db.query(models.User).filter(models.User.id.in_(access_result["allowed"])).all()
            }
        orphaned = _managers_orphaning_active_kindergartens(db, target_users.values())
        if orphaned:
            return JSONResponse(
                status_code=status.HTTP_409_CONFLICT,
                content={
                    "message": "Active kindergartens must retain an active manager",
                    "errors": [{"user_id": user_id, "error": "active_kindergarten_requires_manager"} for user_id in orphaned],
                    "correlation_id": get_correlation_id(),
                },
            )

    # Execute update (reuse batch-loaded users if available, else batch-load)
    succeeded = []
    failed = []
    errors = []

    if 'target_users' not in locals():
        target_users = {
            u.id: u for u in
            db.query(models.User).filter(models.User.id.in_(access_result["allowed"])).all()
        }

    for user_id in access_result["allowed"]:
        try:
            user = target_users.get(user_id)
            if user:
                before_state = model_to_dict(user)
                user.status = bulk_data.new_status
                succeeded.append(user_id)
        except (AttributeError, ValueError, TypeError) as e:  # pragma: no cover — assigning a validated enum to an ORM attribute never raises; purely defensive
            failed.append(user_id)
            errors.append({"user_id": user_id, "error": str(e)})

    try:
        log_audit_event(
            db, AuditAction.BULK_STATUS_UPDATE, current_user, "User",
            target_ids=succeeded,
            metadata={
                "new_status": bulk_data.new_status.value,
                "succeeded_count": len(succeeded),
                "failed_count": len(failed) + len(access_result["forbidden"]) + len(access_result["not_found"])
            },
            sensitivity_level=3
        )
        db.commit()
    except Exception:
        db.rollback()
        raise

    return {
        "message": f"Updated {len(succeeded)} users",
        "succeeded_count": len(succeeded),
        "failed_count": len(failed) + len(access_result["forbidden"]) + len(access_result["not_found"]),
        "succeeded_ids": succeeded,
        "failed_ids": failed + access_result["forbidden"] + access_result["not_found"],
        "errors": errors,
        "forbidden_ids": access_result["forbidden"],
        "not_found_ids": access_result["not_found"],
        "correlation_id": get_correlation_id()
    }


@router.post("/users/bulk-delete")
@limiter.limit(settings.RATE_LIMIT_BULK_DELETE)
def bulk_delete_users(
    request: Request,
    bulk_data: BulkDeleteSchema,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Bulk delete users with confirmation and IDOR protection.

    Features:
    - Requires confirmation token for ALL bulk deletes (destructive)
    - Prevents deleting admin accounts
    - Returns detailed results
    """

    # Check for admin accounts in the list
    admin_users = db.query(models.User).filter(
        models.User.id.in_(bulk_data.user_ids),
        models.User.role == models.UserRole.ADMIN
    ).all()

    if admin_users:
        raise forbidden_error(f"Cannot delete admin accounts: {[u.username for u in admin_users]}")

    # Bulk delete ALWAYS requires confirmation
    if not bulk_data.dry_run:
        if not bulk_data.confirmation_token:
            token = generate_confirmation_token("bulk_delete", bulk_data.user_ids, current_user.id)
            return JSONResponse(
                status_code=status.HTTP_200_OK,
                content={
                    "requires_confirmation": True,
                    "confirmation_token": token,
                    "affected_count": len(bulk_data.user_ids),
                    "warning": "This action is IRREVERSIBLE. Please confirm to proceed.",
                    "correlation_id": get_correlation_id()
                }
            )

        if not verify_confirmation_token(
            bulk_data.confirmation_token,
            "bulk_delete",
            bulk_data.user_ids,
            current_user.id
        ):
            raise validation_error("Invalid confirmation token")

    # Validate access
    access_result = validate_bulk_targets(
        db, current_user, bulk_data.user_ids, models.User,
        can_admin_access_user
    )

    if bulk_data.dry_run:
        return {
            "dry_run": True,
            "would_delete": len(access_result["allowed"]),
            "allowed_ids": access_result["allowed"],
            "forbidden_ids": access_result["forbidden"],
            "not_found_ids": access_result["not_found"],
            "correlation_id": get_correlation_id()
        }

    # Execute soft-delete — batch-load users to avoid N+1
    _now = datetime.now(_JORDAN_TZ)
    deleted_ids = []
    target_users = {
        u.id: u for u in
        db.query(models.User).filter(models.User.id.in_(access_result["allowed"])).all()
    }
    orphaned = _managers_orphaning_active_kindergartens(db, target_users.values())
    if orphaned:
        return JSONResponse(
            status_code=status.HTTP_409_CONFLICT,
            content={
                "message": "Active kindergartens must retain an active manager",
                "errors": [{"user_id": user_id, "error": "active_kindergarten_requires_manager"} for user_id in orphaned],
                "correlation_id": get_correlation_id(),
            },
        )
    for user_id in access_result["allowed"]:
        user = target_users.get(user_id)
        if user:
            user.deleted_at = _now
            user.deleted_by = current_user.id
            user.status = models.UserStatus.INACTIVE
            deleted_ids.append(user_id)

    try:
        log_audit_event(
            db, AuditAction.BULK_USER_DELETE, current_user, "User",
            target_ids=deleted_ids,
            metadata={
                "deleted_count": len(deleted_ids),
                "forbidden_count": len(access_result["forbidden"]),
                "not_found_count": len(access_result["not_found"])
            },
            sensitivity_level=3
        )
        db.commit()
    except Exception:
        db.rollback()
        raise

    return {
        "message": f"Deleted {len(deleted_ids)} users",
        "deleted_count": len(deleted_ids),
        "deleted_ids": deleted_ids,
        "forbidden_ids": access_result["forbidden"],
        "not_found_ids": access_result["not_found"],
        "correlation_id": get_correlation_id()
    }


@router.post("/users/bulk-create")
@limiter.limit(settings.RATE_LIMIT_BULK_CREATE)
def bulk_create_users(
    request: Request,
    bulk_data: BulkCreateSchema,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Bulk create users with per-user validation and error reporting.

    Features:
    - Validates each user against schema
    - Reports per-user errors
    - Supports dry-run mode
    - Prevents creating admin accounts
    """

    # Validate manager assignments for the entire batch
    manager_errors = validate_bulk_manager_assignments(db, [user.__dict__ for user in bulk_data.users])
    if manager_errors:
        return JSONResponse(
            status_code=status.HTTP_200_OK if bulk_data.dry_run else status.HTTP_400_BAD_REQUEST,
            content={
                "dry_run": bulk_data.dry_run,
                "total": len(bulk_data.users),
                "succeeded_count": 0,
                "failed_count": len(manager_errors),
                "errors": manager_errors,
                "message": "Manager validation failed",
                "correlation_id": get_correlation_id()
            }
        )

    succeeded = []
    failed = []
    errors = []

    # Pre-load existing usernames and emails to avoid N+1 queries
    incoming_usernames = {u.username for u in bulk_data.users}
    incoming_emails = {u.email for u in bulk_data.users if u.email is not None}
    existing_usernames = set(
        row[0] for row in db.query(models.User.username)
        .filter(models.User.username.in_(incoming_usernames)).all()
    ) if incoming_usernames else set()
    existing_emails = set(
        row[0] for row in db.query(models.User.email)
        .filter(models.User.email.in_(incoming_emails)).all()
    ) if incoming_emails else set()
    batch_usernames: Set[str] = set()
    batch_emails: Set[str] = set()

    for i, user_data in enumerate(bulk_data.users):
        row_num = i + 1

        if user_data.username in batch_usernames:
            failed.append(row_num)
            errors.append({
                "row": row_num,
                "field": "username",
                "error": "Duplicate username in bulk request"
            })
            continue
        batch_usernames.add(user_data.username)

        if user_data.email is not None:
            if user_data.email in batch_emails:
                failed.append(row_num)
                errors.append({
                    "row": row_num,
                    "field": "email",
                    "error": "Duplicate email in bulk request"
                })
                continue
            batch_emails.add(user_data.email)

        # Prevent creating admins
        if user_data.role == models.UserRole.ADMIN:
            failed.append(row_num)
            errors.append({
                "row": row_num,
                "field": "role",
                "error": "Cannot create admin accounts through bulk create"
            })
            continue

        # Check for existing username/email using pre-loaded sets
        if user_data.username in existing_usernames:
            failed.append(row_num)
            errors.append({
                "row": row_num,
                "field": "username",
                "error": "Username already exists"
            })
            continue

        # Check email uniqueness only if email is provided
        if user_data.email is not None and user_data.email in existing_emails:
                failed.append(row_num)
                errors.append({
                    "row": row_num,
                    "field": "email",
                    "error": "Email already exists"
                })
                continue

        if not bulk_data.dry_run:
            try:
                new_user = models.User(
                    username=user_data.username,
                    email=user_data.email,
                    hashed_password=get_password_hash(user_data.password),
                    role=user_data.role,
                    kindergarten_id=user_data.kindergarten_id,
                    status=models.UserStatus.ACTIVE,
                    must_change_password=user_data.role in {
                        models.UserRole.MANAGER,
                        models.UserRole.SUPERVISOR,
                    },
                )
                for field in (
                    "full_name", "phone_number", "address", "nationality",
                    "national_id", "passport_number",
                ):
                    value = getattr(user_data, field, None)
                    if value is not None:
                        setattr(new_user, field, value)
                db.add(new_user)
                db.flush()
                if new_user.role == models.UserRole.SUPERVISOR:
                    validators.ensure_supervisor_profile(
                        db, new_user, new_user.kindergarten_id
                    )
                db.flush()
                succeeded.append({"row": row_num, "id": new_user.id, "username": new_user.username})
            except IntegrityError as exc:
                logger.warning(
                    "Bulk user import row %s lost a uniqueness race (%s)",
                    row_num,
                    type(exc).__name__,
                )
                failed.append(row_num)
                errors.append({
                    "row": row_num,
                    "field": "conflict",
                    "error": "Username or email became unavailable"
                })
            except (SQLAlchemyError, AttributeError, ValueError, KeyError) as exc:
                logger.warning(
                    "Bulk user import row %s failed (%s)", row_num, type(exc).__name__
                )
                failed.append(row_num)
                errors.append({
                    "row": row_num,
                    "field": "unknown",
                    "error": "User could not be created"
                })
        else:
            succeeded.append({"row": row_num, "username": user_data.username})

    if not bulk_data.dry_run:
        try:
            log_audit_event(
                db, AuditAction.BULK_USER_CREATE, current_user, "User",
                target_ids=[s["id"] for s in succeeded if "id" in s],
                metadata={
                    "created_count": len(succeeded),
                    "failed_count": len(failed)
                },
                sensitivity_level=3
            )
            db.commit()
        except Exception:
            db.rollback()
            raise

    return {
        "dry_run": bulk_data.dry_run,
        "message": f"{'Would create' if bulk_data.dry_run else 'Created'} {len(succeeded)} users",
        "total": len(bulk_data.users),
        "succeeded_count": len(succeeded),
        "failed_count": len(failed),
        "succeeded": succeeded,
        "errors": errors,
        "correlation_id": get_correlation_id()
    }


# =============================================================================
# CSV Import with Per-Row Validation
# =============================================================================

@router.post("/users/import-csv")
@limiter.limit(settings.RATE_LIMIT_CSV_IMPORT)
async def import_users_csv(
    request: Request,
    file: UploadFile = File(...),
    dry_run: bool = Query(False, description="Preview import without applying"),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Import users from CSV with comprehensive validation.

    Features:
    - Per-row validation with detailed error reporting
    - CSV injection protection
    - Dry-run mode for preview
    - Downloadable error report
    """

    if not file.filename.endswith('.csv'):
        raise validation_error("File must be a CSV")

    # P1-C: Enforce a 20 MB file-size limit before reading into memory.
    # file.size is set by the ASGI layer when the Content-Length header is
    # present; we also enforce the limit after reading to guard against chunked
    # uploads that omit the header.
    _MAX_CSV_BYTES = 20 * 1024 * 1024  # 20 MB
    if file.size is not None and file.size > _MAX_CSV_BYTES:
        raise validation_error(
            f"CSV file is too large ({file.size / 1_048_576:.1f} MB). "
            f"Maximum allowed size is 20 MB."
        )

    # Read and decode CSV
    try:
        contents = await file.read()
        # Second guard: catches chunked uploads that omit Content-Length.
        if len(contents) > _MAX_CSV_BYTES:  # pragma: no cover — requires a >20 MB streaming upload without Content-Length header; first guard handles normal cases
            raise validation_error(
                f"CSV file exceeds the 20 MB maximum ({len(contents) / 1_048_576:.1f} MB)."
            )
        decoded = contents.decode('utf-8-sig')  # Handle BOM
    except (UnicodeDecodeError, OSError) as e:
        raise validation_error(f"Could not read file: {str(e)}")

    # Parse CSV
    reader = csv.DictReader(io.StringIO(decoded))

    total_rows = 0
    succeeded = []
    failed = []
    errors: List[CSVRowError] = []
    created_ids = []

    # Bound validation, hashing, and ORM work independently of compressed size.
    _MAX_CSV_ROWS = 1_000
    _MAX_CSV_COLUMNS = 50
    _MAX_CSV_CELL_CHARS = 10_000
    required_fields = {'username', 'email', 'password', 'role'}
    header_fields = set(reader.fieldnames or [])

    if len(reader.fieldnames or []) > _MAX_CSV_COLUMNS:
        raise validation_error(
            f"CSV header exceeds the {_MAX_CSV_COLUMNS} column maximum."
        )

    # Validate headers
    missing_fields = required_fields - header_fields
    if missing_fields:
        raise validation_error(
            f"Missing required columns: {', '.join(missing_fields)}",
            {field: "Column required" for field in missing_fields}
        )

    all_rows = []
    try:
        for row_number, row in enumerate(reader, start=1):
            if row_number > _MAX_CSV_ROWS:
                raise validation_error(
                    f"CSV contains more than the {_MAX_CSV_ROWS:,} row maximum."
                )
            actual_column_count = len(reader.fieldnames or []) + len(row.get(None) or [])
            if actual_column_count > _MAX_CSV_COLUMNS:
                raise validation_error(
                    f"CSV row {row_number + 1} exceeds the {_MAX_CSV_COLUMNS} column maximum."
                )
            if any(len(str(value or "")) > _MAX_CSV_CELL_CHARS for value in row.values()):
                raise validation_error(
                    f"CSV row {row_number + 1} contains a cell longer than "
                    f"{_MAX_CSV_CELL_CHARS:,} characters."
                )
            all_rows.append(row)
    except csv.Error as exc:
        raise validation_error(f"Could not parse CSV: {exc}")

    # Batch duplicate check — one query for all usernames/emails in the file
    all_usernames = [sanitize_csv_cell(str(r.get('username', '')).strip()) for r in all_rows if r.get('username')]
    all_emails = [sanitize_csv_cell(str(r.get('email', '')).strip()) for r in all_rows if r.get('email')]
    existing_conflict_users = db.query(models.User).filter(
        or_(
            models.User.username.in_(all_usernames),
            models.User.email.in_(all_emails),
        )
    ).all() if (all_usernames or all_emails) else []
    taken_usernames = {u.username for u in existing_conflict_users}
    taken_emails = {u.email for u in existing_conflict_users}

    # Batch active-manager check — one query for all manager KG IDs in the file
    csv_manager_kg_ids = set()
    for r in all_rows:
        if r.get('role', '').strip().upper() == 'MANAGER' and r.get('kindergarten_id'):
            try:
                csv_manager_kg_ids.add(int(r['kindergarten_id']))
            except (ValueError, TypeError):
                pass
    existing_manager_by_kg: dict = {}
    if csv_manager_kg_ids:
        for mgr in db.query(models.User).filter(
            models.User.kindergarten_id.in_(csv_manager_kg_ids),
            models.User.role == models.UserRole.MANAGER,
            models.User.status == models.UserStatus.ACTIVE,
        ).all():
            if mgr.kindergarten_id not in existing_manager_by_kg:
                existing_manager_by_kg[mgr.kindergarten_id] = mgr

    manager_kgs_in_csv = set()

    for row in all_rows:
        total_rows += 1
        row_num = total_rows + 1  # Account for header row

        # Sanitize all cells
        sanitized_row = {k: sanitize_csv_cell(str(v).strip()) if v else '' for k, v in row.items()}

        # Validate role
        role_str = sanitized_row.get('role', '').upper()
        if role_str not in ['MANAGER', 'SUPERVISOR', 'PARENT']:
            errors.append(CSVRowError(
                row_number=row_num,
                field='role',
                error_code='INVALID_ROLE',
                message=f"Invalid role: {role_str}. Must be MANAGER, SUPERVISOR, or PARENT"
            ))
            failed.append(row_num)
            continue

        # Validate and create
        try:
            user_data = UserCreateSchema(
                username=sanitized_row.get('username', ''),
                email=sanitized_row.get('email', ''),
                password=sanitized_row.get('password', ''),
                role=models.UserRole(role_str),
                kindergarten_id=int(sanitized_row['kindergarten_id']) if sanitized_row.get('kindergarten_id') else None
            )
        except (ValueError, AttributeError, KeyError, TypeError) as e:
            if hasattr(e, 'errors'):
                for err in e.errors():
                    errors.append(CSVRowError(
                        row_number=row_num,
                        field='.'.join(str(loc) for loc in err.get('loc', [])),
                        error_code='VALIDATION_ERROR',
                        message=err.get('msg', str(err))
                    ))
            else:
                errors.append(CSVRowError(
                    row_number=row_num,
                    error_code='PARSE_ERROR',
                    message=str(e)
                ))
            failed.append(row_num)
            continue

        # Check for existing — use pre-fetched sets (no per-row query)
        dup_field = None
        if user_data.username in taken_usernames:
            dup_field = 'username'
        elif user_data.email in taken_emails:
            dup_field = 'email'
        if dup_field:
            errors.append(CSVRowError(
                row_number=row_num,
                field=dup_field,
                error_code='DUPLICATE',
                message=f"{dup_field.capitalize()} already exists"
            ))
            failed.append(row_num)
            continue

        if user_data.role == models.UserRole.SUPERVISOR and not user_data.kindergarten_id:
            errors.append(CSVRowError(
                row_number=row_num,
                field='kindergarten_id',
                error_code='VALIDATION_ERROR',
                message="Supervisor must belong to a kindergarten",
            ))
            failed.append(row_num)
            continue

        if user_data.role == models.UserRole.MANAGER:
            if user_data.kindergarten_id in manager_kgs_in_csv:
                errors.append(CSVRowError(
                    row_number=row_num,
                    field='kindergarten_id',
                    error_code='CONFLICT',
                    message=f"Multiple managers assigned to kindergarten {user_data.kindergarten_id} in this CSV"
                ))
                failed.append(row_num)
                continue

            # Check for existing active manager — use pre-fetched dict (no per-row query)
            if user_data.kindergarten_id is None:
                errors.append(CSVRowError(
                    row_number=row_num,
                    field='kindergarten_id',
                    error_code='VALIDATION_ERROR',
                    message="Manager must be assigned to a kindergarten",
                ))
                failed.append(row_num)
                continue
            existing_mgr = existing_manager_by_kg.get(user_data.kindergarten_id)
            if existing_mgr:
                errors.append(CSVRowError(
                    row_number=row_num,
                    field='kindergarten_id',
                    error_code='CONFLICT',
                    message=(
                        f"Each kindergarten can have only one active manager. "
                        f"Kindergarten {user_data.kindergarten_id} already has an active manager (ID: {existing_mgr.id})."
                    ),
                ))
                failed.append(row_num)
                continue

            manager_kgs_in_csv.add(user_data.kindergarten_id)

        if not dry_run:
            new_user = models.User(
                username=user_data.username,
                email=user_data.email,
                hashed_password=get_password_hash(user_data.password),
                role=user_data.role,
                kindergarten_id=user_data.kindergarten_id,
                status=models.UserStatus.ACTIVE,
                must_change_password=(user_data.role in [
                    models.UserRole.MANAGER, models.UserRole.SUPERVISOR
                ]),
            )
            db.add(new_user)
            db.flush()
            created_ids.append(new_user.id)

        succeeded.append(row_num)

    if not dry_run and created_ids:
        try:
            log_audit_event(
                db, AuditAction.CSV_IMPORT, current_user, "User",
                target_ids=created_ids,
                metadata={
                    "filename": file.filename,
                    "total_rows": total_rows,
                    "succeeded": len(succeeded),
                    "failed": len(failed)
                },
                sensitivity_level=3
            )
            db.commit()
        except Exception:
            db.rollback()
            raise

    result = CSVImportResult(
        total_rows=total_rows,
        succeeded=len(succeeded),
        failed=len(failed),
        errors=errors,
        created_ids=created_ids if not dry_run else [],
        dry_run=dry_run
    )

    return {
        "total_rows": result.total_rows,
        "succeeded": result.succeeded,
        "failed": result.failed,
        "errors": [e.model_dump() for e in result.errors],
        "created_ids": result.created_ids,
        "dry_run": result.dry_run,
        "correlation_id": get_correlation_id()
    }


def _escape_csv_formula(value: Any) -> str:
    return escape_csv_formula(value)


class _CSVErrorReportBody(BaseModel):
    errors: List[Dict[str, Any]] = Field(..., description="Error list from CSV import response")


@router.post("/users/import-csv/error-report")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def download_csv_error_report(
    request: Request,
    body: _CSVErrorReportBody,
    current_user: models.User = Depends(require_admin),
):

    error_list = body.errors
    headers = ['Row Number', 'Field', 'Error Code', 'Message']
    rows = []
    for err in error_list:
        rows.append([
            err.get('row_number', ''),
            err.get('field', ''),
            err.get('error_code', ''),
            err.get('message', ''),
        ])

    return export_service.generate_csv_response(headers, rows, f"import_errors_{datetime.now(_JORDAN_TZ).date()}.csv")


# =============================================================================
# Contact Messages  (P1-D: previously missing — template existed, no backend)
# =============================================================================

class ContactMessageItem(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    name: str
    email: str
    phone: Optional[str] = None
    subject: Optional[str] = None
    message: str
    is_resolved: bool
    submitted_at: Optional[str] = None
    resolved_at: Optional[str] = None


class ContactMessagesListResponse(BaseModel):
    messages: List[ContactMessageItem]
    total: int
    page: int
    page_size: int
    total_pages: int


@router.get("/contact-messages", response_model=ContactMessagesListResponse)
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def list_contact_messages(
    request: Request,
    q: Optional[str] = Query(None, max_length=100, description="Search by name or subject"),
    status_filter: Optional[str] = Query(None, pattern="^(open|resolved)$",
                                         description="Filter: 'open' or 'resolved'"),
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=100),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    List contact-form submissions with optional search and status filter.
    Admin-only.  Pagination enforced.
    """
    query = db.query(models.ContactMessage)

    if q:
        term = f"%{q}%"
        query = query.filter(
            or_(
                models.ContactMessage.name.ilike(term),
                models.ContactMessage.subject.ilike(term),
                models.ContactMessage.email.ilike(term),
            )
        )

    if status_filter == "open":
        query = query.filter(models.ContactMessage.is_resolved.is_(False))
    elif status_filter == "resolved":
        query = query.filter(models.ContactMessage.is_resolved.is_(True))

    total = query.count()
    offset = (page - 1) * page_size
    rows = (
        query.order_by(models.ContactMessage.submitted_at.desc())
        .offset(offset)
        .limit(page_size)
        .all()
    )

    def _row(m: models.ContactMessage) -> ContactMessageItem:
        return ContactMessageItem(
            id=m.id,
            name=m.name,
            email=m.email,
            phone=m.phone,
            subject=m.subject,
            message=m.message,
            is_resolved=m.is_resolved,
            submitted_at=m.submitted_at.isoformat() if m.submitted_at else None,
            resolved_at=m.resolved_at.isoformat() if m.resolved_at else None,
        )

    return ContactMessagesListResponse(
        messages=[_row(m) for m in rows],
        total=total,
        page=page,
        page_size=page_size,
        total_pages=max(1, (total + page_size - 1) // page_size),
    )


@router.post("/contact-messages/{message_id}/resolve",
             status_code=status.HTTP_200_OK)
@limiter.limit(settings.RATE_LIMIT_ADMIN_WRITE)
def resolve_contact_message(
    request: Request,
    message_id: int,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Mark a contact-form submission as resolved.
    Admin-only.  Idempotent — resolving an already-resolved message is a no-op.
    """

    msg = db.query(models.ContactMessage).filter(
        models.ContactMessage.id == message_id
    ).first()
    if not msg:
        raise not_found_error("Contact message not found")

    if not msg.is_resolved:
        msg.is_resolved = True
        msg.resolved_by_id = current_user.id
        msg.resolved_at = datetime.now(_JORDAN_TZ)
        try:
            log_audit_event(
                db, AuditAction.CONTACT_MESSAGE_RESOLVED, current_user, "ContactMessage",
                target_ids=message_id,
                metadata={"message_id": message_id},
                sensitivity_level=1,
            )
            db.commit()
        except Exception:
            db.rollback()
            raise

    return {"message": "Success"}


# ---------------------------------------------------------------------------
# Safety analytics
# ---------------------------------------------------------------------------

@router.get("/safety/analytics")
def safety_analytics(
    incident_type: Optional[str] = None,
    severity: Optional[str] = None,
    classification: Optional[str] = None,
    parent_informed: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    kindergarten_id: Optional[int] = None,
    child_id: Optional[int] = None,
    governorate: Optional[str] = None,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    validators.validate_admin_role(current_user)

    q = db.query(models.Incident)

    if kindergarten_id:
        q = q.filter(models.Incident.kindergarten_id == kindergarten_id)
    if child_id:
        q = q.filter(models.Incident.child_id == child_id)
    if governorate:
        q = q.join(
            models.Kindergarten, models.Kindergarten.id == models.Incident.kindergarten_id
        ).filter(governorate_filter(models.Kindergarten.governorate, governorate))

    if incident_type:
        try:
            type_val = models.IncidentType(incident_type)
            q = q.filter(models.Incident.type == type_val)
        except ValueError:
            pass

    if severity:
        try:
            sev_val = models.SeverityLevel(severity)
            q = q.filter(models.Incident.severity_level == sev_val)
        except ValueError:
            pass

    if classification:
        q = q.filter(models.Incident.classification == classification)

    if parent_informed is not None and parent_informed != "":
        q = q.filter(models.Incident.parent_informed == (parent_informed.lower() == "true"))

    from_dt = None
    if date_from:
        try:
            from_dt = datetime.fromisoformat(date_from)
            q = q.filter(models.Incident.occurred_at >= from_dt)
        except ValueError:
            pass

    if date_to:
        try:
            to_dt = datetime.fromisoformat(date_to)
            q = q.filter(models.Incident.occurred_at <= to_dt)
        except ValueError:
            pass

    incidents = q.all()
    total = len(incidents)

    by_severity: dict[str, int] = {}
    by_type: dict[str, int] = {}
    by_classification: dict[str, int] = {}
    open_count = 0
    closed_count = 0
    parent_informed_count = 0
    parent_not_informed_count = 0
    by_month: dict[str, int] = {}
    kg_incident_ids: dict[int, list[int]] = {}
    child_incident_counts: dict[int, int] = {}

    for inc in incidents:
        sev_key = inc.severity_level.value if inc.severity_level else "UNKNOWN"
        by_severity[sev_key] = by_severity.get(sev_key, 0) + 1
        type_key = inc.type.value if inc.type else "UNKNOWN"
        by_type[type_key] = by_type.get(type_key, 0) + 1
        cls_key = inc.classification or "UNKNOWN"
        by_classification[cls_key] = by_classification.get(cls_key, 0) + 1

        if inc.closed_at is not None:
            closed_count += 1
        else:
            open_count += 1

        if inc.parent_informed:
            parent_informed_count += 1
        else:
            parent_not_informed_count += 1

        # Jordan month, not UTC: an incident at 22:00 Jordan on the 31st is 19:00Z
        # the same day, but one at 23:30 Jordan on the 1st is 20:30Z on the 1st —
        # while 00:30 Jordan on the 1st is 21:30Z on the *previous* month's last day.
        # Bucketing on the raw UTC value silently moves month-boundary incidents.
        month_key = to_jordan_date(inc.occurred_at).strftime("%Y-%m-01")
        by_month[month_key] = by_month.get(month_key, 0) + 1

        kg_incident_ids.setdefault(inc.kindergarten_id, []).append(inc.id)
        child_incident_counts[inc.child_id] = child_incident_counts.get(inc.child_id, 0) + 1

    trend = [{"month": month, "count": count} for month, count in sorted(by_month.items())]

    by_kindergarten: list[dict[str, Any]] = []
    if kg_incident_ids:
        kgs = db.query(models.Kindergarten).filter(
            models.Kindergarten.id.in_(kg_incident_ids.keys())
        ).all()
        kg_counts = {kg_id: len(ids) for kg_id, ids in kg_incident_ids.items()}
        avg_per_kg = sum(kg_counts.values()) / len(kg_counts)
        critical_kg_ids = {
            inc.kindergarten_id for inc in incidents
            if inc.severity_level == models.SeverityLevel.CRITICAL
        }
        for kg in kgs:
            count = kg_counts.get(kg.id, 0)
            by_kindergarten.append({
                "kindergarten_id": kg.id,
                "name_ar": kg.name_ar,
                "name_en": kg.name_en,
                "count": count,
                "is_high_risk": count > avg_per_kg or kg.id in critical_kg_ids,
            })
        by_kindergarten.sort(key=lambda x: x["count"], reverse=True)

    repeated_children: list[dict[str, Any]] = []
    repeated_ids = [cid for cid, count in child_incident_counts.items() if count > 1]
    if repeated_ids:
        children = db.query(models.Child).filter(models.Child.id.in_(repeated_ids)).all()
        for child in children:
            full_name = f"{child.first_name} {child.last_name}".strip()
            repeated_children.append({
                "id": child.id,
                "name_ar": full_name,
                "name_en": full_name,
                "count": child_incident_counts[child.id],
            })
        repeated_children.sort(key=lambda x: x["count"], reverse=True)

    return {
        "total": total,
        "by_severity": by_severity,
        "by_type": by_type,
        "by_classification": by_classification,
        "open": open_count,
        "closed": closed_count,
        "parent_informed": parent_informed_count,
        "parent_not_informed": parent_not_informed_count,
        "trend": trend,
        "by_kindergarten": by_kindergarten,
        "repeated_children": repeated_children,
    }


# =============================================================================
# Admin Messaging (Targeted Announcements)
# =============================================================================

class AdminRecipientRole(str, enum.Enum):
    MANAGER = "MANAGER"
    SUPERVISOR = "SUPERVISOR"
    PARENT = "PARENT"


class AdminMessageTargetMode(str, enum.Enum):
    ALL_USERS = "ALL_USERS"
    ALL_MANAGERS = "ALL_MANAGERS"
    ALL_SUPERVISORS = "ALL_SUPERVISORS"
    ALL_PARENTS = "ALL_PARENTS"
    GOVERNORATE = "GOVERNORATE"
    KINDERGARTENS = "KINDERGARTENS"


class AdminMessageTarget(BaseModel):
    mode: AdminMessageTargetMode
    roles: Optional[List[AdminRecipientRole]] = None
    governorates: Optional[List[str]] = None
    kindergarten_ids: Optional[List[int]] = None
    search: Optional[str] = None

    model_config = ConfigDict(extra="ignore")


class AdminMessageCreate(BaseModel):
    subject: str
    message_body: str
    target: AdminMessageTarget
    allow_replies: bool = True

    model_config = ConfigDict(extra="ignore")


class AdminMessagePreviewRequest(BaseModel):
    subject: Optional[str] = None
    message_body: Optional[str] = None
    target: AdminMessageTarget
    allow_replies: bool = True
    page: int = 1
    page_size: int = 10

    model_config = ConfigDict(extra="ignore")


class AdminRecipientSummary(BaseModel):
    id: int
    display_name: str
    role: str
    email: Optional[str] = None
    phone: Optional[str] = None
    kindergarten_id: Optional[int] = None
    kindergarten_name: Optional[str] = None
    governorate: Optional[str] = None


class AdminPaginationMeta(BaseModel):
    page: int
    page_size: int
    total: int
    total_pages: int
    has_next: bool
    has_prev: bool


class AdminRecipientListResponse(BaseModel):
    items: List[AdminRecipientSummary]
    pagination: AdminPaginationMeta


class AdminMessageResponse(BaseModel):
    id: int
    thread_type: str
    subject: Optional[str]
    message_body: str
    created_at: datetime
    recipient_count: int
    warnings: List[str] = []


def _dedupe_int_list(values: Optional[List[int]]) -> List[int]:
    cleaned: List[int] = []
    for value in values or []:
        try:
            cleaned.append(int(value))
        except (TypeError, ValueError):
            continue
    return list(dict.fromkeys([v for v in cleaned if v]))


def _validate_jordan_governorates(governorates: Optional[List[str]]) -> List[str]:
    """Validate and deduplicate governorates, returning canonical Arabic forms."""
    canonical: List[str] = []
    for gov in governorates or []:
        if not gov:
            continue
        try:
            canonical.append(validators.validate_jordan_governorate(gov))
        except validators.ValidationError:
            raise APIError(
                status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                code=ErrorCode.VALIDATION_ERROR,
                message="Invalid governorate",
                fields={"governorates": "invalid"},
            )
    return list(dict.fromkeys(canonical))




def _normalize_governorates(governorates: Optional[List[str]]) -> List[str]:
    normalized = _validate_jordan_governorates(governorates)
    for ar_value in list(normalized):
        if ar_value in settings.JORDAN_GOVERNORATES:
            idx = settings.JORDAN_GOVERNORATES.index(ar_value)
            if idx < len(settings.JORDAN_GOVERNORATES_ENGLISH):
                normalized.append(settings.JORDAN_GOVERNORATES_ENGLISH[idx])
    return list(dict.fromkeys(normalized))




def _canonical_governorates(governorates: Optional[List[str]]) -> List[str]:
    return _validate_jordan_governorates(governorates)


def _build_search_filter(search: Optional[str], columns: List[Any]):
    """Return a compound OR/AND filter across columns.

    For each whitespace-separated token in the query we generate an OR clause
    that spans all supplied columns.  Arabic search variants (alef normalisation,
    definite-article stripping, etc.) are added automatically via
    ``build_arabic_search_terms`` so that, for example, searching for
    "العربي" also matches rows containing "عربي" and vice-versa.
    """
    if not search or not columns:
        return None

    # Expand tokens to include Arabic-normalised variants
    tokens = build_arabic_search_terms(search)
    if not tokens:
        return None

    token_clauses = []
    for token in tokens:
        pattern = f"%{token}%"
        token_clauses.append(or_(*(column.ilike(pattern) for column in columns)))
    # All original tokens must match (AND logic); normalised variants are OR-ed
    # inside each token clause above, so a 2-word query still requires both words.
    return and_(*token_clauses)


def _build_staff_recipient_query(
    db: Session,
    roles: List[models.UserRole],
    governorates: List[str],
    kindergarten_ids: List[int],
    search_term: Optional[str]
) -> Any:
    staff_roles = [role for role in roles if role in {models.UserRole.MANAGER, models.UserRole.SUPERVISOR}]
    query = db.query(models.User.id).filter(
        models.User.status == models.UserStatus.ACTIVE,
        models.User.role.in_(staff_roles)
    )
    joined_kindergartens = False

    if kindergarten_ids:
        query = query.filter(models.User.kindergarten_id.in_(kindergarten_ids))

    if governorates:
        query = query.join(
            models.Kindergarten,
            models.User.kindergarten_id == models.Kindergarten.id
        ).filter(models.Kindergarten.governorate.in_(governorates))
        joined_kindergartens = True

    search_cols = [models.User.username, models.User.email]
    if search_term:
        if not joined_kindergartens:
            query = query.outerjoin(
                models.Kindergarten,
                models.User.kindergarten_id == models.Kindergarten.id,
            )
        search_cols.extend([
            models.Kindergarten.name_ar,
            models.Kindergarten.name_en
        ])
        search_filter = _build_search_filter(search_term, search_cols)
        if search_filter is not None:
            query = query.filter(search_filter)

    return query.distinct()


def _build_parent_recipient_query(
    db: Session,
    governorates: List[str],
    kindergarten_ids: List[int],
    search_term: Optional[str]
) -> Any:
    query = (
        db.query(models.User.id)
        .join(models.ParentProfile, models.ParentProfile.user_id == models.User.id)
        .filter(
            models.User.status == models.UserStatus.ACTIVE,
            models.User.role == models.UserRole.PARENT
        )
    )

    search_cols = [
        models.User.username,
        models.User.email,
        models.ParentProfile.first_name,
        models.ParentProfile.second_name,
        models.ParentProfile.last_name,
        models.ParentProfile.first_name_en,
        models.ParentProfile.last_name_en,
        models.ParentProfile.phone_number
    ]

    if search_term:
        query = query.outerjoin(
            models.Child,
            models.Child.parent_id == models.ParentProfile.id
        ).outerjoin(
            models.EnrollmentApplication,
            models.EnrollmentApplication.child_id == models.Child.id
        ).outerjoin(
            models.Kindergarten,
            models.Kindergarten.id == models.EnrollmentApplication.kindergarten_id
        )
        search_cols.extend([
            models.Kindergarten.name_ar,
            models.Kindergarten.name_en
        ])
        search_filter = _build_search_filter(search_term, search_cols)
        if search_filter is not None:
            query = query.filter(search_filter)

    if kindergarten_ids or governorates:
        enrollment_query = (
            db.query(models.ParentProfile.user_id)
            .join(
                models.Child,
                models.Child.parent_id == models.ParentProfile.id
            )
            .join(
                models.EnrollmentApplication,
                models.EnrollmentApplication.child_id == models.Child.id
            )
            .join(
                models.Kindergarten,
                models.Kindergarten.id == models.EnrollmentApplication.kindergarten_id
            )
            .filter(models.EnrollmentApplication.status.in_(ACTIVE_ENROLLMENT_STATUSES))
        )

        if kindergarten_ids:
            enrollment_query = enrollment_query.filter(
                models.EnrollmentApplication.kindergarten_id.in_(kindergarten_ids)
            )
        if governorates:
            enrollment_query = enrollment_query.filter(models.Kindergarten.governorate.in_(governorates))

        enrolled_parent_ids = enrollment_query.distinct().subquery()
        enrolled_parent_ids_select = select(enrolled_parent_ids.c.user_id)

        if governorates and not kindergarten_ids:
            active_parent_ids = (
                db.query(models.ParentProfile.user_id)
                .join(
                    models.Child,
                    models.Child.parent_id == models.ParentProfile.id
                )
                .join(
                    models.EnrollmentApplication,
                    models.EnrollmentApplication.child_id == models.Child.id
                )
                .filter(
                    models.EnrollmentApplication.status.in_(ACTIVE_ENROLLMENT_STATUSES)
                )
                .distinct()
                .subquery()
            )
            active_parent_ids_select = select(active_parent_ids.c.user_id)
            query = query.filter(or_(
                models.ParentProfile.user_id.in_(enrolled_parent_ids_select),
                and_(
                    ~models.ParentProfile.user_id.in_(active_parent_ids_select),
                    models.ParentProfile.home_governorate.in_(governorates)
                )
            ))
        else:
            query = query.filter(models.ParentProfile.user_id.in_(enrolled_parent_ids_select))

    return query.distinct()


def _count_admin_recipients(
    db: Session,
    roles: List[models.UserRole],
    governorates: List[str],
    kindergarten_ids: List[int],
    search: Optional[str]
) -> int:
    recipient_query = _build_admin_recipient_id_query(
        db, roles, governorates, kindergarten_ids, search
    )
    count_stmt = select(func.count()).select_from(
        recipient_query.order_by(None).subquery()
    )
    return db.execute(count_stmt).scalar_one()


def _build_admin_recipient_id_query(
    db: Session,
    roles: List[models.UserRole],
    governorates: List[str],
    kindergarten_ids: List[int],
    search: Optional[str],
) -> Any:
    """Build the one canonical recipient-ID relation used by count/page/send."""
    search_term = (search or "").strip()
    statements = []
    if any(
        role in {models.UserRole.MANAGER, models.UserRole.SUPERVISOR}
        for role in roles
    ):
        statements.append(
            _build_staff_recipient_query(
                db, roles, governorates, kindergarten_ids, search_term
            ).statement
        )
    if models.UserRole.PARENT in roles:
        statements.append(
            _build_parent_recipient_query(
                db, governorates, kindergarten_ids, search_term
            ).statement
        )

    if not statements:
        return db.query(models.User.id).filter(models.User.id.is_(None))
    combined = (
        union_all(*statements).subquery()
        if len(statements) > 1
        else statements[0].subquery()
    )
    return db.query(combined.c.id).distinct().order_by(combined.c.id)


def _resolve_parent_governorates(db: Session, parent_ids: List[int]) -> Dict[int, List[str]]:
    """Return a map of parent user IDs to the governorates associated with them."""
    if not parent_ids:
        return {}

    governorates_map: Dict[int, List[str]] = defaultdict(list)

    enrollment_rows = db.query(
        models.ParentProfile.user_id,
        models.Kindergarten.governorate
    ).join(
        models.Child,
        models.Child.parent_id == models.ParentProfile.id
    ).join(
        models.EnrollmentApplication,
        models.EnrollmentApplication.child_id == models.Child.id
    ).join(
        models.Kindergarten,
        models.Kindergarten.id == models.EnrollmentApplication.kindergarten_id
    ).filter(
        models.ParentProfile.user_id.in_(parent_ids),
        models.EnrollmentApplication.status.in_(ACTIVE_ENROLLMENT_STATUSES)
    ).distinct().all()

    for user_id, governorate in enrollment_rows:
        if governorate:
            governorates_map[user_id].append(governorate)

    profile_rows = db.query(
        models.ParentProfile.user_id,
        models.ParentProfile.home_governorate
    ).filter(
        models.ParentProfile.user_id.in_(parent_ids)
    ).all()

    for user_id, home_governorate in profile_rows:
        if home_governorate:
            governorates_map[user_id].append(home_governorate)

    return {
        user_id: list(dict.fromkeys(governorates))
        for user_id, governorates in governorates_map.items()
    }


def _parent_matches_governorates(
    db: Session,
    parent_user_id: int,
    target_governorates: List[str]
) -> bool:
    """
    Checks if a parent user is associated with any of the target governorates.
    A parent is associated if their home_governorate matches or any of their
    active enrollments' kindergartens are in the target governorates.
    """
    if not target_governorates:
        return False

    parent_profile = db.query(models.ParentProfile).filter(
        models.ParentProfile.user_id == parent_user_id
    ).first()

    if not parent_profile:
        return False

    # Check home_governorate
    if parent_profile.home_governorate and parent_profile.home_governorate in target_governorates:
        return True

    # Check enrolled kindergartens' governorates
    kindergarten_governorates_count = db.query(models.Kindergarten.governorate).join(
        models.EnrollmentApplication,
        models.EnrollmentApplication.kindergarten_id == models.Kindergarten.id
    ).join(
        models.Child,
        models.Child.id == models.EnrollmentApplication.child_id
    ).filter(
        models.Child.parent_id == parent_profile.id,
        models.EnrollmentApplication.status.in_(ACTIVE_ENROLLMENT_STATUSES),
        models.Kindergarten.governorate.in_(target_governorates)
    ).count()

    return kindergarten_governorates_count > 0


def _build_recipient_breakdowns(
    db: Session,
    user_ids: List[int]
) -> Tuple[Dict[str, int], Dict[str, int], Dict[int, int]]:
    """Return role, governorate, and kindergarten counts for the given users."""
    if not user_ids:
        return {}, {}, {}

    rows = db.query(
        models.User.id,
        models.User.role,
        models.User.kindergarten_id
    ).filter(models.User.id.in_(user_ids)).all()

    role_counts: Counter = Counter()
    governorate_counts: Counter = Counter()
    kindergarten_counts: Counter = Counter()
    parent_ids: List[int] = []
    kindergarten_ids: Set[int] = set()

    for user_id, role, kg_id in rows:
        role_counts[role.value] += 1
        if role == models.UserRole.PARENT:
            parent_ids.append(user_id)
        if kg_id:
            kindergarten_counts[kg_id] += 1
            kindergarten_ids.add(kg_id)

    parent_governorates = _resolve_parent_governorates(db, parent_ids) if parent_ids else {}
    for parent_id, gov_list in parent_governorates.items():
        seen: Set[str] = set()
        for gov in gov_list:
            if not gov or gov in seen:  # pragma: no cover — _resolve_parent_governorates pre-filters empty govs and deduplicates with dict.fromkeys; kept for defensive safety
                continue
            governorate_counts[gov] += 1
            seen.add(gov)

    if kindergarten_ids:
        kindergarten_map = {
            kg.id: kg
            for kg in db.query(models.Kindergarten).filter(models.Kindergarten.id.in_(kindergarten_ids)).all()
        }
        for kg_id, count in list(kindergarten_counts.items()):
            kg = kindergarten_map.get(kg_id)
            if kg and kg.governorate:
                governorate_counts[kg.governorate] += count

    return (
        dict(role_counts),
        dict(governorate_counts),
        dict(kindergarten_counts)
    )


def _normalize_recipient_roles(roles: Optional[List[AdminRecipientRole]]) -> List[models.UserRole]:
    normalized: List[models.UserRole] = []
    for role in roles or []:
        role_value = role.value if hasattr(role, "value") else str(role)
        role_value = role_value.strip().upper()
        try:
            role_enum = models.UserRole(role_value)
        except ValueError:
            raise validation_error("Invalid role", fields={"roles": "invalid"})
        if role_enum == models.UserRole.ADMIN:
            raise validation_error("Invalid role", fields={"roles": "invalid"})
        normalized.append(role_enum)
    return list(dict.fromkeys(normalized))


def _target_roles_for_mode(target: AdminMessageTarget) -> List[models.UserRole]:
    if target.mode == AdminMessageTargetMode.ALL_USERS:
        return [models.UserRole.MANAGER, models.UserRole.SUPERVISOR, models.UserRole.PARENT]
    if target.mode == AdminMessageTargetMode.ALL_MANAGERS:
        return [models.UserRole.MANAGER]
    if target.mode == AdminMessageTargetMode.ALL_SUPERVISORS:
        return [models.UserRole.SUPERVISOR]
    if target.mode == AdminMessageTargetMode.ALL_PARENTS:
        return [models.UserRole.PARENT]
    roles = _normalize_recipient_roles(target.roles)
    if not roles:
        raise validation_error("Recipient roles are required", fields={"roles": "required"})
    return roles


def _resolve_admin_recipient_ids(
    db: Session,
    roles: List[models.UserRole],
    governorates: Optional[List[str]] = None,
    kindergarten_ids: Optional[List[int]] = None,
    search: Optional[str] = None,
    limit: Optional[int] = None,
    offset: int = 0,
) -> List[int]:
    governorates = governorates or []
    kindergarten_ids = kindergarten_ids or []
    query = _build_admin_recipient_id_query(
        db, roles, governorates, kindergarten_ids, search
    )
    if offset > 0:
        query = query.offset(offset)
    if limit is not None and limit > 0:
        query = query.limit(limit)
    return [row[0] for row in query.all()]


def _fetch_admin_recipient_summaries(db: Session, user_ids: List[int]) -> List[AdminRecipientSummary]:
    if not user_ids:
        return []

    users = db.query(models.User).filter(models.User.id.in_(user_ids)).all()
    user_map = {user.id: user for user in users}

    parent_ids = [user.id for user in users if user.role == models.UserRole.PARENT]
    parent_profiles = {}
    if parent_ids:
        parent_profiles = {
            profile.user_id: profile
            for profile in db.query(models.ParentProfile).filter(models.ParentProfile.user_id.in_(parent_ids)).all()
        }

    parent_kindergartens: Dict[int, List[str]] = {}
    parent_governorates = _resolve_parent_governorates(db, parent_ids) if parent_ids else {}
    if parent_ids:
        rows = db.query(
            models.ParentProfile.user_id,
            models.Kindergarten.name_ar,
            models.Kindergarten.name_en
        ).join(
            models.Child,
            models.Child.parent_id == models.ParentProfile.id
        ).join(
            models.EnrollmentApplication,
            models.EnrollmentApplication.child_id == models.Child.id
        ).join(
            models.Kindergarten,
            models.Kindergarten.id == models.EnrollmentApplication.kindergarten_id
        ).filter(
            models.ParentProfile.user_id.in_(parent_ids),
            models.EnrollmentApplication.status.in_(ACTIVE_ENROLLMENT_STATUSES)
        ).all()

        for user_id, name_ar, name_en in rows:
            label = name_ar or name_en or ""
            if label:
                parent_kindergartens.setdefault(user_id, []).append(label)

    kindergarten_ids = {user.kindergarten_id for user in users if user.kindergarten_id}
    kindergarten_map = {}
    if kindergarten_ids:
        kindergarten_map = {
            kg.id: kg
            for kg in db.query(models.Kindergarten).filter(models.Kindergarten.id.in_(kindergarten_ids)).all()
        }

    summaries: List[AdminRecipientSummary] = []
    for user_id in user_ids:
        user = user_map.get(user_id)
        if not user:
            continue
        display_name = user.username
        phone = None
        kindergarten_id = user.kindergarten_id
        kindergarten_name = None
        governorate = None

        if user.role == models.UserRole.PARENT:
            profile = parent_profiles.get(user.id)
            if profile:
                name_parts = [profile.first_name, profile.second_name, profile.last_name]
                display_name = " ".join(part for part in name_parts if part)
                phone = profile.phone_number
                governorates = parent_governorates.get(user.id) or []
                governorate = ", ".join(dict.fromkeys(governorates)) if governorates else profile.home_governorate
                kg_names = parent_kindergartens.get(user.id) or []
                kindergarten_name = ", ".join(dict.fromkeys(kg_names)) if kg_names else None
        else:
            kg = kindergarten_map.get(user.kindergarten_id)
            if kg:
                kindergarten_name = kg.name_ar or kg.name_en
                governorate = kg.governorate

        summaries.append(AdminRecipientSummary(
            id=user.id,
            display_name=display_name or user.username,
            role=user.role.value,
            email=user.email,
            phone=phone,
            kindergarten_id=kindergarten_id,
            kindergarten_name=kindergarten_name,
            governorate=governorate
        ))

    return summaries


@router.get("/message-recipients", response_model=AdminRecipientListResponse)
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def list_message_recipients(
    request: Request,
    roles: Optional[List[str]] = Query(None),
    governorates: Optional[List[str]] = Query(None),
    kindergarten_ids: Optional[List[int]] = Query(None),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(settings.DEFAULT_PAGE_SIZE, ge=1, le=settings.MAX_PAGE_SIZE),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    role_values = _normalize_recipient_roles(roles)
    if not role_values:
        role_values = [models.UserRole.MANAGER, models.UserRole.SUPERVISOR, models.UserRole.PARENT]

    governorate_values = _normalize_governorates(governorates)
    kindergarten_id_values = _dedupe_int_list(kindergarten_ids)
    if kindergarten_id_values:
        ensure_kindergartens_exist(db, kindergarten_id_values)

    search_term = (search or "").strip()
    if not search_term:
        search_term = None

    total = _count_admin_recipients(
        db=db,
        roles=role_values,
        governorates=governorate_values,
        kindergarten_ids=kindergarten_id_values,
        search=search_term
    )
    total_pages = max(1, (total + page_size - 1) // page_size)
    offset = (page - 1) * page_size
    page_ids = _resolve_admin_recipient_ids(
        db=db,
        roles=role_values,
        governorates=governorate_values,
        kindergarten_ids=kindergarten_id_values,
        search=search_term,
        limit=page_size,
        offset=offset,
    )
    items = _fetch_admin_recipient_summaries(db, page_ids)

    return AdminRecipientListResponse(
        items=items,
        pagination=AdminPaginationMeta(
            page=page,
            page_size=page_size,
            total=total,
            total_pages=total_pages,
            has_next=page < total_pages,
            has_prev=page > 1
        )
    )


@router.post("/messages", status_code=status.HTTP_201_CREATED, response_model=AdminMessageResponse)
@limiter.limit(settings.RATE_LIMIT_ADMIN_WRITE)
def create_admin_message(
    request: Request,
    payload: AdminMessageCreate,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):

    subject = (payload.subject or "").strip()
    message_body = (payload.message_body or "").strip()
    if not subject:
        raise validation_error("Subject is required", fields={"subject": "required"})
    if not message_body:
        raise validation_error("Message body is required", fields={"message_body": "required"})

    subject = validators.sanitize_input(subject)
    message_body = validators.sanitize_input(message_body)

    target = payload.target
    roles = _target_roles_for_mode(target)
    governorate_values = _normalize_governorates(target.governorates)
    canonical_governorates = _canonical_governorates(target.governorates)
    kindergarten_id_values = _dedupe_int_list(target.kindergarten_ids)

    if target.mode == AdminMessageTargetMode.GOVERNORATE:
        if not canonical_governorates:
            raise validation_error("Governorate is required", fields={"governorates": "required"})
    if target.mode == AdminMessageTargetMode.KINDERGARTENS:
        if not kindergarten_id_values:
            raise validation_error("Kindergarten selection is required", fields={"kindergarten_ids": "required"})
    if kindergarten_id_values:
        ensure_kindergartens_exist(db, kindergarten_id_values)

    search_term = (target.search or "").strip()
    if not search_term:
        search_term = None

    total_recipients = _count_admin_recipients(
        db=db,
        roles=roles,
        governorates=governorate_values,
        kindergarten_ids=kindergarten_id_values,
        search=search_term
    )

    if total_recipients == 0:
        raise validation_error("No matching recipients found", fields={"recipients": "empty"})
    if total_recipients > settings.MAX_MESSAGE_RECIPIENTS:
        raise validation_error(
            f"Too many recipients ({total_recipients}). Maximum allowed: {settings.MAX_MESSAGE_RECIPIENTS}",
            fields={"recipients": "too_many"}
        )

    recipient_ids = _resolve_admin_recipient_ids(
        db=db,
        roles=roles,
        governorates=governorate_values,
        kindergarten_ids=kindergarten_id_values,
        search=search_term,
        limit=settings.MAX_MESSAGE_RECIPIENTS
    )

    target_kindergarten_id = None
    if target.mode == AdminMessageTargetMode.KINDERGARTENS and len(kindergarten_id_values) == 1:
        target_kindergarten_id = kindergarten_id_values[0]

    message = models.Message(
        thread_type=models.MessageThreadType.ANNOUNCEMENT,
        sender_id=current_user.id,
        kindergarten_id=target_kindergarten_id,
        subject=subject,
        message_body=message_body,
        recipient_id=None,
        allow_replies=payload.allow_replies,
        target_mode=target.mode.value,
        target_roles=[role.value for role in roles],
        target_governorates=canonical_governorates or None,
        target_kindergarten_ids=kindergarten_id_values or None,
        target_search=search_term,
        recipient_count=len(recipient_ids)
    )

    warnings: List[str] = []
    staged_notifications: List[models.Notification] = []
    recipient_users = db.query(models.User).filter(models.User.id.in_(recipient_ids)).all()
    chunk_size = 500
    try:
        db.add(message)
        db.flush()
        message.thread_id = message.id
        recipients = [
            models.MessageRecipient(
                message_id=message.id,
                recipient_user_id=recipient_id,
                status="queued"
            )
            for recipient_id in recipient_ids
        ]
        for start in range(0, len(recipients), chunk_size):
            db.bulk_save_objects(recipients[start:start + chunk_size])
        log_audit_event(
            db=db,
            action=AuditAction.ADMIN_MESSAGE_SENT,
            actor=current_user,
            target_type="Message",
            target_ids=message.id,
            metadata={
                "recipient_count": len(recipient_ids),
                "target": {
                    "mode": target.mode.value,
                    "roles": [role.value for role in roles],
                    "governorates": canonical_governorates,
                    "kindergarten_ids": kindergarten_id_values,
                    "search": search_term
                }
            },
            sensitivity_level=2
        )

        notification_result = create_message_notifications(
            db,
            message,
            recipient_users,
            caller_owns_transaction=True,
        )
        if notification_result and recipient_users:
            if isinstance(notification_result, list):
                staged_notifications = notification_result
            log_audit_event(
                db=db,
                action=AuditAction.MESSAGE_NOTIFICATIONS_QUEUED,
                actor=current_user,
                target_type="Message",
                target_ids=message.id,
                metadata={"recipient_count": len(recipient_users)},
                sensitivity_level=1,
            )
        else:
            warnings.append("Message notifications are disabled; status will be reviewed later.")
            log_audit_event(
                db=db,
                action=AuditAction.MESSAGE_NOTIFICATIONS_SKIPPED,
                actor=current_user,
                target_type="Message",
                target_ids=message.id,
                metadata={"reason": "notifications_disabled"},
                sensitivity_level=1,
            )
        db.commit()
    except Exception as exc:
        db.rollback()
        logger.error("Failed to persist admin message: %s", exc)
        raise APIError(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            code=ErrorCode.INTERNAL_ERROR,
            message="Failed to create message",
            details={"reason": "Unable to write message to the database"}
        )
    else:
        db.refresh(message)

    if staged_notifications:
        try:
            dispatch_message_notification_tasks(staged_notifications)
        except Exception as exc:
            error_type = type(exc).__name__
            warnings.append("Notification dispatch is pending retry.")
            try:
                for notification in staged_notifications:
                    notification.error_message = (
                        f"Dispatch scheduling failed ({error_type}); pending retry"
                    )
                db.commit()
            except Exception as record_exc:
                db.rollback()
                logger.error(
                    "Failed to record Admin notification dispatch error for message %s error_type=%s",
                    message.id,
                    type(record_exc).__name__,
                )
            logger.warning(
                "Failed to dispatch Admin notifications for message %s; %s rows remain pending: %s",
                message.id,
                len(staged_notifications),
                error_type,
            )

    return AdminMessageResponse(
        id=message.id,
        thread_type=message.thread_type.value,
        subject=message.subject,
        message_body=message.message_body,
        created_at=message.created_at,
        recipient_count=len(recipient_ids),
        warnings=warnings
    )


# =============================================================================
# Admin Messaging - Additional Endpoints (Preview & Options)
# =============================================================================

class AdminRecipientPreviewResponse(BaseModel):
    """Response for recipient preview endpoint"""
    total_count: int
    has_more: bool = False
    sample_recipients: List[AdminRecipientSummary] = []
    by_role: Dict[str, int] = {}
    by_governorate: Optional[Dict[str, int]] = None
    by_kindergarten: Optional[Dict[int, int]] = None

    model_config = ConfigDict(json_schema_extra={
        "example": {
            "total_count": 150,
            "has_more": True,
            "by_role": {"PARENT": 120, "MANAGER": 15, "SUPERVISOR": 15},
            "by_governorate": {"عمان": 80, "إربد": 70},
            "by_kindergarten": {"12": 40, "21": 50},
            "sample_recipients": [
                {"id": 42, "display_name": "أحمد أحمد", "role": "PARENT", "email": "ahmed@test.com", "governorate": "عمان"}
            ]
        }
    })


class GovernorateOption(BaseModel):
    """Governorate option for dropdown"""
    id: str
    name_ar: str
    name_en: Optional[str] = None


class GovernorateOptionsResponse(BaseModel):
    """Response containing governorate options"""
    governorates: List[GovernorateOption]


@router.get("/options/governorates", response_model=GovernorateOptionsResponse)
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def list_governorate_options(
    request: Request,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Get list of available governorates for message targeting.
    Admin only endpoint.
    """
    try:
        from services.jordan_locations import get_all_governorates
    except (ImportError, AttributeError):
        source_options = [
            {
                "name_ar": gov,
                "name_en": (
                    settings.JORDAN_GOVERNORATES_ENGLISH[idx]
                    if idx < len(settings.JORDAN_GOVERNORATES_ENGLISH)
                    else gov
                ),
            }
            for idx, gov in enumerate(settings.JORDAN_GOVERNORATES)
        ]
    else:
        source_options = get_all_governorates()

    options = []
    try:
        for g in source_options:
            options.append(GovernorateOption(
                id=g["name_ar"],
                name_ar=g["name_ar"],
                name_en=g["name_en"]
            ))
    except (ValidationError, KeyError, TypeError, ValueError) as exc:
        raise APIError(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            code=ErrorCode.VALIDATION_ERROR,
            message="Governorate options are invalid",
            fields={"governorates": "invalid"},
        ) from exc

    options.sort(key=lambda opt: opt.name_ar)
    return GovernorateOptionsResponse(governorates=options)


@router.get("/message-recipients/preview", response_model=AdminRecipientPreviewResponse)
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def preview_message_recipients(
    request: Request,
    mode: AdminMessageTargetMode,
    roles: Optional[List[str]] = Query(None),
    governorates: Optional[List[str]] = Query(None),
    kindergarten_ids: Optional[List[int]] = Query(None),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=settings.MAX_PAGE_SIZE),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Preview recipient count and sample recipients before sending a message.
    Admin only endpoint.
    """
    if mode == AdminMessageTargetMode.ALL_USERS:
        role_values = [models.UserRole.MANAGER, models.UserRole.SUPERVISOR, models.UserRole.PARENT]
    elif mode == AdminMessageTargetMode.ALL_MANAGERS:
        role_values = [models.UserRole.MANAGER]
    elif mode == AdminMessageTargetMode.ALL_SUPERVISORS:
        role_values = [models.UserRole.SUPERVISOR]
    elif mode == AdminMessageTargetMode.ALL_PARENTS:
        role_values = [models.UserRole.PARENT]
    elif roles:
        role_values = _normalize_recipient_roles(roles)
    else:
        raise validation_error("Roles are required for this targeting mode", fields={"roles": "required"})

    governorate_values = _normalize_governorates(governorates) if governorates else []
    kindergarten_id_values = _dedupe_int_list(kindergarten_ids) if kindergarten_ids else []

    if mode == AdminMessageTargetMode.KINDERGARTENS and not kindergarten_id_values:
        raise validation_error("Kindergarten selection is required", fields={"kindergarten_ids": "required"})
    if mode == AdminMessageTargetMode.GOVERNORATE and not governorate_values:
        raise validation_error("Governorate is required", fields={"governorates": "required"})

    if kindergarten_id_values:
        ensure_kindergartens_exist(db, kindergarten_id_values)

    search_term = (search or "").strip() if search else None
    total_count = _count_admin_recipients(
        db=db,
        roles=role_values,
        governorates=governorate_values,
        kindergarten_ids=kindergarten_id_values,
        search=search_term
    )

    all_recipient_ids = _resolve_admin_recipient_ids(
        db=db,
        roles=role_values,
        governorates=governorate_values if governorate_values else None,
        kindergarten_ids=kindergarten_id_values if kindergarten_id_values else None,
        search=search_term,
    )

    sample_ids = _resolve_admin_recipient_ids(
        db=db,
        roles=role_values,
        governorates=governorate_values,
        kindergarten_ids=kindergarten_id_values,
        search=search_term,
        limit=min(page_size, 5),
        offset=(page - 1) * page_size,
    )
    sample_recipients = _fetch_admin_recipient_summaries(db, sample_ids)

    role_breakdown, governorate_breakdown, kindergarten_breakdown = _build_recipient_breakdowns(
        db, all_recipient_ids
    )

    target_metadata = {
        "mode": mode.value,
        "roles": [role.value for role in role_values],
        "governorates": governorate_values or None,
        "kindergarten_ids": kindergarten_id_values or None,
        "search": search_term
    }

    log_audit_event(
        db=db,
        action=AuditAction.ADMIN_MESSAGE_PREVIEW,
        actor=current_user,
        target_type="Preview",
        target_ids=None,
        metadata={
            "recipient_count": total_count,
            "target": target_metadata
        },
        sensitivity_level=1
    )
    db.commit()

    return AdminRecipientPreviewResponse(
        total_count=total_count,
        has_more=page * page_size < total_count,
        sample_recipients=sample_recipients,
        by_role=role_breakdown,
        by_governorate=governorate_breakdown or None,
        by_kindergarten=kindergarten_breakdown or None
    )


@router.post("/messages/preview", response_model=AdminRecipientListResponse)
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def preview_admin_message_post(
    request: Request,
    payload: AdminMessagePreviewRequest,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):

    target = payload.target
    roles = _target_roles_for_mode(target)
    governorate_values = _normalize_governorates(target.governorates)
    canonical_governorates = _canonical_governorates(target.governorates)
    kindergarten_id_values = _dedupe_int_list(target.kindergarten_ids)

    if target.mode == AdminMessageTargetMode.GOVERNORATE:
        if not canonical_governorates:
            raise validation_error("Governorate is required", fields={"governorates": "required"})
    if target.mode == AdminMessageTargetMode.KINDERGARTENS and not kindergarten_id_values:
        raise validation_error("Kindergarten selection is required", fields={"kindergarten_ids": "required"})

    if kindergarten_id_values:
        ensure_kindergartens_exist(db, kindergarten_id_values)

    search_term = (target.search or "").strip()
    if not search_term:
        search_term = None

    total = _count_admin_recipients(
        db=db,
        roles=roles,
        governorates=governorate_values,
        kindergarten_ids=kindergarten_id_values,
        search=search_term,
    )

    page = max(1, payload.page)
    page_size = max(1, min(payload.page_size, settings.MAX_PAGE_SIZE))
    offset = (page - 1) * page_size
    page_ids = _resolve_admin_recipient_ids(
        db=db,
        roles=roles,
        governorates=governorate_values if governorate_values else None,
        kindergarten_ids=kindergarten_id_values if kindergarten_id_values else None,
        search=search_term,
        limit=page_size,
        offset=offset,
    )

    total_pages = max(1, (total + page_size - 1) // page_size) if total else 1
    items = _fetch_admin_recipient_summaries(db, page_ids)

    target_metadata = {
        "mode": target.mode.value,
        "roles": [role.value for role in roles],
        "governorates": governorate_values or None,
        "kindergarten_ids": kindergarten_id_values or None,
        "search": search_term
    }

    log_audit_event(
        db=db,
        action=AuditAction.ADMIN_MESSAGE_PREVIEW,
        actor=current_user,
        target_type="Preview",
        target_ids=None,
        metadata={
            "recipient_count": total,
            "target": target_metadata
        },
        sensitivity_level=1
    )
    db.commit()

    return AdminRecipientListResponse(
        items=items,
        pagination=AdminPaginationMeta(
            page=page,
            page_size=page_size,
            total=total,
            total_pages=max(1, total_pages),
            has_next=page < total_pages,
            has_prev=page > 1
        )
    )


@router.get("/options/kindergartens")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def list_kindergarten_options(
    request: Request,
    governorate: Optional[str] = Query(None),
    status: Optional[models.KindergartenStatus] = None,
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(500, ge=1, le=500),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Get list of kindergartens for message targeting.
    Supports filtering by governorate, status, and search term.
    Admin only endpoint.
    """
    query = db.query(models.Kindergarten).filter(
        models.Kindergarten.status == (status or models.KindergartenStatus.ACTIVE)
    )

    if governorate:
        try:
            gov_normalized = validators.validate_jordan_governorate(governorate)
            query = query.filter(governorate_filter(models.Kindergarten.governorate, gov_normalized))
        except validators.ValidationError:
            pass  # Ignore invalid governorate in filter

    if search:
        search = search[:100]
        search_term = f"%{search}%"
        query = query.filter(or_(
            models.Kindergarten.name_ar.ilike(search_term),
            models.Kindergarten.name_en.ilike(search_term),
            models.Kindergarten.district.ilike(search_term),
            models.Kindergarten.contact_phone.ilike(search_term)
        ))

    # Get total count
    total = query.count()

    # Apply pagination
    offset = (page - 1) * page_size
    kindergartens = query.order_by(
        models.Kindergarten.governorate,
        models.Kindergarten.name_ar
    ).offset(offset).limit(page_size).all()

    return {
        "kindergartens": [
            {
                "id": kg.id,
                "name": kg.name_ar or kg.name_en or f"حضانة {kg.id}",
                "name_ar": kg.name_ar,
                "name_en": kg.name_en,
                "governorate": kg.governorate,
                "district": kg.district,
                "status": kg.status.value,
                "contact_phone": kg.contact_phone
            }
            for kg in kindergartens
        ],
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": max(1, (total + page_size - 1) // page_size)
        }
    }


# =============================================================================
# Performance Monitoring Endpoints
# =============================================================================

@router.get("/performance/metrics")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def get_performance_metrics(
    request: Request,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Get comprehensive performance metrics."""
    from performance_monitor import get_performance_report

    try:
        return get_performance_report()
    except (RuntimeError, AttributeError, TypeError, ValueError) as e:
        logger.error(f"Failed to get performance metrics: {e}")
        raise HTTPException(status_code=500, detail="Failed to retrieve performance metrics")


@router.get("/performance/requests")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def get_request_metrics(
    request: Request,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Get request performance metrics."""
    from performance_monitor import performance_monitor

    try:
        return performance_monitor.get_request_metrics()
    except (RuntimeError, AttributeError, TypeError, ValueError) as e:
        logger.error(f"Failed to get request metrics: {e}")
        raise HTTPException(status_code=500, detail="Failed to retrieve request metrics")


@router.get("/performance/database")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def get_database_metrics(
    request: Request,
    limit: int = Query(100, ge=1, le=1000, description="Number of recent queries to return"),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Get database query performance metrics."""
    from performance_monitor import performance_monitor

    try:
        return {
            "recent_queries": performance_monitor.get_db_metrics(limit),
            "slow_queries": performance_monitor.get_slow_queries(2.0),
        }
    except (RuntimeError, AttributeError, TypeError, ValueError) as e:
        logger.error(f"Failed to get database metrics: {e}")
        raise HTTPException(status_code=500, detail="Failed to retrieve database metrics")


@router.get("/performance/system")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def get_system_metrics(
    request: Request,
    limit: int = Query(50, ge=1, le=500, description="Number of recent system metrics to return"),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Get system performance metrics."""
    from performance_monitor import performance_monitor

    try:
        return performance_monitor.get_system_metrics(limit)
    except (RuntimeError, AttributeError, TypeError, ValueError) as e:
        logger.error(f"Failed to get system metrics: {e}")
        raise HTTPException(status_code=500, detail="Failed to retrieve system metrics")


# =============================================================================
# Admin Dashboard Models
# =============================================================================

class KPITrendMeta(BaseModel):
    """Trend/comparison metadata for a single dashboard KPI, keyed to match `kpis`."""
    value: float
    previous_value: float
    change: float
    change_pct: Optional[float] = None
    trend: str = "flat"
    # good | warning | critical | neutral | unavailable.
    # "neutral" = a raw count with no target, so no good/bad judgment is implied.
    status: str = "good"
    # False when the previous period had no data (previous_value == 0), so a
    # change like "+635" is not a real trend and must not be shown as one.
    baseline_available: bool = True
    # False when the metric could not be computed (e.g. data quality with zero
    # eligible kindergartens) — distinct from a genuine 0 value.
    measurable: bool = True


class DataQualityReason(BaseModel):
    """A single derivable reason contributing to a low data_quality_score."""
    id: str
    label_ar: str
    label_en: str
    count: int

class DashboardSummary(BaseModel):
    """Summary statistics for admin dashboard"""
    attendance_today: int = 0
    pending_applications: int = 0
    pending_daily_reports: int = 0
    recent_incidents: int = 0
    attendance_rate: float = 0.0

class DashboardSystemOverview(BaseModel):
    """System overview statistics"""
    total_kindergartens: int = 0
    active_kindergartens: int = 0
    total_users: int = 0

class DashboardKindergarten(BaseModel):
    """Kindergarten data for dashboard table"""
    id: int
    name_ar: Optional[str] = None
    name_en: Optional[str] = None
    status: str
    license_status: str
    governorate: Optional[str] = None
    district: Optional[str] = None
    enrollments: int = 0
    attendance_today: int = 0
    pending_reports: int = 0
    capacity_utilization: float = 0.0
    total_children: int = 0
    active_children: int = 0
    last_report_date: Optional[str] = None

class DashboardChartPoint(BaseModel):
    """Data point for dashboard charts"""
    date: str
    value: Union[int, float]
    label: Optional[str] = None

class DashboardCharts(BaseModel):
    """Charts data for dashboard"""
    attendance: List[DashboardChartPoint] = []
    enrollment: Dict[str, Any] = {}
    incidents: List[DashboardChartPoint] = []

class DashboardAlert(BaseModel):
    """System alert for dashboard — bilingual title/message fields are canonical; title/message are Arabic fallbacks."""
    id: str
    title: str
    message: str
    title_ar: str = ""
    title_en: str = ""
    message_ar: str = ""
    message_en: str = ""
    severity: str
    timestamp: str
    category: Optional[str] = None
    type: Optional[str] = None
    priority: Optional[str] = None
    kindergarten_id: Optional[int] = None

class ActivityItem(BaseModel):
    """Recent audit-log activity item — bilingual."""
    type: str = "system_update"
    message_ar: str = ""
    message_en: str = ""
    timestamp: str
    user_name: Optional[str] = None
    user_role: Optional[str] = None
    module_ar: Optional[str] = None
    module_en: Optional[str] = None
    entity_type: Optional[str] = None
    entity_label_ar: Optional[str] = None
    entity_label_en: Optional[str] = None
    status: str = "success"
    severity: str = "low"

class AdminDashboardResponse(BaseModel):
    """Complete admin dashboard response"""
    summary: DashboardSummary
    system_overview: DashboardSystemOverview
    kindergartens: List[DashboardKindergarten]
    charts: DashboardCharts
    alerts: List[DashboardAlert]
    kpis: Dict[str, Optional[float]]
    kpi_trends: Dict[str, KPITrendMeta]
    data_quality_reasons: List[DataQualityReason]
    recent_activity: List[ActivityItem]
    generated_at: str


# =============================================================================
# Recent Activity Taxonomy — shared by the dashboard's recent-activity feed
# and the filterable /api/admin/dashboard/activity endpoint (Phase 4).
# =============================================================================

_SIDEBAR_MODULE_LABELS: Dict[str, tuple] = {
    "management":         ("إدارة النظام", "System Management"),
    "operations":         ("العمليات", "Operations"),
    "reports-analytics":  ("التقارير والتحليل", "Reports & Analytics"),
    "governance":         ("الحوكمة", "Governance"),
    "settings":           ("الإعدادات", "Settings"),
}

# action -> (message_ar, message_en, activity_type, module_id)
_ACTIVITY_MAP: Dict[str, tuple] = {
    AuditAction.LOGIN_SUCCESS:              ("تسجيل دخول ناجح إلى النظام", "Successful login",              "user_login",       "settings"),
    AuditAction.LOGIN_FAILED:               ("محاولة تسجيل دخول فاشلة",   "Failed login attempt",           "user_login",       "settings"),
    AuditAction.LOGOUT:                     ("تسجيل خروج من النظام",      "Logged out",                     "user_logout",      "settings"),
    AuditAction.ACCESS_DENIED:               ("رفض الوصول",                "Access denied",                  "user_login",       "settings"),
    AuditAction.USER_CREATED:               ("إضافة مستخدم جديد",         "New user added",                 "user_create",      "management"),
    AuditAction.USER_UPDATED:               ("تحديث بيانات مستخدم",       "User data updated",              "user_update",      "management"),
    AuditAction.USER_DELETED:               ("حذف مستخدم",                "User deleted",                   "user_delete",      "management"),
    AuditAction.BULK_USER_CREATE:           ("إضافة مستخدمين بالجملة",    "Bulk users added",               "user_create",      "management"),
    AuditAction.BULK_USER_DELETE:           ("حذف مستخدمين بالجملة",      "Bulk users deleted",             "user_delete",      "management"),
    AuditAction.KINDERGARTEN_CREATED:       ("إضافة حضانة جديدة",          "New kindergarten added",         "data_create",      "management"),
    AuditAction.KINDERGARTEN_UPDATED:       ("تحديث بيانات حضانة",         "Kindergarten data updated",      "data_update",      "management"),
    AuditAction.KINDERGARTEN_DELETED:       ("حذف حضانة",                  "Kindergarten deleted",           "data_delete",      "management"),
    AuditAction.DAILY_REPORT_CREATED:       ("إنشاء تقرير يومي",          "Daily report created",           "data_create",      "reports-analytics"),
    AuditAction.DAILY_REPORT_EDITED:        ("تعديل تقرير يومي",          "Daily report edited",            "data_update",      "reports-analytics"),
    AuditAction.DAILY_REPORT_DELETED:       ("حذف تقرير يومي",            "Daily report deleted",           "data_delete",      "reports-analytics"),
    AuditAction.DAILY_REPORT_SUBMITTED:     ("تقديم تقرير يومي",          "Daily report submitted",         "data_submit",      "reports-analytics"),
    AuditAction.ADMIN_MESSAGE_SENT:         ("إرسال رسالة إدارية",        "Admin message sent",             "message_sent",     "management"),
    AuditAction.MESSAGE_SENT:               ("إرسال رسالة",               "Message sent",                   "message_sent",     "management"),
    AuditAction.INCIDENT_RESOLVED:          ("إغلاق حادثة",               "Incident resolved",              "incident_log",     "operations"),
    AuditAction.USER_EXPORT:                ("تصدير بيانات المستخدمين",   "User data exported",             "report_export",    "management"),
    AuditAction.AUDIT_LOG_EXPORT:           ("تصدير سجل التدقيق",         "Audit log exported",             "report_export",    "settings"),
    AuditAction.ANALYTICS_EXPORT_DOWNLOADED:("تنزيل تقرير تحليلي",        "Analytics report downloaded",    "report_export",    "reports-analytics"),
    AuditAction.ADMIN_PROFILE_UPDATED:      ("تحديث إعدادات النظام",      "System settings updated",        "settings_change",  "settings"),
    AuditAction.ENROLLMENT_ACCEPTED:        ("قبول طلب تسجيل",            "Enrollment application accepted","data_update",      "operations"),
    AuditAction.ENROLLMENT_REJECTED:        ("رفض طلب تسجيل",             "Enrollment application rejected","data_update",      "operations"),
    # Governance reminders are audited (see the /governance reminder endpoint) but
    # were never classified, which left the feed's `governance` module filter with
    # no rows to match and hid these events from the feed entirely.
    AuditAction.GOVERNANCE_REMINDER_SENT:   ("إرسال تذكير حوكمة",          "Governance reminder sent",       "message_sent",     "governance"),
}

# High-risk actions escalated to "critical" regardless of their sensitivity_level.
_CRITICAL_SEVERITY_ACTIONS = frozenset({
    AuditAction.USER_DELETED,
    AuditAction.BULK_USER_DELETE,
    AuditAction.KINDERGARTEN_DELETED,
    AuditAction.MFA_BYPASS_INITIATED,
    AuditAction.IMPERSONATION_START,
    AuditAction.AUDIT_LOG_CLEANUP,
    AuditAction.AUDIT_LOG_EXPORT,
    AuditAction.ANALYTICS_EXPORT_DOWNLOADED,
})

# Actions that represent a failed/denied attempt rather than a completed action.
_FAILURE_ACTIONS = frozenset({
    AuditAction.LOGIN_FAILED,
    AuditAction.ACCESS_DENIED,
    AuditAction.BULK_ACCESS_DENIED,
    AuditAction.MFA_BYPASS_FAILED_AUTH,
    AuditAction.IMPERSONATION_ATTEMPT_FAILED,
})

_ENTITY_TYPE_LABELS: Dict[str, tuple] = {
    "Auth":                   ("المصادقة", "Authentication"),
    "User":                   ("مستخدم", "User"),
    "Kindergarten":           ("حضانة", "Kindergarten"),
    "DailyReport":            ("تقرير يومي", "Daily Report"),
    "Dashboard":              ("لوحة التحكم", "Dashboard"),
    "EnrollmentApplication":  ("طلب تسجيل", "Enrollment Application"),
    "Message":                ("رسالة", "Message"),
    "Incident":               ("حادثة", "Incident"),
    "GovernanceReminder":     ("تذكير حوكمة", "Governance Reminder"),
}


# =============================================================================
# Custom period validation — shared by the admin endpoints that accept
# `period=custom`. Invalid custom input previously fell back silently to the
# endpoint's default window, so a caller asking for one range was answered with
# a different one and had no way to tell.
# =============================================================================

# A year of daily rows is the widest window these dashboards aggregate without
# pre-rollups; beyond it the query cost stops being interactive.
MAX_CUSTOM_PERIOD_DAYS = 366

# Accepted `period` values, per endpoint. Anything else is a 422 — these are the
# only windows the endpoints can actually resolve.
_ACTIVITY_PERIODS = ("today", "24h", "7d", "30d", "month", "custom")
_KG_OVERVIEW_PERIODS = ("today", "week", "month", "custom")


class CustomPeriodWindow(BaseModel):
    """A validated inclusive [start_date, end_date] window."""

    model_config = ConfigDict(frozen=True)

    start_date: date
    end_date: date

    @model_validator(mode="after")
    def _validate_window(self) -> "CustomPeriodWindow":
        if self.start_date > self.end_date:
            raise ValueError("start_date must be on or before end_date")
        if self.inclusive_days > MAX_CUSTOM_PERIOD_DAYS:
            raise ValueError(
                f"custom period must not exceed {MAX_CUSTOM_PERIOD_DAYS} days "
                f"(requested {self.inclusive_days})"
            )
        return self

    @property
    def inclusive_days(self) -> int:
        return (self.end_date - self.start_date).days + 1


def _resolve_period(
    period: Optional[str],
    start_date: Optional[date],
    end_date: Optional[date],
    *,
    allowed: tuple,
) -> Optional[CustomPeriodWindow]:
    """Validate `period` and return the window when it is `custom`.

    Returns None for the non-custom presets, which each endpoint resolves itself.
    Raises 422 rather than silently substituting a different window.
    """
    if period is None:
        return None
    if period not in allowed:
        raise HTTPException(
            status_code=422,
            detail=f"invalid period '{period}'; expected one of: {', '.join(allowed)}",
        )
    if period != "custom":
        return None
    if not start_date or not end_date:
        raise HTTPException(
            status_code=422,
            detail="custom period requires both start_date and end_date",
        )
    try:
        return CustomPeriodWindow(start_date=start_date, end_date=end_date)
    except ValidationError as exc:
        raise HTTPException(
            status_code=422,
            detail="; ".join(e["msg"].removeprefix("Value error, ") for e in exc.errors()),
        )


# `permission_change` is a *derived* activity type: it is not an AuditAction, so it
# cannot appear in _ACTIVITY_MAP. It is a USER_UPDATED row whose old_data/new_data
# disagree on `role`. The render path and the query filter must derive it from the
# same definition — when they drift, the filter silently returns the wrong rows.
PERMISSION_CHANGE_TYPE = "permission_change"

# Base types that a derived type can steal rows from, keyed by the action involved.
_DERIVED_FROM_ACTION = AuditAction.USER_UPDATED


def _is_role_change(old_data, new_data) -> bool:
    """Python classifier for PERMISSION_CHANGE_TYPE — mirrored in SQL by
    _role_change_clause(). Both must agree or filtering breaks."""
    old_role = (old_data or {}).get("role")
    new_role = (new_data or {}).get("role")
    return old_role is not None and new_role is not None and old_role != new_role


def _role_change_clause():
    """SQL mirror of _is_role_change().

    AuditLog.old_data/new_data are generic JSON columns, so SQLAlchemy renders
    this as `->>` on PostgreSQL and `json_extract` on SQLite — a role change is
    therefore filterable in SQL, without persisting extra audit metadata, which
    keeps `total` and pagination correct.
    """
    old_role = models.AuditLog.old_data["role"].as_string()
    new_role = models.AuditLog.new_data["role"].as_string()
    return and_(old_role.isnot(None), new_role.isnot(None), old_role != new_role)


def _severity_for(action: str, sensitivity_level: Optional[int]) -> str:
    """Map an AuditLog row to a low/medium/high/critical severity tier."""
    if action in _CRITICAL_SEVERITY_ACTIONS:
        return "critical"
    return {1: "low", 2: "medium", 3: "high"}.get(sensitivity_level or 2, "medium")


def _to_jordan_iso(value) -> str:
    """Render a stored timestamp as an ISO string in Jordan time.

    Thin wrapper over ``utils.time_utils.to_jordan_iso`` that returns "" rather than
    None, because ``ActivityItem.timestamp`` is a non-optional str. The conversion
    itself lives in the shared helper so there is one implementation, not one per
    call site (N16).
    """
    return to_jordan_iso(value) or ""


def _activity_item_from_log(log: "models.AuditLog", actor_username: Optional[str] = None) -> Optional["ActivityItem"]:
    """Build a bilingual, enriched ActivityItem from a raw AuditLog row.
    USER_UPDATED is special-cased: a role change in old_data/new_data is
    surfaced as a permission-change message rather than a generic update.
    `actor_username` must come from a caller-side batched User lookup —
    AuditLog has no `user` relationship, so looking it up here would N+1."""
    action_str = str(log.action.value if hasattr(log.action, "value") else log.action)
    mapping = _ACTIVITY_MAP.get(action_str)
    if not mapping:
        return None
    msg_ar, msg_en, act_type, module_id = mapping

    if action_str == _DERIVED_FROM_ACTION and _is_role_change(log.old_data, log.new_data):
        msg_ar, msg_en, act_type = "تعديل صلاحيات مستخدم", "User permissions updated", PERMISSION_CHANGE_TYPE

    module_ar, module_en = _SIDEBAR_MODULE_LABELS.get(module_id, ("", ""))
    entity_ar, entity_en = _ENTITY_TYPE_LABELS.get(log.entity_type or "", (log.entity_type or "", log.entity_type or ""))

    return ActivityItem(
        type=act_type,
        message_ar=msg_ar,
        message_en=msg_en,
        # Rendered in Jordan time, not UTC. The activity feed is Jordan-facing and the
        # UI derives a calendar date from this string; a UTC timestamp puts anything
        # after 21:00 Jordan on the previous day. Storage is UTC (db_types.UTCDateTime)
        # — only the presentation is localised.
        timestamp=_to_jordan_iso(log.created_at),
        user_name=actor_username,
        user_role=(log.actor_role or None),
        module_ar=module_ar,
        module_en=module_en,
        entity_type=log.entity_type,
        entity_label_ar=entity_ar,
        entity_label_en=entity_en,
        status="failed" if action_str in _FAILURE_ACTIONS else "success",
        severity=_severity_for(action_str, log.sensitivity_level),
    )


# =============================================================================
# Admin Health Endpoint
# =============================================================================

@router.get("/health")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def admin_health_check(
    request: Request,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Admin-scoped health check: verifies DB connectivity and returns service status."""
    from sqlalchemy import text as _text

    checks: Dict[str, Any] = {}

    try:
        db.execute(_text("SELECT 1"))
        checks["database"] = "ok"
    except Exception as exc:
        logger.error("Admin health check: DB error: %s", exc)
        checks["database"] = "error"

    try:
        from cache_service import cache_service as _cs
        _cs.get("__health_probe__")
        checks["cache"] = "ok"
    except Exception:
        checks["cache"] = "degraded"

    overall = "ok" if all(v == "ok" for v in checks.values()) else "degraded"
    return {
        "status": overall,
        "checks": checks,
        "admin_id": current_user.id,
        "timestamp": datetime.now(_JORDAN_TZ).isoformat(),
    }


# =============================================================================
# Admin Dashboard Endpoint
# =============================================================================

@router.get("/stats", response_model=Dict[str, Any])
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def get_admin_stats(
    request: Request,
    period_days: int = Query(30, description="Number of days to analyze", ge=1, le=90),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Compatibility stats endpoint for admin clients that expect /api/admin/stats."""
    now = datetime.now(_JORDAN_TZ)
    today = now.date()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)

    total_users = db.query(func.count(models.User.id)).filter(
        models.User.deleted_at.is_(None)
    ).scalar() or 0
    total_kindergartens = db.query(func.count(models.Kindergarten.id)).filter(
        models.Kindergarten.deleted_at.is_(None)
    ).scalar() or 0
    active_kindergartens = db.query(func.count(models.Kindergarten.id)).filter(
        models.Kindergarten.status == models.KindergartenStatus.ACTIVE,
        models.Kindergarten.deleted_at.is_(None),
    ).scalar() or 0
    active_users_today = db.query(func.count(func.distinct(models.AuditLog.user_id))).filter(
        models.AuditLog.action == "LOGIN_SUCCESS",
        models.AuditLog.user_id.isnot(None),
        models.AuditLog.created_at >= today_start,
    ).scalar() or 0
    pending_applications = db.query(func.count(models.EnrollmentApplication.id)).filter(
        models.EnrollmentApplication.status == models.EnrollmentStatus.PENDING_REVIEW,
        models.EnrollmentApplication.deleted_at.is_(None),
    ).scalar() or 0
    pending_reports = db.query(func.count(models.DailyReport.id)).filter(
        models.DailyReport.status == models.DailyReportStatus.SUBMITTED,
    ).scalar() or 0
    recent_incidents = db.query(func.count(models.Incident.id)).filter(
        models.Incident.occurred_at >= jordan_day_bounds(today - timedelta(days=7))[0],
        models.Incident.deleted_at.is_(None),
    ).scalar() or 0
    attendance_today = db.query(func.count(models.AttendanceLog.id)).filter(
        models.AttendanceLog.date == today,
        models.AttendanceLog.status == models.AttendanceStatus.PRESENT,
    ).scalar() or 0
    active_enrollments = db.query(func.count(models.EnrollmentApplication.id)).filter(
        models.EnrollmentApplication.status == models.EnrollmentStatus.ACTIVE,
        models.EnrollmentApplication.deleted_at.is_(None),
    ).scalar() or 0
    attendance_rate = min((attendance_today / active_enrollments * 100.0) if active_enrollments > 0 else 0.0, 100.0)

    data_quality_score = 0.0
    if active_kindergartens > 0:
        active_kg_with_recent_report = db.query(
            func.count(func.distinct(models.DailyReport.kindergarten_id))
        ).join(
            models.Kindergarten,
            models.Kindergarten.id == models.DailyReport.kindergarten_id,
        ).filter(
            models.Kindergarten.status == models.KindergartenStatus.ACTIVE,
            models.DailyReport.date >= today - timedelta(days=7),
        ).scalar() or 0
        data_quality_score = round(
            (active_kg_with_recent_report / active_kindergartens * 100.0),
            1,
        )

    log_audit_event(
        db=db,
        action=AuditAction.ADMIN_DASHBOARD_VIEWED,
        actor=current_user,
        target_type="Dashboard",
        target_ids=None,
        metadata={"period_days": period_days, "endpoint": "stats"},
        sensitivity_level=2,
    )
    db.commit()

    return {
        "generated_at": now.isoformat(),
        "period_days": period_days,
        "summary": {
            "attendance_today": attendance_today,
            "pending_applications": pending_applications,
            "pending_daily_reports": pending_reports,
            "recent_incidents": recent_incidents,
            "attendance_rate": round(attendance_rate, 1),
        },
        "system_overview": {
            "total_kindergartens": total_kindergartens,
            "active_kindergartens": active_kindergartens,
            "total_users": total_users,
        },
        "kpis": {
            "total_users": float(total_users),
            "active_users": float(active_users_today),
            "total_kindergartens": float(total_kindergartens),
            "active_kindergartens": float(active_kindergartens),
            "total_submissions": float(pending_reports),
            "pending_submissions": float(pending_reports),
            "data_quality_score": data_quality_score,
        },
        "alerts": {
            "pending_applications": pending_applications,
            "recent_incidents": recent_incidents,
            "pending_reports": pending_reports,
        },
    }


@router.get("/dashboard", response_model=AdminDashboardResponse)
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def get_admin_dashboard(
    request: Request,
    period_days: int = Query(30, description="Number of days to analyze", ge=1, le=90),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Get comprehensive admin dashboard data with system overview, KPIs, charts, and alerts."""
    now = datetime.now(_JORDAN_TZ)
    today = now.date()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    cache_key = f"dashboard:admin:v4:period_{period_days}:date_{today.isoformat()}"
    cached_payload = _admin_dashboard_cache_get(cache_key)
    if isinstance(cached_payload, dict):
        log_audit_event(
            db=db,
            action=AuditAction.ADMIN_DASHBOARD_VIEWED,
            actor=current_user,
            target_type="Dashboard",
            target_ids=None,
            metadata={"period_days": period_days, "cache_hit": True},
            sensitivity_level=2,
        )
        db.commit()
        return AdminDashboardResponse(**cached_payload)

    week_ago = today - timedelta(days=7)

    total_users = db.query(func.count(models.User.id)).filter(
        models.User.deleted_at.is_(None)
    ).scalar() or 0
    total_kindergartens = db.query(func.count(models.Kindergarten.id)).filter(
        models.Kindergarten.deleted_at.is_(None)
    ).scalar() or 0
    active_kindergartens = db.query(func.count(models.Kindergarten.id)).filter(
        models.Kindergarten.status == models.KindergartenStatus.ACTIVE,
        models.Kindergarten.deleted_at.is_(None)
    ).scalar() or 0
    active_users_today = db.query(func.count(func.distinct(models.AuditLog.user_id))).filter(
        models.AuditLog.action == "LOGIN_SUCCESS",
        models.AuditLog.user_id.isnot(None),
        models.AuditLog.created_at >= today_start,
    ).scalar() or 0

    pending_applications = db.query(func.count(models.EnrollmentApplication.id)).filter(
        models.EnrollmentApplication.status == models.EnrollmentStatus.PENDING_REVIEW,
        models.EnrollmentApplication.deleted_at.is_(None)
    ).scalar() or 0
    pending_reports = db.query(func.count(models.DailyReport.id)).filter(
        models.DailyReport.status == models.DailyReportStatus.SUBMITTED
    ).scalar() or 0
    recent_incidents = db.query(func.count(models.Incident.id)).filter(
        models.Incident.occurred_at >= jordan_day_bounds(week_ago)[0],
        models.Incident.deleted_at.is_(None)
    ).scalar() or 0
    attendance_today = db.query(func.count(models.AttendanceLog.id)).filter(
        models.AttendanceLog.date == today,
        models.AttendanceLog.status == models.AttendanceStatus.PRESENT,
    ).scalar() or 0
    active_enrollments = db.query(func.count(models.EnrollmentApplication.id)).filter(
        models.EnrollmentApplication.status == models.EnrollmentStatus.ACTIVE,
        models.EnrollmentApplication.deleted_at.is_(None)
    ).scalar() or 0

    attendance_rate = min((attendance_today / active_enrollments * 100.0) if active_enrollments > 0 else 0.0, 100.0)

    total_reports_in_period = db.query(func.count(models.DailyReport.id)).filter(
        models.DailyReport.date >= today - timedelta(days=period_days),
        models.DailyReport.date <= today,
    ).scalar() or 0

    # Data quality: % of active KGs that submitted any report in the last 7 days
    active_kg_with_recent_report = 0
    if active_kindergartens > 0:
        active_kg_with_recent_report = db.query(
            func.count(func.distinct(models.DailyReport.kindergarten_id))
        ).join(
            models.Kindergarten,
            models.Kindergarten.id == models.DailyReport.kindergarten_id,
        ).filter(
            models.Kindergarten.status == models.KindergartenStatus.ACTIVE,
            models.DailyReport.date >= today - timedelta(days=7),
        ).scalar() or 0
    data_quality_score = round(
        (active_kg_with_recent_report / active_kindergartens * 100.0) if active_kindergartens > 0 else 0.0, 1
    )

    summary = DashboardSummary(
        attendance_today=attendance_today,
        pending_applications=pending_applications,
        pending_daily_reports=pending_reports,
        recent_incidents=recent_incidents,
        attendance_rate=round(attendance_rate, 1),
    )

    system_overview = DashboardSystemOverview(
        total_kindergartens=total_kindergartens,
        active_kindergartens=active_kindergartens,
        total_users=total_users,
    )

    # Kindergarten detail table is served by a dedicated endpoint; not computed here.
    kindergartens_list: List[DashboardKindergarten] = []

    chart_days = max(7, min(period_days, 90))
    chart_start_date = today - timedelta(days=chart_days - 1)
    # Single query for all chart days instead of N+1
    daily_attendance_counts = dict(
        db.query(
            models.AttendanceLog.date,
            func.count(models.AttendanceLog.id),
        ).filter(
            models.AttendanceLog.date >= chart_start_date,
            models.AttendanceLog.date <= today,
            models.AttendanceLog.status == models.AttendanceStatus.PRESENT,
        ).group_by(models.AttendanceLog.date).all()
    )
    attendance_chart: List[DashboardChartPoint] = []
    for i in range(chart_days):
        day_value = today - timedelta(days=(chart_days - 1 - i))
        day_count = daily_attendance_counts.get(day_value, 0)
        attendance_chart.append(DashboardChartPoint(date=day_value.isoformat(), value=day_count))

    enrollment_stats = db.query(
        models.EnrollmentApplication.status,
        func.count(models.EnrollmentApplication.id),
    ).filter(
        models.EnrollmentApplication.deleted_at.is_(None)
    ).group_by(models.EnrollmentApplication.status).all()
    enrollment_pie = {
        (status.value if hasattr(status, "value") else str(status)).upper(): count
        for status, count in enrollment_stats
    }

    # Build incidents trend chart (last 30 days)
    daily_incident_counts = dict(
        db.query(
            func.date(models.Incident.occurred_at),
            func.count(models.Incident.id),
        ).filter(
            *jordan_date_range_filter(models.Incident.occurred_at, chart_start_date, today),
            models.Incident.deleted_at.is_(None),
        ).group_by(func.date(models.Incident.occurred_at)).all()
    )
    incidents_trend: List[DashboardChartPoint] = []
    for i in range(chart_days):
        day_value = today - timedelta(days=(chart_days - 1 - i))
        day_count = daily_incident_counts.get(day_value, 0)
        incidents_trend.append(DashboardChartPoint(date=day_value.isoformat(), value=day_count))

    charts = DashboardCharts(
        attendance=attendance_chart,
        enrollment=enrollment_pie,
        incidents=incidents_trend,
    )

    alerts: List[DashboardAlert] = []
    if pending_applications > 0:
        _msg_ar = f"يوجد {pending_applications} طلب تسجيل يحتاج المراجعة."
        _msg_en = f"{pending_applications} enrollment application(s) awaiting review."
        alerts.append(
            DashboardAlert(
                id="pending_applications",
                title="طلبات تسجيل بانتظار المراجعة",
                message=_msg_ar,
                title_ar="طلبات تسجيل بانتظار المراجعة",
                title_en="Enrollment Applications Pending Review",
                message_ar=_msg_ar,
                message_en=_msg_en,
                severity="warning",
                timestamp=now.isoformat(),
                category="applications",
                type="applications",
                priority="high" if pending_applications > 10 else "medium",
            )
        )

    if recent_incidents > 0:
        _msg_ar = f"تم تسجيل {recent_incidents} حوادث خلال آخر 7 أيام."
        _msg_en = f"{recent_incidents} incident(s) recorded in the last 7 days."
        alerts.append(
            DashboardAlert(
                id="recent_incidents",
                title="حوادث مسجلة حديثاً",
                message=_msg_ar,
                title_ar="حوادث مسجلة حديثاً",
                title_en="Recent Incidents",
                message_ar=_msg_ar,
                message_en=_msg_en,
                severity="error" if recent_incidents > 5 else "warning",
                timestamp=now.isoformat(),
                category="safety",
                type="safety",
                priority="high" if recent_incidents > 5 else "medium",
            )
        )

    expiring_licenses = db.query(models.Kindergarten).filter(
        models.Kindergarten.status == models.KindergartenStatus.ACTIVE,
        models.Kindergarten.license_valid_until.isnot(None),
        models.Kindergarten.license_valid_until <= today + timedelta(days=30),
    ).all()
    for kg in expiring_licenses:
        days_to_expiry = (kg.license_valid_until - today).days
        expired = days_to_expiry < 0
        _msg_ar = (
            f"انتهت صلاحية ترخيص {kg.name_ar}."
            if expired
            else f"تنتهي صلاحية ترخيص {kg.name_ar} خلال {days_to_expiry} يوم."
        )
        _msg_en = (
            f"License for {kg.name_en or kg.name_ar} has expired."
            if expired
            else f"License for {kg.name_en or kg.name_ar} expires in {days_to_expiry} day(s)."
        )
        alerts.append(
            DashboardAlert(
                id=f"license_expiry_{kg.id}",
                title="تنبيه صلاحية الترخيص",
                message=_msg_ar,
                title_ar="تنبيه صلاحية الترخيص",
                title_en="License Expiry Alert",
                message_ar=_msg_ar,
                message_en=_msg_en,
                severity="error" if expired else "warning",
                timestamp=now.isoformat(),
                category="compliance",
                type="compliance",
                priority="critical" if expired else "high",
                kindergarten_id=kg.id,
            )
        )

    alerts = sorted(
        alerts,
        key=lambda alert: {"critical": 0, "high": 1, "medium": 2, "low": 3}.get(alert.priority or "medium", 4),
    )

    # Data quality reasons: concrete, derivable causes behind a low data_quality_score.
    # Computed live from existing columns — no new table, mirrors data_quality_score itself.
    missing_report_count = max(active_kindergartens - active_kg_with_recent_report, 0)
    missing_contact_count = db.query(func.count(models.Kindergarten.id)).filter(
        models.Kindergarten.status == models.KindergartenStatus.ACTIVE,
        or_(models.Kindergarten.contact_email.is_(None), models.Kindergarten.contact_email == ""),
    ).scalar() or 0
    missing_geo_count = db.query(func.count(models.Kindergarten.id)).filter(
        models.Kindergarten.status == models.KindergartenStatus.ACTIVE,
        or_(models.Kindergarten.latitude.is_(None), models.Kindergarten.longitude.is_(None)),
    ).scalar() or 0
    expired_license_count = sum(1 for kg in expiring_licenses if (kg.license_valid_until - today).days < 0)

    data_quality_reasons: List[DataQualityReason] = []
    if missing_report_count > 0:
        data_quality_reasons.append(DataQualityReason(
            id="missing_recent_report",
            label_ar=f"{missing_report_count} حضانة نشطة بدون تقرير خلال آخر 7 أيام",
            label_en=f"{missing_report_count} active kindergarten(s) without a report in the last 7 days",
            count=missing_report_count,
        ))
    if missing_contact_count > 0:
        data_quality_reasons.append(DataQualityReason(
            id="missing_contact_email",
            label_ar=f"{missing_contact_count} حضانة نشطة بدون بريد إلكتروني للتواصل",
            label_en=f"{missing_contact_count} active kindergarten(s) missing a contact email",
            count=missing_contact_count,
        ))
    if missing_geo_count > 0:
        data_quality_reasons.append(DataQualityReason(
            id="missing_geo_coordinates",
            label_ar=f"{missing_geo_count} حضانة نشطة بدون إحداثيات موقع",
            label_en=f"{missing_geo_count} active kindergarten(s) missing map coordinates",
            count=missing_geo_count,
        ))
    if expired_license_count > 0:
        data_quality_reasons.append(DataQualityReason(
            id="expired_license",
            label_ar=f"{expired_license_count} حضانة بترخيص منتهي الصلاحية",
            label_en=f"{expired_license_count} kindergarten(s) with an expired license",
            count=expired_license_count,
        ))

    # Recent activity from audit log — enriched with actor/role/module/entity/severity.
    recent_audit_logs_with_users = (
        db.query(models.AuditLog, models.User.username)
        .outerjoin(models.User, models.AuditLog.user_id == models.User.id)
        .filter(
            models.AuditLog.action.in_(list(_ACTIVITY_MAP.keys())),
            models.AuditLog.created_at >= now - timedelta(days=7),
        )
        .order_by(models.AuditLog.created_at.desc())
        .limit(10)
        .all()
    )
    recent_activity: List[ActivityItem] = [
        item for item in (
            _activity_item_from_log(log, username) for log, username in recent_audit_logs_with_users
        ) if item is not None
    ]

    kpis: Dict[str, Optional[float]] = {
        "total_users":          float(total_users),
        "active_users":         float(active_users_today),
        "total_kindergartens":  float(total_kindergartens),
        "active_kindergartens": float(active_kindergartens),
        "total_submissions":    float(total_reports_in_period),
        "pending_submissions":  float(pending_reports),
        "data_quality_score":   data_quality_score,
    }

    # KPI trend/comparison metadata: current value vs the equivalent previous period.
    period_start = today - timedelta(days=period_days - 1)
    prev_start, prev_end = KPIService._compute_previous_period(period_start, today)
    prev_boundary = datetime.combine(prev_end + timedelta(days=1), datetime.min.time(), tzinfo=_JORDAN_TZ)
    yesterday = today - timedelta(days=1)
    yesterday_start = datetime.combine(yesterday, datetime.min.time(), tzinfo=_JORDAN_TZ)

    prev_total_users = db.query(func.count(models.User.id)).filter(
        models.User.created_at < prev_boundary
    ).scalar() or 0
    prev_total_kindergartens = db.query(func.count(models.Kindergarten.id)).filter(
        models.Kindergarten.created_at < prev_boundary
    ).scalar() or 0
    # Approximation: applies current status retroactively (no status-history table exists).
    prev_active_kindergartens = db.query(func.count(models.Kindergarten.id)).filter(
        models.Kindergarten.created_at < prev_boundary,
        models.Kindergarten.status == models.KindergartenStatus.ACTIVE,
        models.Kindergarten.deleted_at.is_(None),
    ).scalar() or 0
    prev_active_users = db.query(func.count(func.distinct(models.AuditLog.user_id))).filter(
        models.AuditLog.action == "LOGIN_SUCCESS",
        models.AuditLog.user_id.isnot(None),
        models.AuditLog.created_at >= yesterday_start,
        models.AuditLog.created_at < today_start,
    ).scalar() or 0
    prev_total_submissions = db.query(func.count(models.DailyReport.id)).filter(
        models.DailyReport.date >= prev_start,
        models.DailyReport.date <= prev_end,
    ).scalar() or 0
    prev_pending_submissions = db.query(func.count(models.DailyReport.id)).filter(
        models.DailyReport.status == models.DailyReportStatus.SUBMITTED,
        models.DailyReport.date >= prev_start,
        models.DailyReport.date <= prev_end,
    ).scalar() or 0
    prev_active_kg_with_recent_report = 0
    if prev_active_kindergartens > 0:
        prev_active_kg_with_recent_report = db.query(
            func.count(func.distinct(models.DailyReport.kindergarten_id))
        ).join(
            models.Kindergarten,
            models.Kindergarten.id == models.DailyReport.kindergarten_id,
        ).filter(
            models.Kindergarten.status == models.KindergartenStatus.ACTIVE,
            models.Kindergarten.deleted_at.is_(None),
            models.DailyReport.date >= prev_end - timedelta(days=6),
            models.DailyReport.date <= prev_end,
        ).scalar() or 0
    prev_data_quality_score = round(
        (prev_active_kg_with_recent_report / prev_active_kindergartens * 100.0) if prev_active_kindergartens > 0 else 0.0, 1
    )

    def _kpi_trend(key: str, current: float, previous: float, measurable: bool = True) -> KPITrendMeta:
        direction, change = KPIService._trend_from_values(current, previous)
        baseline_available = previous > 0
        change_pct = round((change / previous * 100.0), 1) if previous else None
        if not measurable:
            # Metric could not be computed — no good/bad claim, no trend.
            status = "unavailable"
        elif key == "pending_submissions":
            # Actionable backlog: anything pending needs attention.
            status = "warning" if current > 0 else "good"
        elif key == "data_quality_score":
            # Threshold-based; a genuine 0 is a governance emergency, not merely "low".
            status = "critical" if current <= 0 else ("good" if current >= 70 else "warning")
        else:
            # Raw counts (users, kindergartens, submissions) have no target, so a
            # green "Good" badge would be meaningless — stay neutral.
            status = "neutral"
        return KPITrendMeta(
            value=current,
            previous_value=previous,
            change=change,
            change_pct=change_pct,
            trend=direction,
            status=status,
            baseline_available=baseline_available,
            measurable=measurable,
        )

    kpi_trends: Dict[str, KPITrendMeta] = {
        "total_users":          _kpi_trend("total_users", total_users, prev_total_users),
        "active_users":         _kpi_trend("active_users", active_users_today, prev_active_users),
        "total_kindergartens":  _kpi_trend("total_kindergartens", total_kindergartens, prev_total_kindergartens),
        "active_kindergartens": _kpi_trend("active_kindergartens", active_kindergartens, prev_active_kindergartens),
        "total_submissions":    _kpi_trend("total_submissions", total_reports_in_period, prev_total_submissions),
        "pending_submissions":  _kpi_trend("pending_submissions", pending_reports, prev_pending_submissions),
        "data_quality_score":   _kpi_trend("data_quality_score", data_quality_score, prev_data_quality_score, measurable=active_kindergartens > 0),
    }

    # Log dashboard access
    log_audit_event(
        db=db,
        action=AuditAction.ADMIN_DASHBOARD_VIEWED,
        actor=current_user,
        target_type="Dashboard",
        target_ids=None,
        metadata={
            "period_days": period_days,
            "kpi_count": len(kpis)
        },
        sensitivity_level=2,
    )
    db.commit()

    response_payload = AdminDashboardResponse(
        summary=summary,
        system_overview=system_overview,
        kindergartens=kindergartens_list,
        charts=charts,
        alerts=alerts,
        kpis=kpis,
        kpi_trends=kpi_trends,
        data_quality_reasons=data_quality_reasons,
        recent_activity=recent_activity,
        generated_at=now.isoformat(),
    )
    _admin_dashboard_cache_set(cache_key, response_payload.model_dump(mode="json"))
    return response_payload


class DashboardActivityResponse(BaseModel):
    """Paginated, filtered recent-activity feed."""
    total: int
    items: List[ActivityItem]
    page: int
    page_size: int


@router.get("/dashboard/activity", response_model=DashboardActivityResponse)
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def get_admin_dashboard_activity(
    request: Request,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: Optional[str] = None,
    period: Optional[str] = Query(None, description="today|24h|7d|30d|month|custom"),
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    activity_type: Optional[str] = None,
    user_id: Optional[int] = None,
    role: Optional[models.UserRole] = None,
    module: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    severity: Optional[str] = None,
    entity_type: Optional[str] = None,
    governorate: Optional[str] = None,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Filterable, paginated recent-activity feed backing the dashboard's activity filter bar."""
    allowed_activity_types = {mapping[2] for mapping in _ACTIVITY_MAP.values()} | {
        PERMISSION_CHANGE_TYPE
    }
    allowed_modules = {mapping[3] for mapping in _ACTIVITY_MAP.values()}
    if activity_type is not None and activity_type not in allowed_activity_types:
        raise HTTPException(
            status_code=422,  # 422 literal: Starlette deprecated the ENTITY constant name
            detail=f"Unsupported activity_type: {activity_type}",
        )
    if module is not None and module not in allowed_modules:
        raise HTTPException(
            status_code=422,  # 422 literal: Starlette deprecated the ENTITY constant name
            detail=f"Unsupported module: {module}",
        )
    if status_filter is not None and status_filter not in {"success", "failed"}:
        raise HTTPException(
            status_code=422,  # 422 literal: Starlette deprecated the ENTITY constant name
            detail=f"Unsupported status: {status_filter}",
        )
    if severity is not None and severity not in {"low", "medium", "high", "critical"}:
        raise HTTPException(
            status_code=422,  # 422 literal: Starlette deprecated the ENTITY constant name
            detail=f"Unsupported severity: {severity}",
        )

    page, page_size, offset = enforce_pagination(page, page_size)
    custom_window = _resolve_period(
        period, start_date, end_date, allowed=_ACTIVITY_PERIODS
    )
    now = datetime.now(_JORDAN_TZ)
    today = now.date()

    query = db.query(models.AuditLog, models.User.username).outerjoin(
        models.User, models.AuditLog.user_id == models.User.id
    ).filter(
        models.AuditLog.action.in_(list(_ACTIVITY_MAP.keys()))
    )

    if period == "today":
        query = query.filter(models.AuditLog.created_at >= datetime.combine(today, datetime.min.time(), tzinfo=_JORDAN_TZ))
    elif period == "24h":
        query = query.filter(models.AuditLog.created_at >= now - timedelta(hours=24))
    elif period == "7d":
        query = query.filter(models.AuditLog.created_at >= now - timedelta(days=7))
    elif period == "30d":
        query = query.filter(models.AuditLog.created_at >= now - timedelta(days=30))
    elif period == "month":
        query = query.filter(models.AuditLog.created_at >= datetime.combine(today.replace(day=1), datetime.min.time(), tzinfo=_JORDAN_TZ))
    elif custom_window is not None:
        query = query.filter(
            models.AuditLog.created_at >= datetime.combine(custom_window.start_date, datetime.min.time(), tzinfo=_JORDAN_TZ),
            models.AuditLog.created_at <= datetime.combine(custom_window.end_date, datetime.max.time(), tzinfo=_JORDAN_TZ),
        )

    if activity_type:
        if activity_type == PERMISSION_CHANGE_TYPE:
            # Derived type: USER_UPDATED rows where the role actually changed.
            query = query.filter(
                models.AuditLog.action == _DERIVED_FROM_ACTION,
                _role_change_clause(),
            )
        else:
            base_actions = [a for a, m in _ACTIVITY_MAP.items() if m[2] == activity_type]
            query = query.filter(models.AuditLog.action.in_(base_actions))
            if _DERIVED_FROM_ACTION in base_actions:
                # Role changes render as `permission_change`, so they must not also
                # come back under `user_update` — the filter would disagree with the
                # `type` on every returned item.
                query = query.filter(~_role_change_clause())

    if module:
        query = query.filter(models.AuditLog.action.in_(
            [a for a, m in _ACTIVITY_MAP.items() if m[3] == module]
        ))

    if entity_type:
        query = query.filter(models.AuditLog.entity_type == entity_type)

    if user_id:
        query = query.filter(models.AuditLog.user_id == user_id)

    if role or governorate:
        scoped_users = db.query(models.User.id)
        if role:
            scoped_users = scoped_users.filter(models.User.role == role)
        if governorate:
            scoped_users = scoped_users.join(
                models.Kindergarten, models.Kindergarten.id == models.User.kindergarten_id
            ).filter(governorate_filter(models.Kindergarten.governorate, governorate))
        query = query.filter(models.AuditLog.user_id.in_([r[0] for r in scoped_users.all()]))

    if search:
        search_term = f"%{search[:100]}%"
        matching_user_ids = [
            r[0] for r in db.query(models.User.id).filter(models.User.username.ilike(search_term)).all()
        ]
        query = query.filter(or_(
            models.AuditLog.user_id.in_(matching_user_ids),
            models.AuditLog.details.ilike(search_term),
        ))

    # status/severity are derived (not raw columns) — push the same logic used by
    # _severity_for/_FAILURE_ACTIONS down into SQL so pagination/total stay correct.
    if status_filter == "failed":
        query = query.filter(models.AuditLog.action.in_(list(_FAILURE_ACTIONS)))
    elif status_filter == "success":
        query = query.filter(~models.AuditLog.action.in_(list(_FAILURE_ACTIONS)))

    if severity == "critical":
        query = query.filter(models.AuditLog.action.in_(list(_CRITICAL_SEVERITY_ACTIONS)))
    elif severity in ("low", "medium", "high"):
        target_level = {"low": 1, "medium": 2, "high": 3}[severity]
        non_critical = ~models.AuditLog.action.in_(list(_CRITICAL_SEVERITY_ACTIONS))
        if severity == "medium":
            query = query.filter(non_critical, or_(
                models.AuditLog.sensitivity_level == target_level,
                models.AuditLog.sensitivity_level.is_(None),
            ))
        else:
            query = query.filter(non_critical, models.AuditLog.sensitivity_level == target_level)

    total = query.count()
    logs_with_users = (
        query.order_by(models.AuditLog.created_at.desc())
        .offset(offset)
        .limit(page_size)
        .all()
    )

    items: List[ActivityItem] = [
        item for item in (
            _activity_item_from_log(log, username) for log, username in logs_with_users
        ) if item is not None
    ]

    return DashboardActivityResponse(total=total, items=items, page=page, page_size=page_size)


# =============================================================================
# Kindergarten Overview Endpoint
# =============================================================================

class KgKpiDetail(BaseModel):
    trend: Optional[str] = None
    note_ar: Optional[str] = None
    note_en: Optional[str] = None
    active_kindergartens: Optional[int] = None
    pending_placement: Optional[int] = None


class KgKpiCard(BaseModel):
    title_ar: str
    title_en: str
    value: Union[int, float]
    unit: Optional[str] = None
    trend: Optional[str] = None
    target: Optional[Union[int, float]] = None
    status: Optional[str] = None
    details: Optional[KgKpiDetail] = None


class KgKindergartenCard(BaseModel):
    id: int
    name_ar: str
    name_en: Optional[str] = None
    governorate: str
    district: str
    children_count: int
    attendance_rate: float
    attendance_status: str
    capacity_utilization: float
    capacity_status: str
    teachers_count: int
    supervisor_gap: int
    children_per_supervisor: float
    open_alerts: int
    health_score: str
    health_label_ar: str
    health_label_en: str
    recommended_action_ar: str
    recommended_action_en: str
    last_report_date: Optional[str] = None
    teacher_data_status: Optional[str] = None


class KgChartPoint(BaseModel):
    name: str
    name_ar: str
    value: float
    meta: Optional[Dict[str, Any]] = None


class KgChartDataset(BaseModel):
    label_ar: str
    label_en: str
    data: List[KgChartPoint]


class KgAlertSummary(BaseModel):
    id: int
    severity: str
    message: str
    kindergarten_id: Optional[int] = None
    kindergarten_name_ar: Optional[str] = None
    governorate: Optional[str] = None
    metric: str
    current_value: float
    triggered_at: str
    status: str
    age_hours: Optional[float] = None


class KgExecutiveHealth(BaseModel):
    critical_alerts: int
    near_capacity_kgs: int
    below_target_attendance: int
    data_quality_issues: int


class KgOverviewResponse(BaseModel):
    generated_at: str
    kpis: List[KgKpiCard]
    kindergartens: List[KgKindergartenCard]
    charts: Dict[str, Any]
    alerts: List[KgAlertSummary]
    executive_health: KgExecutiveHealth
    filters: Dict[str, Any]


def _compute_health_score(attendance_rate: float, occupancy_rate: float, open_alerts: int, data_ok: bool) -> str:
    score = 0.0
    score += min(attendance_rate / 100.0, 1.0) * 40
    score += max(0.0, 1.0 - occupancy_rate / 100.0) * 20
    score += max(0.0, 1.0 - min(open_alerts, 20) / 20.0) * 30
    score += (1.0 if data_ok else 0.0) * 10
    if score >= 85:
        return "excellent"
    if score >= 70:
        return "good"
    if score >= 55:
        return "needs_attention"
    if score >= 35:
        return "at_risk"
    return "critical"


_HEALTH_LABELS = {
    "excellent": ("ممتاز", "Excellent"),
    "good": ("جيد", "Good"),
    "needs_attention": ("يحتاج متابعة", "Needs Attention"),
    "at_risk": ("معرّض للخطر", "At Risk"),
    "critical": ("حرج", "Critical"),
}

_ATTENDANCE_TARGET = 70.0
_OCCUPANCY_SAFE = 70.0
_OCCUPANCY_MONITOR = 85.0
_OCCUPANCY_NEAR = 95.0


@router.get("/kg-overview", response_model=KgOverviewResponse)
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def get_kg_overview(
    request: Request,
    period: str = Query("month", description="Filter period: today, week, month, custom"),
    start_date: Optional[date] = Query(None, description="Start date (required when period=custom)"),
    end_date: Optional[date] = Query(None, description="End date (required when period=custom)"),
    governorate: Optional[str] = Query(None, description="Filter by governorate"),
    city: Optional[str] = Query(None, description="Filter by city (mapped to district)"),
    kindergarten_id: Optional[int] = Query(None, description="Filter by specific kindergarten"),
    class_id: Optional[int] = Query(None, description="Filter by specific class"),
    risk_level: Optional[str] = Query(None, description="Filter by risk level"),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Get comprehensive kindergarten overview with KPIs, health cards, charts, and alerts."""
    custom_window = _resolve_period(
        period, start_date, end_date, allowed=_KG_OVERVIEW_PERIODS
    )
    now = datetime.now(_JORDAN_TZ)
    today = now.date()

    if custom_window is not None:
        date_start = custom_window.start_date
        date_end = custom_window.end_date
    elif period == "today":
        date_start = today
        date_end = today
    elif period == "week":
        date_start = today - timedelta(days=6)
        date_end = today
    else:  # "month"
        date_start = today - timedelta(days=29)
        date_end = today

    query = db.query(models.Kindergarten).filter(models.Kindergarten.deleted_at.is_(None))

    if governorate:
        gov_normalized = settings.JORDAN_GOVERNORATE_ALIASES.get(governorate, governorate)
        query = query.filter(models.Kindergarten.governorate == gov_normalized)
    
    if city:
        city_normalized = city.split("_")[-1] if "_" in city else city
        query = query.filter(models.Kindergarten.district == city_normalized)

    if kindergarten_id:
        query = query.filter(models.Kindergarten.id == kindergarten_id)

    if class_id:
        cls = db.query(models.Class).filter(models.Class.id == class_id).first()
        if cls:
            query = query.filter(models.Kindergarten.id == cls.kindergarten_id)
        else:
            query = query.filter(models.Kindergarten.id == -1)

    all_kgs = query.order_by(models.Kindergarten.name_ar).all()
    kg_ids = [kg.id for kg in all_kgs]
    active_kgs = [
        kg for kg in all_kgs
        if kg.status == models.KindergartenStatus.ACTIVE
    ]

    kg_ids = [kg.id for kg in all_kgs]
    active_kgs = [
        kg for kg in all_kgs
        if kg.status == models.KindergartenStatus.ACTIVE
        and kg.deleted_at is None
    ]

    # Batch queries
    active_enrollments = db.query(
        models.EnrollmentApplication.kindergarten_id,
        func.count(models.EnrollmentApplication.id)
    ).filter(
        models.EnrollmentApplication.kindergarten_id.in_(kg_ids),
        models.EnrollmentApplication.status.in_([
            models.EnrollmentStatus.ACTIVE,
            models.EnrollmentStatus.ACCEPTED
        ])
    ).group_by(models.EnrollmentApplication.kindergarten_id).all()
    children_by_kg = {kid: cnt for kid, cnt in active_enrollments}

    # Attendance rate comes from the authoritative KPI engine, in bulk (3 queries for
    # every kindergarten, not 4 each).
    #
    # It used to be computed here as `present_rows / active_children * 100`, which
    # divides a count of PRESENT rows *across the whole window* by a *single-day*
    # headcount. The two have different dimensions, so the result scaled with the
    # window: one child present 5 days reported 100% over a 1-day window and 500% over
    # a 10-day one — every band still "on_target", because 500 >= the target. The
    # denominator has to be expected child-days, which means respecting working days
    # (Sun–Thu plus OperatingCalendar) and each enrolment's own date range. That is
    # exactly what KPIService already does, and duplicating it here is what let this
    # number drift from the KPI dashboard's in the first place (CLAUDE.md: KPI
    # computations belong in kpi_service.py).
    attendance_components_by_kg = KPIService.compute_attendance_components_bulk(
        db, list(kg_ids), date_start, date_end
    )
    attendance_rate_by_kg = {
        kg_id: (round((attended / expected) * 100, 2) if expected else 0.0)
        for kg_id, (attended, expected) in attendance_components_by_kg.items()
    }

    classes = db.query(models.Class).filter(models.Class.kindergarten_id.in_(kg_ids)).all()
    capacity_by_kg: Dict[int, Dict[str, int]] = defaultdict(lambda: {"total_capacity": 0, "total_enrolled": 0})
    for cls in classes:
        capacity_by_kg[cls.kindergarten_id]["total_capacity"] += cls.capacity_total or 0
        capacity_by_kg[cls.kindergarten_id]["total_enrolled"] += cls.enrolled_children_count or 0

    # Teachers: MANAGER + SUPERVISOR linked to kindergartens
    teachers_by_kg: Dict[int, int] = defaultdict(int)
    staff_users = db.query(models.User).filter(
        models.User.kindergarten_id.in_(kg_ids),
        models.User.role.in_([models.UserRole.MANAGER, models.UserRole.SUPERVISOR]),
        models.User.status == models.UserStatus.ACTIVE
    ).all()
    for user in staff_users:
        if user.kindergarten_id is not None:
            teachers_by_kg[user.kindergarten_id] += 1

    # Alerts per kg (scope is KG / GOVERNORATE / ALL)
    alerts_base_query = db.query(models.ActiveAlert).filter(
        models.ActiveAlert.status.in_([models.AlertStatus.ACTIVE, models.AlertStatus.ACKNOWLEDGED])
    )
    if governorate:
        alerts_base_query = alerts_base_query.filter(
            or_(
                models.ActiveAlert.scope_type == "KINDERGARTEN",
                models.ActiveAlert.scope_type == "GOVERNORATE",
                models.ActiveAlert.scope_type == "ALL",
            )
        )
    alerts_by_kg: Dict[int, List[models.ActiveAlert]] = defaultdict(list)
    gov_alerts: List[models.ActiveAlert] = []
    network_alerts: List[models.ActiveAlert] = []
    for alert in alerts_base_query.limit(500).all():
        if alert.scope_type == "KINDERGARTEN" and alert.scope_id:
            try:
                kid = int(alert.scope_id)
                if kid in kg_ids:
                    alerts_by_kg[kid].append(alert)
            except (TypeError, ValueError):
                pass
        elif alert.scope_type == "GOVERNORATE":
            gov_alerts.append(alert)
        elif alert.scope_type == "ALL":
            network_alerts.append(alert)

    # Governorate mapping for network alerts
    gov_to_kg: Dict[str, List[int]] = defaultdict(list)
    for kg in all_kgs:
        gov_to_kg[kg.governorate].append(kg.id)

    for alert in gov_alerts:
        for kid in gov_to_kg.get(alert.scope_id or "", []):
            alerts_by_kg[kid].append(alert)

    # Daily report status per kg (for last report date and data quality)
    latest_reports = db.query(
        models.DailyReport.kindergarten_id,
        func.max(models.DailyReport.date)
    ).filter(
        models.DailyReport.kindergarten_id.in_(kg_ids)
    ).group_by(models.DailyReport.kindergarten_id).all()
    last_report_by_kg = {kid: dte.isoformat() for kid, dte in latest_reports}

    # Build kindergarten cards
    kg_cards: List[KgKindergartenCard] = []
    for kg in all_kgs:
        kids = children_by_kg.get(kg.id, 0)
        att_rate = round(attendance_rate_by_kg.get(kg.id, 0.0), 1)
        cap = capacity_by_kg.get(kg.id, {"total_capacity": 0, "total_enrolled": 0})
        teachers = teachers_by_kg.get(kg.id, 0)
        open_alerts = len(alerts_by_kg.get(kg.id, []))

        if att_rate >= _ATTENDANCE_TARGET:
            att_status = "on_target"
        elif att_rate >= 55.0:
            att_status = "below_target"
        else:
            att_status = "critical_low"

        if cap["total_capacity"] > 0:
            cap_util = round((cap["total_enrolled"] / cap["total_capacity"] * 100.0), 1)
        else:
            cap_util = 0.0

        if cap_util < _OCCUPANCY_SAFE:
            cap_status = "safe"
        elif cap_util < _OCCUPANCY_MONITOR:
            cap_status = "monitor"
        elif cap_util < _OCCUPANCY_NEAR:
            cap_status = "near_capacity"
        else:
            cap_status = "full"

        # Calculate supervisor KPIs
        req_supervisors = math.ceil(kids / 20.0) if kids > 0 else 0  # Assuming 1 per 20
        sup_gap = req_supervisors - teachers
        sup_gap = max(0, sup_gap)
        cps = round(kids / teachers, 1) if teachers > 0 else 0.0

        teacher_data_ok = teachers > 0
        data_ok = teacher_data_ok

        health = _compute_health_score(att_rate, cap_util, open_alerts, data_ok)
        health_ar, health_en = _HEALTH_LABELS.get(health, ("غير معروف", "Unknown"))

        if health == "critical":
            action_ar = "مراجعة مشكلة الحضور فوراً"
            action_en = "Review attendance issue immediately"
        elif health == "at_risk":
            action_ar = "يحتاج متابعة عاجلة"
            action_en = "Requires urgent follow-up"
        elif health == "needs_attention":
            action_ar = "مراقبة القسم"
            action_en = "Monitor the department"
        elif health == "good":
            action_ar = "الحالة مستقرة"
            action_en = "Status is stable"
        else:
            action_ar = "لا توجد إجراءات مطلوبة"
            action_en = "No actions required"

        teacher_data_status = "updated" if teacher_data_ok else "needs_update"

        if risk_level and health != risk_level:
            continue

        kg_cards.append(KgKindergartenCard(
            id=kg.id,
            name_ar=kg.name_ar,
            name_en=kg.name_en,
            governorate=kg.governorate,
            district=kg.district,
            children_count=kids,
            attendance_rate=att_rate,
            attendance_status=att_status,
            capacity_utilization=cap_util,
            capacity_status=cap_status,
            teachers_count=teachers,
            supervisor_gap=sup_gap,
            children_per_supervisor=cps,
            open_alerts=open_alerts,
            health_score=health,
            health_label_ar=health_ar,
            health_label_en=health_en,
            recommended_action_ar=action_ar,
            recommended_action_en=action_en,
            last_report_date=last_report_by_kg.get(kg.id),
            teacher_data_status=teacher_data_status,
        ))

    # Sort: critical first, then by open alerts desc, then by name
    severity_order = {"critical": 0, "at_risk": 1, "needs_attention": 2, "good": 3, "excellent": 4}
    kg_cards.sort(key=lambda x: (severity_order.get(x.health_score, 5), -x.open_alerts, x.name_ar))

    # Aggregates (based on filtered kg_cards when governorate/risk_level is set)
    total_children = sum(c.children_count for c in kg_cards)
    kg_card_ids = [c.id for c in kg_cards]
    # Network attendance is sum(attended)/sum(expected) over the filtered set, not the
    # mean of the per-kindergarten percentages: averaging rates would weight a 3-child
    # kindergarten the same as a 300-child one. It divided window-wide PRESENT rows by
    # a single-day headcount, so it scaled with the window exactly like the per-card
    # rate did.
    _attended_total = sum(attendance_components_by_kg.get(kid, (0, 0))[0] for kid in kg_card_ids)
    _expected_total = sum(attendance_components_by_kg.get(kid, (0, 0))[1] for kid in kg_card_ids)
    avg_attendance = round((_attended_total / _expected_total * 100.0), 1) if _expected_total > 0 else 0.0
    total_teachers = sum(c.teachers_count for c in kg_cards)
    total_alerts = sum(c.open_alerts for c in kg_cards)

    # Alerts count by severity for KPIs — use only alerts that belong to the filtered set
    filtered_alert_objs: List[models.ActiveAlert] = []
    for kg in kg_cards:
        filtered_alert_objs.extend(alerts_by_kg.get(kg.id, []))
    for alert in gov_alerts:
        if not governorate or alert.scope_id == governorate:
            filtered_alert_objs.append(alert)
    for alert in network_alerts:
        filtered_alert_objs.append(alert)

    alert_severity_counts = Counter()
    alert_type_counts = Counter()
    for a in filtered_alert_objs:
        sev = a.severity.value if hasattr(a.severity, "value") else str(a.severity)
        alert_severity_counts[sev] += 1
        alert_type_counts[a.metric_type] += 1

    # Governorate comparison — attendance weighted by expected child-days
    # (sum(attended) / sum(expected) * 100), so a small kindergarten does not swing the
    # governorate the same as a large one.
    gov_data: Dict[str, Dict[str, Any]] = defaultdict(lambda: {
        "kindergartens": 0,
        "children": 0,
        "attended_sum": 0,
        "expected_sum": 0,
        "alerts": 0,
        "occupancy_sum": 0.0,
        "capacity_sum": 0,
        "enrolled_sum": 0,
    })
    for kg in active_kgs:
        gov_data[kg.governorate]["kindergartens"] += 1
        gov_data[kg.governorate]["children"] += children_by_kg.get(kg.id, 0)
        # Same story as the per-card and network rates: attended/expected child-days,
        # summed across the governorate. `present_sum / children` divided window-wide
        # rows by a single-day headcount and scaled with the window.
        _att, _exp = attendance_components_by_kg.get(kg.id, (0, 0))
        gov_data[kg.governorate]["attended_sum"] += _att
        gov_data[kg.governorate]["expected_sum"] += _exp
        gov_data[kg.governorate]["alerts"] += len(alerts_by_kg.get(kg.id, []))
        cap = capacity_by_kg.get(kg.id, {"total_capacity": 0, "total_enrolled": 0})
        gov_data[kg.governorate]["occupancy_sum"] += cap["total_enrolled"]
        gov_data[kg.governorate]["capacity_sum"] += cap["total_capacity"]
        gov_data[kg.governorate]["enrolled_sum"] += cap["total_enrolled"]

    governorate_chart = []
    for gov_name, d in sorted(gov_data.items()):
        att = round((d["attended_sum"] / d["expected_sum"] * 100.0), 1) if d["expected_sum"] > 0 else 0.0
        occ = round((d["enrolled_sum"] / d["capacity_sum"] * 100.0), 1) if d["capacity_sum"] > 0 else 0.0
        governorate_chart.append({
            "name": gov_name,
            "name_ar": gov_name,
            "kindergartens": d["kindergartens"],
            "children": d["children"],
            "attendance_rate": att,
            "occupancy_rate": occ,
            "alerts": d["alerts"],
        })

    # Charts
    attendance_chart = [
        {
            "name": c.name_ar,
            "name_ar": c.name_ar,
            "value": c.attendance_rate,
            "meta": {"status": c.attendance_status, "id": c.id}
        }
        for c in sorted(kg_cards, key=lambda x: x.attendance_rate)
    ]

    occupancy_chart = [
        {
            "name": c.name_ar,
            "name_ar": c.name_ar,
            "value": c.capacity_utilization,
            "meta": {"status": c.capacity_status, "id": c.id}
        }
        for c in sorted(kg_cards, key=lambda x: x.capacity_utilization, reverse=True)
    ]

    severity_chart = [
        {"name": sev, "name_ar": {"CRITICAL": "حرج", "HIGH": "عالي", "MEDIUM": "متوسط", "LOW": "منخفض"}.get(sev, sev), "value": cnt}
        for sev, cnt in sorted(alert_severity_counts.items())
    ]

    type_chart = [
        {"name": k, "name_ar": k, "value": v}
        for k, v in sorted(alert_type_counts.items(), key=lambda x: x[1], reverse=True)
    ]

    # Alert summaries for the alerts list
    alert_summaries: List[KgAlertSummary] = []
    for kg in kg_cards[:20]:
        for a in alerts_by_kg.get(kg.id, [])[:5]:
            sev = a.severity.value if hasattr(a.severity, "value") else str(a.severity)
            age = None
            if a.triggered_at:
                age = (now - a.triggered_at.replace(tzinfo=timezone.utc)).total_seconds() / 3600.0
            alert_summaries.append(KgAlertSummary(
                id=a.id,
                severity=sev,
                message=a.message,
                kindergarten_id=kg.id if a.scope_type == "KINDERGARTEN" else None,
                kindergarten_name_ar=kg.name_ar if a.scope_type == "KINDERGARTEN" else None,
                governorate=kg.governorate,
                metric=a.metric_type,
                current_value=float(a.current_value) if a.current_value is not None else 0.0,
                triggered_at=a.triggered_at.isoformat() if a.triggered_at else "",
                status=a.status.value if hasattr(a.status, "value") else str(a.status),
                age_hours=round(age, 1) if age is not None else None,
            ))

    alert_summaries.sort(key=lambda x: {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}.get(x.severity, 4))

    # Executive health counters
    critical_alerts = sum(1 for c in kg_cards if c.health_score == "critical")
    near_capacity_kgs = sum(1 for c in kg_cards if c.capacity_status == "near_capacity")
    below_target = sum(1 for c in kg_cards if c.attendance_status in ("below_target", "critical_low"))
    data_quality_issues = sum(1 for c in kg_cards if c.teacher_data_status == "needs_update")

    # KPI cards
    kpis = [
        KgKpiCard(
            title_ar="إجمالي الأطفال",
            title_en="Total Children",
            value=total_children,
            trend=_period_trend(period),
            target=None,
            status="good" if total_children > 0 else "warning",
            details=KgKpiDetail(
                active_kindergartens=len(active_kgs),
                note_ar=f"{len(active_kgs)} حضانات نشطة",
                note_en=f"{len(active_kgs)} active kindergartens",
            )
        ),
        KgKpiCard(
            title_ar="نسبة الحضور",
            title_en="Attendance Rate",
            value=avg_attendance,
            unit="%",
            trend=None,
            target=_ATTENDANCE_TARGET,
            status="critical" if avg_attendance < _ATTENDANCE_TARGET else ("warning" if avg_attendance < 80 else "good"),
            details=KgKpiDetail(
                note_ar=f"{below_target} حضانات أقل من الحد الأدنى {_ATTENDANCE_TARGET}%",
                note_en=f"{below_target} kindergartens below target {_ATTENDANCE_TARGET}%",
            )
        ),
        KgKpiCard(
            title_ar="المعلمات",
            title_en="Teachers",
            value=total_teachers,
            trend=None,
            target=None,
            status="good" if total_teachers > 0 else "warning",
            details=KgKpiDetail(
                note_ar="تحقق من نسبة الأطفال إلى المعلمات حسب الحضانة",
                note_en="Verify child-to-teacher ratio per kindergarten",
            )
        ),
        KgKpiCard(
            title_ar="التنبيهات",
            title_en="Alerts",
            value=total_alerts,
            trend=_period_trend(period),
            target=0,
            status="critical" if total_alerts > 20 else ("warning" if total_alerts > 5 else "good"),
            details=KgKpiDetail(
                note_ar=f"{alert_severity_counts.get('CRITICAL', 0)} حرجة، {alert_severity_counts.get('HIGH', 0)} عالية، {alert_severity_counts.get('MEDIUM', 0)} متوسطة",
                note_en=f"{alert_severity_counts.get('CRITICAL', 0)} critical, {alert_severity_counts.get('HIGH', 0)} high, {alert_severity_counts.get('MEDIUM', 0)} medium",
            )
        ),
    ]

    log_audit_event(
        db=db,
        action=AuditAction.ADMIN_DASHBOARD_VIEWED,
        actor=current_user,
        target_type="KindergartenOverview",
        target_ids=None,
        metadata={"kindergarten_count": len(kg_cards), "alert_count": total_alerts},
        sensitivity_level=2,
    )
    db.commit()

    return KgOverviewResponse(
        generated_at=now.isoformat(),
        kpis=kpis,
        kindergartens=kg_cards,
        charts={
            "attendance_by_kg": attendance_chart,
            "occupancy_pressure": occupancy_chart,
            "alerts_by_severity": severity_chart,
            "alerts_by_type": type_chart,
            "governorate_comparison": governorate_chart,
        },
        alerts=alert_summaries[:50],
        executive_health=KgExecutiveHealth(
            critical_alerts=critical_alerts,
            near_capacity_kgs=near_capacity_kgs,
            below_target_attendance=below_target,
            data_quality_issues=data_quality_issues,
        ),
        filters={"period": period, "governorate": governorate, "risk_level": risk_level},
    )


def _period_trend(period: str) -> str:
    if period == "today":
        return "مقارنة بالأمس"
    if period == "week":
        return "مقارنة بالأسبوع السابق"
    if period == "month":
        return "مقارنة بالشهر السابق"
    return "مقارنة بالفترة السابقة"


# =============================================================================
# Backup Management Endpoints
# =============================================================================

def _record_backup_failure(
    db: Session,
    actor: models.User,
    operation: str,
    **safe_metadata: Any,
) -> None:
    """Best-effort durable failure audit without storing raw exception details."""
    db.rollback()
    try:
        log_audit_event(
            db,
            AuditAction.BACKUP_FAILED,
            actor,
            "Backup",
            metadata={"operation": operation, **safe_metadata},
            sensitivity_level=3,
        )
        db.commit()
    except Exception as audit_exc:
        db.rollback()
        logger.error("Failed to persist backup failure audit: %s", audit_exc)

@router.post("/backup/create")
@limiter.limit(settings.RATE_LIMIT_ADMIN_WRITE)
def create_backup(
    request: Request,
    backup_type: str = Query("manual", description="Type of backup (manual, automated)"),
    include_uploads: bool = Query(True, description="Include uploaded files in backup"),
    include_config: bool = Query(True, description="Include configuration files in backup"),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Enqueue a backup job and return immediately (Admin only)."""

    from backup_tasks import run_backup
    external_succeeded = False

    try:
        log_audit_event(
            db,
            AuditAction.BACKUP_ENQUEUE_ATTEMPTED,
            current_user,
            "Backup",
            metadata={"backup_type": backup_type},
            sensitivity_level=2,
        )
        db.commit()
        task = run_backup.delay(
            backup_type=backup_type,
            include_uploads=include_uploads,
            include_config=include_config,
            triggered_by_user_id=current_user.id,
        )
        external_succeeded = True
        log_audit_event(
            db, AuditAction.BACKUP_ENQUEUED, current_user, "Backup",
            metadata={"backup_type": backup_type, "task_id": task.id},
            sensitivity_level=2,
        )
        db.commit()
        return {
            "message": f"{backup_type.title()} backup enqueued",
            "task_id": task.id,
            "status": "pending",
        }

    except HTTPException:
        raise
    except Exception as e:
        if external_succeeded:
            db.rollback()
            logger.error("Backup enqueue completed but its outcome audit failed: %s", e)
            raise HTTPException(
                status_code=500,
                detail="Backup was enqueued but audit recording failed; do not retry",
            )
        _record_backup_failure(db, current_user, "enqueue", backup_type=backup_type)
        logger.error("Failed to enqueue backup task: %s", e)
        raise HTTPException(status_code=500, detail="Failed to enqueue backup")


@router.get("/backup/list")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def list_backups(
    request: Request,
    backup_type: Optional[str] = Query(None, description="Filter by backup type (database, uploads, config)"),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """List all backups (Admin only)"""
    from backup_manager import backup_manager

    try:
        return backup_manager.list_backups(backup_type)
    except (OSError, IOError, ValueError, KeyError) as e:
        logger.error(f"Failed to list backups: {e}")
        raise HTTPException(status_code=500, detail="Failed to list backups")


class _RestoreConfirmBody(BaseModel):
    confirmation_token: Optional[str] = None


@router.post("/backup/restore/{backup_name}")
@limiter.limit(settings.RATE_LIMIT_ADMIN_WRITE)
def restore_backup(
    request: Request,
    backup_name: str,
    body: _RestoreConfirmBody = _RestoreConfirmBody(),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Restore from a backup (DANGER: overwrites current data).

    First call without a token returns a confirmation token and a warning.
    Second call with the token executes the restore.
    """

    # Sanitize backup_name: reject path separators, null bytes, and parent-directory
    # references. os.path.basename catches separator-based traversal; the ".."
    # substring check rejects names like "..evilfile" that look like traversal attempts.
    if not backup_name or ".." in backup_name or os.path.basename(backup_name) != backup_name:
        raise HTTPException(status_code=400, detail="Invalid backup name")

    from backup_manager import backup_manager
    external_succeeded = False

    try:
        if not backup_manager.validate_backup(backup_name):
            raise HTTPException(status_code=400, detail="Invalid or corrupted backup")

        metadata = backup_manager.get_backup_info(backup_name)
        if not metadata:
            raise HTTPException(status_code=404, detail="Backup not found")

        if metadata["type"] != "database":
            raise HTTPException(
                status_code=400,
                detail="Only database backups can be restored via API. Contact system administrator for other restore types.",
            )

        # First call: issue confirmation token
        if not body.confirmation_token:
            token = generate_confirmation_token("backup_restore", [backup_name], current_user.id)
            return {
                "requires_confirmation": True,
                "confirmation_token": token,
                "warning": (
                    f"Restoring backup '{backup_name}' will OVERWRITE all current data. "
                    "Re-submit with the confirmation_token to proceed."
                ),
            }

        # Second call: verify token then execute
        if not verify_confirmation_token(
            body.confirmation_token,
            "backup_restore",
            [backup_name],
            current_user.id,
        ):
            raise HTTPException(status_code=400, detail="Invalid or expired confirmation token")

        log_audit_event(
            db,
            AuditAction.BACKUP_RESTORE_ATTEMPTED,
            current_user,
            "Backup",
            metadata={"backup_name": backup_name},
            sensitivity_level=3,
        )
        db.commit()
        success = backup_manager.restore_database_backup(backup_name)

        if success:
            external_succeeded = True
            log_audit_event(
                db,
                AuditAction.BACKUP_RESTORED,
                current_user,
                "Backup",
                metadata={"backup_name": backup_name},
                sensitivity_level=3,
            )
            db.commit()
            return {"message": f"Database successfully restored from backup: {backup_name}"}
        else:
            _record_backup_failure(db, current_user, "restore", backup_name=backup_name)
            raise HTTPException(status_code=500, detail="Restore operation failed")

    except HTTPException:
        raise
    except Exception as e:
        if external_succeeded:
            db.rollback()
            logger.error("Backup restore completed but its outcome audit failed: %s", e)
            raise HTTPException(
                status_code=500,
                detail="Restore completed but audit recording failed; do not retry",
            )
        _record_backup_failure(db, current_user, "restore", backup_name=backup_name)
        logger.error(f"Backup restore failed: {e}")
        raise HTTPException(status_code=500, detail="Restore failed due to an internal error")


@router.delete("/backup/{backup_name}")
@limiter.limit(settings.RATE_LIMIT_ADMIN_WRITE)
def delete_backup(
    request: Request,
    backup_name: str,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Delete a backup file (Admin only)"""

    # Sanitize backup_name: reject path separators, null bytes, and parent-directory
    # references. os.path.basename catches separator-based traversal; the ".."
    # substring check rejects names like "..evilfile" that look like traversal attempts.
    if not backup_name or ".." in backup_name or os.path.basename(backup_name) != backup_name:
        raise HTTPException(status_code=400, detail="Invalid backup name")

    from backup_manager import backup_manager
    external_succeeded = False

    try:
        metadata = backup_manager.get_backup_info(backup_name)
        if not metadata:
            raise HTTPException(status_code=404, detail="Backup not found")

        log_audit_event(
            db,
            AuditAction.BACKUP_DELETE_ATTEMPTED,
            current_user,
            "Backup",
            metadata={"backup_name": backup_name},
            sensitivity_level=3,
        )
        db.commit()

        backup_path = metadata.get("backup_path")
        if backup_path and os.path.exists(backup_path):
            os.remove(backup_path)

        if backup_name in backup_manager.metadata:
            del backup_manager.metadata[backup_name]
            backup_manager._save_metadata()

        external_succeeded = True
        # Log the deletion
        log_audit_event(
            db, AuditAction.BACKUP_DELETED, current_user, "Backup",
            metadata={"backup_name": backup_name},
            sensitivity_level=2
        )
        db.commit()

        return {"message": f"Backup {backup_name} deleted successfully"}

    except HTTPException:
        raise
    except Exception as e:
        if external_succeeded:
            db.rollback()
            logger.error("Backup deletion completed but its outcome audit failed: %s", e)
            raise HTTPException(
                status_code=500,
                detail="Deletion completed but audit recording failed; do not retry",
            )
        _record_backup_failure(db, current_user, "delete", backup_name=backup_name)
        logger.error(f"Backup deletion failed: {e}")
        raise HTTPException(status_code=500, detail="Deletion failed")


@router.get("/backup/info/{backup_name}")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def get_backup_info(
    request: Request,
    backup_name: str,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Get detailed information about a specific backup (Admin only)"""
    # Sanitize backup_name: reject path separators, null bytes, and parent-directory
    # references. os.path.basename catches separator-based traversal; the ".."
    # substring check rejects names like "..evilfile" that look like traversal attempts.
    if not backup_name or ".." in backup_name or os.path.basename(backup_name) != backup_name:
        raise HTTPException(status_code=400, detail="Invalid backup name")

    from backup_manager import backup_manager

    try:
        metadata = backup_manager.get_backup_info(backup_name)
        if not metadata:
            raise HTTPException(status_code=404, detail="Backup not found")

        # Add validation status
        metadata_copy = metadata.copy()
        metadata_copy["is_valid"] = backup_manager.validate_backup(backup_name)

        return metadata_copy
    except HTTPException:
        raise
    except (OSError, IOError, KeyError, ValueError) as e:
        logger.error(f"Failed to get backup info: {e}")
        raise HTTPException(status_code=500, detail="Failed to retrieve backup info")


@router.post("/backup/cleanup")
@limiter.limit(settings.RATE_LIMIT_ADMIN_WRITE)
def cleanup_old_backups(
    request: Request,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Clean up old backups beyond retention period (Admin only)"""

    from backup_manager import backup_manager
    external_succeeded = False

    try:
        log_audit_event(
            db,
            AuditAction.BACKUP_CLEANUP_ATTEMPTED,
            current_user,
            "Backup",
            metadata={"operation": "cleanup_old_backups"},
            sensitivity_level=2,
        )
        db.commit()
        backup_manager.cleanup_old_backups()
        external_succeeded = True

        # Log the cleanup
        log_audit_event(
            db, AuditAction.BACKUP_CLEANUP, current_user, "Backup",
            metadata={"action": "cleanup_old_backups"},
            sensitivity_level=1
        )
        db.commit()

        return {"message": "Old backups cleaned up successfully"}

    except HTTPException:
        raise
    except Exception as e:
        if external_succeeded:
            db.rollback()
            logger.error("Backup cleanup completed but its outcome audit failed: %s", e)
            raise HTTPException(
                status_code=500,
                detail="Cleanup completed but audit recording failed; do not retry",
            )
        _record_backup_failure(db, current_user, "cleanup")
        logger.error(f"Backup cleanup failed: {e}")
        raise HTTPException(status_code=500, detail="Cleanup failed")


@router.post("/backup/validate/{backup_name}")
@limiter.limit(settings.RATE_LIMIT_ADMIN_WRITE)
def validate_backup(
    request: Request,
    backup_name: str,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Validate a backup file integrity (Admin only)"""

    # Sanitize backup_name: reject path separators, null bytes, and parent-directory
    # references. os.path.basename catches separator-based traversal; the ".."
    # substring check rejects names like "..evilfile" that look like traversal attempts.
    if not backup_name or ".." in backup_name or os.path.basename(backup_name) != backup_name:
        raise HTTPException(status_code=400, detail="Invalid backup name")

    from backup_manager import backup_manager

    try:
        is_valid = backup_manager.validate_backup(backup_name)

        return {
            "backup_name": backup_name,
            "is_valid": is_valid,
            "message": "Backup is valid" if is_valid else "Backup is corrupted or missing"
        }

    except HTTPException:
        raise
    except (OSError, IOError, ValueError, KeyError) as e:
        logger.error(f"Backup validation failed: {e}")
        raise HTTPException(status_code=500, detail="Validation failed")


# =============================================================================
# Kindergarten Excel Import
# =============================================================================

class KindergartenImportResult(BaseModel):
    """Result summary for kindergarten Excel import."""
    inserted: int = 0
    skipped_duplicate: int = 0
    skipped_empty: int = 0
    errors: List[Dict[str, Any]] = []
    total_rows: int = 0
    inserted_records: List[Dict[str, Any]] = []


@router.post("/kindergartens/import-excel", response_model=KindergartenImportResult)
@limiter.limit(settings.RATE_LIMIT_CSV_IMPORT)
def import_kindergartens_from_excel(
    request: Request,
    file: UploadFile = File(...),
    dry_run: bool = Query(False, description="Preview without writing to DB"),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Import kindergartens from an Excel (.xlsx) file (Admin only).

    Expected columns in the first sheet:
      - Column A: اسم الحضانة (عربي)  → name_ar
      - Column B: اسم الحضانة (إنجليزي) → name_en
      - Column C: المحافظة           → governorate
      - Column D: اللواء             → district
      - Column E: المنطقة            → area
      - Column F: العنوان التفصيلي    → address_line
      - Column G: رقم الهاتف         → contact_phone
    """


    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")

    try:
        import openpyxl
    except ImportError:  # pragma: no cover — openpyxl is a required dependency; kept for optional-dependency safety in minimal environments
        raise HTTPException(status_code=500, detail="openpyxl is not installed on the server")

    # Enforce file size limit (10 MB)
    MAX_UPLOAD_SIZE = settings.MAX_UPLOAD_SIZE_MB * 1024 * 1024
    if file.size is not None and file.size > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=413, detail="File too large")
    try:
        contents = file.file.read()
        if len(contents) > MAX_UPLOAD_SIZE:  # pragma: no cover — requires a >10 MB upload; content-length header check fires first for normal clients
            raise HTTPException(status_code=413, detail="File too large")
        validate_xlsx_archive(contents, max_compressed_bytes=MAX_UPLOAD_SIZE)
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(contents), read_only=True)
        ws = wb.worksheets[0]  # Use first sheet
        if ws.max_column > 100:
            wb.close()
            raise HTTPException(status_code=413, detail="Workbook exceeds the 100-column import limit")
        rows = []
        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row_number > 50_001:
                raise HTTPException(status_code=413, detail="Workbook exceeds the 50,000-row import limit")
            if any(len(str(value)) > 10_000 for value in row if value is not None):
                raise HTTPException(status_code=422, detail=f"Cell value is too long at row {row_number}")
            rows.append(row)
        wb.close()
    except (OSError, IOError, KeyError, ValueError, IndexError, BadZipFile):
        logger.exception("Failed to read uploaded Excel file")
        raise HTTPException(status_code=400, detail="Could not read Excel file")

    # Build existing-kindergarten set for dedup
    existing = set()
    for kg in db.query(
        models.Kindergarten.name_ar,
        models.Kindergarten.governorate,
        models.Kindergarten.district,
    ).all():
        existing.add((kg.name_ar, kg.governorate, kg.district))

    result = KindergartenImportResult(total_rows=len(rows))
    row_errors: List[Dict[str, Any]] = []

    def _clean(val) -> str:
        return str(val).strip() if val is not None else ""

    for row_num, row in enumerate(rows, start=2):
        if len(row) < 7:
            row_errors.append({"row": row_num, "error": "Row has fewer than 7 columns"})
            continue

        name_ar = _clean(row[0])
        name_en = _clean(row[1])
        governorate = _clean(row[2])
        district = _clean(row[3])
        area = _clean(row[4]) or "غير محدد"
        address_line = _clean(row[5]) or "غير محدد"
        phone = _clean(row[6])

        # name_ar, governorate, district, and phone are marked "Required" in
        # the admin UI's column guide -- previously only name_ar was actually
        # enforced, and blank governorate/district/phone were silently
        # replaced with placeholder values ("غير محدد"/"غير متوفر") and
        # inserted into production data anyway, contradicting what the UI
        # told admins.
        if not name_ar or not governorate or not district or not phone:
            result.skipped_empty += 1
            continue

        key = (name_ar, governorate, district)
        if key in existing:
            result.skipped_duplicate += 1
            continue

        if not dry_run:
            try:
                kg = models.Kindergarten(
                    name_ar=name_ar,
                    name_en=name_en or None,
                    governorate=governorate,
                    district=district,
                    area=area,
                    address_line=address_line,
                    contact_phone=phone,
                    status=models.KindergartenStatus.DRAFT,
                )
                db.add(kg)
            except (SQLAlchemyError, AttributeError, ValueError, KeyError) as exc:
                logger.error(f"Excel import row {row_num} error: {exc}")
                row_errors.append({"row": row_num, "error": "Failed to insert row"})
                continue

        existing.add(key)
        result.inserted += 1
        result.inserted_records.append({
            "name_ar": name_ar,
            "name_en": name_en or None,
            "governorate": governorate,
            "district": district,
            "area": area,
            "phone": phone,
        })

    result.errors = row_errors
    logger.info(
        "Kindergarten Excel import: inserted=%d, dup=%d, empty=%d, errors=%d, dry_run=%s",
        result.inserted, result.skipped_duplicate, result.skipped_empty,
        len(row_errors), dry_run,
    )

    if not dry_run:
        try:
            log_audit_event(
                db,
                AuditAction.KINDERGARTEN_IMPORT,
                current_user,
                target_type="kindergarten",
                target_ids=[],
                metadata={"imported_count": result.inserted, "errors": len(row_errors)},
                sensitivity_level=2,
            )
            db.add(models.ImportLog(
                file_name=file.filename,
                total_rows=len(rows),
                imported_count=result.inserted,
                skipped_count=result.skipped_duplicate + result.skipped_empty,
                errors_json=row_errors or None,
            ))
            db.commit()
        except Exception:
            db.rollback()
            logger.exception("Failed to commit kindergarten import")
            raise HTTPException(status_code=500, detail="Database commit failed")

    return result


# =============================================================================
# Governance — Daily Report KPI Endpoints
# =============================================================================

from governance_kpi_service import (
    compute_governance_funnel,
    compute_timeliness_metrics,
    compute_quality_metrics,
    compute_consistency_index,
    compute_fair_ranking,
    detect_low_performers,
    check_reminder_cooldown,
    send_governance_reminder,
    compute_full_gqi,
)
from governance_quality_service import GovernanceQualityService as _GovQualitySvc


class GovernanceReminderRequest(BaseModel):
    target_type: str  # "kindergarten" or "supervisor"
    target_id: int
    reminder_type: str = "low_submission_rate"


# ---------------------------------------------------------------------------
# The governance handlers below are deliberately `def`, not `async def`.
#
# They talk to the database through sync SQLAlchemy and contain no `await`.
# Declared `async def`, Starlette runs them directly on the event loop, so each
# blocking query stalls every other request in the process -- and the governance
# page fires five of these at once. Measured on production: /trend answers in
# 0.22s alone but reported 3.03s inside that batch, because it was queued behind
# the other four. The page waited for the sum of the work, not the slowest part
# of it.
#
# As plain `def`, FastAPI runs them in its threadpool and they overlap properly.
# Concurrency is then bounded by the connection pool (DB_POOL_SIZE 10 +
# DB_MAX_OVERFLOW 20) rather than by the single event loop; a request past that
# waits on DB_POOL_TIMEOUT instead of erroring, and 30 connections sits well
# under this cluster's max_connections of 100.
#
# Adding an `await` to any of these means making it `async def` again -- and
# then the blocking DB work has to move off the loop some other way.
# tests/test_admin_governance.py::TestGovernanceEndpointsRunInThreadpool
# guards this.
# ---------------------------------------------------------------------------


@router.get("/governance/kpis")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def get_governance_kpis(
    request: Request,
    start_date: date = Query(..., description="Start date YYYY-MM-DD"),
    end_date: date = Query(..., description="End date YYYY-MM-DD"),
    kindergarten_id: Optional[int] = Query(None, description="Filter by kindergarten"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    """Get governance funnel KPIs for daily report compliance monitoring."""
    if end_date < start_date:
        raise HTTPException(status_code=400, detail="end_date must be greater than or equal to start_date")

    funnel = compute_governance_funnel(db, start_date, end_date, kindergarten_id)
    timeliness = compute_timeliness_metrics(db, start_date, end_date, kindergarten_id)
    quality = compute_quality_metrics(db, start_date, end_date, kindergarten_id)
    consistency = compute_consistency_index(db, start_date, end_date, kindergarten_id)
    gqi = compute_full_gqi(db, start_date, end_date, kindergarten_id)

    _qsvc = _GovQualitySvc()
    rejection = _qsvc.report_rejection_rate(db, kindergarten_id)
    morning = _qsvc.submission_timing_distribution(db, kindergarten_id)

    return {
        "start_date": str(start_date),
        "end_date": str(end_date),
        "funnel": funnel,
        "timeliness": timeliness,
        "quality": quality,
        "consistency": consistency,
        "gqi": gqi,
        "rejection_rate": rejection.get("rejection_rate", 0.0),
        "morning_rate": morning.get("morning_rate", 0.0),
    }


@router.get("/governance/leaderboard")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def get_governance_leaderboard(
    request: Request,
    start_date: date = Query(..., description="Start date YYYY-MM-DD"),
    end_date: date = Query(..., description="End date YYYY-MM-DD"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    """Get Bayesian-ranked kindergarten leaderboard for daily report compliance."""
    if end_date < start_date:
        raise HTTPException(status_code=400, detail="end_date must be greater than or equal to start_date")

    funnel = compute_governance_funnel(db, start_date, end_date)
    ranked = compute_fair_ranking(funnel, db)
    low_performers = detect_low_performers(funnel, db)
    consistency = compute_consistency_index(db, start_date, end_date)

    kg_ids = [e["kindergarten_id"] for e in ranked]

    if kg_ids:
        # Last report date per KG
        _last_date_rows = (
            db.query(models.DailyReport.kindergarten_id, func.max(models.DailyReport.date))
            .filter(models.DailyReport.kindergarten_id.in_(kg_ids))
            .group_by(models.DailyReport.kindergarten_id)
            .all()
        )
        _last_dates = {kg_id: d for kg_id, d in _last_date_rows}

        # Reminder count per KG (all time)
        _reminder_rows = (
            db.query(models.GovernanceReminder.target_id, func.count(models.GovernanceReminder.id))
            .filter(
                models.GovernanceReminder.target_type == "kindergarten",
                models.GovernanceReminder.target_id.in_(kg_ids),
            )
            .group_by(models.GovernanceReminder.target_id)
            .all()
        )
        _reminder_counts = {kg_id: c for kg_id, c in _reminder_rows}

        # Rejection rate per KG (30 days)
        _cutoff = datetime.now(_JORDAN_TZ) - timedelta(days=30)
        _total_rows = (
            db.query(models.DailyReport.kindergarten_id, func.count(models.DailyReport.id))
            .filter(
                models.DailyReport.kindergarten_id.in_(kg_ids),
                models.DailyReport.created_at >= _cutoff,
                models.DailyReport.status.in_([
                    models.DailyReportStatus.SUBMITTED,
                    models.DailyReportStatus.APPROVED,
                    models.DailyReportStatus.REJECTED,
                    models.DailyReportStatus.RETURNED,
                    models.DailyReportStatus.SENT_TO_PARENT,
                ]),
            )
            .group_by(models.DailyReport.kindergarten_id)
            .all()
        )
        _rejected_rows = (
            db.query(models.DailyReport.kindergarten_id, func.count(models.DailyReport.id))
            .filter(
                models.DailyReport.kindergarten_id.in_(kg_ids),
                models.DailyReport.created_at >= _cutoff,
                models.DailyReport.status.in_([
                    models.DailyReportStatus.REJECTED,
                    models.DailyReportStatus.RETURNED,
                ]),
            )
            .group_by(models.DailyReport.kindergarten_id)
            .all()
        )
        _total_submitted = {kg_id: c for kg_id, c in _total_rows}
        _rejected = {kg_id: c for kg_id, c in _rejected_rows}

        # Morning rate per KG. This used to select (kindergarten_id,
        # created_at) for every report in the window and tally hours in
        # Python. created_at reflects when the row was written, not when the
        # report was filed, so in production all 378k daily_reports fall
        # inside the 30-day cutoff and every one of them crossed the wire on
        # each page load -- ~2.6s, against a ~250ms scan. Aggregate to one row
        # per kindergarten instead.
        #
        # EXTRACT(hour FROM ...) renders natively on PostgreSQL and via
        # STRFTIME on SQLite, so unlike the julianday expression noted below
        # it is safe on both.
        from collections import defaultdict
        _hour = func.extract("hour", models.DailyReport.created_at)
        _timing_rows = (
            db.query(
                models.DailyReport.kindergarten_id,
                func.count(models.DailyReport.id).label("total"),
                func.sum(case((_hour < 10, 1), else_=0)).label("morning"),
            )
            .filter(
                models.DailyReport.kindergarten_id.in_(kg_ids),
                models.DailyReport.created_at >= _cutoff,
                models.DailyReport.created_at.isnot(None),
            )
            .group_by(models.DailyReport.kindergarten_id)
            .all()
        )
        _morning_rates = {
            _kg_id: round(int(_morning or 0) / _total * 100, 1) if _total else 0.0
            for _kg_id, _total, _morning in _timing_rows
        }

        # Average approval hours per KG. An earlier version reached for
        # julianday() unconditionally -- SQLite has it, PostgreSQL does not,
        # and that took the whole governance page down in production. The
        # lesson was to check the dialect, not to avoid SQL: the delta is
        # averaged in the database on PostgreSQL and in Python everywhere
        # else.
        _approval_filters = (
            models.DailyReport.kindergarten_id.in_(kg_ids),
            models.DailyReport.created_at >= _cutoff,
            models.DailyReport.status.in_([
                models.DailyReportStatus.APPROVED,
                models.DailyReportStatus.SENT_TO_PARENT,
            ]),
            models.DailyReport.approved_at.isnot(None),
            models.DailyReport.submitted_at.isnot(None),
        )

        if db.bind is not None and db.bind.dialect.name == "postgresql":
            # 323k of the 378k rows match these filters, so pulling the two
            # timestamps per row to average them in Python cost ~2.8s per
            # page load. EXTRACT(epoch FROM interval) has no SQLite
            # equivalent, hence the explicit dialect branch -- the Python path
            # below stays the reference implementation and is what the test
            # suite exercises.
            _avg_seconds = func.avg(
                func.extract(
                    "epoch",
                    models.DailyReport.approved_at - models.DailyReport.submitted_at,
                )
            )
            _avg_approval = {
                _kg_id: round(float(_secs) / 3600, 1)
                for _kg_id, _secs in (
                    db.query(models.DailyReport.kindergarten_id, _avg_seconds)
                    .filter(*_approval_filters)
                    .group_by(models.DailyReport.kindergarten_id)
                    .all()
                )
                if _secs is not None
            }
        else:
            _approval_rows = (
                db.query(
                    models.DailyReport.kindergarten_id,
                    models.DailyReport.approved_at,
                    models.DailyReport.submitted_at,
                )
                .filter(*_approval_filters)
                .all()
            )
            _approval_hours = defaultdict(list)
            from datetime import timezone as _tz
            for _kg_id, _approved_at, _submitted_at in _approval_rows:
                if _approved_at.tzinfo is None:
                    _approved_at = _approved_at.replace(tzinfo=_tz.utc)
                if _submitted_at.tzinfo is None:
                    _submitted_at = _submitted_at.replace(tzinfo=_tz.utc)
                _approval_hours[_kg_id].append(
                    (_approved_at - _submitted_at).total_seconds() / 3600
                )
            _avg_approval = {
                _kg_id: round(sum(_hours) / len(_hours), 1)
                for _kg_id, _hours in _approval_hours.items()
                if _hours
            }

        # Consistency index per KG from existing computation
        _consistency_per_kg = consistency.get("per_kindergarten", {})

        # Enrich leaderboard entries
        for entry in ranked:
            _kid = entry["kindergarten_id"]
            _tot = _total_submitted.get(_kid, 0)
            _rej = _rejected.get(_kid, 0)
            entry["rejection_rate"] = round(_rej / _tot * 100, 1) if _tot > 0 else 0.0
            entry["morning_rate"] = _morning_rates.get(_kid, 0.0)
            entry["last_report_date"] = _last_dates[_kid].isoformat() if _last_dates.get(_kid) else None
            entry["reminder_count"] = _reminder_counts.get(_kid, 0)
            entry["avg_approval_hours"] = _avg_approval.get(_kid)
            _ci = _consistency_per_kg.get(str(_kid)) or _consistency_per_kg.get(_kid) or {}
            entry["consistency_index"] = round(_ci.get("consistency_index", 0.0) * 100, 1) if _ci else None

    return {
        "start_date": str(start_date),
        "end_date": str(end_date),
        "leaderboard": ranked,
        "low_performers": low_performers,
    }


@router.get("/governance/trend")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def get_governance_trend(
    request: Request,
    days: int = Query(30, ge=7, le=90, description="Number of days for trend (7–90)"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    """Return daily submission rate for the last N days (trend line data)."""
    today = datetime.now(_JORDAN_TZ).date()
    start = today - timedelta(days=days)

    active_count = (
        db.query(func.count(models.Kindergarten.id))
        .filter(models.Kindergarten.status == models.KindergartenStatus.ACTIVE)
        .scalar()
        or 0
    )

    if active_count == 0:
        return {"trend": [], "days": days}

    _submitted_by_date = dict(
        db.query(models.DailyReport.date, func.count(models.DailyReport.id))
        .filter(models.DailyReport.date >= start, models.DailyReport.date <= today)
        .group_by(models.DailyReport.date)
        .all()
    )

    trend = []
    for i in range(days):
        d = start + timedelta(days=i + 1)
        submitted = _submitted_by_date.get(d, 0)
        trend.append({
            "date": d.isoformat(),
            "submitted": submitted,
            "required": active_count,
            "submission_rate": round(submitted / active_count, 4),
        })

    return {"trend": trend, "days": days}


@router.get("/governance/safeguarding")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def get_governance_safeguarding(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    """Return low-performing kindergartens that also have open incidents (safeguarding overlap)."""
    today = datetime.now(_JORDAN_TZ).date()
    week_ago = today - timedelta(days=7)

    funnel = compute_governance_funnel(db, week_ago, today)
    low_performers = detect_low_performers(funnel, db)
    lp_ids = {lp["kindergarten_id"] for lp in low_performers}

    if not lp_ids:
        return {"overlapping": [], "low_performer_count": 0, "open_incident_count": 0}

    _open_rows = (
        db.query(models.Incident.kindergarten_id, func.count(models.Incident.id))
        .filter(
            models.Incident.kindergarten_id.in_(lp_ids),
            models.Incident.closed_at.is_(None),
            models.Incident.deleted_at.is_(None),
        )
        .group_by(models.Incident.kindergarten_id)
        .all()
    )
    _incident_counts = {kg_id: c for kg_id, c in _open_rows}

    overlapping_ids = list(_incident_counts.keys())
    if not overlapping_ids:
        return {
            "overlapping": [],
            "low_performer_count": len(lp_ids),
            "open_incident_count": 0,
        }

    _kg_names = {
        kg.id: {"name_ar": kg.name_ar, "name_en": kg.name_en}
        for kg in db.query(models.Kindergarten).filter(models.Kindergarten.id.in_(overlapping_ids)).all()
    }
    _lp_lookup = {lp["kindergarten_id"]: lp for lp in low_performers}

    overlapping = sorted(
        [
            {
                "kindergarten_id": kg_id,
                "name_ar": _kg_names.get(kg_id, {}).get("name_ar", f"KG#{kg_id}"),
                "name_en": _kg_names.get(kg_id, {}).get("name_en", f"KG#{kg_id}"),
                "submission_rate": _lp_lookup.get(kg_id, {}).get("submission_rate", 0.0),
                "open_incidents": _incident_counts[kg_id],
            }
            for kg_id in overlapping_ids
            if kg_id in _lp_lookup
        ],
        key=lambda x: x["open_incidents"],
        reverse=True,
    )

    return {
        "overlapping": overlapping,
        "low_performer_count": len(lp_ids),
        "open_incident_count": sum(_incident_counts.values()),
    }


@router.post("/governance/reminders")
@limiter.limit(settings.RATE_LIMIT_ADMIN_WRITE)
def send_governance_reminder_endpoint(
    request: Request,
    body: GovernanceReminderRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    """Send a governance reminder to a kindergarten or supervisor."""

    if body.target_type not in ("kindergarten", "supervisor"):
        raise HTTPException(status_code=400, detail="target_type must be 'kindergarten' or 'supervisor'")

    if body.target_type == "kindergarten":
        target_exists = db.query(models.Kindergarten.id).filter(
            models.Kindergarten.id == body.target_id,
            models.Kindergarten.deleted_at.is_(None),
        ).first()
    else:
        target_exists = db.query(models.User.id).filter(
            models.User.id == body.target_id,
            models.User.role == models.UserRole.SUPERVISOR,
            models.User.deleted_at.is_(None),
        ).first()
    if not target_exists:
        raise HTTPException(status_code=404, detail="Reminder target not found")

    can_send, last_sent_at = check_reminder_cooldown(db, body.target_type, body.target_id)
    if not can_send:
        raise HTTPException(
            status_code=429,
            detail={
                "message": (
                    f"Reminder cannot be sent yet due to cooldown. "
                    f"Last sent at {last_sent_at.isoformat() if last_sent_at else 'unknown'}"
                ),
                "cooldown_hours": settings.GOVERNANCE_REMINDER_COOLDOWN_HOURS,
                "last_sent_at": last_sent_at.isoformat() if last_sent_at else None,
            },
        )

    # Build metrics snapshot for the reminder payload
    today = datetime.now(_JORDAN_TZ).date()
    week_ago = today - timedelta(days=7)
    kg_id = body.target_id if body.target_type == "kindergarten" else None
    funnel = compute_governance_funnel(db, week_ago, today, kg_id)
    metrics_snapshot = funnel.get("aggregate", {}) if not kg_id else funnel.get("per_kindergarten", {}).get(kg_id, {})

    try:
        log_audit_event(
            db=db,
            action=AuditAction.GOVERNANCE_REMINDER_ATTEMPTED,
            actor=current_user,
            target_type="GovernanceReminder",
            target_ids=body.target_id,
            after_state={"target_type": body.target_type, "reminder_type": body.reminder_type},
            sensitivity_level=2,
        )
        db.commit()
    except Exception:
        db.rollback()
        raise

    reminder = send_governance_reminder(
        db=db,
        admin_user=current_user,
        target_type=body.target_type,
        target_id=body.target_id,
        reminder_type=body.reminder_type,
        metrics_snapshot=metrics_snapshot,
    )

    log_audit_event(
        db=db,
        action=AuditAction.GOVERNANCE_REMINDER_SENT,
        actor=current_user,
        target_type="GovernanceReminder",
        target_ids=reminder.id,
        after_state={"target_type": body.target_type, "target_id": body.target_id, "reminder_type": body.reminder_type},
        sensitivity_level=2,
    )
    db.commit()

    return {
        "id": reminder.id,
        "target_type": reminder.target_type,
        "target_id": reminder.target_id,
        "reminder_type": reminder.reminder_type,
        "sent_at": reminder.sent_at.isoformat(),
        "cooldown_expires_at": reminder.cooldown_expires_at.isoformat(),
    }


@router.get("/governance/reminders")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def list_governance_reminders(
    request: Request,
    target_type: Optional[str] = Query(None),
    target_id: Optional[int] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    """List governance reminders with optional filters and pagination."""
    q = db.query(models.GovernanceReminder).order_by(models.GovernanceReminder.sent_at.desc())

    if target_type:
        q = q.filter(models.GovernanceReminder.target_type == target_type)
    if target_id is not None:
        q = q.filter(models.GovernanceReminder.target_id == target_id)

    total = q.count()
    items = q.offset((page - 1) * page_size).limit(page_size).all()

    # Batch-resolve governorate for both target types (no N+1): kindergarten
    # reminders resolve directly; supervisor reminders resolve via the
    # supervisor's own assigned kindergarten.
    kg_ids = {r.target_id for r in items if r.target_type == "kindergarten"}
    supervisor_ids = {r.target_id for r in items if r.target_type == "supervisor"}
    if supervisor_ids:
        supervisor_kg_by_id = dict(
            db.query(models.User.id, models.User.kindergarten_id)
            .filter(models.User.id.in_(supervisor_ids))
            .all()
        )
        kg_ids |= {kg_id for kg_id in supervisor_kg_by_id.values() if kg_id is not None}
    else:
        supervisor_kg_by_id = {}
    governorate_by_kg_id = dict(
        db.query(models.Kindergarten.id, models.Kindergarten.governorate)
        .filter(models.Kindergarten.id.in_(kg_ids))
        .all()
    ) if kg_ids else {}

    def _resolve_governorate(r):
        if r.target_type == "kindergarten":
            return governorate_by_kg_id.get(r.target_id)
        if r.target_type == "supervisor":
            kg_id = supervisor_kg_by_id.get(r.target_id)
            return governorate_by_kg_id.get(kg_id) if kg_id else None
        return None

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": [
            {
                "id": r.id,
                "target_type": r.target_type,
                "target_id": r.target_id,
                "reminder_type": r.reminder_type,
                "sent_by": r.sent_by,
                "sent_at": r.sent_at.isoformat() if r.sent_at else None,
                "cooldown_expires_at": r.cooldown_expires_at.isoformat() if r.cooldown_expires_at else None,
                "payload": r.payload,
                "governorate": _resolve_governorate(r),
            }
            for r in items
        ],
    }


# (Duplicated import endpoint removed in favor of canonical /import-excel)


@router.get("/kindergartens/imported")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
async def list_imported_kindergartens(
    request: Request,
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    governorate: str = Query(None),
    district: str = Query(None),
    search: str = Query(None),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """List imported kindergartens with filtering and pagination."""

    service = KindergartenImportService(db)
    result = service.get_imported_kindergartens(
        page=page, per_page=per_page,
        governorate=governorate, district=district, search=search
    )

    return result


def _serialize_import_log(log: models.ImportLog) -> dict:
    error_count = len(log.errors_json) if log.errors_json else 0
    status_val = "FAILED" if error_count > 0 and log.imported_count == 0 else (
        "PARTIAL" if error_count > 0 else "SUCCESS"
    )
    return {
        "id": log.id,
        "import_type": "EXCEL_KINDERGARTENS",
        "filename": log.file_name,
        "imported_by": None,
        "total_rows": log.total_rows,
        "success_count": log.imported_count,
        "error_count": error_count,
        "status": status_val,
        "created_at": log.created_at.isoformat() if log.created_at else None,
    }


@router.get("/imports/logs")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
async def list_import_logs(
    request: Request,
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None, alias="from"),
    date_to: Optional[str] = Query(None, alias="to"),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """List import logs with database-side filtering and pagination."""
    query = db.query(models.ImportLog).order_by(models.ImportLog.created_at.desc())

    if date_from:
        try:
            query = query.filter(models.ImportLog.created_at >= datetime.fromisoformat(date_from))
        except ValueError:
            raise HTTPException(status_code=422, detail="Invalid 'from' date; use ISO-8601")
    if date_to:
        try:
            query = query.filter(models.ImportLog.created_at <= datetime.fromisoformat(date_to))
        except ValueError:
            raise HTTPException(status_code=422, detail="Invalid 'to' date; use ISO-8601")

    if type and type != "EXCEL_KINDERGARTENS":
        query = query.filter(models.ImportLog.id == -1)

    has_errors = and_(models.ImportLog.errors_json.is_not(None), models.ImportLog.errors_json != [])
    if status:
        normalized_status = status.upper()
        if normalized_status == "SUCCESS":
            query = query.filter(or_(models.ImportLog.errors_json.is_(None), models.ImportLog.errors_json == []))
        elif normalized_status == "FAILED":
            query = query.filter(has_errors, models.ImportLog.imported_count == 0)
        elif normalized_status == "PARTIAL":
            query = query.filter(has_errors, models.ImportLog.imported_count > 0)
        else:
            raise HTTPException(status_code=422, detail="Invalid status; use SUCCESS, PARTIAL, or FAILED")

    total = query.order_by(None).count()
    offset = (page - 1) * per_page
    rows = query.offset(offset).limit(per_page).all()
    page_items = [_serialize_import_log(log) for log in rows]

    return {
        "logs": page_items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page if total else 0,
    }


@router.get("/imports/logs/{log_id}")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
async def get_import_log_detail(
    request: Request,
    log_id: int,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Get detail for a single import log including per-row errors."""
    log = db.query(models.ImportLog).filter(models.ImportLog.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Import log not found")
    data = _serialize_import_log(log)
    data["errors"] = log.errors_json or []
    return data


@router.get("/governance/reminders/stats")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def get_governance_reminder_stats(
    request: Request,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Summary stats for the governance reminders dashboard."""
    today_start = datetime.now(_JORDAN_TZ).replace(hour=0, minute=0, second=0, microsecond=0)
    sent_today = db.query(func.count(models.GovernanceReminder.id)).filter(
        models.GovernanceReminder.sent_at >= today_start
    ).scalar() or 0
    total_sent = db.query(func.count(models.GovernanceReminder.id)).scalar() or 0
    # Only return values that are genuinely measured. The old response exposed
    # `pending` and `non_compliant` as hard-coded zeroes, which made "unknown"
    # look like a real operational result on the Admin page.
    return {"sent_today": sent_today, "total_sent": total_sent}


# =============================================================================


@router.get("/managers")
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def list_managers_for_impersonation(
    request: Request,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """List all active MANAGER users for the impersonation picker."""
    managers = (
        db.query(models.User)
        .filter(
            models.User.role == models.UserRole.MANAGER,
            models.User.status == models.UserStatus.ACTIVE,
            models.User.deleted_at.is_(None),
        )
        .all()
    )
    kg_ids = [m.kindergarten_id for m in managers if m.kindergarten_id]
    kg_map = {
        kg.id: kg
        for kg in db.query(models.Kindergarten).filter(models.Kindergarten.id.in_(kg_ids)).all()
    } if kg_ids else {}
    result = [
        {
            "id": m.id,
            "username": m.username,
            "name": m.full_name or m.username,
            "kindergarten_name": (
                (kg_map[m.kindergarten_id].name_ar or kg_map[m.kindergarten_id].name_en)
                if m.kindergarten_id and m.kindergarten_id in kg_map else None
            ),
        }
        for m in managers
    ]
    return {"managers": result}




_VALID_SEVERITIES = {s.value for s in models.SeverityLevel}
_VALID_STATUSES   = {s.value for s in models.AlertStatus}


@router.get("/alerts", response_model=AdminAlertsListResponse)
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def get_admin_alerts(
    request: Request,
    severity: Optional[str] = Query(None),
    governorate: Optional[str] = Query(None),
    # Was Query("ACTIVE") — the frontend's "All Statuses" option (value="")
    # omits this param entirely to mean "no filter", exactly like severity/
    # governorate below, but a non-None default silently forced every
    # "All Statuses" request back to ACTIVE-only. RESOLVED/ACKNOWLEDGED
    # alerts could never be shown through that control despite the option
    # text promising otherwise.
    status: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None, description="ISO date — lower bound on triggered_at"),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Get admin alerts with filtering, full threshold & kindergarten details."""
    try:
        query = db.query(models.ActiveAlert)

        if severity:
            if severity.upper() not in _VALID_SEVERITIES:
                raise HTTPException(
                    status_code=422,
                    detail=f"Invalid severity '{severity}'. Use: {', '.join(_VALID_SEVERITIES)}"
                )
            query = query.filter(models.ActiveAlert.severity == models.SeverityLevel(severity.upper()))

        if status:
            if status.upper() not in _VALID_STATUSES:
                raise HTTPException(
                    status_code=422,
                    detail=f"Invalid status '{status}'. Use: {', '.join(_VALID_STATUSES)}"
                )
            query = query.filter(models.ActiveAlert.status == models.AlertStatus(status.upper()))

        if date_from:
            try:
                dt_from = datetime.fromisoformat(date_from)
                query = query.filter(models.ActiveAlert.triggered_at >= dt_from)
            except ValueError:
                raise HTTPException(status_code=422, detail=f"Invalid date_from: '{date_from}'")

        # Governorate filter: match GOVERNORATE-scoped alerts directly, and KINDERGARTEN
        # alerts where the kindergarten sits in that governorate.
        if governorate:
            gov_normalized = settings.JORDAN_GOVERNORATE_ALIASES.get(governorate, governorate)
            kg_id_strings = [
                str(row.id)
                for row in db.query(models.Kindergarten.id).filter(
                    governorate_filter(models.Kindergarten.governorate, gov_normalized)
                ).all()
            ]
            query = query.filter(
                or_(
                    and_(
                        models.ActiveAlert.scope_type == "GOVERNORATE",
                        models.ActiveAlert.scope_id == gov_normalized,
                    ),
                    and_(
                        models.ActiveAlert.scope_type == "KINDERGARTEN",
                        models.ActiveAlert.scope_id.in_(kg_id_strings),
                    ),
                )
            )

        total = query.count()
        alerts = (
            query
            .order_by(
                models.ActiveAlert.severity.desc(),
                models.ActiveAlert.triggered_at.desc(),
            )
            .offset(skip)
            .limit(limit)
            .all()
        )

        # Batch-load kindergartens for KINDERGARTEN-scoped alerts (avoids N+1).
        kg_ids: Set[int] = set()
        threshold_ids: Set[int] = set()
        for alert in alerts:
            threshold_ids.add(alert.threshold_id)
            if alert.scope_type == "KINDERGARTEN" and alert.scope_id:
                try:
                    kg_ids.add(int(alert.scope_id))
                except (TypeError, ValueError):
                    pass

        # threshold_id → threshold_value
        threshold_map: Dict[int, float] = {}
        if threshold_ids:
            for t_id, t_val in db.query(
                models.AlertThreshold.id, models.AlertThreshold.threshold_value
            ).filter(models.AlertThreshold.id.in_(threshold_ids)).all():
                threshold_map[t_id] = float(t_val)

        # kg_id → (name_ar, name_en, governorate)
        kg_map: Dict[int, models.Kindergarten] = {}
        if kg_ids:
            for kg in db.query(models.Kindergarten).filter(
                models.Kindergarten.id.in_(kg_ids)
            ).all():
                kg_map[kg.id] = kg

        alerts_data: List[AdminAlertResponse] = []
        for alert in alerts:
            governorate_name: Optional[str] = None
            kg_name: Optional[str] = None

            if alert.scope_type == "GOVERNORATE":
                governorate_name = alert.scope_id
            elif alert.scope_type == "KINDERGARTEN" and alert.scope_id:
                try:
                    kg = kg_map.get(int(alert.scope_id))
                    if kg:
                        governorate_name = kg.governorate
                        kg_name = kg.name_ar or kg.name_en
                except (TypeError, ValueError):
                    pass

            alerts_data.append(AdminAlertResponse(
                id=alert.id,
                severity=alert.severity.value if hasattr(alert.severity, "value") else str(alert.severity),
                governorate=governorate_name,
                kindergarten_name=kg_name,
                metric=alert.metric_type,
                current_value=float(alert.current_value) if alert.current_value is not None else 0.0,
                threshold=threshold_map.get(alert.threshold_id),
                triggered_at=alert.triggered_at.isoformat() if alert.triggered_at else "",
                acknowledged_at=alert.acknowledged_at.isoformat() if alert.acknowledged_at else None,
                acknowledged_by_id=alert.acknowledged_by,
                status=alert.status.value if hasattr(alert.status, "value") else str(alert.status),
                message=alert.message,
                scope_type=alert.scope_type,
                scope_id=alert.scope_id,
            ))

        return AdminAlertsListResponse(
            alerts=alerts_data,
            total=total,
            page=(skip // limit) + 1,
            page_size=limit,
        )

    except HTTPException:
        raise
    except (SQLAlchemyError, AttributeError, ValueError, TypeError) as e:
        logger.error(f"Failed to get admin alerts: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.patch("/alerts/{alert_id}/acknowledge", response_model=AdminAlertResponse)
@limiter.limit(settings.RATE_LIMIT_ADMIN_WRITE)
def acknowledge_alert(
    request: Request,
    alert_id: int,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Mark an active alert as acknowledged."""

    alert = db.query(models.ActiveAlert).filter(models.ActiveAlert.id == alert_id).first()
    if not alert:
        raise HTTPException(status_code=404, detail="Alert not found")
    if alert.status == models.AlertStatus.ACKNOWLEDGED:
        raise HTTPException(status_code=409, detail="Alert is already acknowledged")

    now = datetime.now(_JORDAN_TZ)
    alert.status = models.AlertStatus.ACKNOWLEDGED
    alert.acknowledged_by = current_user.id
    alert.acknowledged_at = now
    try:
        log_audit_event(
            db,
            AuditAction.ALERT_ACKNOWLEDGED,
            current_user,
            "ActiveAlert",
            target_ids=alert_id,
            metadata={"metric": alert.metric_type, "severity": alert.severity.value if hasattr(alert.severity, "value") else str(alert.severity)},
            sensitivity_level=2,
        )
        db.commit()
    except Exception:
        db.rollback()
        raise
    db.refresh(alert)

    # Build response (mirrors get_admin_alerts logic for the single record)
    threshold_val: Optional[float] = None
    if alert.threshold_id:
        t = db.query(models.AlertThreshold.threshold_value).filter(
            models.AlertThreshold.id == alert.threshold_id
        ).scalar()
        if t is not None:
            threshold_val = float(t)

    governorate_name: Optional[str] = None
    kg_name: Optional[str] = None
    if alert.scope_type == "GOVERNORATE":
        governorate_name = alert.scope_id
    elif alert.scope_type == "KINDERGARTEN" and alert.scope_id:
        try:
            kg = db.query(models.Kindergarten).filter(
                models.Kindergarten.id == int(alert.scope_id)
            ).first()
            if kg:
                governorate_name = kg.governorate
                kg_name = kg.name_ar or kg.name_en
        except (TypeError, ValueError):
            pass

    return AdminAlertResponse(
        id=alert.id,
        severity=alert.severity.value if hasattr(alert.severity, "value") else str(alert.severity),
        governorate=governorate_name,
        kindergarten_name=kg_name,
        metric=alert.metric_type,
        current_value=float(alert.current_value) if alert.current_value is not None else 0.0,
        threshold=threshold_val,
        triggered_at=alert.triggered_at.isoformat() if alert.triggered_at else "",
        acknowledged_at=alert.acknowledged_at.isoformat() if alert.acknowledged_at else None,
        acknowledged_by_id=alert.acknowledged_by,
        status=alert.status.value if hasattr(alert.status, "value") else str(alert.status),
        message=alert.message,
        scope_type=alert.scope_type,
        scope_id=alert.scope_id,
    )


# =============================================================================
# Jordan Heat Map API (legacy alias — preferred entry point is
# `/api/admin/heat-map/*`, mounted from `heatmap.backend.admin_router`).
# This endpoint is kept for backward compatibility with the existing
# analytics page (`/admin/analytics`) which expects the slug-keyed payload.
# =============================================================================

class HeatmapGovernorateData(BaseModel):
    amman: Dict[str, Any] = {}
    irbid: Dict[str, Any] = {}
    zarqa: Dict[str, Any] = {}
    mafraq: Dict[str, Any] = {}
    jerash: Dict[str, Any] = {}
    ajloun: Dict[str, Any] = {}
    balqa: Dict[str, Any] = {}
    madaba: Dict[str, Any] = {}
    karak: Dict[str, Any] = {}
    # Canonical slug is "tafileh" (heatmap.backend.constants.GOVERNORATES). The
    # field was spelled "tafilah", so with extra='allow' the real data landed in
    # an undeclared "tafileh" key while this one stayed permanently {} — Tafilah
    # rendered blank for any client reading the declared field.
    # test_heatmap_data_endpoint pins these names to the canonical slug set.
    tafileh: Dict[str, Any] = {}
    maan: Dict[str, Any] = {}
    aqaba: Dict[str, Any] = {}

    model_config = ConfigDict(extra='allow')


class HeatmapResponse(BaseModel):
    data: HeatmapGovernorateData
    indicators: List[str]
    last_update: str
    summary: Optional[Dict[str, Any]] = None
    risk_legend: Optional[List[Dict[str, Any]]] = None


INDICATOR_LABELS = {
    'nursery_status': {'ar': 'حالة الحضانات', 'en': 'Nursery Status'},
    'children_registration': {'ar': 'الأطفال والتسجيل', 'en': 'Children Registration'},
    'staff_classrooms': {'ar': 'الموظفون والفصول', 'en': 'Staff & Classrooms'},
    'safety_incidents': {'ar': 'السلامة والحوادث', 'en': 'Safety & Incidents'},
    'reports_attendance': {'ar': 'التقارير والحضور', 'en': 'Reports & Attendance'},
    'tasks_governance': {'ar': 'المهام والحوكمة', 'en': 'Tasks & Governance'}
}


@router.get("/heatmap-data", response_model=HeatmapResponse)
@limiter.limit(settings.RATE_LIMIT_ADMIN_READ)
def get_heatmap_data(
    request: Request,
    indicator: Optional[str] = Query(None, description="Main indicator to display"),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Get heat map data for Jordan governorates (legacy slug-keyed payload).

    The new canonical endpoint is `/api/admin/heat-map/data` which returns a
    richer payload; this endpoint is kept for backward compatibility with
    the analytics page that consumes the slug-keyed format.
    """
    try:
        try:
            from heatmap.backend import service as heatmap_service
            overview = heatmap_service.get_map_overview(db)
        except Exception as svc_exc:
            logger.warning("Heatmap service unavailable, falling back to inline query: %s", svc_exc)
            overview = _fallback_map_overview(db)

        data: Dict[str, Dict[str, Any]] = {}
        for g in overview.get("governorates", []):
            indicators = g.get("main_indicators", {}) or {}

            def _num(value, default=0):
                """Coerce an indicator to a number, tolerating unavailable (None).

                dict.get(key, default) returns None when the key is present with a
                None value, so a plain .get(..., 0) does NOT protect the int()
                calls below. children_registration has been unavailable by design
                since 49b85238, which made this endpoint raise TypeError and
                return 500.
                """
                return default if value is None else value

            data[g["slug"]] = {
                "name": g.get("name_en", g["slug"].capitalize()),
                # Counts come from their own fields, not decoded back out of a
                # 0-100 indicator slot: nursery_status is a percentage, and
                # children_registration is an unavailable indicator, never a count.
                "kindergarten_count": int(_num(g.get("kg_count"))),
                "children_count": int(_num(g.get("student_count"))),
                # Unavailable stays unavailable: tasks_governance is legitimately
                # None when no governance score exists, and 0 is the *worst* band —
                # rendering "not measured" as "failing".
                "governance_score": indicators.get("tasks_governance"),
                # Real count from its own field, not decoded from a 0-100 score.
                "incidents_total": int(_num(g.get("incident_count"))),
                "risk_score": g.get("risk_score", 0),
                "last_update": overview.get("last_update"),
                "main_indicators": g.get("main_indicators", {}),
                "risk_level": g.get("risk_level", {}),
            }

        return HeatmapResponse(
            data=data,
            indicators=list(INDICATOR_LABELS.keys()),
            last_update=overview.get("last_update", now_amman().isoformat()),
            summary=overview.get("summary"),
            risk_legend=overview.get("risk_legend"),
        )

    except HTTPException:
        raise
    except (SQLAlchemyError, AttributeError, ValueError, TypeError) as e:
        logger.error(f"Failed to get heatmap data: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")


def _fallback_map_overview(db: Session) -> Dict[str, Any]:
    """
    Inline fallback when the heatmap service is not importable.

    P1-A fix: Kindergarten has NO governance_score column — that column lives on
    the separate GovernanceScore table (final_governance_score).  We join through
    GovernanceScore to compute the per-governorate average.  If GovernanceScore
    has no rows yet we default to 0.0, which produces a risk_score of 50 rather
    than crashing with AttributeError.
    """
    governorates = ['amman', 'irbid', 'zarqa', 'mafraq', 'jerash', 'ajloun',
                    'balqa', 'madaba', 'karak', 'tafileh', 'maan', 'aqaba']

    # Aggregate every metric once with GROUP BY (4 queries total, avoids the
    # previous 36-query N+1 loop of 3 queries per governorate).
    kg_counts: Dict[str, int] = dict(
        db.query(models.Kindergarten.governorate, func.count(models.Kindergarten.id))
        .group_by(models.Kindergarten.governorate)
        .all()
    )

    governance_avgs: Dict[str, float] = dict(
        db.query(
            models.Kindergarten.governorate,
            func.avg(models.GovernanceScore.final_governance_score),
        )
        .join(models.Kindergarten,
              models.GovernanceScore.kindergarten_id == models.Kindergarten.id)
        .group_by(models.Kindergarten.governorate)
        .all()
    )

    incident_counts: Dict[str, int] = dict(
        db.query(models.Kindergarten.governorate, func.count(models.Incident.id))
        .join(models.Kindergarten,
              models.Incident.kindergarten_id == models.Kindergarten.id)
        .group_by(models.Kindergarten.governorate)
        .all()
    )

    # Children registration: distinct active-enrolled children per governorate.
    children_counts: Dict[str, int] = dict(
        db.query(
            models.Kindergarten.governorate,
            func.count(func.distinct(models.EnrollmentApplication.child_id)),
        )
        .join(models.Kindergarten,
              models.EnrollmentApplication.kindergarten_id == models.Kindergarten.id)
        .filter(models.EnrollmentApplication.is_active.is_(True))
        .group_by(models.Kindergarten.governorate)
        .all()
    )

    data: List[Dict[str, Any]] = []
    for gov in governorates:
        gov_name = gov.capitalize()

        total_kgs = int(kg_counts.get(gov_name, 0) or 0)
        # Distinguish "no governance data" from a genuine 0 score: None is carried
        # into the indicator (0 is the worst band, so defaulting to it renders
        # un-assessed governorates as failing), while the risk math still needs a
        # number and treats absent data as neutral 0.
        raw_governance = governance_avgs.get(gov_name)
        avg_governance = None if raw_governance is None else float(raw_governance)
        incident_count = int(incident_counts.get(gov_name, 0) or 0)
        children_count = int(children_counts.get(gov_name, 0) or 0)

        risk_score = calculate_governorate_risk_score(avg_governance or 0.0, incident_count)
        data.append({
            "slug": gov,
            "name_en": gov_name,
            # Raw counts travel in their own fields, matching the heatmap
            # service's payload. They were previously smuggled through the
            # main_indicators slots (which are 0-100 scores), so the reader had
            # to decode a count back out of an indicator — and broke as soon as
            # children_registration became correctly unavailable.
            "kg_count": total_kgs,
            "student_count": children_count,
            "incident_count": incident_count,
            "main_indicators": {
                "tasks_governance": round(avg_governance, 1) if avg_governance is not None else None,
                "nursery_status": total_kgs,
                # children_registration is unavailable by design (no defensible
                # population denominator) — never fabricate it from a count.
                "children_registration": None,
                "safety_incidents": max(0, 100 - incident_count * 5),
            },
            "risk_score": risk_score,
            "risk_level": {
                "key": "low" if risk_score < 25 else "medium" if risk_score < 50 else "high" if risk_score < 75 else "critical",
                "name_en": "Low" if risk_score < 25 else "Medium" if risk_score < 50 else "High" if risk_score < 75 else "Critical",
                "name_ar": "منخفض" if risk_score < 25 else "متوسط" if risk_score < 50 else "مرتفع" if risk_score < 75 else "حرج",
                "color": "#22C55E" if risk_score < 25 else "#F59E0B" if risk_score < 50 else "#F97316" if risk_score < 75 else "#EF4444",
            },
        })
    return {
        "last_update": now_amman().isoformat(),
        "indicators": [{"key": k, **v} for k, v in INDICATOR_LABELS.items()],
        "governorates": data,
        "summary": {
            "total_governorates": len(data),
            "average_risk": round(sum(d["risk_score"] for d in data) / len(data), 1) if data else 0,
            "high_risk_count": sum(1 for d in data if d["risk_score"] >= 50),
            "critical_count": sum(1 for d in data if d["risk_score"] >= 75),
        },
        "risk_legend": [],
    }


def calculate_governorate_risk_score(governance_score: float, incident_count: int) -> int:
    """Calculate risk score for a governorate based on metrics."""
    score = 100 - governance_score
    if incident_count > 10:
        score += 20
    return min(max(int(score), 0), 100)


@router.get("/dashboard/critical-cases")
def get_critical_cases(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin)
):
    """Fetch the top critical kindergartens requiring enforcement based on governance score."""
    try:
        critical_kgs = db.query(models.Kindergarten, models.GovernanceScore).join(
            models.GovernanceScore, models.Kindergarten.id == models.GovernanceScore.kindergarten_id
        ).filter(
            models.GovernanceScore.final_governance_score <= 40
        ).order_by(models.GovernanceScore.final_governance_score.asc()).limit(10).all()

        results = []
        for kg, gov in critical_kgs:
            results.append({
                "id": kg.id,
                "name_ar": kg.name_ar,
                "name_en": kg.name_en,
                "governorate": kg.governorate,
                "risk_score": 100 - gov.final_governance_score if gov and gov.final_governance_score is not None else 100,
                "incident_count": 0,
                "unregistered_children": 0,
                "has_license": True,
                "principal_name": "",
                "phone_number": kg.contact_phone if hasattr(kg, "contact_phone") else "",
            })
    except Exception:
        critical_kgs = db.query(models.Kindergarten).limit(5).all()
        results = [
            {
                "id": kg.id, "name_ar": kg.name_ar, "name_en": kg.name_en,
                "governorate": kg.governorate, "risk_score": 85, "incident_count": 3,
                "unregistered_children": 0, "has_license": False, "principal_name": "",
                "phone_number": kg.contact_phone if hasattr(kg, "contact_phone") else "",
            }
            for kg in critical_kgs
        ]

    return {"critical_cases": results}


# =============================================================================
# Admin self-service profile endpoints
# =============================================================================

class AdminProfileUpdateSchema(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)
    full_name: Optional[str] = Field(None, min_length=1, max_length=255)
    email: Optional[str] = Field(None, min_length=1, max_length=255)
    phone_number: Optional[str] = Field(None, min_length=1, max_length=20)


class AdminPasswordChangeSchema(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)
    current_password: str = Field(..., min_length=1)
    new_password: str = Field(..., min_length=8)
    confirm_password: str = Field(..., min_length=8)


@router.put("/profile")
@limiter.limit("10/minute")
def update_admin_profile(
    request: Request,
    payload: AdminProfileUpdateSchema,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Update the logged-in admin's own profile fields."""

    before = {
        "full_name": current_user.full_name,
        "email": current_user.email,
        "phone_number": current_user.phone_number,
    }
    if payload.full_name is not None:
        current_user.full_name = payload.full_name
    if payload.email is not None:
        current_user.email = payload.email
    if payload.phone_number is not None:
        current_user.phone_number = payload.phone_number

    after = {
        "full_name": current_user.full_name,
        "email": current_user.email,
        "phone_number": current_user.phone_number,
    }
    try:
        log_audit_event(
            db, AuditAction.ADMIN_PROFILE_UPDATED, current_user, "User",
            target_ids=current_user.id,
            before_state=before,
            after_state=after,
            sensitivity_level=2,
        )
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {
        "message": "Profile updated",
        "full_name": current_user.full_name,
        "email": current_user.email,
        "phone_number": current_user.phone_number,
        "correlation_id": get_correlation_id(),
    }


@router.post("/profile/password")
@limiter.limit("5/minute")
def change_admin_password(
    request: Request,
    payload: AdminPasswordChangeSchema,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Allow the logged-in admin to change their own password."""

    if payload.new_password != payload.confirm_password:
        raise validation_error("New passwords do not match")

    if not verify_password(payload.current_password, current_user.hashed_password):
        log_audit_event(
            db, AuditAction.ADMIN_PASSWORD_CHANGE_FAILED, current_user, "User",
            target_ids=current_user.id,
            metadata={"reason": "Current password incorrect"},
            sensitivity_level=3,
        )
        db.commit()
        raise unauthenticated_error("Current password is incorrect")

    change_user_password(db, current_user, payload.new_password, commit=False)
    try:
        log_audit_event(
            db, AuditAction.ADMIN_PASSWORD_CHANGED, current_user, "User",
            target_ids=current_user.id,
            sensitivity_level=3,
        )
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {"message": "Password changed successfully", "correlation_id": get_correlation_id()}


@router.get("/profile/audit-logs")
@limiter.limit("30/minute")
def get_admin_own_audit_log(
    request: Request,
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Return audit log entries for the current admin user."""
    total = db.query(func.count(models.AuditLog.id)).filter(
        models.AuditLog.user_id == current_user.id
    ).scalar() or 0

    rows = (
        db.query(models.AuditLog)
        .filter(models.AuditLog.user_id == current_user.id)
        .order_by(models.AuditLog.created_at.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )

    events = []
    for row in rows:
        ts = row.created_at
        if ts and ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc).astimezone(_JORDAN_TZ)
        elif ts:
            ts = ts.astimezone(_JORDAN_TZ)
        events.append({
            "id": row.id,
            "action": row.action,
            "entity_type": row.entity_type,
            "entity_id": row.entity_id,
            "ip_address": row.ip_address,
            "sensitivity_level": row.sensitivity_level,
            "created_at": ts.isoformat() if ts else None,
        })

    return {"total": total, "events": events}


@router.post("/audit-logs/cleanup")
@limiter.limit(settings.RATE_LIMIT_ADMIN_WRITE)
def cleanup_audit_logs(
    request: Request,
    days: int = Query(90, ge=30, description="Delete audit log entries older than this many days"),
    dry_run: bool = Query(False, description="If true, return count without deleting"),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Delete audit log entries older than *days* days. Minimum 30 days. Admin only."""

    cutoff = datetime.now(_JORDAN_TZ) - timedelta(days=days)
    query = db.query(models.AuditLog).filter(models.AuditLog.created_at < cutoff)
    count = query.count()

    if dry_run:
        return {"dry_run": True, "would_delete": count, "cutoff": cutoff.isoformat()}

    query.delete(synchronize_session=False)
    try:
        log_audit_event(
            db,
            AuditAction.AUDIT_LOG_CLEANUP,
            current_user,
            "AuditLog",
            metadata={"deleted_count": count, "retention_days": days, "cutoff": cutoff.isoformat()},
            sensitivity_level=3,
        )
        db.commit()
    except Exception:
        db.rollback()
        raise

    return {"deleted": count, "cutoff": cutoff.isoformat(), "days": days}
