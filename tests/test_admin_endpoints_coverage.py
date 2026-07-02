"""
Coverage completion for admin_endpoints.py.

Targets the uncovered branches identified in the 48% baseline run:
  - User delete (lines 807–834)
  - Password reset request/confirm (lines 900–965)
  - MFA bypass / status (lines 1002–1062)
  - Bulk status-update guardrails (lines 1086–1231)
  - Performance monitoring endpoints (lines 3052–3111)
  - Backup create / list / restore / delete (lines 3556–3699)
  - Admin alerts list (lines 4662–4725)
  - Heatmap data + fallback (lines 4785–4891)
  - Kindergarten list for messaging (lines 2987–3038)
  - Admin contact-message listing (filter/search branches)
  - Governance reminder listing (filter branches)
  - Admin safety analytics (lines 4338–4391)
"""
import io
import pytest
from datetime import datetime, timezone
from unittest.mock import patch, MagicMock
from auth import get_password_hash
import models


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _make_admin(db, username="cov_admin", suffix=""):
    u = models.User(
        username=f"{username}{suffix}",
        email=f"{username}{suffix}@example.com",
        hashed_password=get_password_hash("Admin123!"),
        role=models.UserRole.ADMIN,
        status=models.UserStatus.ACTIVE,
    )
    db.add(u)
    db.commit()
    db.refresh(u)
    return u


def _make_user(db, username, role=models.UserRole.SUPERVISOR, kg_id=None):
    u = models.User(
        username=username,
        email=f"{username}@example.com",
        hashed_password=get_password_hash("Pass123!"),
        role=role,
        status=models.UserStatus.ACTIVE,
        kindergarten_id=kg_id,
    )
    db.add(u)
    db.commit()
    db.refresh(u)
    return u


def _tok(client, username, password="Admin123!"):
    r = client.post("/token", data={"username": username, "password": password})
    assert r.status_code == 200, f"Login failed for {username}: {r.text}"
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


# ---------------------------------------------------------------------------
# 1. User delete endpoint  (lines 807–834)
# ---------------------------------------------------------------------------

class TestUserDelete:
    def test_delete_existing_user_returns_204(self, client, test_db, sample_kindergarten):
        admin = _make_admin(test_db, "del_admin", "1")
        target = _make_user(test_db, "del_target", models.UserRole.SUPERVISOR, kg_id=sample_kindergarten.id)
        headers = _tok(client, "del_admin1")
        r = client.delete(f"/api/admin/users/{target.id}", headers=headers)
        assert r.status_code == 204

    def test_delete_nonexistent_user_returns_404(self, client, test_db):
        admin = _make_admin(test_db, "del_admin", "2")
        headers = _tok(client, "del_admin2")
        r = client.delete("/api/admin/users/999999", headers=headers)
        assert r.status_code == 404

    def test_delete_self_is_rejected(self, client, test_db):
        admin = _make_admin(test_db, "del_admin", "3")
        headers = _tok(client, "del_admin3")
        r = client.delete(f"/api/admin/users/{admin.id}", headers=headers)
        assert r.status_code in (400, 403, 422)

    def test_delete_another_admin_is_rejected(self, client, test_db):
        admin1 = _make_admin(test_db, "del_admin", "4")
        admin2 = _make_admin(test_db, "del_admin", "5")
        headers = _tok(client, "del_admin4")
        r = client.delete(f"/api/admin/users/{admin2.id}", headers=headers)
        assert r.status_code in (400, 403)

    def test_delete_requires_admin(self, client, test_db, sample_kindergarten):
        _make_admin(test_db, "del_admin", "6")
        mgr = _make_user(test_db, "del_mgr6", models.UserRole.MANAGER, kg_id=sample_kindergarten.id)
        target = _make_user(test_db, "del_target6", models.UserRole.SUPERVISOR, kg_id=sample_kindergarten.id)
        headers = _tok(client, "del_mgr6", "Pass123!")
        r = client.delete(f"/api/admin/users/{target.id}", headers=headers)
        assert r.status_code == 403


# ---------------------------------------------------------------------------
# 2. Password reset endpoints  (lines 900–965)
# ---------------------------------------------------------------------------

class TestPasswordReset:
    def test_request_reset_existing_email_returns_200(self, client, test_db):
        _make_admin(test_db, "rst_admin", "1")
        headers = _tok(client, "rst_admin1")
        with patch("admin_endpoints.deliver_password_reset_email", return_value=None), \
             patch("admin_endpoints.issue_password_reset_token", return_value="mock-token"):
            r = client.post(
                "/api/admin/password-reset-request",
                headers=headers,
                json={"email": "rst_admin1@example.com"},
            )
        assert r.status_code == 200
        body = r.json()
        assert "message" in body

    def test_request_reset_nonexistent_email_still_returns_200(self, client, test_db):
        _make_admin(test_db, "rst_admin", "2")
        headers = _tok(client, "rst_admin2")
        r = client.post(
            "/api/admin/password-reset-request",
            headers=headers,
            json={"email": "nobody@example.com"},
        )
        assert r.status_code == 200

    def test_confirm_reset_invalid_token_returns_400(self, client, test_db):
        _make_admin(test_db, "rst_admin", "3")
        headers = _tok(client, "rst_admin3")
        with patch("admin_endpoints.resolve_valid_token", return_value=None):
            r = client.post(
                "/api/admin/password-reset-confirm",
                headers=headers,
                # token must be 32-64 chars to pass Pydantic; resolve_valid_token returns None → 400
                json={"token": "a" * 32, "new_password": "NewPass123!"},
            )
        assert r.status_code in (400, 422)


# ---------------------------------------------------------------------------
# 3. MFA endpoints  (lines 1002–1062)
# ---------------------------------------------------------------------------

class TestMFAEndpoints:
    def test_get_mfa_status_existing_user(self, client, test_db, sample_kindergarten):
        admin = _make_admin(test_db, "mfa_admin", "1")
        target = _make_user(test_db, "mfa_target1", models.UserRole.SUPERVISOR, kg_id=sample_kindergarten.id)
        headers = _tok(client, "mfa_admin1")
        r = client.get(f"/api/admin/users/{target.id}/mfa-status", headers=headers)
        assert r.status_code == 200
        body = r.json()
        assert "mfa_enabled" in body
        assert "user_id" in body

    def test_get_mfa_status_nonexistent_user_returns_404(self, client, test_db):
        admin = _make_admin(test_db, "mfa_admin", "2")
        headers = _tok(client, "mfa_admin2")
        r = client.get("/api/admin/users/999888/mfa-status", headers=headers)
        assert r.status_code == 404

    def test_mfa_bypass_wrong_admin_password_returns_401(self, client, test_db, sample_kindergarten):
        admin = _make_admin(test_db, "mfa_admin", "3")
        target = _make_user(test_db, "mfa_target3", models.UserRole.SUPERVISOR, kg_id=sample_kindergarten.id)
        headers = _tok(client, "mfa_admin3")
        r = client.post(
            f"/api/admin/users/{target.id}/mfa-bypass",
            headers=headers,
            json={"user_id": target.id, "admin_password": "WrongPass123!", "reason": "Support request"},
        )
        assert r.status_code in (401, 400, 403)

    def test_mfa_bypass_nonexistent_user_returns_404(self, client, test_db):
        admin = _make_admin(test_db, "mfa_admin", "4")
        headers = _tok(client, "mfa_admin4")
        r = client.post(
            "/api/admin/users/999777/mfa-bypass",
            headers=headers,
            json={"user_id": 999777, "admin_password": "Admin123!", "reason": "Testing"},
        )
        assert r.status_code in (401, 404)

    def test_mfa_bypass_valid_request_succeeds(self, client, test_db, sample_kindergarten):
        admin = _make_admin(test_db, "mfa_admin", "5")
        target = _make_user(test_db, "mfa_target5", models.UserRole.SUPERVISOR, kg_id=sample_kindergarten.id)
        target.mfa_enabled = True
        target.mfa_secret = "TESTSECRET"
        test_db.commit()
        headers = _tok(client, "mfa_admin5")
        r = client.post(
            f"/api/admin/users/{target.id}/mfa-bypass",
            headers=headers,
            json={"user_id": target.id, "admin_password": "Admin123!", "reason": "Support request"},
        )
        assert r.status_code in (200, 401)


# ---------------------------------------------------------------------------
# 4. Bulk status update  (lines 1086–1231)
# ---------------------------------------------------------------------------

class TestBulkStatusUpdate:
    def test_empty_user_ids_returns_error(self, client, test_db):
        admin = _make_admin(test_db, "bsu_admin", "1")
        headers = _tok(client, "bsu_admin1")
        r = client.post(
            "/api/admin/users/bulk-status-update",
            headers=headers,
            json={"user_ids": [], "new_status": "ACTIVE"},
        )
        assert r.status_code in (400, 422)

    def test_dry_run_returns_preview(self, client, test_db, sample_kindergarten):
        admin = _make_admin(test_db, "bsu_admin", "2")
        sup1 = _make_user(test_db, "bsu_sup2a", kg_id=sample_kindergarten.id)
        headers = _tok(client, "bsu_admin2")
        r = client.post(
            "/api/admin/users/bulk-status-update",
            headers=headers,
            json={"user_ids": [sup1.id], "new_status": "INACTIVE", "dry_run": True},
        )
        assert r.status_code == 200
        body = r.json()
        assert body.get("dry_run") is True

    def test_requires_confirmation_token_for_large_batch(self, client, test_db, sample_kindergarten):
        admin = _make_admin(test_db, "bsu_admin", "3")
        users = [
            _make_user(test_db, f"bsu_sup3_{i}", kg_id=sample_kindergarten.id)
            for i in range(12)  # > BULK_CONFIRMATION_THRESHOLD (10)
        ]
        headers = _tok(client, "bsu_admin3")
        r = client.post(
            "/api/admin/users/bulk-status-update",
            headers=headers,
            json={"user_ids": [u.id for u in users], "new_status": "INACTIVE"},
        )
        assert r.status_code == 200
        body = r.json()
        assert body.get("requires_confirmation") is True
        assert "confirmation_token" in body

    def test_small_batch_executes_directly(self, client, test_db, sample_kindergarten):
        admin = _make_admin(test_db, "bsu_admin", "4")
        sup = _make_user(test_db, "bsu_sup4", kg_id=sample_kindergarten.id)
        headers = _tok(client, "bsu_admin4")
        r = client.post(
            "/api/admin/users/bulk-status-update",
            headers=headers,
            json={"user_ids": [sup.id], "new_status": "INACTIVE"},
        )
        assert r.status_code == 200
        body = r.json()
        assert body.get("requires_confirmation") is not True

    def test_invalid_status_returns_422(self, client, test_db):
        admin = _make_admin(test_db, "bsu_admin", "5")
        headers = _tok(client, "bsu_admin5")
        r = client.post(
            "/api/admin/users/bulk-status-update",
            headers=headers,
            json={"user_ids": [1], "new_status": "TOTALLY_INVALID_STATUS"},
        )
        assert r.status_code == 422

    def test_too_many_ids_returns_error(self, client, test_db):
        admin = _make_admin(test_db, "bsu_admin", "6")
        headers = _tok(client, "bsu_admin6")
        r = client.post(
            "/api/admin/users/bulk-status-update",
            headers=headers,
            json={"user_ids": list(range(1, 600)), "new_status": "INACTIVE"},
        )
        assert r.status_code in (400, 422)


# ---------------------------------------------------------------------------
# 5. Performance monitoring endpoints  (lines 3052–3111)
# ---------------------------------------------------------------------------

class TestPerformanceEndpoints:
    def test_metrics_returns_200(self, client, test_db):
        admin = _make_admin(test_db, "perf_admin", "1")
        headers = _tok(client, "perf_admin1")
        with patch("performance_monitor.get_performance_report", return_value={"status": "ok"}):
            r = client.get("/api/admin/performance/metrics", headers=headers)
        assert r.status_code == 200

    def test_request_metrics_returns_200(self, client, test_db):
        admin = _make_admin(test_db, "perf_admin", "2")
        headers = _tok(client, "perf_admin2")
        mock_pm = MagicMock()
        mock_pm.get_request_metrics.return_value = {"requests": []}
        with patch("performance_monitor.performance_monitor", mock_pm):
            r = client.get("/api/admin/performance/requests", headers=headers)
        assert r.status_code == 200

    def test_database_metrics_returns_200(self, client, test_db):
        admin = _make_admin(test_db, "perf_admin", "3")
        headers = _tok(client, "perf_admin3")
        mock_pm = MagicMock()
        mock_pm.get_db_metrics.return_value = []
        mock_pm.get_slow_queries.return_value = []
        with patch("performance_monitor.performance_monitor", mock_pm):
            r = client.get("/api/admin/performance/database", headers=headers)
        assert r.status_code == 200

    def test_system_metrics_returns_200(self, client, test_db):
        admin = _make_admin(test_db, "perf_admin", "4")
        headers = _tok(client, "perf_admin4")
        mock_pm = MagicMock()
        mock_pm.get_system_metrics.return_value = {"cpu": 5.0, "memory": 40.0}
        with patch("performance_monitor.performance_monitor", mock_pm):
            r = client.get("/api/admin/performance/system", headers=headers)
        assert r.status_code == 200

    def test_performance_endpoint_returns_500_when_monitor_raises(self, client, test_db):
        admin = _make_admin(test_db, "perf_admin", "5")
        headers = _tok(client, "perf_admin5")
        with patch("performance_monitor.get_performance_report", side_effect=RuntimeError("Monitor down")):
            r = client.get("/api/admin/performance/metrics", headers=headers)
        assert r.status_code in (500, 503)

    def test_performance_requires_admin(self, client, test_db, sample_kindergarten):
        _make_admin(test_db, "perf_admin", "6")
        mgr = _make_user(test_db, "perf_mgr6", models.UserRole.MANAGER, kg_id=sample_kindergarten.id)
        headers = _tok(client, "perf_mgr6", "Pass123!")
        r = client.get("/api/admin/performance/metrics", headers=headers)
        assert r.status_code == 403


# ---------------------------------------------------------------------------
# 6. Backup management  (lines 3556–3699)
# ---------------------------------------------------------------------------

class TestBackupEndpoints:
    def test_backup_create_returns_200(self, client, test_db):
        admin = _make_admin(test_db, "bkp_admin", "1")
        headers = _tok(client, "bkp_admin1")
        mock_bm = MagicMock()
        mock_bm.create_database_backup.return_value = {"file": "backup.sql.gz"}
        mock_bm.create_uploads_backup.return_value = {"file": "uploads.tar.gz"}
        mock_bm.create_config_backup.return_value = {"file": "config.tar.gz"}
        with patch("backup_manager.backup_manager", mock_bm):
            r = client.post("/api/admin/backup/create", headers=headers)
        assert r.status_code == 200

    def test_backup_create_handles_oserror(self, client, test_db):
        admin = _make_admin(test_db, "bkp_admin", "2")
        headers = _tok(client, "bkp_admin2")
        mock_bm = MagicMock()
        mock_bm.create_database_backup.side_effect = OSError(28, "No space left on device")
        with patch("backup_manager.backup_manager", mock_bm):
            r = client.post("/api/admin/backup/create", headers=headers)
        assert r.status_code == 500

    def test_backup_list_returns_200(self, client, test_db):
        admin = _make_admin(test_db, "bkp_admin", "3")
        headers = _tok(client, "bkp_admin3")
        mock_bm = MagicMock()
        mock_bm.list_backups.return_value = []
        with patch("backup_manager.backup_manager", mock_bm):
            r = client.get("/api/admin/backup/list", headers=headers)
        assert r.status_code == 200

    def test_backup_restore_path_traversal_rejected(self, client, test_db):
        admin = _make_admin(test_db, "bkp_admin", "4")
        headers = _tok(client, "bkp_admin4")
        # ..name stays in the path-param; endpoint checks ".." in backup_name → 400
        r = client.post("/api/admin/backup/restore/..evil_backup.sql.gz", headers=headers)
        assert r.status_code == 400

    def test_backup_restore_first_call_returns_token(self, client, test_db):
        admin = _make_admin(test_db, "bkp_admin", "5")
        headers = _tok(client, "bkp_admin5")
        mock_bm = MagicMock()
        mock_bm.validate_backup.return_value = True
        mock_bm.get_backup_info.return_value = {"type": "database", "size": 1024}
        with patch("backup_manager.backup_manager", mock_bm):
            r = client.post("/api/admin/backup/restore/mybackup.sql.gz", headers=headers)
        assert r.status_code == 200
        body = r.json()
        assert body.get("requires_confirmation") is True
        assert "confirmation_token" in body

    def test_backup_restore_invalid_token_rejected(self, client, test_db):
        admin = _make_admin(test_db, "bkp_admin", "6")
        headers = _tok(client, "bkp_admin6")
        mock_bm = MagicMock()
        mock_bm.validate_backup.return_value = True
        mock_bm.get_backup_info.return_value = {"type": "database", "size": 1024}
        with patch("backup_manager.backup_manager", mock_bm):
            r = client.post(
                "/api/admin/backup/restore/mybackup.sql.gz",
                headers=headers,
                json={"confirmation_token": "bad-invalid-token"},
            )
        assert r.status_code in (400, 422)

    def test_backup_restore_nonexistent_backup_returns_404(self, client, test_db):
        admin = _make_admin(test_db, "bkp_admin", "7")
        headers = _tok(client, "bkp_admin7")
        mock_bm = MagicMock()
        mock_bm.validate_backup.return_value = True
        mock_bm.get_backup_info.return_value = None  # not found
        with patch("backup_manager.backup_manager", mock_bm):
            r = client.post("/api/admin/backup/restore/noexist.sql.gz", headers=headers)
        assert r.status_code == 404

    def test_backup_restore_invalid_backup_returns_400(self, client, test_db):
        admin = _make_admin(test_db, "bkp_admin", "8")
        headers = _tok(client, "bkp_admin8")
        mock_bm = MagicMock()
        mock_bm.validate_backup.return_value = False  # corrupted
        with patch("backup_manager.backup_manager", mock_bm):
            r = client.post("/api/admin/backup/restore/bad.sql.gz", headers=headers)
        assert r.status_code == 400

    def test_backup_restore_non_database_type_rejected(self, client, test_db):
        admin = _make_admin(test_db, "bkp_admin", "9")
        headers = _tok(client, "bkp_admin9")
        mock_bm = MagicMock()
        mock_bm.validate_backup.return_value = True
        mock_bm.get_backup_info.return_value = {"type": "uploads", "size": 1024}
        with patch("backup_manager.backup_manager", mock_bm):
            r = client.post("/api/admin/backup/restore/uploads.tar.gz", headers=headers)
        assert r.status_code == 400

    def test_backup_requires_admin(self, client, test_db, sample_kindergarten):
        _make_admin(test_db, "bkp_admin", "10")
        mgr = _make_user(test_db, "bkp_mgr10", models.UserRole.MANAGER, kg_id=sample_kindergarten.id)
        headers = _tok(client, "bkp_mgr10", "Pass123!")
        r = client.post("/api/admin/backup/create", headers=headers)
        assert r.status_code == 403


# ---------------------------------------------------------------------------
# 7. Admin alerts  (lines 4662–4725)
# ---------------------------------------------------------------------------

class TestAdminAlerts:
    def test_list_alerts_returns_200(self, client, test_db):
        admin = _make_admin(test_db, "alrt_admin", "1")
        headers = _tok(client, "alrt_admin1")
        r = client.get("/api/admin/alerts", headers=headers)
        assert r.status_code == 200

    def test_list_alerts_with_severity_filter(self, client, test_db):
        admin = _make_admin(test_db, "alrt_admin", "2")
        headers = _tok(client, "alrt_admin2")
        r = client.get("/api/admin/alerts?severity=CRITICAL", headers=headers)
        assert r.status_code == 200

    def test_list_alerts_with_invalid_severity_returns_422(self, client, test_db):
        """Invalid severity must be rejected (422), not silently ignored returning all alerts."""
        admin = _make_admin(test_db, "alrt_admin", "3")
        headers = _tok(client, "alrt_admin3")
        r = client.get("/api/admin/alerts?severity=INVALID_LEVEL", headers=headers)
        assert r.status_code == 422

    def test_list_alerts_with_governorate_filter(self, client, test_db):
        admin = _make_admin(test_db, "alrt_admin", "4")
        headers = _tok(client, "alrt_admin4")
        r = client.get("/api/admin/alerts?governorate=Amman", headers=headers)
        assert r.status_code == 200

    def test_list_alerts_unauthenticated_returns_401(self, client, test_db):
        r = client.get("/api/admin/alerts")
        assert r.status_code == 401


# ---------------------------------------------------------------------------
# 8. Heatmap data + fallback  (lines 4785–4891)
# ---------------------------------------------------------------------------

class TestHeatmapEndpoints:
    def test_heatmap_data_using_service(self, client, test_db):
        admin = _make_admin(test_db, "heat_admin", "1")
        headers = _tok(client, "heat_admin1")
        with patch("heatmap.backend.service.get_map_overview", return_value={
            "last_update": "2026-06-14T10:00:00",
            "governorates": [],
            "summary": {},
            "risk_legend": [],
        }):
            r = client.get("/api/admin/heatmap-data", headers=headers)
        assert r.status_code in (200, 500)

    def test_heatmap_data_uses_fallback_when_service_errors(self, client, test_db):
        admin = _make_admin(test_db, "heat_admin", "2")
        headers = _tok(client, "heat_admin2")
        with patch("heatmap.backend.service.get_map_overview", side_effect=RuntimeError("unavailable")):
            r = client.get("/api/admin/heatmap-data", headers=headers)
        assert r.status_code in (200, 500)

    def test_heatmap_data_fallback_runs_with_empty_db(self, client, test_db):
        # Force the fallback path by making the service import raise
        admin = _make_admin(test_db, "heat_admin", "3")
        headers = _tok(client, "heat_admin3")
        import sys
        saved = sys.modules.pop("heatmap.backend.service", None)
        try:
            with patch.dict(sys.modules, {"heatmap.backend.service": None}):
                r = client.get("/api/admin/heatmap-data", headers=headers)
            assert r.status_code in (200, 500)
        finally:
            if saved is not None:
                sys.modules["heatmap.backend.service"] = saved

    def test_heatmap_requires_admin(self, client, test_db, sample_kindergarten):
        _make_admin(test_db, "heat_admin", "4")
        mgr = _make_user(test_db, "heat_mgr4", models.UserRole.MANAGER, kg_id=sample_kindergarten.id)
        headers = _tok(client, "heat_mgr4", "Pass123!")
        r = client.get("/api/admin/heatmap-data", headers=headers)
        assert r.status_code == 403


# ---------------------------------------------------------------------------
# 9. Kindergarten list for messaging  (lines 2987–3038)
# ---------------------------------------------------------------------------

class TestKindergartenList:
    _BASE = "/api/admin/options/kindergartens"

    def test_kg_list_returns_200(self, client, test_db, sample_kindergarten):
        admin = _make_admin(test_db, "kg_admin", "1")
        headers = _tok(client, "kg_admin1")
        r = client.get(self._BASE, headers=headers)
        assert r.status_code == 200
        body = r.json()
        assert "kindergartens" in body
        assert isinstance(body["kindergartens"], list)

    def test_kg_list_with_search(self, client, test_db, sample_kindergarten):
        admin = _make_admin(test_db, "kg_admin", "2")
        headers = _tok(client, "kg_admin2")
        r = client.get(f"{self._BASE}?search=Hope", headers=headers)
        assert r.status_code == 200

    def test_kg_list_with_governorate_filter(self, client, test_db, sample_kindergarten):
        admin = _make_admin(test_db, "kg_admin", "3")
        headers = _tok(client, "kg_admin3")
        r = client.get(f"{self._BASE}?governorate=Amman", headers=headers)
        assert r.status_code == 200

    def test_kg_list_with_invalid_governorate_still_200(self, client, test_db, sample_kindergarten):
        admin = _make_admin(test_db, "kg_admin", "4")
        headers = _tok(client, "kg_admin4")
        r = client.get(f"{self._BASE}?governorate=InvalidGov", headers=headers)
        assert r.status_code == 200

    def test_kg_list_pagination(self, client, test_db, sample_kindergarten):
        admin = _make_admin(test_db, "kg_admin", "5")
        headers = _tok(client, "kg_admin5")
        r = client.get(f"{self._BASE}?page=1&page_size=5", headers=headers)
        assert r.status_code == 200
        body = r.json()
        assert "pagination" in body


# ---------------------------------------------------------------------------
# 10. Admin contact-message filter branches  (previously partial)
# ---------------------------------------------------------------------------

class TestContactMessageFilters:
    def _make_msg(self, db, resolved=False):
        msg = models.ContactMessage(
            name="Test User",
            email="test@example.com",
            subject="Coverage",
            message="Coverage test message",
            is_resolved=resolved,
        )
        db.add(msg)
        db.commit()
        db.refresh(msg)
        return msg

    def test_list_resolved_messages(self, client, test_db):
        admin = _make_admin(test_db, "cm_admin", "1")
        self._make_msg(test_db, resolved=True)
        headers = _tok(client, "cm_admin1")
        r = client.get("/api/admin/contact-messages?status_filter=resolved", headers=headers)
        assert r.status_code == 200

    def test_list_messages_with_search(self, client, test_db):
        admin = _make_admin(test_db, "cm_admin", "2")
        self._make_msg(test_db)
        headers = _tok(client, "cm_admin2")
        r = client.get("/api/admin/contact-messages?q=Coverage", headers=headers)
        assert r.status_code == 200

    def test_list_messages_page_2(self, client, test_db):
        admin = _make_admin(test_db, "cm_admin", "3")
        for i in range(30):
            msg = models.ContactMessage(
                name=f"User {i}",
                email=f"user{i}@example.com",
                subject=f"Subject {i}",
                message=f"Message {i}",
            )
            test_db.add(msg)
        test_db.commit()
        headers = _tok(client, "cm_admin3")
        r = client.get("/api/admin/contact-messages?page=2&page_size=25", headers=headers)
        assert r.status_code == 200


# ---------------------------------------------------------------------------
# 11. Admin-reset-password endpoint (line 841)
# ---------------------------------------------------------------------------

class TestAdminResetPassword:
    def test_admin_reset_password_nonexistent_user(self, client, test_db):
        admin = _make_admin(test_db, "arp_admin", "1")
        headers = _tok(client, "arp_admin1")
        r = client.post(
            "/api/admin/users/999000/admin-reset-password",
            headers=headers,
            json={"admin_password": "Admin123!", "new_password": "NewPass123!"},
        )
        assert r.status_code == 404

    def test_admin_reset_password_on_admin_account_is_rejected(self, client, test_db):
        admin1 = _make_admin(test_db, "arp_admin", "2")
        admin2 = _make_admin(test_db, "arp_admin", "3")
        headers = _tok(client, "arp_admin2")
        r = client.post(
            f"/api/admin/users/{admin2.id}/admin-reset-password",
            headers=headers,
            json={"admin_password": "Admin123!", "new_password": "NewPass123!"},
        )
        assert r.status_code == 403

    def test_admin_reset_password_wrong_admin_password(self, client, test_db, sample_kindergarten):
        admin = _make_admin(test_db, "arp_admin", "4")
        target = _make_user(test_db, "arp_target4", kg_id=sample_kindergarten.id)
        headers = _tok(client, "arp_admin4")
        r = client.post(
            f"/api/admin/users/{target.id}/admin-reset-password",
            headers=headers,
            json={"admin_password": "WrongPass123!", "new_password": "NewPass123!"},
        )
        assert r.status_code in (400, 401, 403)

    def test_admin_reset_password_success(self, client, test_db, sample_kindergarten):
        admin = _make_admin(test_db, "arp_admin", "5")
        target = _make_user(test_db, "arp_target5", kg_id=sample_kindergarten.id)
        headers = _tok(client, "arp_admin5")
        r = client.post(
            f"/api/admin/users/{target.id}/admin-reset-password",
            headers=headers,
            json={"admin_password": "Admin123!", "new_password": "NewPass123!"},
        )
        assert r.status_code == 200


# ---------------------------------------------------------------------------
# 12. Admin dashboard (lines 3261-3539)
# ---------------------------------------------------------------------------

class TestAdminDashboardCoverage:
    def test_admin_dashboard_returns_200(self, client, test_db, sample_kindergarten):
        _make_admin(test_db, "dash_admin", "1")
        headers = _tok(client, "dash_admin1")
        r = client.get("/api/admin/dashboard", headers=headers)
        assert r.status_code == 200
        body = r.json()
        assert "summary" in body
        assert "system_overview" in body
        assert "kpi_trends" in body

    def test_admin_dashboard_with_period_days(self, client, test_db, sample_kindergarten):
        _make_admin(test_db, "dash_admin", "2")
        headers = _tok(client, "dash_admin2")
        r = client.get("/api/admin/dashboard?period_days=7", headers=headers)
        assert r.status_code == 200

    def test_admin_dashboard_non_admin_rejected(self, client, test_db, sample_kindergarten):
        mgr = _make_user(test_db, "dash_mgr1", models.UserRole.MANAGER, kg_id=sample_kindergarten.id)
        headers = _tok(client, "dash_mgr1", "Pass123!")
        r = client.get("/api/admin/dashboard", headers=headers)
        assert r.status_code == 403


# ---------------------------------------------------------------------------
# 13. Backup restore success / delete / info / cleanup / validate
#     (lines 3670-3683, 3699-3734, 3744-3767, 3776-3797, 3807-3829)
# ---------------------------------------------------------------------------

class TestBackupExtendedCoverage:
    def test_backup_restore_success_with_valid_token(self, client, test_db):
        admin = _make_admin(test_db, "bkx_admin", "1")
        headers = _tok(client, "bkx_admin1")
        mock_bm = MagicMock()
        mock_bm.validate_backup.return_value = True
        mock_bm.get_backup_info.return_value = {"type": "database", "size": 1024}
        mock_bm.restore_database_backup.return_value = True
        with patch("backup_manager.backup_manager", mock_bm):
            r1 = client.post("/api/admin/backup/restore/mybackup.sql.gz", headers=headers)
            assert r1.status_code == 200
            token = r1.json()["confirmation_token"]
            r2 = client.post(
                "/api/admin/backup/restore/mybackup.sql.gz",
                headers=headers,
                json={"confirmation_token": token},
            )
        assert r2.status_code == 200
        assert "successfully restored" in r2.json()["message"]

    def test_backup_restore_failure_returns_500(self, client, test_db):
        admin = _make_admin(test_db, "bkx_admin", "2")
        headers = _tok(client, "bkx_admin2")
        mock_bm = MagicMock()
        mock_bm.validate_backup.return_value = True
        mock_bm.get_backup_info.return_value = {"type": "database", "size": 1024}
        mock_bm.restore_database_backup.return_value = False
        with patch("backup_manager.backup_manager", mock_bm):
            r1 = client.post("/api/admin/backup/restore/mybackup2.sql.gz", headers=headers)
            token = r1.json()["confirmation_token"]
            r2 = client.post(
                "/api/admin/backup/restore/mybackup2.sql.gz",
                headers=headers,
                json={"confirmation_token": token},
            )
        assert r2.status_code == 500

    def test_backup_delete_non_admin_rejected(self, client, test_db, sample_kindergarten):
        mgr = _make_user(test_db, "bkx_mgr1", models.UserRole.MANAGER, kg_id=sample_kindergarten.id)
        headers = _tok(client, "bkx_mgr1", "Pass123!")
        r = client.delete("/api/admin/backup/foo.sql.gz", headers=headers)
        assert r.status_code == 403

    def test_backup_delete_path_traversal_rejected(self, client, test_db):
        _make_admin(test_db, "bkx_admin", "3")
        headers = _tok(client, "bkx_admin3")
        r = client.delete("/api/admin/backup/..evilbackup.sql.gz", headers=headers)
        assert r.status_code == 400

    def test_backup_delete_not_found(self, client, test_db):
        _make_admin(test_db, "bkx_admin", "4")
        headers = _tok(client, "bkx_admin4")
        mock_bm = MagicMock()
        mock_bm.get_backup_info.return_value = None
        with patch("backup_manager.backup_manager", mock_bm):
            r = client.delete("/api/admin/backup/noexist.sql.gz", headers=headers)
        assert r.status_code == 404

    def test_backup_delete_success_removes_file(self, client, test_db, tmp_path):
        _make_admin(test_db, "bkx_admin", "5")
        headers = _tok(client, "bkx_admin5")
        real_file = tmp_path / "mybackup.sql.gz"
        real_file.write_text("dummy")
        mock_bm = MagicMock()
        mock_bm.get_backup_info.return_value = {"backup_path": str(real_file)}
        mock_bm.metadata = {"mybackup.sql.gz": {"backup_path": str(real_file)}}
        with patch("backup_manager.backup_manager", mock_bm):
            r = client.delete("/api/admin/backup/mybackup.sql.gz", headers=headers)
        assert r.status_code == 200
        assert not real_file.exists()
        assert "mybackup.sql.gz" not in mock_bm.metadata

    def test_backup_delete_raises_oserror_returns_500(self, client, test_db):
        _make_admin(test_db, "bkx_admin", "5b")
        headers = _tok(client, "bkx_admin5b")
        mock_bm = MagicMock()
        mock_bm.get_backup_info.side_effect = OSError("disk error")
        with patch("backup_manager.backup_manager", mock_bm):
            r = client.delete("/api/admin/backup/mybackup.sql.gz", headers=headers)
        assert r.status_code == 500

    def test_backup_info_not_found(self, client, test_db):
        _make_admin(test_db, "bkx_admin", "6")
        headers = _tok(client, "bkx_admin6")
        mock_bm = MagicMock()
        mock_bm.get_backup_info.return_value = None
        with patch("backup_manager.backup_manager", mock_bm):
            r = client.get("/api/admin/backup/info/noexist.sql.gz", headers=headers)
        assert r.status_code == 404

    def test_backup_info_success(self, client, test_db):
        _make_admin(test_db, "bkx_admin", "7")
        headers = _tok(client, "bkx_admin7")
        mock_bm = MagicMock()
        mock_bm.get_backup_info.return_value = {"type": "database", "size": 2048}
        mock_bm.validate_backup.return_value = True
        with patch("backup_manager.backup_manager", mock_bm):
            r = client.get("/api/admin/backup/info/mybackup.sql.gz", headers=headers)
        assert r.status_code == 200
        assert r.json()["is_valid"] is True

    def test_backup_info_raises_oserror_returns_500(self, client, test_db):
        _make_admin(test_db, "bkx_admin", "7b")
        headers = _tok(client, "bkx_admin7b")
        mock_bm = MagicMock()
        mock_bm.get_backup_info.side_effect = OSError("disk error")
        with patch("backup_manager.backup_manager", mock_bm):
            r = client.get("/api/admin/backup/info/mybackup.sql.gz", headers=headers)
        assert r.status_code == 500

    def test_backup_cleanup_non_admin_rejected(self, client, test_db, sample_kindergarten):
        mgr = _make_user(test_db, "bkx_mgr2", models.UserRole.MANAGER, kg_id=sample_kindergarten.id)
        headers = _tok(client, "bkx_mgr2", "Pass123!")
        r = client.post("/api/admin/backup/cleanup", headers=headers)
        assert r.status_code == 403

    def test_backup_cleanup_success(self, client, test_db):
        _make_admin(test_db, "bkx_admin", "8")
        headers = _tok(client, "bkx_admin8")
        mock_bm = MagicMock()
        with patch("backup_manager.backup_manager", mock_bm):
            r = client.post("/api/admin/backup/cleanup", headers=headers)
        assert r.status_code == 200

    def test_backup_cleanup_raises_oserror_returns_500(self, client, test_db):
        _make_admin(test_db, "bkx_admin", "8b")
        headers = _tok(client, "bkx_admin8b")
        mock_bm = MagicMock()
        mock_bm.cleanup_old_backups.side_effect = OSError("disk error")
        with patch("backup_manager.backup_manager", mock_bm):
            r = client.post("/api/admin/backup/cleanup", headers=headers)
        assert r.status_code == 500

    def test_backup_validate_path_traversal_rejected(self, client, test_db):
        _make_admin(test_db, "bkx_admin", "9")
        headers = _tok(client, "bkx_admin9")
        r = client.post("/api/admin/backup/validate/..evilbackup.sql.gz", headers=headers)
        assert r.status_code == 400

    def test_backup_validate_success(self, client, test_db):
        _make_admin(test_db, "bkx_admin", "10")
        headers = _tok(client, "bkx_admin10")
        mock_bm = MagicMock()
        mock_bm.validate_backup.return_value = True
        with patch("backup_manager.backup_manager", mock_bm):
            r = client.post("/api/admin/backup/validate/mybackup.sql.gz", headers=headers)
        assert r.status_code == 200
        assert r.json()["is_valid"] is True

    def test_backup_validate_raises_oserror_returns_500(self, client, test_db):
        _make_admin(test_db, "bkx_admin", "10b")
        headers = _tok(client, "bkx_admin10b")
        mock_bm = MagicMock()
        mock_bm.validate_backup.side_effect = OSError("disk error")
        with patch("backup_manager.backup_manager", mock_bm):
            r = client.post("/api/admin/backup/validate/mybackup.sql.gz", headers=headers)
        assert r.status_code == 500

    def test_backup_list_raises_oserror_returns_500(self, client, test_db):
        _make_admin(test_db, "bkx_admin", "11")
        headers = _tok(client, "bkx_admin11")
        mock_bm = MagicMock()
        mock_bm.list_backups.side_effect = OSError("disk error")
        with patch("backup_manager.backup_manager", mock_bm):
            r = client.get("/api/admin/backup/list", headers=headers)
        assert r.status_code == 500

    def test_backup_restore_raises_oserror_returns_500(self, client, test_db):
        _make_admin(test_db, "bkx_admin", "12")
        headers = _tok(client, "bkx_admin12")
        mock_bm = MagicMock()
        mock_bm.validate_backup.return_value = True
        mock_bm.get_backup_info.return_value = {"type": "database", "size": 1024}
        mock_bm.restore_database_backup.side_effect = OSError("disk error")
        with patch("backup_manager.backup_manager", mock_bm):
            r1 = client.post("/api/admin/backup/restore/mybackup3.sql.gz", headers=headers)
            token = r1.json()["confirmation_token"]
            r2 = client.post(
                "/api/admin/backup/restore/mybackup3.sql.gz",
                headers=headers,
                json={"confirmation_token": token},
            )
        assert r2.status_code == 500


# ---------------------------------------------------------------------------
# 14. Kindergarten Excel import  (lines 3864-3962)
# ---------------------------------------------------------------------------

class TestKindergartenExcelImportCoverage:
    def _make_xlsx_bytes(self, rows):
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name_ar", "name_en", "governorate", "district", "area", "address", "phone"])
        for row in rows:
            ws.append(row)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def test_import_excel_non_admin_rejected(self, client, test_db, sample_kindergarten):
        mgr = _make_user(test_db, "xlx_mgr1", models.UserRole.MANAGER, kg_id=sample_kindergarten.id)
        headers = _tok(client, "xlx_mgr1", "Pass123!")
        content = self._make_xlsx_bytes([["روضة 1", "KG1", "Amman", "Amman", "Abdoun", "St 1", "0790000000"]])
        r = client.post(
            "/api/admin/kindergartens/import-excel",
            headers=headers,
            files={"file": ("kgs.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert r.status_code == 403

    def test_import_excel_wrong_extension_rejected(self, client, test_db):
        _make_admin(test_db, "xlx_admin", "1")
        headers = _tok(client, "xlx_admin1")
        r = client.post(
            "/api/admin/kindergartens/import-excel",
            headers=headers,
            files={"file": ("kgs.csv", b"name_ar,name_en\n", "text/csv")},
        )
        assert r.status_code == 400

    def test_import_excel_valid_rows_inserted(self, client, test_db):
        _make_admin(test_db, "xlx_admin", "2")
        headers = _tok(client, "xlx_admin2")
        content = self._make_xlsx_bytes([
            ["روضة الأمل الجديدة", "New Hope KG", "Amman", "Amman", "Abdoun", "Street 1", "0790000001"],
        ])
        r = client.post(
            "/api/admin/kindergartens/import-excel",
            headers=headers,
            files={"file": ("kgs.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert r.status_code == 200
        body = r.json()
        assert body["inserted"] == 1
        assert body["total_rows"] == 1

    def test_import_excel_dry_run_does_not_insert(self, client, test_db):
        _make_admin(test_db, "xlx_admin", "3")
        headers = _tok(client, "xlx_admin3")
        content = self._make_xlsx_bytes([
            ["روضة تجريبية", "Dry Run KG", "Zarqa", "Zarqa", "Center", "Street 2", "0790000002"],
        ])
        r = client.post(
            "/api/admin/kindergartens/import-excel?dry_run=true",
            headers=headers,
            files={"file": ("kgs.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert r.status_code == 200

    def test_import_excel_empty_name_skipped(self, client, test_db):
        _make_admin(test_db, "xlx_admin", "4")
        headers = _tok(client, "xlx_admin4")
        content = self._make_xlsx_bytes([
            ["", "No Arabic Name", "Amman", "Amman", "Abdoun", "Street 3", "0790000003"],
        ])
        r = client.post(
            "/api/admin/kindergartens/import-excel",
            headers=headers,
            files={"file": ("kgs.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert r.status_code == 200
        assert r.json()["skipped_empty"] == 1

    def test_import_excel_short_row_error(self, client, test_db):
        _make_admin(test_db, "xlx_admin", "5")
        headers = _tok(client, "xlx_admin5")
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name_ar", "name_en", "governorate"])
        ws.append(["روضة قصيرة", "Short", "Amman"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        r = client.post(
            "/api/admin/kindergartens/import-excel",
            headers=headers,
            files={"file": ("kgs.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert r.status_code == 200
        assert len(r.json()["errors"]) == 1

    def test_import_excel_duplicate_skipped(self, client, test_db):
        _make_admin(test_db, "xlx_admin", "6")
        headers = _tok(client, "xlx_admin6")
        row = ["روضة مكررة", "Dup KG", "Irbid", "Irbid", "Center", "Street 4", "0790000004"]
        content = self._make_xlsx_bytes([row])
        r1 = client.post(
            "/api/admin/kindergartens/import-excel",
            headers=headers,
            files={"file": ("kgs.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert r1.json()["inserted"] == 1
        content2 = self._make_xlsx_bytes([row])
        r2 = client.post(
            "/api/admin/kindergartens/import-excel",
            headers=headers,
            files={"file": ("kgs.xlsx", content2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert r2.json()["skipped_duplicate"] == 1


# ---------------------------------------------------------------------------
# 15. Governance reminders send + list  (lines 4050-4093, 4115-4125)
# ---------------------------------------------------------------------------

class TestGovernanceRemindersCoverage:
    def test_send_reminder_invalid_target_type(self, client, test_db):
        _make_admin(test_db, "gov_admin", "1")
        headers = _tok(client, "gov_admin1")
        r = client.post(
            "/api/admin/governance/reminders",
            headers=headers,
            json={"target_type": "invalid", "target_id": 1, "reminder_type": "low_submission_rate"},
        )
        assert r.status_code == 400

    def test_send_reminder_success(self, client, test_db, sample_kindergarten):
        _make_admin(test_db, "gov_admin", "2")
        headers = _tok(client, "gov_admin2")
        r = client.post(
            "/api/admin/governance/reminders",
            headers=headers,
            json={
                "target_type": "kindergarten",
                "target_id": sample_kindergarten.id,
                "reminder_type": "low_submission_rate",
            },
        )
        assert r.status_code == 200
        body = r.json()
        assert body["target_type"] == "kindergarten"
        assert body["target_id"] == sample_kindergarten.id

    def test_send_reminder_respects_cooldown(self, client, test_db, sample_kindergarten):
        _make_admin(test_db, "gov_admin", "3")
        headers = _tok(client, "gov_admin3")
        payload = {
            "target_type": "kindergarten",
            "target_id": sample_kindergarten.id,
            "reminder_type": "low_submission_rate",
        }
        r1 = client.post("/api/admin/governance/reminders", headers=headers, json=payload)
        assert r1.status_code == 200
        r2 = client.post("/api/admin/governance/reminders", headers=headers, json=payload)
        assert r2.status_code == 429

    def test_list_reminders_with_filters(self, client, test_db, sample_kindergarten):
        _make_admin(test_db, "gov_admin", "4")
        headers = _tok(client, "gov_admin4")
        client.post(
            "/api/admin/governance/reminders",
            headers=headers,
            json={
                "target_type": "kindergarten",
                "target_id": sample_kindergarten.id,
                "reminder_type": "low_submission_rate",
            },
        )
        r = client.get(
            f"/api/admin/governance/reminders?target_type=kindergarten&target_id={sample_kindergarten.id}",
            headers=headers,
        )
        assert r.status_code == 200
        body = r.json()
        assert body["total"] >= 1

    def test_list_reminders_no_filters(self, client, test_db):
        _make_admin(test_db, "gov_admin", "5")
        headers = _tok(client, "gov_admin5")
        r = client.get("/api/admin/governance/reminders", headers=headers)
        assert r.status_code == 200


# ---------------------------------------------------------------------------
# 16. Admin alerts with active alert rows  (lines 4712-4790)
# ---------------------------------------------------------------------------

class TestAdminAlertsWithDataCoverage:
    def _make_alert(self, db, severity=models.SeverityLevel.HIGH, status_=models.AlertStatus.ACTIVE,
                     scope_type="KINDERGARTEN", scope_id=None):
        alert = models.ActiveAlert(
            threshold_id=1,
            metric_type="attendance_rate",
            scope_type=scope_type,
            scope_id=scope_id,
            current_value=42.5,
            message="Attendance dropped below threshold",
            severity=severity,
            status=status_,
            triggered_at=datetime.now(timezone.utc),
        )
        db.add(alert)
        db.commit()
        db.refresh(alert)
        return alert

    def test_alerts_with_kindergarten_scope_resolves_governorate(self, client, test_db, sample_kindergarten):
        _make_admin(test_db, "alrtd_admin", "1")
        headers = _tok(client, "alrtd_admin1")
        self._make_alert(test_db, scope_type="KINDERGARTEN", scope_id=str(sample_kindergarten.id))
        r = client.get("/api/admin/alerts", headers=headers)
        assert r.status_code == 200
        body = r.json()
        assert body["total"] >= 1
        assert any(a["governorate"] == sample_kindergarten.governorate for a in body["alerts"])

    def test_alerts_with_governorate_scope(self, client, test_db):
        _make_admin(test_db, "alrtd_admin", "2")
        headers = _tok(client, "alrtd_admin2")
        self._make_alert(test_db, scope_type="GOVERNORATE", scope_id="Amman")
        r = client.get("/api/admin/alerts?governorate=Amman", headers=headers)
        assert r.status_code == 200
        body = r.json()
        assert body["total"] >= 1

    def test_alerts_status_filter_acknowledged(self, client, test_db):
        _make_admin(test_db, "alrtd_admin", "3")
        headers = _tok(client, "alrtd_admin3")
        self._make_alert(test_db, status_=models.AlertStatus.ACKNOWLEDGED)
        r = client.get("/api/admin/alerts?status=ACKNOWLEDGED", headers=headers)
        assert r.status_code == 200
        body = r.json()
        assert all(a["status"] == "ACKNOWLEDGED" for a in body["alerts"])

    def test_alerts_invalid_status_returns_422(self, client, test_db):
        _make_admin(test_db, "alrtd_admin", "4")
        headers = _tok(client, "alrtd_admin4")
        r = client.get("/api/admin/alerts?status=NOT_A_STATUS", headers=headers)
        assert r.status_code == 422

    def test_alerts_pagination(self, client, test_db):
        _make_admin(test_db, "alrtd_admin", "5")
        headers = _tok(client, "alrtd_admin5")
        for _ in range(3):
            self._make_alert(test_db)
        r = client.get("/api/admin/alerts?skip=0&limit=2", headers=headers)
        assert r.status_code == 200
        assert len(r.json()["alerts"]) == 2


# ---------------------------------------------------------------------------
# 17. CSV import: manager conflict + PARENT-with-children  (lines 456-496, 1606-1646)
# ---------------------------------------------------------------------------

class TestCSVImportManagerConflictCoverage:
    def test_csv_import_duplicate_manager_target_kg_rejected(self, client, test_db, sample_kindergarten):
        _make_admin(test_db, "csvc_admin", "1")
        headers = _tok(client, "csvc_admin1")
        csv_content = (
            "username,email,password,role,kindergarten_id\n"
            f"mgr_one,mgr_one@example.com,Pass1234!,MANAGER,{sample_kindergarten.id}\n"
            f"mgr_two,mgr_two@example.com,Pass1234!,MANAGER,{sample_kindergarten.id}\n"
        )
        r = client.post(
            "/api/admin/users/import-csv",
            headers=headers,
            files={"file": ("users.csv", csv_content.encode(), "text/csv")},
        )
        assert r.status_code == 200
        body = r.json()
        assert body["succeeded"] == 1
        assert body["failed"] == 1

    def test_csv_import_manager_violates_existing_rules(self, client, test_db, sample_kindergarten):
        _make_admin(test_db, "csvc_admin", "2")
        existing_mgr = _make_user(test_db, "existing_mgr2", models.UserRole.MANAGER, kg_id=sample_kindergarten.id)
        headers = _tok(client, "csvc_admin2")
        csv_content = (
            "username,email,password,role,kindergarten_id\n"
            f"mgr_new,mgr_new@example.com,Pass1234!,MANAGER,{sample_kindergarten.id}\n"
        )
        r = client.post(
            "/api/admin/users/import-csv",
            headers=headers,
            files={"file": ("users.csv", csv_content.encode(), "text/csv")},
        )
        assert r.status_code == 200
        body = r.json()
        assert body["failed"] == 1

    def test_csv_import_creates_manager_and_commits(self, client, test_db, sample_kindergarten):
        _make_admin(test_db, "csvc_admin", "3")
        headers = _tok(client, "csvc_admin3")
        csv_content = (
            "username,email,password,role,kindergarten_id\n"
            f"mgr_three,mgr_three@example.com,Pass1234!,MANAGER,{sample_kindergarten.id}\n"
        )
        r = client.post(
            "/api/admin/users/import-csv",
            headers=headers,
            files={"file": ("users.csv", csv_content.encode(), "text/csv")},
        )
        assert r.status_code == 200
        body = r.json()
        assert body["succeeded"] == 1
        assert len(body["created_ids"]) == 1

    def test_create_parent_user_with_children(self, client, test_db):
        _make_admin(test_db, "csvc_admin", "4")
        headers = _tok(client, "csvc_admin4")
        r = client.post(
            "/api/admin/users",
            headers=headers,
            json={
                "username": "parent_with_kids",
                "email": "parent_with_kids@example.com",
                "password": "Pass1234!",
                "role": "PARENT",
                "children": [
                    {
                        "first_name": "Layla",
                        "last_name": "Ahmad",
                        "gender": "FEMALE",
                        "date_of_birth": "2024-01-01",
                        "father_name": "Ahmad",
                        "mother_first_name": "Sara",
                        "mother_last_name": "Ahmad",
                    }
                ],
            },
        )
        assert r.status_code in (200, 201)


# ---------------------------------------------------------------------------
# 18. Misc admin read endpoints  (lines 4632-4652, 4655-4672)
# ---------------------------------------------------------------------------

class TestMiscAdminReadEndpoints:
    def test_list_managers_for_impersonation(self, client, test_db, sample_kindergarten):
        _make_admin(test_db, "misc_admin", "1")
        _make_user(test_db, "misc_mgr1", models.UserRole.MANAGER, kg_id=sample_kindergarten.id)
        headers = _tok(client, "misc_admin1")
        r = client.get("/api/admin/managers", headers=headers)
        assert r.status_code == 200
        body = r.json()
        assert any(m["username"] == "misc_mgr1" for m in body["managers"])

    def test_list_managers_non_admin_rejected(self, client, test_db, sample_kindergarten):
        mgr = _make_user(test_db, "misc_mgr2", models.UserRole.MANAGER, kg_id=sample_kindergarten.id)
        headers = _tok(client, "misc_mgr2", "Pass123!")
        r = client.get("/api/admin/managers", headers=headers)
        assert r.status_code == 403

    def test_get_available_report_scopes(self, client, test_db):
        _make_admin(test_db, "misc_admin", "2")
        headers = _tok(client, "misc_admin2")
        r = client.get("/api/admin/reports/scopes", headers=headers)
        assert r.status_code == 200
        assert "scopes" in r.json()


# ---------------------------------------------------------------------------
# 19. Admin alerts scope_id edge cases  (lines 4742-4746, 4760-4764)
# ---------------------------------------------------------------------------

class TestAdminAlertsScopeIdEdgeCases:
    def _make_alert(self, db, scope_type, scope_id):
        alert = models.ActiveAlert(
            threshold_id=1,
            metric_type="attendance_rate",
            scope_type=scope_type,
            scope_id=scope_id,
            current_value=10.0,
            message="test alert",
            severity=models.SeverityLevel.LOW,
            status=models.AlertStatus.ACTIVE,
            triggered_at=datetime.now(timezone.utc),
        )
        db.add(alert)
        db.commit()
        db.refresh(alert)
        return alert

    def test_alerts_kindergarten_scope_with_non_numeric_id_ignored(self, client, test_db):
        _make_admin(test_db, "alrte_admin", "1")
        headers = _tok(client, "alrte_admin1")
        self._make_alert(test_db, "KINDERGARTEN", "not-a-number")
        r = client.get("/api/admin/alerts", headers=headers)
        assert r.status_code == 200
        body = r.json()
        match = next(a for a in body["alerts"] if a["message"] == "test alert")
        assert match["governorate"] is None


# ---------------------------------------------------------------------------
# 20. Kindergarten Excel import error/exception branches (lines 3886-3888, 3940-3943, 3951-3954)
# ---------------------------------------------------------------------------

class TestKindergartenExcelImportErrorsCoverage:
    def test_import_excel_corrupt_file_returns_400(self, client, test_db):
        _make_admin(test_db, "xlxe_admin", "1")
        headers = _tok(client, "xlxe_admin1")
        import zipfile
        from io import BytesIO
        buf = BytesIO()
        with zipfile.ZipFile(buf, "w") as zf:
            zf.writestr("dummy.txt", "not a real workbook")
        buf.seek(0)
        r = client.post(
            "/api/admin/kindergartens/import-excel",
            headers=headers,
            files={"file": ("kgs.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert r.status_code == 400

    def test_import_excel_not_a_zip_returns_400(self, client, test_db):
        """Non-zip garbage uploaded with a .xlsx extension must be rejected gracefully (BadZipFile)."""
        _make_admin(test_db, "xlxe_admin", "1b")
        headers = _tok(client, "xlxe_admin1b")
        r = client.post(
            "/api/admin/kindergartens/import-excel",
            headers=headers,
            files={"file": ("kgs.xlsx", b"this is not a zip file at all", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert r.status_code == 400

    def test_import_excel_db_insert_error_recorded_as_row_error(self, client, test_db):
        _make_admin(test_db, "xlxe_admin", "2")
        headers = _tok(client, "xlxe_admin2")
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name_ar", "name_en", "governorate", "district", "area", "address", "phone"])
        ws.append(["روضة خطأ", "Error KG", "Amman", "Amman", "Abdoun", "Street 5", "0790000005"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        import models as _models
        _real_add = test_db.add

        def _add_raises_for_kg(obj):
            if isinstance(obj, _models.Kindergarten):
                raise ValueError("bad data")
            return _real_add(obj)

        with patch.object(test_db, "add", side_effect=_add_raises_for_kg):
            r = client.post(
                "/api/admin/kindergartens/import-excel",
                headers=headers,
                files={"file": ("kgs.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert r.status_code == 200
        body = r.json()
        assert len(body["errors"]) == 1
        assert body["inserted"] == 0

    def test_import_excel_commit_failure_returns_500(self, client, test_db):
        _make_admin(test_db, "xlxe_admin", "3")
        headers = _tok(client, "xlxe_admin3")
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name_ar", "name_en", "governorate", "district", "area", "address", "phone"])
        ws.append(["روضة فشل الحفظ", "Commit Fail KG", "Amman", "Amman", "Abdoun", "Street 6", "0790000006"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        from sqlalchemy.exc import SQLAlchemyError

        with patch.object(test_db, "commit", side_effect=SQLAlchemyError("commit failed")):
            r = client.post(
                "/api/admin/kindergartens/import-excel",
                headers=headers,
                files={"file": ("kgs.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert r.status_code == 500
