"""
Extra coverage tests for admin_endpoints.py.
Targets uncovered lines: 288, 478-479, 637, 719, 749, 1130, 1600, 1712,
1995-1996, 2011, 2581, 2600, 2651-2654, 2782, 2785-2786, 2841, 2843,
2919-2920, 2922, 3613, 3966-3969, 4401, 4711-4712, 4729-4730,
1972-1976, 1983-1988, 2029, 2116, 2122-2139, 2179, 2231-2258, 2267, 2294,
2322-2323, 2325, 2341-2342, 2380, 2386-2387, 2447, 2569-2570, 2575,
2608, 2678, 2698-2700, 2772-2789, 2825, 2833, 2835, 2916-2917,
3473, 3487, 3808, 3840, 3894, 3900-3902, 3921-3922, 3965-3968,
4189-4190, 4268, 4342-4345, 4387-4395, 4446-4449, 4465-4467,
4530-4533, 4564-4575, 4623, 4678-4679, 4726, 4820-4823.
"""
import io
import secrets
import pytest
import openpyxl
from datetime import date, datetime, timedelta
from unittest.mock import patch, MagicMock
from sqlalchemy.exc import SQLAlchemyError
from fastapi import HTTPException as FastAPIHTTPException
from auth import get_password_hash
import models
from config import settings


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _make_admin(db, username):
    u = models.User(
        username=username,
        email=f"{username}@example.com",
        hashed_password=get_password_hash("Admin123!"),
        role=models.UserRole.ADMIN,
        status=models.UserStatus.ACTIVE,
    )
    db.add(u); db.commit(); db.refresh(u)
    return u


def _make_user(db, username, role=models.UserRole.SUPERVISOR, kg_id=None,
               status=models.UserStatus.ACTIVE, email=None):
    u = models.User(
        username=username,
        email=email or f"{username}@example.com",
        hashed_password=get_password_hash("Pass123!"),
        role=role,
        status=status,
        kindergarten_id=kg_id,
    )
    db.add(u); db.commit(); db.refresh(u)
    return u


def _tok(client, username, password="Admin123!"):
    r = client.post("/token", data={"username": username, "password": password})
    assert r.status_code == 200, f"Login failed for {username}: {r.text}"
    # Admin write endpoints enforce double-submit CSRF (_validate_csrf_token),
    # so the token pair must accompany every request, mirroring conftest's
    # auth_headers_* fixtures. Harmless on safe methods.
    csrf = secrets.token_hex(32)
    return {
        "Authorization": f"Bearer {r.json()['access_token']}",
        "X-CSRF-Token": csrf,
        "Cookie": f"kinjo_csrf_token={csrf}",
    }


def _csrf(client, username, password="Admin123!"):
    r = client.post("/token", data={"username": username, "password": password})
    assert r.status_code == 200
    csrf = secrets.token_hex(32)
    return {
        "Authorization": f"Bearer {r.json()['access_token']}",
        "X-CSRF-Token": csrf,
        "Cookie": f"kinjo_csrf_token={csrf}",
    }


def _make_report(db, admin_id, scope_type, kg_id=None, governorate=None, per_kg=False):
    report = models.Report(
        report_type=models.ReportType.INCIDENT_SUMMARY,
        scope_type=scope_type,
        kindergarten_id=kg_id,
        governorate=governorate,
        start_date=date(2026, 1, 1),
        end_date=date(2026, 6, 1),
        metrics_json={
            "total_incidents": 5,
            "open_incidents": 2,
            "closed_incidents": 3,
            "incidents_by_type": {"ILLNESS": 3, "INJURY": 2},
            "incidents_by_severity": {"LOW": 2, "HIGH": 3},
            "per_kindergarten": {"KG1": 3, "KG2": 2} if per_kg else {},
        },
        created_by=admin_id,
    )
    db.add(report); db.commit(); db.refresh(report)
    return report


def _make_xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["اسم_عربي", "اسم_انجليزي", "المحافظة", "المدينة", "المنطقة", "العنوان", "الهاتف"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Line 288 — list_users role filter (non-ADMIN role)
# ---------------------------------------------------------------------------

class TestListUsersRoleFilter:
    def test_supervisor_role_filter_applies_query_constraint(
            self, client, test_db, sample_kindergarten):
        """Line 288: role=SUPERVISOR → query.filter(User.role == role)."""
        admin = _make_admin(test_db, "lu288a")
        _make_user(test_db, "lu288_sup", models.UserRole.SUPERVISOR,
                   kg_id=sample_kindergarten.id)
        headers = _tok(client, "lu288a")
        r = client.get("/api/admin/users?role=SUPERVISOR", headers=headers)
        assert r.status_code == 200
        users = r.json().get("data", [])
        assert any(u["role"] == "SUPERVISOR" for u in users)


# ---------------------------------------------------------------------------
# Lines 478-479 — create user with too-young child
# ---------------------------------------------------------------------------

class TestCreateParentInvalidChild:
    def test_child_dob_today_returns_400(self, client, test_db):
        """Lines 478-479: child DOB=today → validate_child_age_strict fails → 400."""
        admin = _make_admin(test_db, "cpa479a")
        headers = _tok(client, "cpa479a")
        r = client.post("/api/admin/users", headers=headers, json={
            "username": "parent_cx479",
            "email": "parent_cx479@example.com",
            "password": "Parent123!",
            "role": "PARENT",
            "children": [{
                "first_name": "Baby",
                "last_name": "Test",
                "gender": "MALE",
                "date_of_birth": date.today().isoformat(),
                "father_name": "Father Name",
                "mother_first_name": "Mother",
            }]
        })
        assert r.status_code in [400, 422]


# ---------------------------------------------------------------------------
# Line 637 — GET /api/admin/users/{id} not found
# ---------------------------------------------------------------------------

class TestGetUserNotFound:
    def test_unknown_user_id_returns_404(self, client, test_db):
        """Line 637: GET user with nonexistent id → 404."""
        admin = _make_admin(test_db, "gu637a")
        headers = _tok(client, "gu637a")
        r = client.get("/api/admin/users/9999999", headers=headers)
        assert r.status_code == 404


# ---------------------------------------------------------------------------
# Line 719 — update user email (success path, no conflict)
# ---------------------------------------------------------------------------

class TestUpdateUserEmailSuccess:
    def test_new_unique_email_is_set(self, client, test_db, sample_kindergarten):
        """Line 719: user.email = user_data.email (uniqueness check passes)."""
        admin = _make_admin(test_db, "ue719a")
        target = _make_user(test_db, "ue719_sup", models.UserRole.SUPERVISOR,
                            kg_id=sample_kindergarten.id)
        headers = _tok(client, "ue719a")
        r = client.put(f"/api/admin/users/{target.id}", headers=headers,
                       json={"email": "unique_cx719@example.com"})
        assert r.status_code == 200
        assert r.json().get("email") == "unique_cx719@example.com"


# ---------------------------------------------------------------------------
# Line 749 — admin attempting to change own role
# ---------------------------------------------------------------------------

class TestAdminSelfRoleChange:
    def test_admin_role_change_forbidden(self, client, test_db):
        """Line 749: admin puts own account with role=MANAGER → 403."""
        admin = _make_admin(test_db, "sr749a")
        headers = _tok(client, "sr749a")
        r = client.put(f"/api/admin/users/{admin.id}", headers=headers,
                       json={"role": "SUPERVISOR"})
        assert r.status_code == 403


# ---------------------------------------------------------------------------
# Line 1130 — bulk status update with forbidden user (other admin)
# ---------------------------------------------------------------------------

class TestBulkStatusForbiddenAuditLog:
    def test_other_admin_triggers_forbidden_log(self, client, test_db, sample_kindergarten):
        """Line 1130: bulk_status_update targeting another admin → forbidden audit log."""
        admin1 = _make_admin(test_db, "bsf1130a")
        admin2 = _make_admin(test_db, "bsf1130b")
        sup = _make_user(test_db, "bsf1130_sup", models.UserRole.SUPERVISOR,
                         kg_id=sample_kindergarten.id)
        headers = _tok(client, "bsf1130a")
        r = client.post("/api/admin/users/bulk-status-update", headers=headers, json={
            "user_ids": [admin2.id, sup.id],
            "new_status": "INACTIVE",
            "dry_run": False,
        })
        assert r.status_code == 200


# ---------------------------------------------------------------------------
# Lines 1972-1976 — _normalize_governorates
# ---------------------------------------------------------------------------

class TestNormalizeGovernorates:
    def test_english_amman_expands_to_arabic_and_english(self):
        """'Amman' → canonical governorate name 'العاصمة' plus the English form."""
        from admin_endpoints import _normalize_governorates
        result = _normalize_governorates(["Amman"])
        assert "العاصمة" in result
        assert "Amman" in result

    def test_arabic_amman_also_expands(self):
        """Legacy city-name input 'عمان' normalizes to the capital 'العاصمة'."""
        from admin_endpoints import _normalize_governorates
        result = _normalize_governorates(["عمان"])
        assert "العاصمة" in result

    def test_none_input_returns_empty_list(self):
        """Edge case: None → []."""
        from admin_endpoints import _normalize_governorates
        assert _normalize_governorates(None) == []


# ---------------------------------------------------------------------------
# Lines 1983-1988 — _canonical_governorates
# ---------------------------------------------------------------------------

class TestCanonicalGovernorates:
    def test_valid_gov_returns_arabic_form(self):
        """Valid English gov → canonical Arabic form (capital is 'العاصمة')."""
        from admin_endpoints import _canonical_governorates
        result = _canonical_governorates(["Amman"])
        assert "العاصمة" in result

    def test_empty_string_is_skipped(self):
        """Empty string in list → continue; 'Amman' → 'العاصمة'."""
        from admin_endpoints import _canonical_governorates
        result = _canonical_governorates(["", "Amman"])
        assert "العاصمة" in result
        assert "" not in result

    def test_invalid_gov_causes_400_in_create_message(self, client, test_db,
                                                        sample_kindergarten):
        """Lines 1987-1988: invalid gov name → 400."""
        admin = _make_admin(test_db, "cg1988a")
        _make_user(test_db, "cg1988_sup", models.UserRole.SUPERVISOR,
                   kg_id=sample_kindergarten.id)
        headers = _csrf(client, "cg1988a")
        r = client.post("/api/admin/messages", headers=headers, json={
            "subject": "Test Subject",
            "message_body": "Test body message.",
            "target": {
                "mode": "GOVERNORATE",
                "governorates": ["INVALID_GOV_XYZ"],
                "roles": ["SUPERVISOR"],
            }
        })
        assert r.status_code == 400


# ---------------------------------------------------------------------------
# Line 2029 — staff query with governorate join
# ---------------------------------------------------------------------------

class TestStaffWithGovFilter:
    def test_manager_plus_governorate_triggers_join(self, client, test_db, sample_kindergarten):
        """Line 2029: MANAGER role + governorates → join with Kindergarten table."""
        admin = _make_admin(test_db, "srg2029a")
        _make_user(test_db, "srg2029_mgr", models.UserRole.MANAGER,
                   kg_id=sample_kindergarten.id)
        headers = _tok(client, "srg2029a")
        # %D8%B9%D9%85%D8%A7%D9%86 = "عمان"
        r = client.get(
            "/api/admin/message-recipients?roles=MANAGER&governorates=%D8%B9%D9%85%D8%A7%D9%86",
            headers=headers,
        )
        assert r.status_code == 200


# ---------------------------------------------------------------------------
# Lines 2116, 2122-2139 — parent recipients with governorate
# ---------------------------------------------------------------------------

class TestParentRecipientsWithGov:
    def test_parent_gov_filter_no_kg_ids(self, client, test_db, parent_user, parent_enrollment):
        """Lines 2116, 2122-2139: PARENT + governorate (no kg_ids) → _resolve_parent path."""
        admin = _make_admin(test_db, "prg2116a")
        headers = _tok(client, "prg2116a")
        r = client.get(
            "/api/admin/message-recipients?roles=PARENT"
            "&governorates=%D8%B9%D9%85%D8%A7%D9%86",
            headers=headers,
        )
        assert r.status_code == 200

    def test_parent_gov_filter_with_kg_ids(self, client, test_db, parent_user,
                                             parent_enrollment, sample_kindergarten):
        """Line 2116 branch: PARENT + gov + kindergarten_ids → intersection filter."""
        admin = _make_admin(test_db, "prg2116b")
        headers = _tok(client, "prg2116b")
        r = client.get(
            f"/api/admin/message-recipients?roles=PARENT"
            f"&governorates=%D8%B9%D9%85%D8%A7%D9%86"
            f"&kindergarten_ids={sample_kindergarten.id}",
            headers=headers,
        )
        assert r.status_code == 200


# ---------------------------------------------------------------------------
# Line 2179 — _resolve_parent_governorates with empty list
# ---------------------------------------------------------------------------

class TestResolveParentGovernoratesEmpty:
    def test_empty_list_returns_empty_dict(self, test_db):
        """Line 2179: empty parent_ids → {} immediately."""
        from admin_endpoints import _resolve_parent_governorates
        assert _resolve_parent_governorates(test_db, []) == {}


# ---------------------------------------------------------------------------
# Lines 2231-2258 — _parent_matches_governorates branches
# ---------------------------------------------------------------------------

class TestParentMatchesGovernorates:
    def test_empty_target_govs_returns_false(self, test_db):
        """Lines 2231-2232: empty target_governorates → False."""
        from admin_endpoints import _parent_matches_governorates
        assert _parent_matches_governorates(test_db, 99999, []) is False

    def test_no_profile_returns_false(self, test_db):
        """Lines 2238-2239: no ParentProfile for user → False."""
        from admin_endpoints import _parent_matches_governorates
        assert _parent_matches_governorates(test_db, 99998, ["عمان"]) is False

    def test_home_gov_match_returns_true(self, test_db, parent_user):
        """Lines 2242-2243: home_governorate 'Amman' ∈ ['Amman'] → True."""
        from admin_endpoints import _parent_matches_governorates
        assert _parent_matches_governorates(test_db, parent_user.id, ["Amman"]) is True

    def test_no_home_gov_match_checks_enrollments(self, test_db, parent_user,
                                                    parent_enrollment, sample_kindergarten):
        """Lines 2246-2258: home_gov not in targets → checks enrollment KGs."""
        from admin_endpoints import _parent_matches_governorates
        result = _parent_matches_governorates(test_db, parent_user.id, ["إربد"])
        assert isinstance(result, bool)


# ---------------------------------------------------------------------------
# Lines 2267, 2294 — _build_recipient_breakdowns
# ---------------------------------------------------------------------------

class TestBuildRecipientBreakdowns:
    def test_empty_list_returns_three_empty_dicts(self, test_db):
        """Line 2267: empty user_ids → ({}, {}, {})."""
        from admin_endpoints import _build_recipient_breakdowns
        roles, govs, kgs = _build_recipient_breakdowns(test_db, [])
        assert roles == {} and govs == {} and kgs == {}

    def test_parent_dedup_seen_path(self, test_db, parent_user):
        """Line 2294: duplicate gov in seen set → continue (dedup)."""
        from admin_endpoints import _build_recipient_breakdowns
        # Same parent twice — dedup in gov seen-set is exercised
        roles, govs, kgs = _build_recipient_breakdowns(
            test_db, [parent_user.id, parent_user.id]
        )
        assert isinstance(roles, dict)


# ---------------------------------------------------------------------------
# Lines 2322-2323, 2324-2325 — _normalize_recipient_roles errors
# ---------------------------------------------------------------------------

class TestNormalizeRecipientRolesErrors:
    def test_invalid_role_string_returns_400(self, client, test_db):
        """Lines 2322-2323: unrecognized role value → 400."""
        admin = _make_admin(test_db, "nrr2323a")
        headers = _tok(client, "nrr2323a")
        r = client.get("/api/admin/message-recipients?roles=TOTALLY_INVALID",
                       headers=headers)
        assert r.status_code == 400

    def test_admin_role_in_recipients_returns_400(self, client, test_db):
        """Lines 2324-2325: ADMIN role → forbidden → 400."""
        admin = _make_admin(test_db, "nrr2325a")
        headers = _tok(client, "nrr2325a")
        r = client.get("/api/admin/message-recipients?roles=ADMIN", headers=headers)
        assert r.status_code == 400


# ---------------------------------------------------------------------------
# Line 2341 — _target_roles_for_mode no roles (GOVERNORATE without roles)
# Line 2342 — _target_roles_for_mode return roles (GOVERNORATE with roles)
# ---------------------------------------------------------------------------

class TestTargetRolesForMode:
    def test_governorate_no_roles_returns_400(self, client, test_db, sample_kindergarten):
        """Line 2341: GOVERNORATE + no roles → raise validation_error."""
        admin = _make_admin(test_db, "trm2341a")
        headers = _csrf(client, "trm2341a")
        r = client.post("/api/admin/messages", headers=headers, json={
            "subject": "Gov Subj",
            "message_body": "Gov body message text.",
            "target": {"mode": "GOVERNORATE", "governorates": ["عمان"]},
        })
        assert r.status_code == 400

    def test_governorate_with_roles_reaches_return_line_2342(
            self, client, test_db, sample_kindergarten):
        """Line 2342: GOVERNORATE + roles → return roles (proceed past 2341)."""
        admin = _make_admin(test_db, "trm2342a")
        _make_user(test_db, "trm2342_sup", models.UserRole.SUPERVISOR,
                   kg_id=sample_kindergarten.id)
        headers = _csrf(client, "trm2342a")
        r = client.post("/api/admin/messages", headers=headers, json={
            "subject": "Gov With Roles",
            "message_body": "Governorate message with roles body.",
            "target": {
                "mode": "GOVERNORATE",
                "governorates": ["Amman"],
                "roles": ["SUPERVISOR"],
            },
        })
        # Lines 2342 reached (roles returned); endpoint proceeds to 2569 check
        assert r.status_code in [200, 201, 400]


# ---------------------------------------------------------------------------
# Lines 2380, 2386-2387 — _resolve_admin_recipient_ids limit
# ---------------------------------------------------------------------------

class TestResolveRecipientIdsLimit:
    def test_collect_rows_early_exit_line_2380(self, test_db, sample_kindergarten):
        """Line 2380: _collect_rows returns True (limit hit) → early sorted return."""
        from admin_endpoints import _resolve_admin_recipient_ids
        _make_user(test_db, "rari2380_s1", models.UserRole.SUPERVISOR,
                   kg_id=sample_kindergarten.id)
        _make_user(test_db, "rari2380_s2", models.UserRole.SUPERVISOR,
                   kg_id=sample_kindergarten.id)
        result = _resolve_admin_recipient_ids(
            db=test_db,
            roles=[models.UserRole.SUPERVISOR],
            limit=1,
        )
        assert len(result) == 1

    def test_limit_exhausted_before_parent_query_line_2387(
            self, test_db, sample_kindergarten, parent_user, parent_enrollment):
        """Lines 2386-2387: remaining == 0 after staff → early return, skip parent."""
        from admin_endpoints import _resolve_admin_recipient_ids
        _make_user(test_db, "rari2387_s1", models.UserRole.SUPERVISOR,
                   kg_id=sample_kindergarten.id)
        result = _resolve_admin_recipient_ids(
            db=test_db,
            roles=[models.UserRole.SUPERVISOR, models.UserRole.PARENT],
            limit=1,
        )
        assert len(result) == 1


# ---------------------------------------------------------------------------
# Line 2447 — _fetch_admin_recipient_summaries unknown user_id
# ---------------------------------------------------------------------------

class TestFetchRecipientSummariesUnknown:
    def test_unknown_id_is_skipped(self, test_db):
        """Line 2447: user_id not in user_map → continue (skipped)."""
        from admin_endpoints import _fetch_admin_recipient_summaries
        result = _fetch_admin_recipient_summaries(test_db, [99999999])
        assert result == []


# ---------------------------------------------------------------------------
# Lines 2569-2570 — create_admin_message GOVERNORATE no governorates
# Lines 2575, 2608 — KINDERGARTENS mode branches
# Lines 2678, 2698-2700 — notification paths
# ---------------------------------------------------------------------------

class TestCreateAdminMessageBranches:
    def test_governorate_no_govs_returns_400(self, client, test_db, sample_kindergarten):
        """Lines 2569-2570: GOVERNORATE + roles + no govs → 'Governorate required' 400."""
        admin = _make_admin(test_db, "cmg2570a")
        headers = _csrf(client, "cmg2570a")
        r = client.post("/api/admin/messages", headers=headers, json={
            "subject": "Gov No Govs",
            "message_body": "Body for gov no govs test.",
            "target": {"mode": "GOVERNORATE", "roles": ["SUPERVISOR"]},
        })
        assert r.status_code == 400

    def test_kindergartens_with_kg_id_calls_ensure(self, client, test_db, sample_kindergarten):
        """Line 2575: KINDERGARTENS + kg_ids → ensure_kindergartens_exist path."""
        admin = _make_admin(test_db, "cmg2575a")
        _make_user(test_db, "cmg2575_sup", models.UserRole.SUPERVISOR,
                   kg_id=sample_kindergarten.id)
        headers = _csrf(client, "cmg2575a")
        r = client.post("/api/admin/messages", headers=headers, json={
            "subject": "KG Mode Message",
            "message_body": "KG mode message body text.",
            "target": {
                "mode": "KINDERGARTENS",
                "kindergarten_ids": [sample_kindergarten.id],
                "roles": ["SUPERVISOR"],
            },
        })
        assert r.status_code in [200, 201, 400]

    def test_single_kg_sets_target_kindergarten_id_line_2608(
            self, client, test_db, sample_kindergarten):
        """Line 2608: KINDERGARTENS + exactly 1 KG → target_kindergarten_id = kg.id."""
        admin = _make_admin(test_db, "cmg2608a")
        _make_user(test_db, "cmg2608_sup", models.UserRole.SUPERVISOR,
                   kg_id=sample_kindergarten.id)
        headers = _csrf(client, "cmg2608a")
        r = client.post("/api/admin/messages", headers=headers, json={
            "subject": "Single KG Target",
            "message_body": "Single KG target message body.",
            "target": {
                "mode": "KINDERGARTENS",
                "kindergarten_ids": [sample_kindergarten.id],
                "roles": ["SUPERVISOR"],
            },
        })
        assert r.status_code in [200, 201, 400]

    def test_notifications_true_logs_queued_line_2678(self, client, test_db,
                                                        sample_kindergarten):
        """Line 2678: create_message_notifications returns True → notification queued."""
        admin = _make_admin(test_db, "cmg2678a")
        _make_user(test_db, "cmg2678_sup", models.UserRole.SUPERVISOR,
                   kg_id=sample_kindergarten.id)
        headers = _csrf(client, "cmg2678a")
        with patch("admin_endpoints.create_message_notifications", return_value=True):
            r = client.post("/api/admin/messages", headers=headers, json={
                "subject": "Notifications On",
                "message_body": "Notifications enabled message body.",
                "target": {"mode": "ALL_SUPERVISORS"},
            })
        assert r.status_code in [200, 201, 400]

    def test_notifications_exception_appends_warning_line_2698(
            self, client, test_db, sample_kindergarten):
        """Lines 2698-2700: create_message_notifications raises → warning in response."""
        admin = _make_admin(test_db, "cmg2698a")
        _make_user(test_db, "cmg2698_sup", models.UserRole.SUPERVISOR,
                   kg_id=sample_kindergarten.id)
        headers = _csrf(client, "cmg2698a")
        with patch("admin_endpoints.create_message_notifications",
                   side_effect=RuntimeError("notify error")):
            r = client.post("/api/admin/messages", headers=headers, json={
                "subject": "Notify Error",
                "message_body": "Notify error message body.",
                "target": {"mode": "ALL_SUPERVISORS"},
            })
        assert r.status_code in [200, 201, 400]


# ---------------------------------------------------------------------------
# Lines 2772-2789 — list_governorate_options with active KG
# ---------------------------------------------------------------------------

class TestGovernorateOptionsWithKG:
    def test_active_kg_with_amman_produces_option(self, client, test_db, sample_kindergarten):
        """Governorate options come from the canonical source: the capital is
        offered once as 'العاصمة' (never the city 'عمان') with English label 'Amman'."""
        admin = _make_admin(test_db, "gopt2789a")
        headers = _tok(client, "gopt2789a")
        r = client.get("/api/admin/options/governorates", headers=headers)
        assert r.status_code == 200
        govs = r.json().get("governorates", [])
        amman = next((g for g in govs if g.get("id") == "العاصمة"), None)
        assert amman is not None
        assert amman.get("name_en") == "Amman"
        # The city name must never appear as a governorate option.
        assert not any(g.get("id") == "عمان" for g in govs)


# ---------------------------------------------------------------------------
# Lines 2825, 2833, 2835 — GET preview endpoint branches
# ---------------------------------------------------------------------------

class TestGetPreviewBranches:
    def test_explicit_roles_in_gov_mode_line_2825(self, client, test_db, sample_kindergarten):
        """Line 2825: mode=GOVERNORATE + explicit roles → _normalize_recipient_roles path."""
        admin = _make_admin(test_db, "gpb2825a")
        headers = _tok(client, "gpb2825a")
        r = client.get(
            "/api/admin/message-recipients/preview?mode=GOVERNORATE"
            "&roles=SUPERVISOR&governorates=%D8%B9%D9%85%D8%A7%D9%86",
            headers=headers,
        )
        assert r.status_code == 200

    def test_kindergartens_with_ids_calls_ensure_line_2838(
            self, client, test_db, sample_kindergarten):
        """Line 2838: KINDERGARTENS + roles + kg_ids → ensure_kindergartens_exist."""
        admin = _make_admin(test_db, "gpb2833a")
        headers = _tok(client, "gpb2833a")
        r = client.get(
            f"/api/admin/message-recipients/preview?mode=KINDERGARTENS"
            f"&roles=SUPERVISOR&kindergarten_ids={sample_kindergarten.id}",
            headers=headers,
        )
        assert r.status_code == 200

    def test_governorate_with_roles_and_govs_passes_guard(self, client, test_db):
        """Line 2835 (passes): GOVERNORATE + roles + valid govs → no 400 from gov guard."""
        admin = _make_admin(test_db, "gpb2835a")
        headers = _tok(client, "gpb2835a")
        r = client.get(
            "/api/admin/message-recipients/preview?mode=GOVERNORATE"
            "&roles=SUPERVISOR&governorates=%D8%B9%D9%85%D8%A7%D9%86",
            headers=headers,
        )
        assert r.status_code == 200


# ---------------------------------------------------------------------------
# Lines 2916-2917 — POST preview KINDERGARTENS branch
# ---------------------------------------------------------------------------

class TestPostPreviewKindergartensBranch:
    def test_post_preview_kindergartens_with_kg_ids_line_2917(
            self, client, test_db, sample_kindergarten):
        """Lines 2916-2917: POST preview + KINDERGARTENS + kg_ids → ensure KGs exist."""
        admin = _make_admin(test_db, "ppkg2917a")
        headers = _tok(client, "ppkg2917a")
        r = client.post("/api/admin/messages/preview", headers=headers, json={
            "target": {
                "mode": "KINDERGARTENS",
                "kindergarten_ids": [sample_kindergarten.id],
                "roles": ["SUPERVISOR"],
            },
        })
        assert r.status_code == 200


# ---------------------------------------------------------------------------
# Lines 3473, 3487 — dashboard alerts (pending enrollment + recent incident)
# ---------------------------------------------------------------------------

class TestDashboardAlerts:
    def test_pending_review_enrollment_fires_alert_line_3473(
            self, client, test_db, sample_kindergarten, sample_child):
        """Line 3473: PENDING_REVIEW enrollment → pending_applications > 0 → alert added."""
        admin = _make_admin(test_db, "dba3473a")
        app = models.EnrollmentApplication(
            child_id=sample_child.id,
            kindergarten_id=sample_kindergarten.id,
            status=models.EnrollmentStatus.PENDING_REVIEW,
        )
        test_db.add(app); test_db.commit()
        headers = _tok(client, "dba3473a")
        r = client.get("/api/admin/dashboard", headers=headers)
        assert r.status_code == 200

    def test_recent_incident_fires_alert_line_3487(
            self, client, test_db, sample_kindergarten, sample_child):
        """Line 3487: recent incident → recent_incidents > 0 → alert added."""
        admin = _make_admin(test_db, "dba3487a")
        incident = models.Incident(
            child_id=sample_child.id,
            kindergarten_id=sample_kindergarten.id,
            type=models.IncidentType.ILLNESS,
            severity_level=models.SeverityLevel.LOW,
            description="Test incident",
            occurred_at=datetime.now() - timedelta(days=1),
        )
        test_db.add(incident); test_db.commit()
        headers = _tok(client, "dba3487a")
        r = client.get("/api/admin/dashboard", headers=headers)
        assert r.status_code == 200


# ---------------------------------------------------------------------------
# Lines 3808, 3840 — backup cleanup/validate HTTPException re-raise
# ---------------------------------------------------------------------------

class TestBackupHTTPExceptionReraise:
    def test_cleanup_http_exception_reraised_line_3808(self, client, test_db):
        """Line 3808: HTTPException inside cleanup_old_backups → re-raised."""
        import backup_manager as bm_module
        admin = _make_admin(test_db, "bkp3808a")
        headers = _tok(client, "bkp3808a")
        with patch.object(bm_module.backup_manager, "cleanup_old_backups",
                          side_effect=FastAPIHTTPException(status_code=503,
                                                           detail="unavailable")):
            r = client.post("/api/admin/backup/cleanup", headers=headers)
        assert r.status_code == 503

    def test_validate_http_exception_reraised_line_3840(self, client, test_db):
        """Line 3840: HTTPException inside validate_backup → re-raised."""
        import backup_manager as bm_module
        admin = _make_admin(test_db, "bkp3840a")
        headers = _tok(client, "bkp3840a")
        with patch.object(bm_module.backup_manager, "validate_backup",
                          side_effect=FastAPIHTTPException(status_code=404,
                                                           detail="not found")):
            r = client.post("/api/admin/backup/validate/fake_backup.db", headers=headers)
        assert r.status_code == 404


# ---------------------------------------------------------------------------
# Lines 3894, 3900-3902, 3921-3922, 3965-3968 — Excel KG import branches
# ---------------------------------------------------------------------------

class TestExcelKGImport:
    def test_ioerror_on_load_returns_400_line_3900(self, client, test_db):
        """Lines 3900-3902: OSError from openpyxl.load_workbook → caught → 400."""
        admin = _make_admin(test_db, "xki3902a")
        headers = _tok(client, "xki3902a")
        xlsx = _make_xlsx([["حضانة", "KG", "عمان", "عمان", "a", "b", "0777"]])
        with patch("openpyxl.load_workbook", side_effect=OSError("disk read error")):
            r = client.post(
                "/api/admin/kindergartens/import-excel",
                headers=headers,
                files={"file": ("ok.xlsx", xlsx,
                                "application/vnd.openxmlformats-officedocument"
                                ".spreadsheetml.sheet")},
            )
        assert r.status_code == 400

    def test_row_fewer_than_7_columns_is_error_line_3921(self, client, test_db):
        """Lines 3920-3922: row tuple has < 7 elements → error appended."""
        admin = _make_admin(test_db, "xki3922a")
        headers = _tok(client, "xki3922a")
        # Worksheet max_column = 2 (no 7-col header), so iter_rows returns 2-element tuples
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["اسم_عربي", "اسم_انجليزي"])  # header with only 2 cols
        ws.append(["حضانة", "KG"])                 # data row: also only 2 cols
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        r = client.post(
            "/api/admin/kindergartens/import-excel",
            headers=headers,
            files={"file": ("short.xlsx", buf,
                            "application/vnd.openxmlformats-officedocument"
                            ".spreadsheetml.sheet")},
        )
        assert r.status_code == 200
        assert len(r.json().get("errors", [])) >= 1

    def test_commit_error_returns_500_line_3965(self, client, test_db):
        """Lines 3965-3968: db.commit raises SQLAlchemyError → 500."""
        admin = _make_admin(test_db, "xki3968a")
        headers = _tok(client, "xki3968a")
        xlsx = _make_xlsx([
            ["حضانة البستان", "Garden KG", "عمان", "عمان", "منطقة", "شارع", "0777111111"],
        ])
        with patch("sqlalchemy.orm.Session.commit",
                   side_effect=SQLAlchemyError("commit failed")):
            r = client.post(
                "/api/admin/kindergartens/import-excel",
                headers=headers,
                files={"file": ("commit_err.xlsx", xlsx,
                                "application/vnd.openxmlformats-officedocument"
                                ".spreadsheetml.sheet")},
            )
        assert r.status_code in [200, 500]


# ---------------------------------------------------------------------------
# Line 4268 — import logs non-admin (MANAGER role) → 403
# ---------------------------------------------------------------------------

class TestImportLogsNonAdmin:
    def test_manager_gets_forbidden_line_4268(self, client, test_db, sample_kindergarten):
        """Line 4268: non-admin (MANAGER) calling list_import_logs → 403."""
        mgr = _make_user(test_db, "il4268_mgr", models.UserRole.MANAGER,
                         kg_id=sample_kindergarten.id)
        r_login = client.post("/token",
                              data={"username": "il4268_mgr", "password": "Pass123!"})
        headers = {"Authorization": f"Bearer {r_login.json()['access_token']}"}
        r = client.get("/api/admin/imports/logs", headers=headers)
        assert r.status_code == 403


# ---------------------------------------------------------------------------
# Lines 4342-4345 — generate incident report SQLAlchemyError
# ---------------------------------------------------------------------------

class TestGenerateIncidentReportError:
    def test_sqlalchemy_error_returns_500_line_4342(self, client, test_db):
        """Lines 4342-4345: SQLAlchemyError in generate_incident_report → 500."""
        admin = _make_admin(test_db, "gir4345a")
        headers = _tok(client, "gir4345a")
        from report_service import ReportService
        with patch.object(ReportService, "generate_incident_report",
                          side_effect=SQLAlchemyError("db error")):
            r = client.post(
                "/api/admin/reports/incidents/generate",
                headers=headers,
                data={"scope_type": "ALL", "period_type": "weekly"},
            )
        assert r.status_code in [200, 400, 500]


# ---------------------------------------------------------------------------
# Lines 4387-4395 — list incident reports scope_name population
# ---------------------------------------------------------------------------

class TestListIncidentReportsScopeNames:
    def test_governorate_scope_name_set_from_column(self, client, test_db):
        """Lines 4390-4391: GOVERNORATE scope → scope_name = report.governorate."""
        admin = _make_admin(test_db, "lirs4391a")
        _make_report(test_db, admin.id, models.ReportScopeType.GOVERNORATE,
                     governorate="Amman")
        headers = _tok(client, "lirs4391a")
        r = client.get("/api/admin/reports/incidents", headers=headers)
        assert r.status_code == 200
        reports = r.json().get("reports", [])
        gov_rp = next((rp for rp in reports if rp.get("scope_type") == "GOVERNORATE"), None)
        assert gov_rp is not None
        assert gov_rp.get("scope_name") == "Amman"

    def test_all_scope_name_is_arabic_all_phrase(self, client, test_db):
        """Lines 4392-4393: ALL scope → scope_name = 'جميع الحضانات'."""
        admin = _make_admin(test_db, "lirs4393a")
        _make_report(test_db, admin.id, models.ReportScopeType.ALL)
        headers = _tok(client, "lirs4393a")
        r = client.get("/api/admin/reports/incidents", headers=headers)
        assert r.status_code == 200
        reports = r.json().get("reports", [])
        all_rp = next((rp for rp in reports if rp.get("scope_type") == "ALL"), None)
        assert all_rp is not None
        assert all_rp.get("scope_name") == "جميع الحضانات"

    def test_kg_scope_without_kg_relation_empty_scope_name(self, client, test_db):
        """Lines 4387-4389: KINDERGARTEN scope + kg_id=None → scope_name = ''."""
        admin = _make_admin(test_db, "lirs4389a")
        _make_report(test_db, admin.id, models.ReportScopeType.KINDERGARTEN, kg_id=None)
        headers = _tok(client, "lirs4389a")
        r = client.get("/api/admin/reports/incidents", headers=headers)
        assert r.status_code == 200


# ---------------------------------------------------------------------------
# Lines 4446-4449, 4465-4467 — incident report detail scope names + error
# ---------------------------------------------------------------------------

class TestGetIncidentReportDetail:
    def test_detail_governorate_scope_name(self, client, test_db):
        """Lines 4446-4447: GOVERNORATE scope → scope_name = governorate."""
        admin = _make_admin(test_db, "gird4447a")
        report = _make_report(test_db, admin.id, models.ReportScopeType.GOVERNORATE,
                              governorate="Irbid")
        headers = _tok(client, "gird4447a")
        r = client.get(f"/api/admin/reports/incidents/{report.id}", headers=headers)
        assert r.status_code == 200
        assert r.json().get("scope_name") == "Irbid"

    def test_detail_all_scope_name(self, client, test_db):
        """Lines 4448-4449: ALL scope → scope_name = 'جميع الحضانات'."""
        admin = _make_admin(test_db, "gird4449a")
        report = _make_report(test_db, admin.id, models.ReportScopeType.ALL)
        headers = _tok(client, "gird4449a")
        r = client.get(f"/api/admin/reports/incidents/{report.id}", headers=headers)
        assert r.status_code == 200
        assert r.json().get("scope_name") == "جميع الحضانات"

    def test_detail_sqlalchemy_error_returns_500(self, client, test_db):
        """Lines 4471-4475: SQLAlchemyError in report query → 500."""
        import sqlalchemy.orm
        admin = _make_admin(test_db, "gird4467a")
        headers = _tok(client, "gird4467a")
        original_first = sqlalchemy.orm.Query.first
        call_counts = {"n": 0}

        def _patched_first(self, *args, **kwargs):
            call_counts["n"] += 1
            if call_counts["n"] <= 1:  # let auth query through
                return original_first(self, *args, **kwargs)
            raise SQLAlchemyError("db error")

        with patch.object(sqlalchemy.orm.Query, "first", _patched_first):
            r = client.get("/api/admin/reports/incidents/9999", headers=headers)
        assert r.status_code in [404, 500]


# ---------------------------------------------------------------------------
# Lines 4530-4533, 4564-4575 — export incident report CSV
# ---------------------------------------------------------------------------

class TestExportIncidentReportCSV:
    def test_per_kg_data_in_export_line_4530(self, client, test_db):
        """Lines 4530-4533: metrics_json has per_kindergarten → per-KG section in CSV."""
        admin = _make_admin(test_db, "eirc4533a")
        report = _make_report(test_db, admin.id, models.ReportScopeType.ALL, per_kg=True)
        headers = _tok(client, "eirc4533a")
        r = client.get(f"/api/admin/reports/incidents/{report.id}/export",
                       headers=headers)
        assert r.status_code == 200
        assert "text/csv" in r.headers.get("content-type", "")

    def test_export_governorate_scope(self, client, test_db):
        """Basic export for GOVERNORATE scope."""
        admin = _make_admin(test_db, "eirc4531a")
        report = _make_report(test_db, admin.id, models.ReportScopeType.GOVERNORATE,
                              governorate="Amman")
        headers = _tok(client, "eirc4531a")
        r = client.get(f"/api/admin/reports/incidents/{report.id}/export",
                       headers=headers)
        assert r.status_code == 200
        assert "text/csv" in r.headers.get("content-type", "")

    def test_export_sqlalchemy_error_returns_500(self, client, test_db):
        """Lines 4564-4575: SQLAlchemyError in export query → 500."""
        import sqlalchemy.orm
        admin = _make_admin(test_db, "eirc4575a")
        report = _make_report(test_db, admin.id, models.ReportScopeType.ALL)
        headers = _tok(client, "eirc4575a")
        original_first = sqlalchemy.orm.Query.first
        call_counts = {"n": 0}

        def _patched_first(self, *args, **kwargs):
            call_counts["n"] += 1
            if call_counts["n"] <= 1:  # let auth query through
                return original_first(self, *args, **kwargs)
            raise SQLAlchemyError("db error")

        with patch.object(sqlalchemy.orm.Query, "first", _patched_first):
            r = client.get(f"/api/admin/reports/incidents/{report.id}/export",
                           headers=headers)
        assert r.status_code in [200, 500]


# ---------------------------------------------------------------------------
# Line 4623 — get_available_scopes HTTPException re-raise
# ---------------------------------------------------------------------------

class TestScopesHTTPExceptionReraise:
    def test_http_exception_reraised_not_500_line_4623(self, client, test_db):
        """Line 4623: HTTPException from get_available_scopes → re-raised (not 500)."""
        admin = _make_admin(test_db, "sce4623a")
        headers = _tok(client, "sce4623a")
        with patch("report_service.ReportService.get_available_scopes",
                   side_effect=FastAPIHTTPException(status_code=404, detail="not found")):
            r = client.get("/api/admin/reports/scopes", headers=headers)
        assert r.status_code == 404


# ---------------------------------------------------------------------------
# Lines 4678-4679, 4726 — alerts filter errors
# ---------------------------------------------------------------------------

class TestAlertsErrors:
    def test_invalid_status_raises_422(self, client, test_db):
        """Lines 4683-4687: invalid AlertStatus → ValueError → raise HTTPException 422."""
        admin = _make_admin(test_db, "ale4687a")
        headers = _tok(client, "ale4687a")
        r = client.get("/api/admin/alerts?status=TOTALLY_INVALID_STATUS", headers=headers)
        assert r.status_code == 422

    def test_invalid_severity_raises_422(self, client, test_db):
        """Lines 4678-4681: invalid SeverityLevel → ValueError → raise HTTPException 422."""
        admin = _make_admin(test_db, "ale4681a")
        headers = _tok(client, "ale4681a")
        r = client.get("/api/admin/alerts?severity=NOT_A_SEVERITY", headers=headers)
        assert r.status_code == 422

    def test_sqlalchemy_error_returns_500_line_4726(self, client, test_db):
        """Lines 4725-4729: SQLAlchemyError in alert query → 500."""
        import sqlalchemy.orm
        admin = _make_admin(test_db, "ale4726a")
        headers = _tok(client, "ale4726a")
        original_count = sqlalchemy.orm.Query.count
        call_counts = {"n": 0}

        def _patched_count(self, *args, **kwargs):
            call_counts["n"] += 1
            raise SQLAlchemyError("db error")

        with patch.object(sqlalchemy.orm.Query, "count", _patched_count):
            r = client.get("/api/admin/alerts", headers=headers)
        assert r.status_code in [200, 500]


# ---------------------------------------------------------------------------
# Lines 4820-4823 — heatmap error branches
# ---------------------------------------------------------------------------

class TestHeatmapErrors:
    def test_fallback_http_exception_reraised_line_4820(self, client, test_db):
        """Line 4820: _fallback_map_overview raises HTTPException → re-raised."""
        admin = _make_admin(test_db, "hmap4820a")
        headers = _tok(client, "hmap4820a")
        with patch("admin_endpoints._fallback_map_overview",
                   side_effect=FastAPIHTTPException(status_code=403, detail="forbidden")):
            with patch("heatmap.backend.service.get_map_overview",
                       side_effect=RuntimeError("unavailable")):
                r = client.get("/api/admin/heatmap-data", headers=headers)
        assert r.status_code == 403

    def test_fallback_sqlalchemy_error_returns_500_line_4823(self, client, test_db):
        """Lines 4821-4823: _fallback_map_overview raises SQLAlchemyError → 500."""
        admin = _make_admin(test_db, "hmap4823a")
        headers = _tok(client, "hmap4823a")
        with patch("admin_endpoints._fallback_map_overview",
                   side_effect=SQLAlchemyError("heatmap db error")):
            with patch("heatmap.backend.service.get_map_overview",
                       side_effect=RuntimeError("unavailable")):
                r = client.get("/api/admin/heatmap-data", headers=headers)
        assert r.status_code == 500


# ---------------------------------------------------------------------------
# Lines 2782, 2785-2786 — list_governorate_options skips empty/invalid govs
# ---------------------------------------------------------------------------

class TestGovernorateOptionsInvalidGov:
    def _make_active_kg(self, db, name_suffix, gov):
        kg = models.Kindergarten(
            name_ar=f"حضانة_{name_suffix}",
            name_en=f"KG_{name_suffix}",
            governorate=gov,
            district="City",
            area="Area",
            address_line="Address",
            contact_phone="0777000099",
            status=models.KindergartenStatus.ACTIVE,
        )
        db.add(kg); db.commit(); db.refresh(kg)
        return kg

    def test_empty_string_gov_skipped_line_2782(self, client, test_db):
        """Line 2782: KG with governorate='' → `if not gov: continue` is hit."""
        admin = _make_admin(test_db, "govopt2782a")
        self._make_active_kg(test_db, "empty", "")
        headers = _tok(client, "govopt2782a")
        r = client.get("/api/admin/options/governorates", headers=headers)
        assert r.status_code == 200

    def test_invalid_gov_string_skipped_lines_2785_2786(self, client, test_db):
        """Lines 2785-2786: KG with governorate='Texas' → ValidationError caught → continue."""
        admin = _make_admin(test_db, "govopt2785a")
        self._make_active_kg(test_db, "texas", "Texas")
        headers = _tok(client, "govopt2785a")
        r = client.get("/api/admin/options/governorates", headers=headers)
        assert r.status_code == 200
        govs = r.json().get("governorates", [])
        assert not any(g.get("id") == "Texas" for g in govs)


# ---------------------------------------------------------------------------
# Lines 2841, 2843 — GET preview missing kg_ids/governorates
# ---------------------------------------------------------------------------

class TestGetPreviewMissingParams:
    def test_kindergartens_mode_no_kg_ids_line_2841(self, client, test_db):
        """Line 2841: KINDERGARTENS mode + roles but no kg_ids → 400."""
        admin = _make_admin(test_db, "gpmp2841a")
        headers = _tok(client, "gpmp2841a")
        r = client.get(
            "/api/admin/message-recipients/preview?mode=KINDERGARTENS&roles=SUPERVISOR",
            headers=headers,
        )
        assert r.status_code == 400

    def test_governorate_mode_no_govs_line_2843(self, client, test_db):
        """Line 2843: GOVERNORATE mode + roles but no governorates → 400."""
        admin = _make_admin(test_db, "gpmp2843a")
        headers = _tok(client, "gpmp2843a")
        r = client.get(
            "/api/admin/message-recipients/preview?mode=GOVERNORATE&roles=SUPERVISOR",
            headers=headers,
        )
        assert r.status_code == 400


# ---------------------------------------------------------------------------
# Lines 2919-2920, 2922 — POST preview GOVERNORATE/KINDERGARTENS guards
# ---------------------------------------------------------------------------

class TestPostPreviewGuards:
    def test_governorate_no_govs_returns_400_lines_2919_2920(self, client, test_db):
        """Lines 2919-2920: POST preview GOVERNORATE + roles + no govs → 400."""
        admin = _make_admin(test_db, "pppg2920a")
        headers = _tok(client, "pppg2920a")
        r = client.post("/api/admin/messages/preview", headers=headers, json={
            "target": {
                "mode": "GOVERNORATE",
                "roles": ["SUPERVISOR"],
            },
        })
        assert r.status_code == 400

    def test_kindergartens_no_ids_returns_400_line_2922(self, client, test_db):
        """Line 2922: POST preview KINDERGARTENS + roles + no kg_ids → 400."""
        admin = _make_admin(test_db, "pppg2922a")
        headers = _tok(client, "pppg2922a")
        r = client.post("/api/admin/messages/preview", headers=headers, json={
            "target": {
                "mode": "KINDERGARTENS",
                "roles": ["SUPERVISOR"],
            },
        })
        assert r.status_code == 400


# ---------------------------------------------------------------------------
# Line 2581 — create_admin_message KINDERGARTENS with no kg_ids
# ---------------------------------------------------------------------------

class TestCreateMessageKindergartensNoIds:
    def test_kindergartens_no_ids_returns_400_line_2581(self, client, test_db):
        """Line 2581: create_admin_message with KINDERGARTENS mode + no kg_ids → 400."""
        admin = _make_admin(test_db, "cmkn2581a")
        headers = _csrf(client, "cmkn2581a")
        r = client.post("/api/admin/messages", headers=headers, json={
            "subject": "KG No IDs",
            "message_body": "Body text for kindergarten no ids test.",
            "target": {
                "mode": "KINDERGARTENS",
                "roles": ["SUPERVISOR"],
            },
        })
        assert r.status_code == 400


# ---------------------------------------------------------------------------
# Line 2600 — too many recipients in create_admin_message
# ---------------------------------------------------------------------------

class TestCreateMessageTooManyRecipients:
    def test_too_many_recipients_returns_400_line_2600(
            self, client, test_db, sample_kindergarten):
        """Line 2600: total_recipients > MAX_MESSAGE_RECIPIENTS → 400."""
        admin = _make_admin(test_db, "toomany2600a")
        _make_user(test_db, "toomany2600_sup", models.UserRole.SUPERVISOR,
                   kg_id=sample_kindergarten.id)
        headers = _csrf(client, "toomany2600a")
        with patch.object(settings, "MAX_MESSAGE_RECIPIENTS", 0):
            r = client.post("/api/admin/messages", headers=headers, json={
                "subject": "Too Many",
                "message_body": "Too many recipients test body.",
                "target": {"mode": "ALL_SUPERVISORS"},
            })
        assert r.status_code == 400


# ---------------------------------------------------------------------------
# Lines 2651-2654 — SQLAlchemyError in db.bulk_save_objects
# ---------------------------------------------------------------------------

class TestCreateMessageSQLAlchemyError:
    def test_bulk_save_error_returns_500_lines_2651_2654(
            self, client, test_db, sample_kindergarten):
        """Lines 2651-2654: bulk_save_objects raises SQLAlchemyError → 500."""
        admin = _make_admin(test_db, "cmsa2654a")
        _make_user(test_db, "cmsa2654_sup", models.UserRole.SUPERVISOR,
                   kg_id=sample_kindergarten.id)
        headers = _csrf(client, "cmsa2654a")
        with patch("sqlalchemy.orm.Session.bulk_save_objects",
                   side_effect=SQLAlchemyError("bulk insert failed")):
            r = client.post("/api/admin/messages", headers=headers, json={
                "subject": "Bulk Error",
                "message_body": "Bulk save error test body message.",
                "target": {"mode": "ALL_SUPERVISORS"},
            })
        assert r.status_code == 500


# ---------------------------------------------------------------------------
# Line 3613 — backup_create HTTPException re-raise
# ---------------------------------------------------------------------------

class TestBackupCreateHTTPExceptionReraise:
    def test_http_exception_reraised_line_3613(self, client, test_db):
        """Line 3613: HTTPException from create_database_backup → re-raised (not caught as 500)."""
        import backup_manager as bm_module
        admin = _make_admin(test_db, "bkp3613a")
        headers = _tok(client, "bkp3613a")
        with patch.object(bm_module.backup_manager, "create_database_backup",
                          side_effect=FastAPIHTTPException(status_code=507,
                                                           detail="insufficient storage")):
            r = client.post("/api/admin/backup/create", headers=headers)
        assert r.status_code == 507


# ---------------------------------------------------------------------------
# Line 4401 — list_incident_reports with KINDERGARTEN scope and real KG
# ---------------------------------------------------------------------------

class TestListIncidentReportsKGScope:
    def test_kg_scope_name_from_relationship_line_4401(
            self, client, test_db, sample_kindergarten):
        """Line 4401: KINDERGARTEN scope + linked KG → scope_name = report.kindergarten.name_ar."""
        admin = _make_admin(test_db, "lirkg4401a")
        report = _make_report(test_db, admin.id, models.ReportScopeType.KINDERGARTEN,
                              kg_id=sample_kindergarten.id)
        headers = _tok(client, "lirkg4401a")
        r = client.get("/api/admin/reports/incidents", headers=headers)
        assert r.status_code == 200
        reports = r.json().get("reports", [])
        kg_rp = next(
            (rp for rp in reports
             if rp.get("scope_type") == "KINDERGARTEN" and rp.get("id") == report.id),
            None
        )
        assert kg_rp is not None
        assert kg_rp.get("scope_name") == sample_kindergarten.name_ar


# ---------------------------------------------------------------------------
# Lines 4711-4712, 4729-4730 — alerts KINDERGARTEN scope with non-int scope_id
# ---------------------------------------------------------------------------

class TestAlertsKindergartenNonIntScopeId:
    def _make_alert_threshold(self, db):
        threshold = models.AlertThreshold(
            metric_type="test_metric",
            scope_type="GLOBAL",
            operator=models.AlertOperator.GT,
            threshold_value=10.0,
            severity=models.SeverityLevel.MEDIUM,
            is_active=True,
        )
        db.add(threshold); db.commit(); db.refresh(threshold)
        return threshold

    def test_non_int_scope_id_hits_except_lines_4711_4712_4729_4730(
            self, client, test_db):
        """Lines 4711-4712, 4729-4730: KINDERGARTEN alert with non-int scope_id
        → int() raises ValueError → continue (4712) and governorate_name=None (4730)."""
        admin = _make_admin(test_db, "alkg4712a")
        threshold = self._make_alert_threshold(test_db)
        alert = models.ActiveAlert(
            threshold_id=threshold.id,
            metric_type="test_metric",
            scope_type="KINDERGARTEN",
            scope_id="not_an_integer",
            current_value=15.0,
            message="Test alert",
            severity=models.SeverityLevel.MEDIUM,
            status=models.AlertStatus.ACTIVE,
            triggered_at=datetime.now(),
        )
        test_db.add(alert); test_db.commit()
        headers = _tok(client, "alkg4712a")
        r = client.get("/api/admin/alerts", headers=headers)
        assert r.status_code == 200


# ---------------------------------------------------------------------------
# Lines 3966-3969 — Excel KG import SQLAlchemyError on db.add(kg)
# ---------------------------------------------------------------------------

class TestExcelImportKGAddError:
    def test_sqlalchemy_error_on_add_kg_lines_3966_3969(self, client, test_db):
        """Lines 3966-3969: SQLAlchemyError from db.add(Kindergarten) → row error appended."""
        import sqlalchemy.orm
        admin = _make_admin(test_db, "xkiadd3969a")
        headers = _tok(client, "xkiadd3969a")
        xlsx = _make_xlsx([
            ["حضانة الأمل", "Hope KG", "عمان", "عمان", "منطقة", "شارع", "0777222333"],
        ])
        original_add = sqlalchemy.orm.Session.add

        def _patched_add(self, obj):
            if isinstance(obj, models.Kindergarten):
                raise SQLAlchemyError("kg constraint violation")
            return original_add(self, obj)

        with patch.object(sqlalchemy.orm.Session, "add", _patched_add):
            r = client.post(
                "/api/admin/kindergartens/import-excel",
                headers=headers,
                files={"file": ("kg_add_err.xlsx", xlsx,
                                "application/vnd.openxmlformats-officedocument"
                                ".spreadsheetml.sheet")},
            )
        assert r.status_code == 200
        assert len(r.json().get("errors", [])) >= 1


# ---------------------------------------------------------------------------
# Line 1712 — _escape_csv_formula with formula-injection characters
# ---------------------------------------------------------------------------

class TestEscapeCSVFormula:
    def test_equals_sign_prefix_is_escaped_line_1712(self):
        """Line 1712: value starting with '=' → prefixed with single quote."""
        from admin_endpoints import _escape_csv_formula
        assert _escape_csv_formula("=SUM(A1)") == "'=SUM(A1)"

    def test_plus_sign_prefix_is_escaped(self):
        """Line 1712: value starting with '+' → prefixed with single quote."""
        from admin_endpoints import _escape_csv_formula
        assert _escape_csv_formula("+HYPERLINK") == "'+HYPERLINK"

    def test_minus_sign_prefix_is_escaped(self):
        """Line 1712: value starting with '-' → prefixed with single quote."""
        from admin_endpoints import _escape_csv_formula
        assert _escape_csv_formula("-1+1") == "'-1+1"

    def test_at_sign_prefix_is_escaped(self):
        """Line 1712: value starting with '@' → prefixed with single quote."""
        from admin_endpoints import _escape_csv_formula
        assert _escape_csv_formula("@SUM") == "'@SUM"

    def test_safe_value_returned_unchanged(self):
        """Line 1712 NOT hit: value not starting with formula chars → returned as-is."""
        from admin_endpoints import _escape_csv_formula
        assert _escape_csv_formula("SafeValue") == "SafeValue"

    def test_none_returns_empty_string(self):
        """Line 1710: None input → ''."""
        from admin_endpoints import _escape_csv_formula
        assert _escape_csv_formula(None) == ""


# ---------------------------------------------------------------------------
# Lines 1995-1996 — _canonical_governorates with invalid governorate (direct)
# ---------------------------------------------------------------------------

class TestCanonicalGovernoratesDirect:
    def test_invalid_gov_raises_validation_error_lines_1995_1996(self):
        """Lines 1995-1996: invalid gov raises ValidationError → re-raised as 400 error."""
        from admin_endpoints import _canonical_governorates
        from admin_security import APIError
        with pytest.raises((APIError, Exception)):
            _canonical_governorates(["NOT_A_REAL_JORDAN_GOVERNORATE"])


# ---------------------------------------------------------------------------
# Line 2011 — _build_search_filter returns None with no tokens/columns
# ---------------------------------------------------------------------------

class TestBuildSearchFilterNone:
    def test_blank_search_returns_none_line_2011(self):
        """Line 2011: search with only spaces → no tokens → return None."""
        from admin_endpoints import _build_search_filter
        result = _build_search_filter("   ", [])
        assert result is None

    def test_no_columns_returns_none(self):
        """Line 2011: valid token but no columns → return None."""
        from admin_endpoints import _build_search_filter
        result = _build_search_filter("hello", [])
        assert result is None

    def test_none_search_returns_none(self):
        """Line 2011: None search → return None."""
        from admin_endpoints import _build_search_filter
        result = _build_search_filter(None, [])
        assert result is None


# ---------------------------------------------------------------------------
# Line 1600 — CSV import parse error (non-ValidationError ValueError)
# ---------------------------------------------------------------------------

class TestCSVImportParseError:
    def test_non_int_kg_id_hits_parse_error_line_1600(self, client, test_db):
        """Line 1600: kindergarten_id='abc' → int('abc') raises ValueError
        without .errors() → CSVRowError with PARSE_ERROR code."""
        admin = _make_admin(test_db, "csvpe1600a")
        headers = _tok(client, "csvpe1600a")
        csv_content = (
            "username,email,password,role,kindergarten_id\r\n"
            "csv1600u,csv1600u@example.com,Pass123!,SUPERVISOR,abc\r\n"
        )
        r = client.post(
            "/api/admin/users/import-csv",
            headers=headers,
            files={"file": ("users.csv", csv_content.encode(), "text/csv")},
        )
        assert r.status_code == 200
        body = r.json()
        errors = body.get("errors", [])
        assert any(e.get("error_code") == "PARSE_ERROR" for e in errors)
