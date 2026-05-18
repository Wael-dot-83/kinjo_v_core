"""
Import all kindergartens from Dataset Excel files into KinJo DB.
- Reads all .xlsx files in the Dataset folder
- Deduplicates across files by normalized Arabic name
- Skips records already in DB (match by normalized name_ar)
- Inserts only new, complete records
- Standalone: bypasses config/settings to avoid env var conflicts
"""
import os
import sys
import unicodedata
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import re
import openpyxl
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

# ── Config ────────────────────────────────────────────────────────────────────
DATASET_PATH = r"C:\Users\waelj\OneDrive - zuj.edu.jo\Desktop\Dataset"
DB_URL = "sqlite:///./data/kinjo.db"

# Priority: most complete/updated files first
PRIORITY_FILES = [
    "merged_all_uploads.xlsx",          # ~1293 rows — largest merged set
    "روضات_وحضانات_محدث.xlsx",
    "_محدث روضات_وحضانات_محدث.xlsx",
    "روضات_وحضانات_دمج_كامل_محدث_873.xlsx",
    "روضات_وحضانات_دمج_كامل_محدثzxz.xlsx",
    "روضات_وحضانات_دمج_كامل.xlsx",
    "روضات_وحضانات_تحديث_الانامل_الذهبية4.xlsx",
    "lk.xlsx",
]

COL_AR   = "اسم الروضة (عربي)"
COL_EN   = "اسم الروضة (إنجليزي)"
COL_GOV  = "المحافظة"
COL_CITY = "المدينة"
COL_AREA = "المنطقة"
COL_ADDR = "العنوان التفصيلي"
COL_PHONE= "رقم الهاتف"
ALL_COLS = [COL_AR, COL_EN, COL_GOV, COL_CITY, COL_AREA, COL_ADDR, COL_PHONE]

PHONE_RE = re.compile(r"^(\+962|00962|0)[0-9]{9}$")


# ── Helpers ───────────────────────────────────────────────────────────────────
def norm(s) -> str:
    """Normalize Arabic text for dedup comparison."""
    if not s:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKC", s)
    return " ".join(s.split()).lower()


def clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s in ("None", "nan") else s


def normalize_phone(ph: str) -> str:
    if not ph:
        return ""
    ph = ph.replace(" ", "").replace("-", "")
    if PHONE_RE.match(ph):
        return ph
    # Handle numeric float like 790225163.0
    if "." in ph:
        ph = ph.split(".")[0]
    if ph.isdigit() and len(ph) == 9:
        ph = "0" + ph
        if PHONE_RE.match(ph):
            return ph
    if ph.isdigit() and len(ph) == 10 and ph.startswith("0"):
        if PHONE_RE.match(ph):
            return ph
    return ph  # keep raw if unrecognized


# ── Excel reader ──────────────────────────────────────────────────────────────
def read_excel_files() -> list[dict]:
    all_files = [f for f in os.listdir(DATASET_PATH) if f.endswith(".xlsx")]
    ordered = PRIORITY_FILES + [f for f in all_files if f not in PRIORITY_FILES]

    seen_norm: set[str] = set()
    records: list[dict] = []

    for fname in ordered:
        fpath = os.path.join(DATASET_PATH, fname)
        if not os.path.exists(fpath):
            continue
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception as e:
            print(f"  [SKIP] {fname}: {e}")
            continue

        for sheet_name in wb.sheetnames:
            if "Summary" in sheet_name or sheet_name.strip() in ("ورقة1", "Sheet1"):
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Find header row (scan first 3 rows for Arabic name column)
            hdr_idx = None
            for i, row in enumerate(rows[:3]):
                if row and any(v and COL_AR in str(v) for v in row):
                    hdr_idx = i
                    break
            if hdr_idx is None:
                continue

            header = [str(v).strip() if v else "" for v in rows[hdr_idx]]
            col_map: dict[str, int] = {}
            for j, h in enumerate(header):
                for wanted in ALL_COLS:
                    if wanted in h and wanted not in col_map:
                        col_map[wanted] = j

            if COL_AR not in col_map:
                continue

            added = 0
            for row in rows[hdr_idx + 1:]:
                if not row:
                    continue
                idx_ar = col_map[COL_AR]
                raw_ar = row[idx_ar] if idx_ar < len(row) else None
                name_ar = clean(raw_ar)
                if not name_ar or name_ar.startswith("اسم"):
                    continue

                key = norm(name_ar)
                if key in seen_norm:
                    continue
                seen_norm.add(key)

                def gcol(col_name: str) -> str:
                    idx = col_map.get(col_name)
                    if idx is None or idx >= len(row):
                        return ""
                    return clean(row[idx])

                records.append({
                    "name_ar":      name_ar,
                    "name_en":      gcol(COL_EN) or None,
                    "governorate":  gcol(COL_GOV) or "غير محدد",
                    "city":         gcol(COL_CITY) or "غير محدد",
                    "area":         gcol(COL_AREA) or "غير محدد",
                    "address_line": gcol(COL_ADDR) or "غير محدد",
                    "contact_phone": normalize_phone(gcol(COL_PHONE)) or "غير محدد",
                })
                added += 1

        wb.close()
        print(f"  [{fname.encode('ascii','replace').decode()}]: {added} unique records collected")

    return records


# ── DB import ─────────────────────────────────────────────────────────────────
def run():
    engine = create_engine(DB_URL, connect_args={"check_same_thread": False})
    Session = sessionmaker(bind=engine)
    db = Session()

    try:
        # Get existing normalized names from DB
        existing_raw = db.execute(text("SELECT name_ar FROM kindergartens WHERE name_ar IS NOT NULL")).fetchall()
        existing_norm: set[str] = {norm(r[0]) for r in existing_raw}
        print(f"\nExisting in DB : {len(existing_norm)} kindergartens")

        print("\nReading Excel files...")
        excel_records = read_excel_files()
        print(f"\nUnique from Excel: {len(excel_records)} kindergartens")

        to_insert = [r for r in excel_records if norm(r["name_ar"]) not in existing_norm]
        print(f"New to insert  : {len(to_insert)} kindergartens")

        if not to_insert:
            print("\nNothing to insert — all records already in database.")
            return

        inserted = 0
        errors = 0
        for rec in to_insert:
            try:
                db.execute(text("""
                    INSERT INTO kindergartens
                        (name_ar, name_en, governorate, city, area, address_line,
                         contact_phone, contact_email, status,
                         created_at, updated_at)
                    VALUES
                        (:name_ar, :name_en, :governorate, :city, :area, :address_line,
                         :contact_phone, NULL, 'ACTIVE',
                         CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                """), rec)
                inserted += 1
            except Exception as e:
                print(f"  [ERROR] '{rec['name_ar']}': {e}")
                errors += 1

        db.commit()

        total = db.execute(text("SELECT COUNT(*) FROM kindergartens")).scalar()
        print(f"\n✓ Inserted {inserted} new kindergartens ({errors} errors)")
        print(f"✓ Total kindergartens in DB: {total}")

    except Exception as e:
        db.rollback()
        print(f"\n[FATAL] {e}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    run()
