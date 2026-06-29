"""
Seed kindergartens from Excel dataset files.
Reads all .xlsx files in the Dataset folder, deduplicates by Arabic name and coordinates,
then inserts only records not already present in the database.

Data validation & web search capability:
- Missing latitude/longitude: falls back to Jordan governorate centers
- For precise coordinates, consider using web search/geocoding for missing data
"""
import openpyxl
import os
import sys

sys.path.insert(0, os.path.dirname(__file__))

from database import SessionLocal
import models

DATASET_PATH = r"C:\Users\waelj\OneDrive - zuj.edu.jo\Desktop\Project-Kinjo-seed\DATS"

HEADERS = [
    "اسم الروضة (عربي)",
    "اسم الروضة (إنجليزي)",
    "المحافظة",
    "المدينة",
    "المنطقة",
    "العنوان التفصيلي",
    "رقم الهاتف",
    "خط العرض",
    "خط الطول",
]

JORDAN_GOVERNORATE_CENTERS = {
    "Amman": {"latitude": 31.95, "longitude": 35.95},
    "Irbid": {"latitude": 32.55, "longitude": 35.85},
    "Zarqa": {"latitude": 32.07, "longitude": 36.10},
    "Mafraq": {"latitude": 32.34, "longitude": 36.20},
    "Jerash": {"latitude": 32.28, "longitude": 35.90},
    "Ajloun": {"latitude": 32.33, "longitude": 35.75},
    "Balqa": {"latitude": 32.04, "longitude": 35.78},
    "Madaba": {"latitude": 31.72, "longitude": 35.79},
    "Karak": {"latitude": 31.18, "longitude": 35.70},
    "Tafileh": {"latitude": 30.83, "longitude": 35.60},
    "Ma'an": {"latitude": 30.20, "longitude": 35.73},
    "Aqaba": {"latitude": 29.53, "longitude": 35.00},
}

# Files to process (most complete/updated first for priority)
PRIORITY_FILES = [
    "روضات_وحضانات_محدث.xlsx",
    "merged_all_uploads.xlsx",
    "_محدث روضات_وحضانات_محدث.xlsx",
]


def _parse_coord(value) -> float | None:
    """Parse coordinate value, returning None if invalid/empty."""
    if value is None:
        return None
    try:
        coord = float(str(value).strip())
        if coord == 0.0:
            return None
        return coord
    except (ValueError, TypeError):
        return None


def _normalize_governorate(gov: str) -> str:
    """Normalize governorate name to match JORDAN_GOVERNORATE_CENTERS keys."""
    if not gov:
        return ""
    gov_clean = gov.strip()
    mapping = {
        "عمان": "Amman",
        "إربد": "Irbid",
        "الزرقاء": "Zarqa",
        "المفرق": "Mafraq",
        "جرش": "Jerash",
        "عجلون": "Ajloun",
        "البلقاء": "Balqa",
        "المداببة": "Madaba",
        "الكرك": "Karak",
        "الطفيلة": "Tafileh",
        "معان": "Ma'an",
        "العقبة": "Aqaba",
    }
    if gov_clean in mapping:
        return mapping[gov_clean]
    if gov_clean in JORDAN_GOVERNORATE_CENTERS:
        return gov_clean
    return gov_clean


def _get_fallback_coords(governorate: str) -> tuple[float | None, float | None]:
    """Get fallback coordinates for a governorate."""
    gov_key = _normalize_governorate(governorate)
    if gov_key in JORDAN_GOVERNORATE_CENTERS:
        center = JORDAN_GOVERNORATE_CENTERS[gov_key]
        return center["latitude"], center["longitude"]
    return None, None


def collect_from_excel() -> list[dict]:
    """Read all Excel files and return deduplicated list of kindergartens."""
    all_rows: list[dict] = []
    seen: set[tuple[str, float | None, float | None]] = set()

    all_files = [f for f in os.listdir(DATASET_PATH) if f.endswith(".xlsx")]
    ordered = PRIORITY_FILES + [f for f in all_files if f not in PRIORITY_FILES]

    for fname in ordered:
        fpath = os.path.join(DATASET_PATH, fname)
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception as e:
            print(f"  [SKIP] Cannot open {fname}: {e}")
            continue

        for sheet_name in wb.sheetnames:
            if "Summary" in sheet_name or "ورقة" in sheet_name:
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header_row_idx = None
            for i, row in enumerate(rows[:3]):
                if row and any(
                    v and "اسم الروضة" in str(v) for v in row
                ):
                    header_row_idx = i
                    break
            if header_row_idx is None:
                continue

            header = [
                str(v).strip() if v else "" for v in rows[header_row_idx]
            ]

            col_map: dict[str, int] = {}
            for j, h in enumerate(header):
                for wanted in HEADERS:
                    if wanted in h:
                        col_map[wanted] = j
                        break

            if "اسم الروضة (عربي)" not in col_map:
                print(f"  [SKIP] {fname}::{sheet_name} — cannot map Arabic name column")
                wb.close()
                continue

            def get_col(row_data, col_name: str) -> str:
                idx = col_map.get(col_name)
                if idx is None or idx >= len(row_data):
                    return ""
                v = row_data[idx]
                return str(v).strip() if v and str(v).strip() not in ("None", "") else ""

            added_from_sheet = 0
            for row_data in rows[header_row_idx + 1 :]:
                if not row_data:
                    continue

                idx_ar = col_map["اسم الروضة (عربي)"]
                raw_ar = row_data[idx_ar] if idx_ar < len(row_data) else None
                name_ar = str(raw_ar).strip() if raw_ar and str(raw_ar).strip() != "None" else ""

                if not name_ar or name_ar.startswith("اسم"):
                    continue

                governorate = get_col(row_data, "المحافظة")
                lat_raw = None
                lon_raw = None
                if "خط العرض" in col_map:
                    idx_lat = col_map["خط العرض"]
                    lat_raw = row_data[idx_lat] if idx_lat < len(row_data) else None
                if "خط الطول" in col_map:
                    idx_lon = col_map["خط الطول"]
                    lon_raw = row_data[idx_lon] if idx_lon < len(row_data) else None
                lat = _parse_coord(lat_raw)
                lon = _parse_coord(lon_raw)

                if lat is None or lon is None:
                    lat, lon = _get_fallback_coords(governorate)

                key = (name_ar, lat, lon)
                if key in seen:
                    continue
                seen.add(key)

                all_rows.append(
                    {
                        "name_ar": name_ar,
                        "name_en": get_col(row_data, "اسم الروضة (إنجليزي)") or None,
                        "governorate": get_col(row_data, "المحافظة") or "غير محدد",
                        "city": get_col(row_data, "المدينة") or "غير محدد",
                        "area": get_col(row_data, "المنطقة") or "غير محدد",
                        "address_line": get_col(row_data, "العنوان التفصيلي") or "غير محدد",
                        "contact_phone": get_col(row_data, "رقم الهاتف") or "غير محدد",
                        "latitude": lat,
                        "longitude": lon,
                    }
                )
                added_from_sheet += 1

            print(f"  {fname}::{sheet_name} → {added_from_sheet} new unique records")

        wb.close()

    return all_rows


def seed():
    db = SessionLocal()
    try:
        # Get all existing kindergartens (name_ar + latitude + longitude) for deduplication
        existing_keys: set[tuple[str, float | None, float | None]] = set()
        for kg in db.query(
            models.Kindergarten.name_ar,
            models.Kindergarten.latitude,
            models.Kindergarten.longitude,
        ).all():
            key = (kg[0].strip() if kg[0] else "", kg[1], kg[2])
            existing_keys.add(key)
        print(f"\nExisting in DB: {len(existing_keys)} kindergartens")

        excel_records = collect_from_excel()
        print(f"\nUnique from Excel: {len(excel_records)} kindergartens")

        to_insert = [r for r in excel_records if (r["name_ar"], r.get("latitude"), r.get("longitude")) not in existing_keys]
        print(f"New to insert:     {len(to_insert)} kindergartens")

        if not to_insert:
            print("\nNothing to insert — all records already in database.")
            return

        inserted = 0
        skipped = 0
        for rec in to_insert:
            try:
                kg = models.Kindergarten(
                    name_ar=rec["name_ar"],
                    name_en=rec["name_en"],
                    governorate=rec["governorate"],
                    district=rec["city"],
                    area=rec["area"],
                    address_line=rec["address_line"],
                    contact_phone=rec["contact_phone"],
                    contact_email=None,
                    status=models.KindergartenStatus.ACTIVE,
                    latitude=rec.get("latitude"),
                    longitude=rec.get("longitude"),
                )
                db.add(kg)
                inserted += 1
            except Exception as e:
                print(f"  [ERROR] Cannot insert '{rec['name_ar']}': {e}")
                skipped += 1

        db.commit()
        print(f"\n✓ Inserted {inserted} new kindergartens ({skipped} errors)")
        total = db.query(models.Kindergarten).count()
        print(f"✓ Total kindergartens in DB now: {total}")

    except Exception as e:
        db.rollback()
        print(f"\n[FATAL] {e}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    seed()
